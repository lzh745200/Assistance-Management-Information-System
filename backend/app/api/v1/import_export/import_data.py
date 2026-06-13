"""
数据导入API端点

Task 2.6: 实现导入API端点
- POST /api / v1 / import / template - 下载模板
- POST /api / v1 / import / villages - 导入帮扶村数据
- GET /api / v1 / import / history - 导入历史
- POST /api / v1 / import / validate - 验证导入数据（不执行导入）

Requirements: 1.1, 1.4, 1.8, 20.1
"""

from datetime import datetime
from typing import Any, Dict, List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.permission_utils import is_superuser
from app.core.security import get_current_user
from app.models.user import User
from app.services.data_validator_service import DataValidatorService
from app.services.entity_import_validator import EntityImportValidator
from app.services.excel_importer_service import ExcelImporterService, ImportMode
from app.services.excel_template_service import ExcelTemplateService

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# 支持的实体类型
VALID_ENTITY_TYPES = frozenset({"supported_village", "project", "fund", "school"})

router = APIRouter(prefix="/import", tags=["数据导入"])


# ==================== Response Schemas ====================


class ImportErrorResponse(BaseModel):
    """导入错误响应"""

    row_number: int = Field(..., description="错误行号")
    field_name: str = Field(..., description="错误字段")
    error_code: str = Field(..., description="错误代码")
    message: str = Field(..., description="错误信息")
    value: Optional[str] = Field(None, description="错误值")


class ImportResultResponse(BaseModel):
    """导入结果响应"""

    success: bool = Field(..., description="是否成功")
    total_rows: int = Field(..., description="总行数")
    success_rows: int = Field(..., description="成功行数")
    failed_rows: int = Field(..., description="失败行数")
    skipped_rows: int = Field(..., description="跳过行数")
    error_count: int = Field(..., description="错误数量")
    errors: List[ImportErrorResponse] = Field(default_factory=list, description="错误列表")
    created_ids: List[int] = Field(default_factory=list, description="创建的记录ID")
    import_history_id: Optional[int] = Field(None, description="导入历史ID")


class ValidationSummaryResponse(BaseModel):
    """验证摘要响应"""

    is_valid: bool = Field(..., description="是否有效")
    total_rows: int = Field(..., description="总行数")
    valid_rows: int = Field(..., description="有效行数")
    invalid_rows: int = Field(..., description="无效行数")
    error_count: int = Field(..., description="错误数量")
    warning_count: int = Field(..., description="警告数量")
    errors_by_type: Dict[str, int] = Field(default_factory=dict, description="按类型分组的错误数量")
    errors_by_field: Dict[str, int] = Field(default_factory=dict, description="按字段分组的错误数量")
    warnings: List[str] = Field(default_factory=list, description="警告列表")
    first_errors: List[Dict[str, Any]] = Field(default_factory=list, description="前10个错误详情")


class ImportHistoryResponse(BaseModel):
    """导入历史响应"""

    id: int
    user_id: int
    file_name: str
    file_size: int
    import_mode: str
    entity_type: str
    status: str
    total_rows: Optional[int]
    success_rows: Optional[int]
    failed_rows: Optional[int]
    started_at: Optional[datetime]
    completed_at: Optional[datetime]
    created_at: datetime

    model_config = ConfigDict(from_attributes=True)


class ImportHistoryListResponse(BaseModel):
    """导入历史列表响应"""

    items: List[ImportHistoryResponse]
    total: int
    page: int
    page_size: int


# ==================== API Endpoints ====================


@router.get("/template")
async def download_import_template(
    entity_type: str = Query(
        "supported_village",
        description="实体类型: supported_village, project, fund, school, policy",
    ),
    current_user: User = Depends(get_current_user),
):
    """
    下载导入模板

    Requirements: 1.1, 1.9

    - entity_type: 实体类型，支持 supported_village、project、fund、school、policy
    - 返回 Excel 模板文件，包含所有字段说明和示例数据
    - 统一样式：军绿主题 (ExcelTemplateService)
    """
    template_service = ExcelTemplateService()
    entity_labels = {
        "supported_village": "帮扶村",
        "project": "项目",
        "fund": "资金",
        "school": "学校",
        "policy": "政策",
    }
    method_map = {
        "supported_village": "generate_village_template",
        "project": "generate_project_template",
        "fund": "generate_fund_template",
        "school": "generate_school_template",
        "policy": "generate_policy_template",
    }

    method_name = method_map.get(entity_type)
    if not method_name:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的实体类型: {entity_type}，支持 {', '.join(method_map.keys())}",
        )

    content = getattr(template_service, method_name)()

    label = entity_labels.get(entity_type, entity_type)
    filename = f"{label}导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return Response(
        content=content,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/villages", response_model=ImportResultResponse)
async def import_villages(
    file: UploadFile = File(..., description="Excel文件 (.xlsx/.xls)"),
    mode: str = Query("incremental", description="导入模式: incremental(增量) / full(全量覆盖)"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    导入帮扶村数据（向后兼容，等价于 entity_type=supported_village）
    """
    return await _import_entities(
        file=file,
        mode=mode,
        entity_type="supported_village",
        current_user=current_user,
        db=db,
    )


@router.post("/entities", response_model=ImportResultResponse)
async def import_entities(
    file: UploadFile = File(..., description="Excel文件 (.xlsx/.xls)"),
    mode: str = Query("incremental", description="导入模式: incremental(增量) / full(全量覆盖)"),
    entity_type: str = Query(
        "supported_village",
        description="实体类型: supported_village, project, fund, school",
    ),
    dry_run: bool = Query(False, description="仅校验不实际导入，返回错误行号+修正建议"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    通用实体导入

    - file: Excel文件，支持 .xlsx 和 .xls 格式
    - mode: 导入模式（incremental / full）
    - entity_type: 实体类型（supported_village, project, fund, school）
    - dry_run: 仅校验不导入，返回错误行号 + 修正建议
    """
    return await _import_entities(
        file=file,
        mode=mode,
        entity_type=entity_type,
        dry_run=dry_run,
        current_user=current_user,
        db=db,
    )


async def _import_entities(
    file: UploadFile,
    mode: str,
    entity_type: str,
    current_user: User,
    db: Session,
    dry_run: bool = False,
) -> ImportResultResponse:
    """通用实体导入内部处理。dry_run=True 仅校验不写入数据库。"""
    try:
        import_mode = ImportMode(mode)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"无效的导入模式: {mode}，支持 incremental 或 full")

    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="不支持的文件格式，请上传 .xlsx 或 .xls 文件")

    if entity_type not in VALID_ENTITY_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的实体类型: {entity_type}，支持 {', '.join(VALID_ENTITY_TYPES)}",
        )

    file_bytes = await file.read()

    if dry_run:
        # 仅校验模式：使用独立 DB 会话执行导入后丢弃，确保无数据泄露
        from app.core.database import SessionLocal
        dry_db = SessionLocal()
        try:
            dry_importer = ExcelImporterService(dry_db)
            result = await dry_importer.import_data_async(
                file_bytes=file_bytes,
                filename=file.filename or "",
                content_type=file.content_type or "",
                user_id=current_user.id,
                mode=import_mode,
                entity_type=entity_type,
            )
        finally:
            dry_db.close()
    else:
        importer = ExcelImporterService(db)
        result = await importer.import_data_async(
            file_bytes=file_bytes,
            filename=file.filename or "",
            content_type=file.content_type or "",
            user_id=current_user.id,
            mode=import_mode,
            entity_type=entity_type,
        )

    return ImportResultResponse(**result.to_dict())


@router.post("/validate", response_model=ValidationSummaryResponse)
async def validate_import_data(
    file: UploadFile = File(..., description="Excel文件 (.xlsx/.xls)"),
    entity_type: str = Query(
        "supported_village",
        description="实体类型: supported_village, project, fund, school",
    ),
    validate_county: bool = Query(True, description="是否验证县 / 市字段"),
    validate_tiered_level: bool = Query(True, description="是否验证梯次等级字段"),
    current_user: User = Depends(get_current_user),
):
    """
    验证导入数据（不执行实际导入）

    Requirements: 20.1 - 数据导入功能验证
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    if entity_type not in VALID_ENTITY_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的实体类型: {entity_type}，支持 {', '.join(VALID_ENTITY_TYPES)}",
        )

    # 根据实体类型选择验证器
    if entity_type == "supported_village":
        validator = DataValidatorService()
        header_parser = validator.parse_excel_headers
        example_markers = ["某某部门", "示例部门"]
    else:
        validator = EntityImportValidator(entity_type)
        header_parser = validator.parse_excel_headers
        example_markers = ["某某村", "示例名称", "某某希望小学", "某某帮扶单位"]

    is_valid, error_msg = validator.validate_file_format(file.filename)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    file_content = await file.read()
    file_size = len(file_content)

    is_valid, error_msg = validator.validate_file_size(file_size)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    try:
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active

        rows = []
        headers = []
        header_mapping = {}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(cell).strip() if cell else "" for cell in row]
                header_mapping = header_parser(headers)
                continue

            if row_idx == 2:
                first_value = row[0] if len(row) > 0 else None
                if first_value and str(first_value).strip() in example_markers:
                    continue

            if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                continue

            row_data = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx in header_mapping:
                    field_name = header_mapping[col_idx]
                    row_data[field_name] = cell_value

            if row_data:
                rows.append(row_data)

        wb.close()

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel文件解析失败: {str(e)}")

    # 执行验证
    if entity_type == "supported_village":
        result = validator.validate_import_data(
            rows=rows,
            validate_county=validate_county,
            validate_tiered_level=validate_tiered_level,
        )
        summary = validator.get_validation_summary(result)
    else:
        result = validator.validate_batch(rows)
        summary = {
            "is_valid": result.is_valid,
            "total_rows": result.total_rows,
            "valid_rows": result.valid_rows,
            "invalid_rows": result.total_rows - result.valid_rows,
            "error_count": len(result.errors),
            "warning_count": len(result.warnings),
            "errors_by_type": {},
            "errors_by_field": {},
            "warnings": result.warnings,
            "first_errors": [e.to_dict() for e in result.errors[:10]],
        }
        # 补充错误分组统计
        error_by_type = {}
        error_by_field = {}
        for error in result.errors:
            code = error.error_code.value if hasattr(error.error_code, "value") else str(error.error_code)
            error_by_type[code] = error_by_type.get(code, 0) + 1
            if error.field_name:
                label = validator._get_field_label(error.field_name)
                error_by_field[label] = error_by_field.get(label, 0) + 1
        summary["errors_by_type"] = error_by_type
        summary["errors_by_field"] = error_by_field

    return ValidationSummaryResponse(**summary)


class PreviewRowResponse(BaseModel):
    """预览行数据"""

    row_number: int
    data: Dict[str, Any]
    has_error: bool = False
    errors: List[ImportErrorResponse] = Field(default_factory=list)
    is_duplicate_in_db: bool = False


class ImportPreviewResponse(BaseModel):
    """导入预览响应"""

    total_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_in_db_rows: int = Field(0, description="与数据库已有记录重复的行数")
    rows: List[PreviewRowResponse] = Field(default_factory=list, description="前50行预览数据")
    warnings: List[str] = Field(default_factory=list)


@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_import_data(
    file: UploadFile = File(..., description="Excel文件 (.xlsx/.xls)"),
    entity_type: str = Query(
        "supported_village",
        description="实体类型: supported_village, project, fund, school",
    ),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    导入预览（解析 Excel 后返回数据但不入库）

    功能：
    - 解析 Excel 文件并返回前 50 行解析结果
    - 对每行执行格式校验、必填字段校验
    - 检测与数据库已有记录的重复
    - 用户确认后再调用 /import/entities 执行实际导入
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    if entity_type not in VALID_ENTITY_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的实体类型: {entity_type}，支持 {', '.join(VALID_ENTITY_TYPES)}",
        )

    if entity_type == "supported_village":
        validator = DataValidatorService()
        duplicate_field = "village_name"
        from app.models.supported_village import SupportedVillage

        existing_names = set(
            name[0].strip().lower() for name in db.query(SupportedVillage.village_name).all() if name[0]
        )
    else:
        validator = EntityImportValidator(entity_type)
        duplicate_field = validator.config.get("duplicate_key", "name")
        if entity_type == "project":
            from app.models.project import Project as EntityModel
        elif entity_type == "fund":
            from app.models.fund import Fund as EntityModel
        elif entity_type == "school":
            from app.models.school import School as EntityModel
        existing_names = set(
            name[0].strip().lower() for name in db.query(getattr(EntityModel, duplicate_field)).all() if name[0]
        )

    is_valid, error_msg = validator.validate_file_format(file.filename)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    file_content = await file.read()
    is_valid, error_msg = validator.validate_file_size(len(file_content))
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    try:
        importer = ExcelImporterService(db)
        rows, _headers = importer.parse_excel(file_content, entity_type=entity_type)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel文件解析失败: {str(e)}")

    preview_rows = []
    valid_count = 0
    invalid_count = 0
    db_dup_count = 0

    for idx, row in enumerate(rows, 1):
        if entity_type == "supported_village":
            row_errors = validator.validate_row(row, idx)
        else:
            row_errors = validator.validate_row(row, idx)

        is_dup = False
        dup_value = (row.get(duplicate_field) or "").strip()
        if dup_value and dup_value.lower() in existing_names:
            is_dup = True
            db_dup_count += 1

        has_err = len(row_errors) > 0
        if has_err:
            invalid_count += 1
        else:
            valid_count += 1

        if idx <= 50:
            preview_rows.append(
                PreviewRowResponse(
                    row_number=idx,
                    data=row,
                    has_error=has_err,
                    errors=[
                        ImportErrorResponse(
                            row_number=e.row_number,
                            field_name=e.field_name,
                            error_code=(e.error_code.value if hasattr(e.error_code, "value") else str(e.error_code)),
                            message=e.message,
                            value=str(e.value) if e.value is not None else None,
                        )
                        for e in row_errors
                    ],
                    is_duplicate_in_db=is_dup,
                )
            )

    warnings = validator._generate_warnings(rows) if rows else []
    if db_dup_count > 0:
        warnings.insert(0, f"发现 {db_dup_count} 条记录与数据库已有记录重复（增量导入时将被跳过）")

    return ImportPreviewResponse(
        total_rows=len(rows),
        valid_rows=valid_count,
        invalid_rows=invalid_count,
        duplicate_in_db_rows=db_dup_count,
        rows=preview_rows,
        warnings=warnings,
    )


@router.get("/history", response_model=ImportHistoryListResponse)
async def get_import_history(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    获取导入历史

    Requirements: 1.8

    - 返回当前用户的导入历史记录
    - 支持分页查询
    """
    importer = ExcelImporterService(db)
    histories, total = importer.get_import_history(user_id=current_user.id, page=page, page_size=page_size)

    return ImportHistoryListResponse(
        items=[ImportHistoryResponse.model_validate(h) for h in histories],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/history/{history_id}", response_model=ImportHistoryResponse)
async def get_import_history_detail(
    history_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    获取导入历史详情

    Requirements: 1.8

    - 返回指定ID的导入历史记录详情
    """
    importer = ExcelImporterService(db)
    history = importer.get_import_history_by_id(history_id)

    if not history:
        raise HTTPException(status_code=404, detail="导入历史记录不存在")

    # 验证权限（只能查看自己的记录，管理员可查看所有）
    if history.user_id != current_user.id and not is_superuser(current_user):
        raise HTTPException(status_code=403, detail="无权查看此记录")

    return ImportHistoryResponse.model_validate(history)

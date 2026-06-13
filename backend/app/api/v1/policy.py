"""
政策法规API
支持政策分类、文档管理、导入导出功能
"""

import io
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from starlette.responses import FileResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session

from ...core.cache import cache_manager
from ...core.database import get_db
from ...core.security import get_current_user
from ...models.policy import Policy, PolicyCategory, PolicyFavorite

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/policies", tags=["政策法规"])

# ==================== 辅助函数 ====================


def _level_display_map() -> Dict[str, str]:
    return {
        "national": "国家级",
        "provincial": "省级",
        "municipal": "市级",
        "county": "县级",
        "military": "军队",
        "central_military": "中央军委",
        "theater": "战区",
        "army": "军",
        "division": "师",
    }


def _status_display_map() -> Dict[str, str]:
    return {"draft": "草稿", "active": "有效", "invalid": "失效"}


def _safe_isoformat(val) -> Optional[str]:
    """安全地将日期值转为 ISO 格式字符串"""
    if val is None:
        return None
    try:
        if val is None:
            return None
        if hasattr(val, "isoformat"):
            return val.isoformat()
        return str(val)
    except (TypeError, AttributeError) as e:
        logger.debug(f"日期格式化失败: {e}")
        return str(val) if val else None


def _policy_to_frontend(policy: Policy) -> Dict[str, Any]:
    """将数据库 Policy 对象转换为前端期望的格式"""
    level_val = str(policy.level) if policy.level else ""
    status_val = str(policy.status) if policy.status else "draft"
    category_val = str(policy.category) if policy.category else ""

    # 前端需要的 category_name / level_name / status_name
    level_names = _level_display_map()
    status_names = _status_display_map()

    return {
        "id": policy.id,
        "title": policy.title,
        "code": policy.code,
        "content": policy.content or "",
        "summary": policy.summary,
        "keywords": policy.keywords,
        # 分类 & 层级
        "category": category_val,
        "category_name": (
            "军队政策" if category_val == "military" else "地方政策" if category_val == "local" else category_val
        ),
        "organization_level": level_val,
        "level": level_val,
        "level_name": level_names.get(level_val, level_val),
        # 状态
        "status": status_val,
        "status_name": status_names.get(status_val, status_val),
        # 日期
        "publish_date": _safe_isoformat(policy.issue_date) or _safe_isoformat(policy.created_at),
        "issue_date": _safe_isoformat(policy.issue_date),
        "effective_date": _safe_isoformat(policy.effective_date),
        # 发布信息
        "issuing_authority": policy.issuing_authority,
        "department": policy.issuing_authority,  # 前端使用 department 显示
        "document_number": policy.code,
        # 附件
        "attachment_urls": [policy.file_path] if policy.file_path else [],
        # 统计
        "view_count": policy.view_count or 0,
        "download_count": policy.download_count or 0,
        # 时间
        "created_at": _safe_isoformat(policy.created_at),
        "updated_at": _safe_isoformat(policy.updated_at),
    }


# ==================== Pydantic模型 ====================


class PolicyCategoryBase(BaseModel):
    name: str
    code: Optional[str] = None
    parent_id: Optional[int] = None
    description: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True


class PolicyCategoryCreate(PolicyCategoryBase):
    pass


class PolicyCategoryResponse(PolicyCategoryBase):
    id: int
    created_at: Optional[datetime] = None

    model_config = ConfigDict(from_attributes=True)


class PolicyCreateRequest(BaseModel):
    """创建/更新政策请求（兼容前端字段）"""

    title: str
    content: Optional[str] = None
    category: Optional[str] = None
    organization_level: Optional[str] = None
    level: Optional[str] = None
    status: Optional[str] = "draft"
    publish_date: Optional[str] = None
    effective_date: Optional[str] = None
    expiry_date: Optional[str] = None
    issuing_authority: Optional[str] = None
    document_number: Optional[str] = None
    code: Optional[str] = None
    summary: Optional[str] = None
    keywords: Optional[str] = None


class PolicyUpdateRequest(BaseModel):
    """更新政策请求"""

    title: Optional[str] = None
    content: Optional[str] = None
    category: Optional[str] = None
    organization_level: Optional[str] = None
    level: Optional[str] = None
    status: Optional[str] = None
    publish_date: Optional[str] = None
    effective_date: Optional[str] = None
    expiry_date: Optional[str] = None
    issuing_authority: Optional[str] = None
    document_number: Optional[str] = None
    code: Optional[str] = None
    summary: Optional[str] = None
    keywords: Optional[str] = None


# ==================== 政策分类API ====================


@router.get("/categories")
async def get_categories(
    parent_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
    response: Any = None,
):
    """获取政策分类列表 —— 兼容前端两种调用方式"""
    try:
        query = db.query(PolicyCategory)
        if parent_id is not None:
            query = query.filter(PolicyCategory.parent_id == parent_id)
        if is_active is not None:
            query = query.filter(PolicyCategory.is_active == is_active)
        cats = query.order_by(PolicyCategory.sort_order, PolicyCategory.id).all()
        # 如果有分类数据，按后端结构返回
        if cats:
            return [
                {
                    "id": c.id,
                    "name": c.name,
                    "code": c.code,
                    "parent_id": c.parent_id,
                    "description": c.description,
                    "sort_order": c.sort_order,
                    "is_active": c.is_active,
                    "created_at": c.created_at.isoformat() if c.created_at else None,
                }
                for c in cats
            ]
    except (ValueError, TypeError, AttributeError) as e:
        logger.warning(f"查询分类表失败: {e}")

    # 返回前端期望的静态分类配置
    return {
        "military": {
            "label": "军队政策",
            "levels": [
                {"value": "central_military", "label": "中央军委"},
                {"value": "theater", "label": "战区"},
                {"value": "army", "label": "军"},
                {"value": "division", "label": "师"},
            ],
        },
        "local": {
            "label": "地方政策",
            "levels": [
                {"value": "national", "label": "国家级"},
                {"value": "provincial", "label": "省级"},
                {"value": "municipal", "label": "市级"},
                {"value": "county", "label": "县级"},
            ],
        },
    }


@router.get("/categories/tree")
async def get_category_tree(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """获取政策分类树形结构"""
    try:
        categories = (
            db.query(PolicyCategory)
            .filter(PolicyCategory.is_active == True)  # noqa: E712
            .order_by(PolicyCategory.sort_order, PolicyCategory.id)
            .all()
        )

        cat_map: Dict[int, Dict[str, Any]] = {
            cat.id: {
                "id": cat.id,
                "name": cat.name,
                "code": cat.code,
                "parent_id": cat.parent_id,
                "children": [],
            }
            for cat in categories
        }

        tree: List[Dict[str, Any]] = []
        for cat in categories:
            node = cat_map[cat.id]
            if cat.parent_id and cat.parent_id in cat_map:
                cat_map[cat.parent_id]["children"].append(node)
            else:
                tree.append(node)
        return tree
    except (ValueError, TypeError, KeyError) as e:
        logger.warning(f"获取分类树失败: {e}")
        return []


@router.post("/categories")
async def create_category(
    data: PolicyCategoryCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建政策分类"""
    if data.code:
        existing = db.query(PolicyCategory).filter(PolicyCategory.code == data.code).first()
        if existing:
            raise HTTPException(status_code=400, detail="分类编码已存在")

    category = PolicyCategory(**data.model_dump())
    db.add(category)
    db.commit()
    db.refresh(category)
    return {"id": category.id, "name": category.name, "code": category.code}


@router.put("/categories/{category_id}")
async def update_category(
    category_id: int,
    data: PolicyCategoryCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新政策分类"""
    category = db.query(PolicyCategory).filter(PolicyCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(category, key, value)
    db.commit()
    db.refresh(category)
    return {"id": category.id, "name": category.name, "code": category.code}


@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除政策分类"""
    category = db.query(PolicyCategory).filter(PolicyCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类不存在")

    children = db.query(PolicyCategory).filter(PolicyCategory.parent_id == category_id).count()
    if children > 0:
        raise HTTPException(status_code=400, detail="该分类下有子分类，无法删除")

    db.delete(category)
    db.commit()
    return {"message": "删除成功"}


# ==================== 统计 API ====================


@router.get("/statistics")
async def get_policy_statistics(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """获取政策统计数据（前端 Category.vue 使用）

    优化：使用数据库聚合查询替代 Python 循环，避免 N+1 查询问题
    """
    from sqlalchemy import func

    # 使用数据库聚合查询，一次性获取所有统计数据
    stats = (
        db.query(Policy.category, Policy.level, func.count(Policy.id).label("count"))
        .group_by(Policy.category, Policy.level)
        .all()
    )

    military_levels: Dict[str, int] = {}
    local_levels: Dict[str, int] = {}
    military_total = 0
    local_total = 0

    for stat in stats:
        cat = str(stat.category) if stat.category else ""
        lvl = str(stat.level) if stat.level else "unknown"
        count = int(stat.count)

        if cat == "military":
            military_total += count
            military_levels[lvl] = military_levels.get(lvl, 0) + count
        else:
            local_total += count
            local_levels[lvl] = local_levels.get(lvl, 0) + count

    return {
        "military": {"total": military_total, "levels": military_levels},
        "local": {"total": local_total, "levels": local_levels},
    }


# ==================== 导入导出API ====================


@router.get("/import/template")
async def download_import_template():
    """下载政策导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_policy_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''policy_import_template.xlsx"},
    )


@router.post("/import")
async def import_policies(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    从 Excel 导入政策
    前端调用 POST /policies/import
    返回格式: { imported: int, errors: list }
    """
    from io import BytesIO

    from openpyxl import load_workbook

    from ...core.upload_security import validate_excel_upload

    try:
        content = await validate_excel_upload(file)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件校验失败: {str(e)}")

    try:
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        level_map = {
            "国家级": "national",
            "省级": "provincial",
            "市级": "municipal",
            "县级": "county",
            "军队": "military",
            "中央军委": "central_military",
            "战区": "theater",
            "军": "army",
            "师": "division",
        }
        status_map = {
            "草稿": "draft",
            "有效": "active",
            "失效": "invalid",
            "已发布": "active",
            "已归档": "invalid",
            "已过期": "invalid",
        }

        imported = 0
        errors: List[Dict[str, Any]] = []
        error_rows: List[int] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            if not row[1]:
                errors.append({"row": row_idx, "title": "", "error": "政策标题不能为空"})
                error_rows.append(row_idx)
                continue

            try:
                # 解析日期 - 使用安全的索引访问
                issue_date = None
                effective_date = None

                if len(row) > 5:
                    try:
                        if row[5]:
                            if isinstance(row[5], datetime):
                                issue_date = row[5]
                            elif isinstance(row[5], str):
                                try:
                                    issue_date = datetime.strptime(row[5].strip(), "%Y-%m-%d")
                                except ValueError:
                                    pass
                    except (TypeError, AttributeError) as e:
                        logger.debug(f"解析发布日期失败 (行{row_idx}): {e}")

                if len(row) > 6:
                    try:
                        if row[6]:
                            if isinstance(row[6], datetime):
                                effective_date = row[6]
                            elif isinstance(row[6], str):
                                try:
                                    effective_date = datetime.strptime(row[6].strip(), "%Y-%m-%d")
                                except ValueError:
                                    pass
                    except (TypeError, AttributeError) as e:
                        logger.debug(f"解析生效日期失败 (行{row_idx}): {e}")

                # 安全地提取字段值
                try:
                    level_str = str(row[3]).strip() if len(row) > 3 and row[3] else None
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取政策级别失败 (行{row_idx}): {e}")
                    level_str = None

                level_val = level_map.get(level_str, "national") if level_str else "national"

                try:
                    status_str = str(row[7]).strip() if len(row) > 7 and row[7] else None
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取状态失败 (行{row_idx}): {e}")
                    status_str = None

                status_val = status_map.get(status_str, "active") if status_str else "active"

                category_val = "military" if level_val == "military" else "local"

                try:
                    content_val = str(row[9]).strip() if len(row) > 9 and row[9] else ""
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取内容失败 (行{row_idx}): {e}")
                    content_val = ""

                try:
                    title_val = str(row[1]).strip() if len(row) > 1 and row[1] else f"未命名政策{row_idx}"
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取标题失败 (行{row_idx}): {e}")
                    title_val = f"未命名政策{row_idx}"

                try:
                    code_val = str(row[2]).strip() if len(row) > 2 and row[2] else None
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取文号失败 (行{row_idx}): {e}")
                    code_val = None

                try:
                    authority_val = str(row[4]).strip() if len(row) > 4 and row[4] else None
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取发布机关失败 (行{row_idx}): {e}")
                    authority_val = None

                try:
                    keywords_val = str(row[8]).strip() if len(row) > 8 and row[8] else None
                except (TypeError, ValueError) as e:
                    logger.debug(f"提取关键词失败 (行{row_idx}): {e}")
                    keywords_val = None

                # 检查重复文号（唯一约束）
                if code_val:
                    existing = db.query(Policy).filter(Policy.code == code_val).first()
                    if existing:
                        errors.append(
                            {
                                "row": row_idx,
                                "title": title_val,
                                "error": f"文号“{code_val}”已存在，已跳过",
                            }
                        )
                        error_rows.append(row_idx)
                        continue

                policy = Policy(
                    title=title_val,
                    code=code_val,
                    level=level_val,
                    issuing_authority=authority_val,
                    issue_date=issue_date,
                    effective_date=effective_date,
                    status=status_val,
                    keywords=keywords_val,
                    category=category_val,
                    content=content_val,
                )
                # 用 savepoint 隔离每行，单行失败不影响其他行
                nested = db.begin_nested()
                try:
                    db.add(policy)
                    db.flush()
                    nested.commit()
                except Exception as row_err:
                    nested.rollback()
                    logger.warning(f"保存政策失败 (行{row_idx}): {row_err}")
                    raise row_err
                imported += 1
            except (ValueError, TypeError, KeyError) as e:
                errors.append(
                    {
                        "row": row_idx,
                        "title": (str(row[1]) if len(row) > 1 and row[1] else f"未命名政策{row_idx}"),
                        "error": str(e),
                    }
                )
                error_rows.append(row_idx)

        db.commit()

        # 前端期望的响应格式: { imported: N, errors: [...], errorRows: [...] }
        return {
            "imported": imported,
            "errors": errors,
            "errorRows": error_rows,
            "total": imported + len(errors),
        }
    except Exception as e:
        db.rollback()
        logger.error(f"导入政策失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# 保留旧路径做兼容


@router.post("/import/excel")
async def import_policies_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入政策（旧路径兼容）"""
    return await import_policies(file=file, current_user=current_user, db=db)


def _build_export_workbook(policies_list: List[Policy]):
    """构建导出 Excel 工作簿（PDF/WPS/Excel 共用）"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    level_display = _level_display_map()
    status_display = _status_display_map()

    wb = Workbook()
    ws = wb.active
    ws.title = "政策法规"

    headers = [
        "序号",
        "政策标题",
        "政策文号",
        "分类",
        "政策级别",
        "发布机关",
        "发布日期",
        "生效日期",
        "状态",
        "关键词",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, policy in enumerate(policies_list, 1):
        level_val = str(policy.level) if policy.level else ""
        status_val = str(policy.status) if policy.status else "draft"
        cat = str(policy.category) if policy.category else ""
        ws.append(
            [
                idx,
                policy.title,
                policy.code or "",
                ("军队政策" if cat == "military" else "地方政策" if cat == "local" else cat),
                level_display.get(level_val, level_val),
                policy.issuing_authority or "",
                policy.issue_date.strftime("%Y-%m-%d") if policy.issue_date else "",
                (policy.effective_date.strftime("%Y-%m-%d") if policy.effective_date else ""),
                status_display.get(status_val, status_val),
                policy.keywords or "",
            ]
        )

    widths = {
        "A": 8,
        "B": 50,
        "C": 20,
        "D": 12,
        "E": 12,
        "F": 20,
        "G": 15,
        "H": 15,
        "I": 12,
        "J": 30,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return wb


def _query_policies_for_export(db: Session, params: dict) -> List[Policy]:
    """根据筛选条件查询要导出的政策"""
    query = db.query(Policy)
    if params.get("category"):
        query = query.filter(Policy.category == params["category"])
    if params.get("organization_level"):
        query = query.filter(Policy.level == params["organization_level"])
    if params.get("status"):
        query = query.filter(Policy.status == params["status"])
    if params.get("search"):
        kw = params["search"]
        query = query.filter((Policy.title.contains(kw)) | (Policy.keywords.contains(kw)))
    return query.order_by(Policy.created_at.desc()).all()


@router.get("/export/excel")
async def export_policies_excel(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策到 Excel"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    wb = _build_export_workbook(policies_list)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=policies.xlsx"},
    )


@router.get("/export/pdf")
async def export_policies_pdf(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策为 PDF（以 Excel 格式替代，前端下载后缀为 .pdf）"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    wb = _build_export_workbook(policies_list)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/octet-stream",
        headers={"Content-Disposition": "attachment; filename=policies.pdf"},
    )


@router.get("/export/wps")
async def export_policies_wps(
    category: Optional[str] = None,
    organization_level: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出政策为 WPS 格式（以 Excel 格式替代）"""
    policies_list = _query_policies_for_export(
        db,
        {
            "category": category,
            "organization_level": organization_level,
            "status": status,
            "search": search,
        },
    )
    wb = _build_export_workbook(policies_list)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/octet-stream",
        headers={"Content-Disposition": "attachment; filename=policies.wps"},
    )


@router.get("/types")
async def get_policy_types(db: Session = Depends(get_db)):
    """获取政策类型选项 — 合并预定义类型与数据库中的实际分类"""
    # 预定义政策类型
    types = [
        {"value": "military", "label": "军队政策"},
        {"value": "local", "label": "地方政策"},
        {"value": "national", "label": "国家政策"},
        {"value": "provincial", "label": "省级政策"},
        {"value": "municipal", "label": "市级政策"},
        {"value": "county", "label": "县级政策"},
        {"value": "other", "label": "其他"},
    ]

    # 尝试从数据库中获取已有的政策分类（如果 PolicyCategory 模型已启用）
    try:
        categories = db.query(PolicyCategory).filter(PolicyCategory.is_active.is_(True)).all()
        if categories:
            db_types = [{"value": c.code or c.name, "label": c.name} for c in categories]
            # 合并去重：数据库分类优先
            existing_codes = {t["value"] for t in types}
            for dt in db_types:
                if dt["value"] not in existing_codes:
                    types.append(dt)
    except Exception:
        logger.debug("获取政策分类时出错，使用预定义类型")

    return {"data": types}


@router.get("/options/levels")
async def get_level_options():
    """获取政策级别选项"""
    return [
        {"value": "national", "label": "国家级"},
        {"value": "provincial", "label": "省级"},
        {"value": "municipal", "label": "市级"},
        {"value": "county", "label": "县级"},
        {"value": "military", "label": "军队"},
    ]


@router.get("/options/statuses")
async def get_status_options():
    """获取政策状态选项"""
    return [
        {"value": "active", "label": "有效"},
        {"value": "invalid", "label": "失效"},
        {"value": "draft", "label": "草稿"},
    ]


# ==================== 文件上传与预览 API ====================


@router.post("/{policy_id}/upload")
async def upload_policy_file(
    policy_id: int,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传政策附件文件（支持 pdf/doc/docx/pptx）"""
    import os

    from app.core.config import settings

    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    # 校验文件类型
    ext = os.path.splitext(file.filename or "")[1].lower().lstrip(".")
    allowed = {"pdf", "doc", "docx", "pptx"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型，仅支持: {', '.join(allowed)}")

    content = await file.read()
    # 限制 50MB
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小超过50MB限制")

    # 存储到 uploads/policies/ 目录
    upload_dir = os.path.join(settings.UPLOAD_DIR, "policies")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"policy_{policy_id}_{int(datetime.now().timestamp())}.{ext}"
    file_path = os.path.join(upload_dir, safe_name)

    with open(file_path, "wb") as f:
        f.write(content)

    # 更新数据库
    setattr(policy, "file_path", file_path)
    setattr(policy, "file_size", len(content))
    setattr(policy, "file_type", ext)
    db.commit()

    return {
        "message": "上传成功",
        "file_path": file_path,
        "file_size": len(content),
        "file_type": ext,
    }


@router.get("/{policy_id}/preview")
async def preview_policy_file(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """预览政策附件文件（返回文件流或HTML）"""
    import os

    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    if not policy.file_path or not os.path.exists(policy.file_path):
        # 没有附件，返回正文内容作HTML预览
        html = f"<html><body><h1>{policy.title}</h1><div>{policy.content or '无内容'}</div></body></html>"
        return StreamingResponse(
            io.BytesIO(html.encode("utf-8")),
            media_type="text/html",
        )

    ext = (policy.file_type or "").lower()

    if ext == "pdf":
        return FileResponse(
            path=policy.file_path,
            media_type="application/pdf",
            filename=os.path.basename(policy.file_path),
        )
    elif ext in ("doc", "docx"):
        # 使用 mammoth 将 docx 转换为 HTML
        try:
            import mammoth

            with open(policy.file_path, "rb") as f:
                result = mammoth.convert_to_html(f)
            html_style = "body{font-family:SimSun,serif;padding:20px;max-width:800px;margin:0 auto}"
            html = (
                f"<html><head><meta charset='utf-8'><style>{html_style}</style></head>"
                f"<body><h2>{policy.title}</h2>{result.value}</body></html>"
            )
            return StreamingResponse(
                io.BytesIO(html.encode("utf-8")),
                media_type="text/html",
            )
        except ImportError:
            # mammoth 未安装，返回下载
            return FileResponse(
                path=policy.file_path,
                media_type="application/octet-stream",
                filename=os.path.basename(policy.file_path),
            )
    else:
        # 其他类型直接下载
        return FileResponse(
            path=policy.file_path,
            media_type="application/octet-stream",
            filename=os.path.basename(policy.file_path),
        )


@router.get("/{policy_id}/download")
async def download_policy_file(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载政策附件"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")
    if not policy.file_path or not os.path.exists(policy.file_path):
        raise HTTPException(status_code=404, detail="附件文件不存在")

    current_count = policy.download_count or 0
    setattr(policy, "download_count", current_count + 1)
    db.commit()

    return FileResponse(
        path=policy.file_path,
        media_type="application/octet-stream",
        filename=os.path.basename(policy.file_path),
    )


# ==================== 批量操作 API ====================


@router.post("/batch-delete")
async def batch_delete_policies(data: dict, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """批量删除政策"""
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="未提供要删除的ID")

    deleted = db.query(Policy).filter(Policy.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    await cache_manager.delete("policies:list")
    return {"message": f"成功删除{deleted}条政策", "deleted": deleted}


# ==================== 政策CRUD API ====================


@router.get("")
async def get_policies(
    # 前端参数 (store 发送)
    skip: Optional[int] = Query(None, description="偏移量"),
    limit: Optional[int] = Query(None, description="数量限制"),
    category: Optional[str] = Query(None, description="分类: military/local"),
    organization_level: Optional[str] = Query(None, description="组织层级"),
    search: Optional[str] = Query(None, description="搜索关键字"),
    order_by: Optional[str] = Query(None, description="排序字段"),
    order_desc: Optional[bool] = Query(None, description="是否降序"),
    # 年度/文号筛选
    year: Optional[int] = Query(None, description="发布年度筛选"),
    document_code: Optional[str] = Query(None, description="文号筛选"),
    # 兼容旧参数
    page: Optional[int] = Query(None, ge=1),
    page_size: Optional[int] = Query(None, ge=1, le=200),
    keyword: Optional[str] = Query(None),
    level: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取政策列表 —— 兼容前端 skip/limit 和旧 page/page_size 参数"""
    _cache_key = "policies:list"
    _no_filter = not any([category, organization_level, search, order_by, year, document_code, keyword, level, status])
    _is_default_page = skip in (None, 0) and limit in (None, 20) and page is None
    if _no_filter and _is_default_page:
        cached = await cache_manager.get(_cache_key)
        if cached is not None:
            return cached

    query = db.query(Policy)

    # 关键字搜索
    kw = search or keyword
    if kw:
        query = query.filter((Policy.title.contains(kw)) | (Policy.code.contains(kw)) | (Policy.keywords.contains(kw)))
    # 分类过滤
    if category:
        query = query.filter(Policy.category == category)
    # 层级过滤
    lvl = organization_level or level
    if lvl:
        query = query.filter(Policy.level == lvl)
    # 状态过滤
    if status:
        query = query.filter(Policy.status == status)
    # 年度筛选
    if year:
        from sqlalchemy import extract

        query = query.filter(extract("year", Policy.issue_date) == year)
    # 文号筛选
    if document_code:
        query = query.filter(Policy.code.contains(document_code))

    total = query.count()

    # 排序

    sort_col: Any = Policy.created_at
    if order_by == "publish_date":
        sort_col = Policy.issue_date
    elif order_by == "title":
        sort_col = Policy.title
    if order_desc is False:
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    # 分页
    if skip is not None:
        offset = max(skip, 0)
        lim = limit or 10
    elif page is not None:
        offset = (page - 1) * (page_size or 20)
        lim = page_size or 20
    else:
        offset = 0
        lim = limit or 20

    items = query.offset(offset).limit(lim).all()

    result = {
        "items": [_policy_to_frontend(p) for p in items],
        "total": total,
        "page": (offset // lim) + 1 if lim else 1,
        "page_size": lim,
    }
    if _no_filter and _is_default_page:
        await cache_manager.set(_cache_key, result, ttl=300)
    return result


@router.get("/{policy_id}/related")
async def get_related_policies(
    policy_id: int,
    limit: int = Query(5, ge=1, le=20),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取相关政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    query = db.query(Policy).filter(Policy.id != policy_id)
    if policy.category:
        query = query.filter(Policy.category == policy.category)
    related = query.order_by(Policy.created_at.desc()).limit(limit).all()
    return [_policy_to_frontend(p) for p in related]


@router.get("/{policy_id}")
async def get_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取政策详情"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    current_count = policy.view_count or 0
    setattr(policy, "view_count", current_count + 1)
    db.commit()

    return _policy_to_frontend(policy)


@router.post("")
async def create_policy(
    data: PolicyCreateRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建政策"""
    try:
        issue_date = None
        effective_date = None
        if data.publish_date:
            try:
                issue_date = datetime.fromisoformat(data.publish_date)
            except (ValueError, TypeError):
                pass
        if data.effective_date:
            try:
                effective_date = datetime.fromisoformat(data.effective_date)
            except (ValueError, TypeError):
                pass

        policy = Policy(
            title=data.title,
            content=data.content or "",
            category=data.category or "local",
            level=data.organization_level or data.level or "national",
            status=data.status or "draft",
            issuing_authority=data.issuing_authority,
            code=data.document_number or data.code,
            issue_date=issue_date,
            effective_date=effective_date,
            summary=data.summary,
            keywords=data.keywords,
        )
        db.add(policy)
        db.commit()
        db.refresh(policy)
        await cache_manager.delete("policies:list")
        return _policy_to_frontend(policy)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建政策失败: {e}")
        raise HTTPException(status_code=500, detail=f"创建政策失败: {str(e)}")


@router.put("/{policy_id}")
async def update_policy(
    policy_id: int,
    data: PolicyUpdateRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    try:
        update_data = data.model_dump(exclude_unset=True)

        field_mapping = {
            "organization_level": "level",
            "document_number": "code",
            "publish_date": "issue_date",
        }

        for frontend_key, db_key in field_mapping.items():
            if frontend_key in update_data:
                update_data[db_key] = update_data.pop(frontend_key)

        # 处理日期字段
        for date_field in ["issue_date", "effective_date"]:
            if date_field in update_data and isinstance(update_data[date_field], str):
                try:
                    update_data[date_field] = datetime.fromisoformat(update_data[date_field])
                except (ValueError, TypeError):
                    update_data.pop(date_field)

        # 移除不属于模型的字段
        valid_columns = {c.name for c in Policy.__table__.columns}
        for key in list(update_data.keys()):
            if key not in valid_columns:
                update_data.pop(key)

        for key, value in update_data.items():
            setattr(policy, key, value)

        db.commit()
        db.refresh(policy)
        await cache_manager.delete("policies:list")
        return _policy_to_frontend(policy)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新政策失败 (id={policy_id}): {e}")
        raise HTTPException(status_code=500, detail=f"更新政策失败: {str(e)}")


@router.delete("/{policy_id}")
async def delete_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    db.delete(policy)
    db.commit()
    await cache_manager.delete("policies:list")
    return {"message": "删除成功"}


@router.post("/{policy_id}/publish")
async def publish_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """发布政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    setattr(policy, "status", "active")
    db.commit()
    await cache_manager.delete("policies:list")
    return {"message": "发布成功"}


@router.post("/{policy_id}/archive")
async def archive_policy(
    policy_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """归档政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    setattr(policy, "status", "invalid")
    db.commit()
    await cache_manager.delete("policies:list")
    return {"message": "归档成功"}


# ==================== 收藏API ====================


@router.post("/{policy_id}/favorite")
async def add_favorite(
    policy_id: int,
    user_id: int = Query(..., description="用户ID"),
    note: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """收藏政策"""
    policy = db.query(Policy).filter(Policy.id == policy_id).first()
    if not policy:
        raise HTTPException(status_code=404, detail="政策不存在")

    existing = (
        db.query(PolicyFavorite)
        .filter(PolicyFavorite.policy_id == policy_id, PolicyFavorite.user_id == user_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="已收藏该政策")

    favorite = PolicyFavorite(policy_id=policy_id, user_id=user_id)
    db.add(favorite)
    db.commit()
    return {"message": "收藏成功"}


@router.delete("/{policy_id}/favorite")
async def remove_favorite(
    policy_id: int,
    user_id: int = Query(..., description="用户ID"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """取消收藏"""
    favorite = (
        db.query(PolicyFavorite)
        .filter(PolicyFavorite.policy_id == policy_id, PolicyFavorite.user_id == user_id)
        .first()
    )
    if not favorite:
        raise HTTPException(status_code=404, detail="未收藏该政策")

    db.delete(favorite)
    db.commit()
    return {"message": "取消收藏成功"}


@router.get("/user/{user_id}/favorites")
async def get_user_favorites(user_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    """获取用户收藏的政策"""
    favorites = db.query(PolicyFavorite).filter(PolicyFavorite.user_id == user_id).all()
    policy_ids = [f.policy_id for f in favorites]
    if not policy_ids:
        return []
    items = db.query(Policy).filter(Policy.id.in_(policy_ids)).all()
    return [_policy_to_frontend(p) for p in items]


# ── FTS5 全文搜索 ──

@router.get("/search")
async def search_policies(
    q: str = Query("", description="搜索关键词"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """全文检索帮扶政策（FTS5 + 关键词高亮）"""
    from app.services.policy_fts_service import search_policies_fts
    results = search_policies_fts(db, q, limit=limit, offset=offset)
    return {"data": results, "total": len(results), "query": q}

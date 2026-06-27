from app.core.permission_utils import is_superuser

"""项目管理 API — 完整 CRUD + 任务管理 + 经费关联 + 统计导出 + 模板导入"""

import logging
import mimetypes
import os
import uuid as _uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy import func
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.errors import AppError
from app.core.exceptions import NotFoundException
from app.core.security import AuditLogService, check_rate_limit, get_client_ip, get_current_user
from app.models.project import Fund, Project, ProjectFile, ProjectStatus, ProjectTask
from app.services.audit_enhancement_service import AuditEnhancementService
from app.models.audit import AuditAction
from app.utils.db_error_handler import handle_db_errors_async
from app.services.work_log_service import write_work_log
from .deps import ADMIN_ROLES

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/projects", tags=["项目管理"])


# 项目关键字段（用于 Diff 留痕）
_PROJECT_KEY_FIELDS = [
    "name",
    "code",
    "type",
    "status",
    "budget",
    "progress",
    "responsible_unit",
    "responsible_person",
    "contact_phone",
    "start_date",
    "end_date",
    "village_id",
    "fund_source",
    "is_delayed",
    "delay_reason",
    "remarks",
]


def _project_to_diff_dict(p: Project) -> dict:
    """提取项目关键字段为 Diff 字典"""
    return {
        "name": p.name,
        "code": p.code,
        "type": p.type,
        "status": p.status,
        "budget": float(p.budget) if p.budget is not None else None,
        "progress": p.progress,
        "responsible_unit": p.responsible_unit,
        "responsible_person": p.responsible_person,
        "contact_phone": p.contact_phone,
        "start_date": p.start_date.isoformat() if p.start_date else None,
        "end_date": p.end_date.isoformat() if p.end_date else None,
        "village_id": p.village_id,
        "fund_source": p.fund_source,
        "is_delayed": bool(p.is_delayed) if p.is_delayed is not None else False,
        "delay_reason": p.delay_reason,
        "remarks": p.remarks,
    }


# ==================== 权限 & 工具 ====================


def _can_modify_project(project: Project, user) -> bool:
    """检查用户是否可以修改项目"""
    # 管理员可修改所有
    if user.role in ADMIN_ROLES:
        return True
    # 普通用户只能修改自己创建的项目
    return getattr(user, "id", None) == project.created_by


def _get_project_or_404(db: Session, project_id: int) -> Project:
    """获取项目或抛出 404"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise NotFoundException("项目不存在")
    return project


# ==================== Pydantic 模型 ====================


class ProjectCreate(BaseModel):
    """创建项目请求体"""

    name: str = Field(..., min_length=1, max_length=200, description="项目名称")
    code: Optional[str] = Field(None, max_length=50, description="项目编号(可选,后端自动生成)")
    type: Optional[str] = Field(None, max_length=50, description="项目类型")
    village_id: Optional[int] = Field(None, ge=1, description="帮扶村ID")
    description: Optional[str] = Field(None, description="项目描述")
    budget: Optional[float] = Field(None, ge=0, description="预算金额(万元)")
    start_date: Optional[str] = Field(None, description="开始日期 YYYY-MM-DD")
    end_date: Optional[str] = Field(None, description="结束日期 YYYY-MM-DD")
    responsible_unit: Optional[str] = Field(None, max_length=200, description="负责单位")
    responsible_person: Optional[str] = Field(None, max_length=50, description="负责人")
    contact_phone: Optional[str] = Field(None, max_length=20, description="联系电话")
    urgency_level: Optional[str] = Field(None, max_length=20, description="紧急程度")
    contract_number: Optional[str] = Field(None, max_length=50, description="合同编号")
    fund_manager: Optional[str] = Field(None, max_length=50, description="资金负责人")
    fund_usage_plan: Optional[str] = Field(None, max_length=50, description="资金使用计划")
    fund_source: Optional[str] = Field(None, max_length=50, description="资金来源")
    payer_account_name: Optional[str] = Field(None, max_length=200, description="拨款账户名称")
    payer_account_number: Optional[str] = Field(None, max_length=50, description="拨款卡号")
    payer_bank: Optional[str] = Field(None, max_length=200, description="拨款开户行")
    payer_handler: Optional[str] = Field(None, max_length=50, description="拨款经办人")
    payer_contact: Optional[str] = Field(None, max_length=50, description="拨款方联系方式")
    payee_account_name: Optional[str] = Field(None, max_length=200, description="收款单位账户名称")
    payee_bank: Optional[str] = Field(None, max_length=200, description="收款开户行")
    payee_handler: Optional[str] = Field(None, max_length=50, description="收款经办人")
    payee_contact: Optional[str] = Field(None, max_length=50, description="收款方联系方式")
    is_delayed: Optional[bool] = Field(None, description="是否延期")
    delay_reason: Optional[str] = Field(None, description="延期原因")
    expected_benefits: Optional[str] = Field(None, description="预期效益")
    achievements: Optional[str] = Field(None, description="项目成果")
    tags: Optional[str] = Field(None, description="项目标签(逗号分隔)")
    remarks: Optional[str] = Field(None, description="备注")

    @field_validator("start_date", "end_date", mode="before")
    @classmethod
    def validate_date_format(cls, v: Optional[str]) -> Optional[str]:
        if v:
            try:
                date.fromisoformat(v)
            except (ValueError, TypeError):
                raise ValueError("日期格式应为 YYYY-MM-DD")
        return v

    @model_validator(mode="after")
    def validate_dates(self):
        if self.start_date and self.end_date:
            if date.fromisoformat(self.end_date) < date.fromisoformat(self.start_date):
                raise ValueError("结束日期不能早于开始日期")
        return self


class ProjectUpdate(BaseModel):
    """更新项目请求体"""

    name: Optional[str] = Field(None, min_length=1, max_length=200)
    type: Optional[str] = Field(None, max_length=50)
    description: Optional[str] = None
    budget: Optional[float] = Field(None, ge=0)
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    progress: Optional[int] = Field(None, ge=0, le=100, description="进度 0-100")
    status: Optional[str] = None
    responsible_unit: Optional[str] = Field(None, max_length=200)
    responsible_person: Optional[str] = Field(None, max_length=50)
    contact_phone: Optional[str] = Field(None, max_length=20)
    urgency_level: Optional[str] = Field(None, max_length=20)
    contract_number: Optional[str] = Field(None, max_length=50)
    fund_manager: Optional[str] = Field(None, max_length=50)
    fund_usage_plan: Optional[str] = Field(None, max_length=50)
    fund_source: Optional[str] = Field(None, max_length=50)
    payer_account_name: Optional[str] = Field(None, max_length=200)
    payer_account_number: Optional[str] = Field(None, max_length=50)
    payer_bank: Optional[str] = Field(None, max_length=200)
    payer_handler: Optional[str] = Field(None, max_length=50)
    payer_contact: Optional[str] = Field(None, max_length=50)
    payee_account_name: Optional[str] = Field(None, max_length=200)
    payee_bank: Optional[str] = Field(None, max_length=200)
    payee_handler: Optional[str] = Field(None, max_length=50)
    payee_contact: Optional[str] = Field(None, max_length=50)
    is_delayed: Optional[bool] = None
    delay_reason: Optional[str] = None
    expected_benefits: Optional[str] = None
    achievements: Optional[str] = None
    tags: Optional[str] = None
    remarks: Optional[str] = None

    @field_validator("start_date", "end_date", mode="before")
    @classmethod
    def validate_date_format(cls, v: Optional[str]) -> Optional[str]:
        if v:
            try:
                date.fromisoformat(v)
            except (ValueError, TypeError):
                raise ValueError("日期格式应为 YYYY-MM-DD")
        return v

    @field_validator("status", mode="before")
    @classmethod
    def validate_status(cls, v: Optional[str]) -> Optional[str]:
        valid = {s.value for s in ProjectStatus}
        if v and v not in valid:
            raise ValueError(f"无效状态，可选值: {', '.join(valid)}")
        return v


class ProjectStatsResponse(BaseModel):
    """项目统计"""

    total: int = 0
    draft: int = 0
    pending: int = 0
    approved: int = 0
    in_progress: int = 0
    completed: int = 0
    cancelled: int = 0
    total_budget: float = 0
    total_invested: float = 0


# ---------- 任务模型 ----------


class TaskCreate(BaseModel):
    """创建任务"""

    name: str = Field(..., min_length=1, max_length=200, description="任务名称")
    description: Optional[str] = None
    status: Optional[str] = Field("pending", description="pending/in_progress/completed")
    priority: Optional[int] = Field(0, ge=0, le=10, description="优先级 0-10")
    assignee: Optional[str] = Field(None, max_length=50)
    due_date: Optional[str] = Field(None, description="截止日期 YYYY-MM-DD")

    @field_validator("due_date", mode="before")
    @classmethod
    def validate_date(cls, v):
        if v:
            try:
                date.fromisoformat(v)
            except (ValueError, TypeError):
                raise ValueError("日期格式应为 YYYY-MM-DD")
        return v


class TaskUpdate(BaseModel):
    """更新任务"""

    name: Optional[str] = Field(None, min_length=1, max_length=200)
    description: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[int] = Field(None, ge=0, le=10)
    assignee: Optional[str] = Field(None, max_length=50)
    due_date: Optional[str] = None


# ---------- 经费子模型 ----------


def _get_fund_health_fields(db: Session, project_id: int) -> dict:
    """获取项目经费健康度、预算执行率、支付偏差率（项目列表用）"""
    funds = db.query(Fund).filter(Fund.project_id == project_id).all()
    if not funds:
        return {
            "fund_health_score": None,
            "budget_execution_rate": None,
            "payment_deviation_rate": None,
        }

    # 取第一条 Fund 的 health_score（批量计算时已同步更新）
    health_score = funds[0].health_score if funds else None

    # 预算执行率
    total_approved = sum(float(f.approved_amount or f.amount or 0) for f in funds)
    total_used = sum(float(f.used_amount or 0) for f in funds)
    execution_rate = round(total_used / total_approved * 100, 2) if total_approved > 0 else 0

    # 支付偏差率（取各 Fund deviation_rate 的最大值）
    deviation_rates = [float(f.deviation_rate or 0) for f in funds]
    max_deviation = max(deviation_rates) if deviation_rates else 0

    return {
        "fund_health_score": health_score,
        "budget_execution_rate": execution_rate,
        "payment_deviation_rate": max_deviation,
    }


class ProjectFundCreate(BaseModel):
    """项目经费创建"""

    name: str = Field(..., min_length=1, max_length=200, description="经费名称")
    amount: float = Field(..., gt=0, description="金额")
    source: Optional[str] = Field(None, max_length=200, description="来源")
    purpose: Optional[str] = Field(None, description="用途")


def _batch_get_fund_health_fields(db: Session, project_ids: List[int]) -> dict:
    """批量获取项目经费健康度字段（解决 N+1 查询问题）

    一次查询获取所有项目的 Fund 记录，在内存中分组计算。

    Returns:
        {project_id: {fund_health_score, budget_execution_rate, payment_deviation_rate}}
    """
    if not project_ids:
        return {}

    funds = db.query(Fund).filter(Fund.project_id.in_(project_ids)).all()

    # 按 project_id 分组
    from collections import defaultdict

    grouped: dict = defaultdict(list)
    for f in funds:
        try:
            pid = int(f.project_id)
        except (TypeError, ValueError):
            continue
        grouped[pid].append(f)

    result = {}
    for pid in project_ids:
        pid_funds = grouped.get(pid, [])
        if not pid_funds:
            result[pid] = {
                "fund_health_score": None,
                "budget_execution_rate": None,
                "payment_deviation_rate": None,
            }
            continue

        health_score = pid_funds[0].health_score if pid_funds else None
        total_approved = sum(float(f.approved_amount or f.amount or 0) for f in pid_funds)
        total_used = sum(float(f.used_amount or 0) for f in pid_funds)
        execution_rate = round(total_used / total_approved * 100, 2) if total_approved > 0 else 0
        deviation_rates = [float(f.deviation_rate or 0) for f in pid_funds]
        max_deviation = max(deviation_rates) if deviation_rates else 0

        result[pid] = {
            "fund_health_score": health_score,
            "budget_execution_rate": execution_rate,
            "payment_deviation_rate": max_deviation,
        }

    return result


# ==================== 工具函数 ====================


def _project_to_dict(p: Project) -> dict:
    """Project ORM → 前端友好 dict（完整字段，用于详情接口）"""
    return {
        "id": p.id,
        "name": p.name,
        "code": p.code,
        "type": p.type,
        "status": p.status,
        "description": p.description,
        "objectives": p.objectives,
        "village_id": p.village_id,
        "organization_id": p.organization_id,
        "budget": float(p.budget) if p.budget else 0,
        "actual_cost": float(p.actual_cost) if p.actual_cost else 0,
        "progress": p.progress or 0,
        "leader": p.leader,
        "contact": p.contact,
        "invested_amount": float(p.invested_amount) if p.invested_amount else 0,
        "responsible_unit": p.responsible_unit,
        "responsible_person": p.responsible_person,
        "contact_phone": p.contact_phone,
        "start_date": p.start_date.isoformat() if p.start_date else None,
        "end_date": p.end_date.isoformat() if p.end_date else None,
        "actual_start_date": (p.actual_start_date.isoformat() if p.actual_start_date else None),
        "actual_end_date": p.actual_end_date.isoformat() if p.actual_end_date else None,
        "urgency_level": p.urgency_level or "normal",
        "contract_number": p.contract_number,
        "fund_manager": p.fund_manager,
        "fund_usage_plan": p.fund_usage_plan,
        "fund_source": p.fund_source,
        "payer_account_name": p.payer_account_name,
        "payer_account_number": p.payer_account_number,
        "payer_bank": p.payer_bank,
        "payer_handler": p.payer_handler,
        "payer_contact": p.payer_contact,
        "payee_account_name": p.payee_account_name,
        "payee_bank": p.payee_bank,
        "payee_handler": p.payee_handler,
        "payee_contact": p.payee_contact,
        "is_delayed": bool(p.is_delayed) if p.is_delayed is not None else False,
        "delay_reason": p.delay_reason,
        "expected_benefits": p.expected_benefits,
        "achievements": p.achievements,
        "tags": p.tags,
        "remarks": p.remarks,
        "created_by": p.created_by,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


def _project_to_list_dict(p: Project) -> dict:
    """Project ORM → 精简 dict（列表接口用，省略长文本字段以减小响应体积）"""
    return {
        "id": p.id,
        "name": p.name,
        "code": p.code,
        "type": p.type,
        "status": p.status,
        "village_id": p.village_id,
        "organization_id": p.organization_id,
        "budget": float(p.budget) if p.budget else 0,
        "actual_cost": float(p.actual_cost) if p.actual_cost else 0,
        "progress": p.progress or 0,
        "invested_amount": float(p.invested_amount) if p.invested_amount else 0,
        "responsible_unit": p.responsible_unit,
        "responsible_person": p.responsible_person,
        "contact_phone": p.contact_phone,
        "start_date": p.start_date.isoformat() if p.start_date else None,
        "end_date": p.end_date.isoformat() if p.end_date else None,
        "urgency_level": p.urgency_level or "normal",
        "fund_source": p.fund_source,
        "is_delayed": bool(p.is_delayed) if p.is_delayed is not None else False,
        "tags": p.tags,
        "created_by": p.created_by,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


def _task_to_dict(t: ProjectTask) -> dict:
    return {
        "id": t.id,
        "project_id": t.project_id,
        "name": t.name,
        "description": t.description,
        "status": t.status,
        "priority": t.priority,
        "assignee": t.assignee,
        "due_date": t.due_date.isoformat() if t.due_date else None,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }


# ==================== 统计 (放在 /{project_id} 之前避免路由冲突) ====================


@router.get("/stats", summary="项目统计概览")
async def get_project_stats(
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    返回各状态项目数量和预算汇总，供前端统计卡片一次调用获取。
    """
    rows = (
        db.query(
            Project.status,
            func.count(Project.id),
            func.coalesce(func.sum(Project.budget), 0),
        )
        .filter(Project.status != ProjectStatus.CANCELLED.value)
        .group_by(Project.status)
        .all()
    )

    stats = ProjectStatsResponse()
    for row_status, count, budget_sum in rows:
        stats.total += count
        stats.total_budget += float(budget_sum)
        if row_status and hasattr(stats, row_status):
            setattr(stats, row_status, count)

    invested = (
        db.query(func.coalesce(func.sum(Project.invested_amount), 0))
        .filter(Project.status != ProjectStatus.CANCELLED.value)
        .scalar()
    )
    stats.total_invested = float(invested)

    return stats


# ==================== 导出 ====================


@router.get("/export", summary="导出项目列表 Excel")
async def export_projects(
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    export_status: Optional[str] = Query(None, alias="status"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出项目列表为 Excel（或 CSV 兜底），上限 10000 条防内存溢出"""
    query = db.query(Project).filter(Project.status != ProjectStatus.CANCELLED.value)
    if keyword:
        query = query.filter(Project.name.contains(keyword))
    if project_type:
        query = query.filter(Project.type == project_type)
    if export_status:
        query = query.filter(Project.status == export_status)

    projects = query.order_by(Project.id.desc()).limit(10000).all()

    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "项目列表"
        headers = [
            "序号",
            "项目编号",
            "项目名称",
            "类型",
            "状态",
            "预算(万元)",
            "已投入(万元)",
            "进度(%)",
            "负责人",
            "负责单位",
            "开始日期",
            "结束日期",
        ]
        ws.append(headers)
        for idx, p in enumerate(projects, 1):
            ws.append(
                [
                    idx,
                    p.code or "",
                    p.name or "",
                    p.type or "",
                    p.status or "",
                    float(p.budget) if p.budget else 0,
                    float(p.invested_amount) if p.invested_amount else 0,
                    p.progress or 0,
                    p.responsible_person or "",
                    p.responsible_unit or "",
                    p.start_date.isoformat() if p.start_date else "",
                    p.end_date.isoformat() if p.end_date else "",
                ]
            )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"项目列表_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )
    except ImportError:
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["项目编号", "项目名称", "类型", "状态", "预算", "进度", "负责人"])
        for p in projects:
            writer.writerow(
                [
                    p.code,
                    p.name,
                    p.type,
                    p.status,
                    float(p.budget) if p.budget else 0,
                    p.progress or 0,
                    p.responsible_person or "",
                ]
            )
        buf = BytesIO(output.getvalue().encode("utf-8-sig"))
        buf.seek(0)
        filename = f"项目列表_{datetime.now().strftime('%Y%m%d')}.csv"
        return StreamingResponse(
            buf,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )


# ==================== 项目 CRUD ====================


@router.get("", summary="获取项目列表")
async def list_projects(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: Optional[str] = None,
    project_type: Optional[str] = None,
    status: Optional[str] = None,
    village_id: Optional[int] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = Query("desc"),
    include_cancelled: bool = Query(False, description="是否包含已取消项目"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取项目分页列表。默认排除已取消的项目。

    数据权限：管理员可看所有数据，普通用户只能看本组织数据。
    """
    query = db.query(Project)

    # N+1 优化：预加载关联数据（selectinload 分次查询，避免超大 JOIN）
    from sqlalchemy.orm import selectinload

    query = query.options(
        selectinload(Project.tasks),
        selectinload(Project.funds),
    )

    if not include_cancelled:
        query = query.filter(Project.status != ProjectStatus.CANCELLED.value)

    # 数据权限过滤：根据用户 data_scope 精确控制可见数据范围
    from app.core.data_permission import filter_by_data_scope

    query = filter_by_data_scope(query, Project, current_user, db=db)

    if keyword:
        query = query.filter((Project.name.contains(keyword)) | (Project.code.contains(keyword)))
    if project_type:
        query = query.filter(Project.type == project_type)
    if status:
        query = query.filter(Project.status == status)
    if village_id:
        query = query.filter(Project.village_id == village_id)

    # 排序
    sort_column = getattr(Project, sort_by, None) if sort_by else None
    if sort_column is not None:
        query = query.order_by(sort_column.desc() if sort_order == "desc" else sort_column.asc())
    else:
        query = query.order_by(Project.id.desc())

    total = query.count()
    projects = query.offset((page - 1) * page_size).limit(page_size).all()

    # 批量查询经费健康度字段（解决 N+1 查询问题）
    project_ids = [p.id for p in projects]
    fund_health_map = _batch_get_fund_health_fields(db, project_ids) if project_ids else {}

    items = []
    for p in projects:
        d = _project_to_list_dict(p)
        d.update(
            fund_health_map.get(
                p.id,
                {
                    "fund_health_score": None,
                    "budget_execution_rate": None,
                    "payment_deviation_rate": None,
                },
            )
        )
        items.append(d)

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


@router.get("/{project_id}", summary="获取项目详情")
async def get_project(
    project_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取项目详情，含关联经费/任务数量。"""
    project = _get_project_or_404(db, project_id)

    funds_total = db.query(func.count(Fund.id)).filter(Fund.project_id == project_id).scalar() or 0
    tasks_total = db.query(func.count(ProjectTask.id)).filter(ProjectTask.project_id == project_id).scalar() or 0

    data = _project_to_dict(project)
    data["funds_count"] = funds_total
    data["tasks_count"] = tasks_total
    return data


@router.post("", status_code=status.HTTP_201_CREATED, summary="创建项目")
@handle_db_errors_async
async def create_project(
    data: ProjectCreate,
    request: Request,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建项目。编号全局唯一，预算≥0，结束日期≥开始日期。"""
    # code自动生成: 如果前端未提供，则后端自动生成
    code = data.code
    if not code:
        import uuid

        code = f"PRJ-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
    if db.query(Project).filter(Project.code == code).first():
        raise AppError.bad_request("项目编码已存在")

    project = Project(
        name=data.name,
        code=code,
        type=data.type,
        village_id=data.village_id,
        description=data.description,
        budget=(Decimal(str(data.budget)) if data.budget is not None else Decimal("0")),
        start_date=date.fromisoformat(data.start_date) if data.start_date else None,
        end_date=date.fromisoformat(data.end_date) if data.end_date else None,
        responsible_unit=data.responsible_unit,
        responsible_person=data.responsible_person,
        contact_phone=data.contact_phone,
        urgency_level=data.urgency_level or "normal",
        contract_number=data.contract_number,
        fund_manager=data.fund_manager,
        fund_usage_plan=data.fund_usage_plan,
        fund_source=data.fund_source,
        payer_account_name=data.payer_account_name,
        payer_account_number=data.payer_account_number,
        payer_bank=data.payer_bank,
        payer_handler=data.payer_handler,
        payer_contact=data.payer_contact,
        payee_account_name=data.payee_account_name,
        payee_bank=data.payee_bank,
        payee_handler=data.payee_handler,
        payee_contact=data.payee_contact,
        is_delayed=data.is_delayed or False,
        delay_reason=data.delay_reason,
        expected_benefits=data.expected_benefits,
        achievements=data.achievements,
        tags=data.tags,
        remarks=data.remarks,
        created_by=getattr(current_user, "id", None),
        organization_id=getattr(current_user, "organization_id", None),
    )
    db.add(project)
    db.flush()

    audit = AuditLogService(db)
    await audit.log(
        db=db,
        action="create_project",
        user_id=getattr(current_user, "id", None),
        username=getattr(current_user, "username", None),
        resource="project",
        resource_id=str(project.id),
        details=f"创建项目: {project.name} ({project.code})",
        ip_address=get_client_ip(request),
    )

    # Diff 留痕
    try:
        audit_log = AuditEnhancementService.create_audit_log(
            db,
            AuditAction.CREATE,
            current_user,
            "project",
            str(project.id),
            detail=f"创建项目: {project.name}",
        )
        AuditEnhancementService.record_changes(
            db,
            audit_log,
            None,
            _project_to_diff_dict(project),
            "create",
            _PROJECT_KEY_FIELDS,
        )
    except Exception as diff_err:
        logger.warning(f"项目创建 Diff 留痕失败: {diff_err}")

    detail_parts = []
    if project.type:
        detail_parts.append(f"类型：{project.type}")
    if project.responsible_unit:
        detail_parts.append(f"负责单位：{project.responsible_unit}")
    write_work_log(
        db,
        "project",
        "create",
        project.id,
        project.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
        detail="，".join(detail_parts),
    )

    db.commit()
    db.refresh(project)

    try:
        from app.api.v1.data.dashboard import invalidate_dashboard_cache

        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")
    logger.info(f"项目创建成功: id={project.id}, name={project.name}")
    return {"id": project.id, "name": project.name, "code": project.code}


# ── 项目更新辅助函数 ──


def _convert_update_fields(update_data: dict) -> dict:
    """日期和预算字段转换"""
    for date_field in ("start_date", "end_date"):
        if date_field in update_data and update_data[date_field]:
            update_data[date_field] = date.fromisoformat(update_data[date_field])
    if "budget" in update_data and update_data["budget"] is not None:
        update_data["budget"] = Decimal(str(update_data["budget"]))
    return update_data


def _validate_update_dates(update_data: dict, project: Project) -> None:
    """日期交叉验证"""
    new_start = update_data.get("start_date", project.start_date)
    new_end = update_data.get("end_date", project.end_date)
    if new_start and new_end and new_end < new_start:
        raise AppError.bad_request("结束日期不能早于开始日期")


def _apply_project_changes(project: Project, update_data: dict) -> List[str]:
    """应用字段变更，返回变更的字段列表"""
    changed_fields = []
    for key, value in update_data.items():
        try:
            old = getattr(project, key)
        except AttributeError:
            logger.warning(f"跳过未知字段 '{key}' (模型无此列)")
            continue
        if old != value:
            changed_fields.append(key)
        setattr(project, key, value)
    return changed_fields


def _handle_status_change_event(db: Session, project_id: int, old_status: str, new_status: str, current_user) -> None:
    """触发经费阶段联动事件"""
    from app.services.fund_event_handler import on_project_status_change

    operator = getattr(current_user, "full_name", None) or getattr(current_user, "username", "")
    on_project_status_change(db, project_id, old_status, new_status, operator)


async def _log_project_update_audit(
    db: Session,
    request: Request,
    project: Project,
    project_id: int,
    changed_fields: List[str],
    old_project_data: dict,
    current_user,
) -> None:
    """记录审计日志、Diff 留痕和工作日志"""
    audit = AuditLogService(db)
    await audit.log(
        db=db,
        action="update_project",
        user_id=getattr(current_user, "id", None),
        username=getattr(current_user, "username", None),
        resource="project",
        resource_id=str(project_id),
        details=f"更新字段: {', '.join(changed_fields)}",
        ip_address=get_client_ip(request),
    )
    try:
        audit_log = AuditEnhancementService.create_audit_log(
            db,
            AuditAction.UPDATE,
            current_user,
            "project",
            str(project_id),
            detail=f"更新项目: {project.name}",
        )
        AuditEnhancementService.record_changes(
            db,
            audit_log,
            old_project_data,
            _project_to_diff_dict(project),
            "update",
            _PROJECT_KEY_FIELDS,
        )
    except Exception as diff_err:
        logger.warning(f"项目更新 Diff 留痕失败: {diff_err}")
    write_work_log(
        db,
        "project",
        "update",
        project.id,
        project.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
    )


def _invalidate_project_cache() -> None:
    """失效仪表盘缓存"""
    try:
        from app.api.v1.data.dashboard import invalidate_dashboard_cache

        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")


@router.put("/{project_id}", summary="更新项目")
@handle_db_errors_async
async def update_project(
    project_id: int,
    data: ProjectUpdate,
    request: Request,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新项目。进度 0-100，状态仅限有效枚举值。"""
    project = _get_project_or_404(db, project_id)

    if not _can_modify_project(project, current_user):
        raise AppError.forbidden("仅管理员或项目创建者可更新项目")

    old_project_data = _project_to_diff_dict(project)
    update_data = data.model_dump(exclude_unset=True)

    update_data = _convert_update_fields(update_data)
    _validate_update_dates(update_data, project)

    old_status = project.status
    changed_fields = _apply_project_changes(project, update_data)

    if "status" in changed_fields and old_status != project.status:
        try:
            _handle_status_change_event(db, project_id, old_status, project.status, current_user)
        except Exception as evt_err:
            db.rollback()
            logger.error(f"经费阶段联动失败，项目更新已回滚: {evt_err}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"项目更新失败: 经费阶段联动异常 ({evt_err})")

    if changed_fields:
        await _log_project_update_audit(
            db, request, project, project_id,
            changed_fields, old_project_data, current_user,
        )

    db.commit()
    db.refresh(project)
    _invalidate_project_cache()

    return {"message": "更新成功", "data": _project_to_dict(project)}


@router.delete("/{project_id}", summary="删除项目")
@handle_db_errors_async
async def delete_project(
    project_id: int,
    request: Request,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """软删除项目。仅管理员或项目创建者可执行。"""
    project = _get_project_or_404(db, project_id)

    is_admin = getattr(current_user, "role", None) in ("admin", "super_admin")
    is_creator = getattr(current_user, "id", None) == project.created_by
    if not (is_admin or is_creator or is_superuser(current_user)):
        raise AppError.forbidden("仅管理员或项目创建者可以删除项目")

    if project.status == ProjectStatus.CANCELLED.value:
        raise HTTPException(status_code=400, detail="项目已被删除")

    # 记录删除前的旧值
    old_project_data = _project_to_diff_dict(project)

    project.status = ProjectStatus.CANCELLED.value

    audit = AuditLogService(db)
    await audit.log(
        db=db,
        action="delete_project",
        user_id=getattr(current_user, "id", None),
        username=getattr(current_user, "username", None),
        resource="project",
        resource_id=str(project_id),
        details=f"删除(软)项目: {project.name}",
        ip_address=get_client_ip(request),
    )

    # Diff 留痕
    try:
        audit_log = AuditEnhancementService.create_audit_log(
            db,
            AuditAction.DELETE,
            current_user,
            "project",
            str(project_id),
            detail=f"删除项目: {project.name}",
        )
        AuditEnhancementService.record_changes(
            db,
            audit_log,
            old_project_data,
            None,
            "delete",
            _PROJECT_KEY_FIELDS,
        )
    except Exception as diff_err:
        logger.warning(f"项目删除 Diff 留痕失败: {diff_err}")

    write_work_log(
        db,
        "project",
        "delete",
        project.id,
        project.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
    )

    db.commit()

    try:
        from app.api.v1.data.dashboard import invalidate_dashboard_cache
        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")

    return {"message": "删除成功"}


@router.get("/{project_id}/history/changes", summary="获取项目变更历史")
async def get_project_change_history(
    project_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取项目关键字段的变更历史（Diff 留痕）"""
    _get_project_or_404(db, project_id)
    history = AuditEnhancementService.get_change_history(db, "project", str(project_id), limit=100)
    return {"items": history}


# ==================== 项目经费 ====================


@router.get("/{project_id}/funds", summary="获取项目经费列表")
async def get_project_funds(
    project_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取指定项目关联的经费记录。"""
    _get_project_or_404(db, project_id)

    query = db.query(Fund).filter(Fund.project_id == project_id)
    total = query.count()
    funds = query.order_by(Fund.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [f.to_dict() for f in funds],
    }


@router.post("/{project_id}/funds", status_code=201, summary="添加项目经费")
async def create_project_fund(
    project_id: int,
    data: ProjectFundCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """为项目添加经费记录（使用请求体传参）。"""
    project = _get_project_or_404(db, project_id)

    try:
        fund = Fund(
            project_id=project_id,
            project_name=project.name,
            name=data.name,
            amount=data.amount,
            source=data.source,
            purpose=data.purpose,
        )
        db.add(fund)
        db.commit()
        db.refresh(fund)
        return {"id": fund.id, "message": "经费添加成功"}
    except Exception as e:
        db.rollback()
        logger.error(f"添加经费失败: project_id={project_id}, error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail="添加经费失败")


# ==================== 项目任务 ====================


@router.get("/{project_id}/tasks", summary="获取项目任务列表")
async def get_project_tasks(
    project_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    task_status: Optional[str] = Query(None, alias="status"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取指定项目下的任务列表。"""
    _get_project_or_404(db, project_id)

    query = db.query(ProjectTask).filter(ProjectTask.project_id == project_id)
    if task_status:
        query = query.filter(ProjectTask.status == task_status)

    total = query.count()
    tasks = (
        query.order_by(ProjectTask.priority.desc(), ProjectTask.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [_task_to_dict(t) for t in tasks],
    }


@router.post("/{project_id}/tasks", status_code=201, summary="创建项目任务")
async def create_project_task(
    project_id: int,
    data: TaskCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """为项目创建任务。"""
    _get_project_or_404(db, project_id)

    try:
        task = ProjectTask(
            project_id=project_id,
            name=data.name,
            description=data.description,
            status=data.status or "pending",
            priority=data.priority or 0,
            assignee=data.assignee,
            due_date=date.fromisoformat(data.due_date) if data.due_date else None,
        )
        db.add(task)
        db.commit()
        db.refresh(task)
        return _task_to_dict(task)
    except Exception as e:
        db.rollback()
        logger.error(f"创建任务失败: project_id={project_id}, error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail="创建任务失败")


@router.put("/{project_id}/tasks/{task_id}", summary="更新项目任务")
async def update_project_task(
    project_id: int,
    task_id: int,
    data: TaskUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新指定任务。"""
    _get_project_or_404(db, project_id)
    task = db.query(ProjectTask).filter(ProjectTask.id == task_id, ProjectTask.project_id == project_id).first()
    if not task:
        raise NotFoundException("任务不存在")

    try:
        update_data = data.model_dump(exclude_unset=True)
        if "due_date" in update_data and update_data["due_date"]:
            update_data["due_date"] = date.fromisoformat(update_data["due_date"])

        for key, value in update_data.items():
            if hasattr(task, key):
                setattr(task, key, value)

        db.commit()
        return _task_to_dict(task)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新任务失败: task_id={task_id}, error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail="更新任务失败")


@router.delete("/{project_id}/tasks/{task_id}", summary="删除项目任务")
async def delete_project_task(
    project_id: int,
    task_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除指定任务（物理删除）。"""
    _get_project_or_404(db, project_id)
    task = db.query(ProjectTask).filter(ProjectTask.id == task_id, ProjectTask.project_id == project_id).first()
    if not task:
        raise NotFoundException("任务不存在")

    try:
        db.delete(task)
        db.commit()
        return {"message": "任务删除成功"}
    except Exception as e:
        db.rollback()
        logger.error(f"删除任务失败: task_id={task_id}, error={e}", exc_info=True)
        raise HTTPException(status_code=500, detail="删除任务失败")


# ==================== 帮扶项目导入模板 & 批量导入 ====================

# 模板字段定义 — 按类别分组
_TPL_SECTIONS = [
    (
        "一、基本信息",
        [
            {
                "col": "项目名称",
                "field": "name",
                "required": True,
                "width": 28,
                "desc": "帮扶项目全称（必填）",
            },
            {
                "col": "项目编号",
                "field": "code",
                "required": False,
                "width": 18,
                "desc": "如不填由系统自动生成",
            },
            {
                "col": "项目类型",
                "field": "type",
                "required": True,
                "width": 14,
                "desc": "基础设施/教育帮扶/产业发展/医疗卫生/农业发展/其他",
            },
            {
                "col": "帮扶村",
                "field": "village_name",
                "required": False,
                "width": 16,
                "desc": "关联帮扶村名称",
            },
            {
                "col": "项目描述",
                "field": "description",
                "required": False,
                "width": 40,
                "desc": "项目内容简介",
            },
            {
                "col": "项目目标",
                "field": "objectives",
                "required": False,
                "width": 40,
                "desc": "预期达成目标",
            },
        ],
    ),
    (
        "二、组织管理",
        [
            {
                "col": "负责单位",
                "field": "responsible_unit",
                "required": True,
                "width": 22,
                "desc": "实施帮扶的责任单位（必填）",
            },
            {
                "col": "负责人",
                "field": "responsible_person",
                "required": True,
                "width": 12,
                "desc": "项目负责人姓名（必填）",
            },
            {
                "col": "联系电话",
                "field": "contact_phone",
                "required": False,
                "width": 16,
                "desc": "负责人联系方式",
            },
            {
                "col": "合同编号",
                "field": "contract_number",
                "required": False,
                "width": 18,
                "desc": "帮扶协议或合同编号",
            },
        ],
    ),
    (
        "三、资金管理",
        [
            {
                "col": "预算金额(万元)",
                "field": "budget",
                "required": True,
                "width": 16,
                "desc": "项目总预算，单位万元（必填）",
            },
            {
                "col": "已投入金额(万元)",
                "field": "invested_amount",
                "required": False,
                "width": 18,
                "desc": "截至填报日期的已投入金额",
            },
            {
                "col": "资金来源",
                "field": "fund_source",
                "required": False,
                "width": 14,
                "desc": "上级拨付/本级预算投入/财政专项/社会捐赠/自筹/其他",
            },
            {
                "col": "资金负责人",
                "field": "fund_manager",
                "required": False,
                "width": 12,
                "desc": "资金管理责任人",
            },
            {
                "col": "资金使用计划",
                "field": "fund_usage_plan",
                "required": False,
                "width": 30,
                "desc": "资金分配与使用安排",
            },
        ],
    ),
    (
        "六、资金拨付去向",
        [
            {
                "col": "拨款账户名称",
                "field": "payer_account_name",
                "required": False,
                "width": 22,
                "desc": "拨款方账户全称",
            },
            {
                "col": "拨款卡号",
                "field": "payer_account_number",
                "required": False,
                "width": 22,
                "desc": "拨款方银行卡号",
            },
            {
                "col": "拨款开户行",
                "field": "payer_bank",
                "required": False,
                "width": 22,
                "desc": "拨款方开户银行",
            },
            {
                "col": "拨款经办人",
                "field": "payer_handler",
                "required": False,
                "width": 12,
                "desc": "拨款方经办人姓名",
            },
            {
                "col": "拨款联系方式",
                "field": "payer_contact",
                "required": False,
                "width": 16,
                "desc": "拨款方联系电话",
            },
            {
                "col": "收款账户名称",
                "field": "payee_account_name",
                "required": False,
                "width": 22,
                "desc": "收款单位账户全称",
            },
            {
                "col": "收款开户行",
                "field": "payee_bank",
                "required": False,
                "width": 22,
                "desc": "收款方开户银行",
            },
            {
                "col": "收款经办人",
                "field": "payee_handler",
                "required": False,
                "width": 12,
                "desc": "收款方经办人姓名",
            },
            {
                "col": "收款联系方式",
                "field": "payee_contact",
                "required": False,
                "width": 16,
                "desc": "收款方联系电话",
            },
        ],
    ),
    (
        "四、进度管理",
        [
            {
                "col": "计划开始日期",
                "field": "start_date",
                "required": False,
                "width": 16,
                "desc": "格式 YYYY-MM-DD",
            },
            {
                "col": "计划结束日期",
                "field": "end_date",
                "required": False,
                "width": 16,
                "desc": "格式 YYYY-MM-DD",
            },
            {
                "col": "当前进度(%)",
                "field": "progress",
                "required": False,
                "width": 14,
                "desc": "0-100 的整数",
            },
            {
                "col": "项目状态",
                "field": "status",
                "required": False,
                "width": 12,
                "desc": "草稿/待审批/已审批/进行中/已完成",
            },
            {
                "col": "紧急程度",
                "field": "urgency_level",
                "required": False,
                "width": 12,
                "desc": "紧急/重要/一般",
            },
            {
                "col": "是否延期",
                "field": "is_delayed",
                "required": False,
                "width": 10,
                "desc": "是/否",
            },
            {
                "col": "延期原因",
                "field": "delay_reason",
                "required": False,
                "width": 30,
                "desc": "如有延期请说明原因",
            },
        ],
    ),
    (
        "五、成效评估",
        [
            {
                "col": "预期效益",
                "field": "expected_benefits",
                "required": False,
                "width": 40,
                "desc": "项目建成后的预期社会/经济效益",
            },
            {
                "col": "项目成果",
                "field": "achievements",
                "required": False,
                "width": 40,
                "desc": "已取得的帮扶成果",
            },
            {
                "col": "项目标签",
                "field": "tags",
                "required": False,
                "width": 20,
                "desc": "多个标签用逗号分隔",
            },
            {
                "col": "备注",
                "field": "remarks",
                "required": False,
                "width": 30,
                "desc": "其他需要说明的事项",
            },
        ],
    ),
]

# 扁平化列定义（供生成/解析使用）
_ALL_TPL_COLS = [col for _, cols in _TPL_SECTIONS for col in cols]


@router.get("/import/template", summary="下载帮扶项目导入模板")
async def download_project_template(
    current_user=Depends(get_current_user),
):
    """下载帮扶项目导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_project_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''project_import_template.xlsx"},
    )


# ── Excel 导入列头映射（中文列名 → 模型字段名）──
_COL_TO_FIELD = {
    "项目名称": "name", "项目类型": "type", "项目状态": "status",
    "项目描述": "description", "项目目标": "objectives",
    "预算金额": "budget", "已投入金额": "invested_amount",
    "项目进度": "progress", "开始日期": "start_date", "结束日期": "end_date",
    "负责单位": "responsible_unit", "负责人": "responsible_person",
    "紧急程度": "urgency_level", "经费来源": "fund_source",
    "项目代码": "code", "所属村庄": "village",
}

# ── Excel 导入值映射（中文显示值 → 内部枚举值）──
_TYPE_MAP = {
    "基础设施": "infrastructure", "产业扶持": "industry", "教育帮扶": "education",
    "医疗帮扶": "medical", "民生改善": "livelihood", "其他": "other",
}
_STATUS_MAP = {
    "规划中": "draft", "待审批": "pending", "进行中": "in_progress",
    "已完成": "completed", "已暂停": "paused", "已取消": "cancelled",
}
_URGENCY_MAP = {"低": "low", "中": "medium", "高": "high", "紧急": "urgent"}
_FUND_SOURCE_MAP = {
    "中央财政": "central", "地方财政": "local", "军队专项": "military",
    "社会资金": "social", "自筹资金": "self_raised", "其他": "other",
}


# ── 项目导入辅助函数 ──


async def _check_import_rate_limit(user_id: int, client_ip: str) -> None:
    """导入速率限制：每用户每小时最多 5 次"""
    user_key = f"import_projects:{user_id}:{client_ip}"
    is_allowed = await check_rate_limit(key=user_key, limit=5, window=3600)
    if not is_allowed:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="导入操作过于频繁，请稍后再试（每小时最多 5 次）",
        )


async def _parse_import_excel(file: UploadFile):
    """验证文件类型并解析 Excel，返回 worksheet"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式文件")
    try:
        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        return wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {e}")


def _detect_import_headers(ws):
    """解析表头，返回 (header_row_number, headers_dict)"""
    for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        cells = [str(c).strip() if c else "" for c in row]
        matched = {}
        for ci, val in enumerate(cells):
            if val in _COL_TO_FIELD:
                matched[ci] = _COL_TO_FIELD[val]
        if len(matched) >= 3:
            return ri, matched
    return None, {}


def _extract_row_data(row: tuple, headers: dict) -> dict:
    """从 Excel 行中提取字段数据"""
    data = {}
    for ci, field in headers.items():
        val = row[ci] if ci < len(row) else None
        if val is not None and isinstance(val, str):
            val = val.strip()
        if val is not None and val != "":
            data[field] = val
    return data


def _build_import_project(db: Session, data: dict, current_user) -> Project:
    """根据提取的数据构建 Project 对象"""
    proj_type = _TYPE_MAP.get(data.get("type", ""), data.get("type", "other")) or "other"
    proj_status = _STATUS_MAP.get(data.get("status", ""), data.get("status", "draft")) or "draft"
    urgency = _URGENCY_MAP.get(data.get("urgency_level", ""), data.get("urgency_level", "normal")) or "normal"

    code = data.get("code") or f"PRJ-{datetime.now().strftime('%Y%m%d')}-{_uuid.uuid4().hex[:6].upper()}"
    if db.query(Project).filter(Project.code == code).first():
        code = f"PRJ-{datetime.now().strftime('%Y%m%d')}-{_uuid.uuid4().hex[:6].upper()}"

    budget_val = data.get("budget")
    invested_val = data.get("invested_amount")
    progress_val = data.get("progress")

    return Project(
        name=data["name"],
        code=code,
        type=proj_type,
        status=proj_status,
        description=data.get("description"),
        objectives=data.get("objectives"),
        budget=(Decimal(str(budget_val)) if budget_val is not None else Decimal("0")),
        invested_amount=(Decimal(str(invested_val)) if invested_val is not None else Decimal("0")),
        progress=int(float(progress_val)) if progress_val is not None else 0,
        start_date=(date.fromisoformat(str(data["start_date"])[:10]) if data.get("start_date") else None),
        end_date=(date.fromisoformat(str(data["end_date"])[:10]) if data.get("end_date") else None),
        responsible_unit=data.get("responsible_unit"),
        responsible_person=data.get("responsible_person"),
        contact_phone=(str(data.get("contact_phone", "")) if data.get("contact_phone") else None),
        urgency_level=urgency,
        contract_number=data.get("contract_number"),
        fund_manager=data.get("fund_manager"),
        fund_usage_plan=data.get("fund_usage_plan"),
        fund_source=(
            _FUND_SOURCE_MAP.get(data.get("fund_source", ""), data.get("fund_source"))
            if data.get("fund_source")
            else None
        ),
        payer_account_name=data.get("payer_account_name"),
        payer_account_number=(
            str(data.get("payer_account_number", "")) if data.get("payer_account_number") else None
        ),
        payer_bank=data.get("payer_bank"),
        payer_handler=data.get("payer_handler"),
        payer_contact=(str(data.get("payer_contact", "")) if data.get("payer_contact") else None),
        payee_account_name=data.get("payee_account_name"),
        payee_bank=data.get("payee_bank"),
        payee_handler=data.get("payee_handler"),
        payee_contact=(str(data.get("payee_contact", "")) if data.get("payee_contact") else None),
        is_delayed=str(data.get("is_delayed", "")).strip() == "是",
        delay_reason=data.get("delay_reason"),
        expected_benefits=data.get("expected_benefits"),
        achievements=data.get("achievements"),
        tags=data.get("tags"),
        remarks=data.get("remarks"),
        created_by=getattr(current_user, "id", None),
        organization_id=getattr(current_user, "organization_id", None),
    )


def _process_import_rows(db: Session, ws, header_row: int, headers: dict, current_user) -> tuple:
    """逐行解析数据并创建项目，返回 (created, failed, errors)"""
    created = 0
    failed = 0
    errors = []
    example_hints = {"XX村饮水安全工程", "某某帮扶单位", "张三"}

    for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        data = _extract_row_data(row, headers)
        if data.get("name") in example_hints or data.get("responsible_unit") in example_hints:
            continue
        if not data.get("name"):
            continue
        try:
            project = _build_import_project(db, data, current_user)
            db.add(project)
            db.flush()
            created += 1
        except Exception as e:
            failed += 1
            errors.append({"row": ri, "name": data.get("name", ""), "error": str(e)})
            logger.warning(f"导入第{ri}行失败: {e}")

    return created, failed, errors


@router.post("/import", summary="批量导入帮扶项目")
async def import_projects(
    file: UploadFile = File(..., description="Excel文件"),
    mode: str = Query("incremental", description="导入模式: incremental / overwrite"),
    request: Request = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量导入帮扶项目数据。"""
    client_ip = get_client_ip(request) if request else "unknown"
    await _check_import_rate_limit(current_user.id, client_ip)

    ws = await _parse_import_excel(file)
    header_row, headers = _detect_import_headers(ws)

    if not header_row:
        raise HTTPException(status_code=400, detail="未找到有效表头，请使用系统提供的模板")

    created, failed, errors = _process_import_rows(db, ws, header_row, headers, current_user)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"数据提交失败: {e}")

    return {
        "success": True,
        "imported": created,
        "failed": failed,
        "errors": errors[:20],
    }


# ==================== 项目文件管理 ====================

_VALID_CATEGORIES = {"research", "approval", "implementation", "acceptance", "photo"}
_CATEGORY_LABEL = {
    "research": "研究论证",
    "approval": "项目立项",
    "implementation": "组织实施",
    "acceptance": "项目验收",
    "photo": "项目照片",
}


def _file_to_dict(f: ProjectFile) -> dict:
    return {
        "id": f.id,
        "project_id": f.project_id,
        "category": f.category,
        "category_label": _CATEGORY_LABEL.get(f.category, f.category),
        "filename": f.filename,
        "file_size": f.file_size,
        "uploaded_by": f.uploaded_by,
        "created_at": f.created_at.isoformat() if f.created_at else None,
        "download_url": f"/api/v1/projects/{f.project_id}/files/{f.id}/download",
        "preview_url": f"/api/v1/projects/{f.project_id}/files/{f.id}/preview",
    }


@router.post("/{project_id}/files", summary="上传项目文件")
async def upload_project_files(
    project_id: int,
    category: str = Form(..., description="分类: research/implementation/acceptance/photo"),
    files: List[UploadFile] = File(..., description="一个或多个文件"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传项目附件，支持批量上传。category 可选 research/implementation/acceptance/photo"""
    if category not in _VALID_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"无效分类，可选: {', '.join(_VALID_CATEGORIES)}")

    project = _get_project_or_404(db, project_id)

    # 权限检查：管理员或项目创建者可上传
    if not _can_modify_project(project, current_user):
        raise HTTPException(status_code=403, detail="仅管理员或项目创建者可上传文件")

    from app.core.config import settings

    upload_dir = os.path.join(settings.UPLOAD_DIR, "projects", str(project_id), category)
    os.makedirs(upload_dir, exist_ok=True)

    uploaded = []
    for f in files:
        # 生成唯一文件名避免冲突
        ext = os.path.splitext(f.filename or "")[1]

        # 安全检查：文件类型白名单校验
        ext_lower = ext.lstrip(".").lower()
        if ext_lower and ext_lower not in settings.allowed_file_types_list:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件类型: .{ext_lower}，允许类型: {', '.join(settings.allowed_file_types_list)}",
            )

        unique_name = f"{_uuid.uuid4().hex}{ext}"
        save_path = os.path.join(upload_dir, unique_name)

        content = await f.read()
        if len(content) > settings.MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"文件 {f.filename} 超过 {settings.MAX_FILE_SIZE // 1048576}MB 限制",
            )

        with open(save_path, "wb") as fp:
            fp.write(content)

        pf = ProjectFile(
            project_id=project.id,
            category=category,
            filename=f.filename or unique_name,
            filepath=save_path,
            file_size=len(content),
            uploaded_by=getattr(current_user, "id", None),
        )
        db.add(pf)
        db.flush()
        uploaded.append(_file_to_dict(pf))

    db.commit()
    return {"success": True, "files": uploaded}


@router.get("/{project_id}/files", summary="获取项目文件列表")
async def list_project_files(
    project_id: int,
    category: Optional[str] = Query(None, description="按分类筛选"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取项目附件列表，可按 category 筛选"""
    _get_project_or_404(db, project_id)
    q = db.query(ProjectFile).filter(ProjectFile.project_id == project_id)
    if category:
        q = q.filter(ProjectFile.category == category)
    items = q.order_by(ProjectFile.created_at.desc()).all()

    # 按分类分组返回
    grouped: dict = {}
    for f in items:
        grouped.setdefault(f.category, []).append(_file_to_dict(f))
    return {"files": [_file_to_dict(f) for f in items], "grouped": grouped}


@router.delete("/{project_id}/files/{file_id}", summary="删除项目文件")
async def delete_project_file(
    project_id: int,
    file_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除项目附件（同时删除磁盘文件）。仅管理员或项目创建者可删除。"""
    project = _get_project_or_404(db, project_id)

    # 权限检查：管理员或项目创建者可删除
    if not _can_modify_project(project, current_user):
        raise HTTPException(status_code=403, detail="仅管理员或项目创建者可删除文件")

    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id, ProjectFile.project_id == project_id).first()
    if not pf:
        raise NotFoundException("文件不存在")

    if pf.filepath and os.path.exists(pf.filepath):
        try:
            os.remove(pf.filepath)
        except OSError:
            pass

    db.delete(pf)
    db.commit()
    return {"success": True}


@router.get("/{project_id}/files/{file_id}/download", summary="下载项目文件")
async def download_project_file(
    project_id: int,
    file_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载项目附件"""
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id, ProjectFile.project_id == project_id).first()
    if not pf:
        raise NotFoundException("文件不存在")

    if not pf.filepath or not os.path.exists(pf.filepath):
        raise NotFoundException("文件不存在")

    return FileResponse(
        path=pf.filepath,
        filename=pf.filename,
        media_type="application/octet-stream",
    )


@router.get("/{project_id}/files/{file_id}/preview", summary="在线预览项目文件")
async def preview_project_file(
    project_id: int,
    file_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """在线预览项目附件（图片、PDF 等直接内嵌显示）"""
    pf = db.query(ProjectFile).filter(ProjectFile.id == file_id, ProjectFile.project_id == project_id).first()
    if not pf or not pf.filepath or not os.path.exists(pf.filepath):
        raise NotFoundException("文件不存在")

    mime_type, _ = mimetypes.guess_type(pf.filename or pf.filepath)
    if not mime_type:
        mime_type = "application/octet-stream"

    safe_name = quote(pf.filename or "file", safe="")
    return FileResponse(
        path=pf.filepath,
        media_type=mime_type,
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{safe_name}"},
    )

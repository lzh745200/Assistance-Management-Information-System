"""
报表模板管理 API
上级可创建模板并分配给下级；下级下载模板、填写、上传
上传时后端解析 Excel（openpyxl），根据 fields 映射自动导入到对应模块数据表
"""

import io
import json
import logging
import uuid
from datetime import datetime, date as date_type
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session

from ...core.database import get_db
from ...core.security import get_current_user
from ...models.report_template import ReportTemplate

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/report-templates", tags=["报表模板管理"])

# ---- Schemas ----


class TemplateCreate(BaseModel):
    name: str
    type: str  # "import" | "export"
    module: str  # "village" | "school" | "fund" | "project" | "rural_work" | "comprehensive"
    fields: Optional[str] = None
    format_config: Optional[str] = None
    description: Optional[str] = None


class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    module: Optional[str] = None
    fields: Optional[str] = None
    format_config: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None


class TemplateOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    name: str
    type: str
    module: str
    fields: Optional[str] = None
    format_config: Optional[str] = None
    description: Optional[str] = None
    file_path: Optional[str] = None
    is_active: bool
    created_by: Optional[int] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _template_to_dict(t) -> dict:
    """安全地将 ORM ReportTemplate 转换为前端需要的字典，避免 datetime 序列化问题"""
    try:
        return {
            "id": t.id,
            "name": t.name or "",
            "type": t.type or "import",
            "module": t.module or "village",
            "fields": t.fields,
            "format_config": t.format_config,
            "description": t.description,
            "file_path": getattr(t, "file_path", None),
            "is_active": t.is_active if t.is_active is not None else True,
            "created_by": t.created_by,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "updated_at": t.updated_at.isoformat() if t.updated_at else None,
        }
    except Exception as e:
        logger.warning("模板序列化失败 (id=%s): %s", getattr(t, "id", "?"), e)
        return {
            "id": getattr(t, "id", 0),
            "name": getattr(t, "name", ""),
            "type": getattr(t, "type", "import"),
            "module": getattr(t, "module", "village"),
            "fields": getattr(t, "fields", None),
            "format_config": getattr(t, "format_config", None),
            "description": getattr(t, "description", None),
            "file_path": getattr(t, "file_path", None),
            "is_active": True,
            "created_by": getattr(t, "created_by", None),
            "created_at": None,
            "updated_at": None,
        }


# ---- 默认模板字段定义 ----


DEFAULT_FIELDS = {
    "village": [
        {
            "excel_col": "A",
            "excel_header": "村名",
            "db_field": "village_name",
            "required": True,
        },
        {
            "excel_col": "B",
            "excel_header": "所属县市",
            "db_field": "county",
            "required": True,
        },
        {
            "excel_col": "C",
            "excel_header": "帮扶单位",
            "db_field": "support_unit",
            "required": False,
        },
        {
            "excel_col": "D",
            "excel_header": "所属部门",
            "db_field": "department",
            "required": False,
        },
        {
            "excel_col": "E",
            "excel_header": "区域范围",
            "db_field": "region_scope",
            "required": False,
        },
        {
            "excel_col": "F",
            "excel_header": "是否三区三州",
            "db_field": "is_three_regions",
            "required": False,
        },
        {
            "excel_col": "G",
            "excel_header": "是否重点县",
            "db_field": "is_key_county",
            "required": False,
        },
    ],
    "school": [
        {
            "excel_col": "A",
            "excel_header": "学校名称",
            "db_field": "name",
            "required": True,
        },
        {
            "excel_col": "B",
            "excel_header": "所在区县",
            "db_field": "district",
            "required": True,
        },
        {
            "excel_col": "C",
            "excel_header": "学校类型",
            "db_field": "type",
            "required": False,
        },
        {
            "excel_col": "D",
            "excel_header": "帮扶单位",
            "db_field": "support_unit",
            "required": False,
        },
        {
            "excel_col": "E",
            "excel_header": "学生人数",
            "db_field": "student_count",
            "required": False,
        },
        {
            "excel_col": "F",
            "excel_header": "教师人数",
            "db_field": "teacher_count",
            "required": False,
        },
    ],
    "fund": [
        {
            "excel_col": "A",
            "excel_header": "名称",
            "db_field": "name",
            "required": True,
        },
        {
            "excel_col": "B",
            "excel_header": "金额(万元)",
            "db_field": "amount",
            "required": True,
        },
        {
            "excel_col": "C",
            "excel_header": "类型",
            "db_field": "type",
            "required": False,
        },
        {
            "excel_col": "D",
            "excel_header": "来源",
            "db_field": "source",
            "required": False,
        },
        {
            "excel_col": "E",
            "excel_header": "用途",
            "db_field": "purpose",
            "required": False,
        },
        {
            "excel_col": "F",
            "excel_header": "日期",
            "db_field": "date",
            "required": False,
        },
    ],
    "project": [
        # 基本信息
        {"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True},
        {"excel_col": "B", "excel_header": "项目编号", "db_field": "code", "required": False},
        {"excel_col": "C", "excel_header": "项目类型", "db_field": "type", "required": True},
        {"excel_col": "D", "excel_header": "帮扶村", "db_field": "village_name", "required": False},
        {"excel_col": "E", "excel_header": "项目描述", "db_field": "description", "required": False},
        {"excel_col": "F", "excel_header": "项目目标", "db_field": "objectives", "required": False},
        # 组织管理
        {"excel_col": "G", "excel_header": "负责单位", "db_field": "responsible_unit", "required": True},
        {"excel_col": "H", "excel_header": "负责人", "db_field": "responsible_person", "required": True},
        {"excel_col": "I", "excel_header": "联系电话", "db_field": "contact_phone", "required": False},
        {"excel_col": "J", "excel_header": "合同编号", "db_field": "contract_number", "required": False},
        # 资金管理
        {"excel_col": "K", "excel_header": "预算金额(万元)", "db_field": "budget", "required": True},
        {"excel_col": "L", "excel_header": "已投入金额(万元)", "db_field": "invested_amount", "required": False},
        {"excel_col": "M", "excel_header": "资金来源", "db_field": "fund_source", "required": False},
        {"excel_col": "N", "excel_header": "资金负责人", "db_field": "fund_manager", "required": False},
        {"excel_col": "O", "excel_header": "资金使用计划", "db_field": "fund_usage_plan", "required": False},
        # 进度管理
        {"excel_col": "P", "excel_header": "计划开始日期", "db_field": "start_date", "required": False},
        {"excel_col": "Q", "excel_header": "计划结束日期", "db_field": "end_date", "required": False},
        {"excel_col": "R", "excel_header": "当前进度(%)", "db_field": "progress", "required": False},
        {"excel_col": "S", "excel_header": "项目状态", "db_field": "status", "required": False},
        {"excel_col": "T", "excel_header": "紧急程度", "db_field": "urgency_level", "required": False},
        # 成果管理
        {"excel_col": "U", "excel_header": "是否延期", "db_field": "is_delayed", "required": False},
        {"excel_col": "V", "excel_header": "延期原因", "db_field": "delay_reason", "required": False},
        {"excel_col": "W", "excel_header": "预期效益", "db_field": "expected_benefits", "required": False},
        {"excel_col": "X", "excel_header": "项目成果", "db_field": "achievements", "required": False},
        {"excel_col": "Y", "excel_header": "备注", "db_field": "remarks", "required": False},
        # 资金拨付
        {"excel_col": "Z", "excel_header": "拨款账户名称", "db_field": "payer_account_name", "required": False},
        {"excel_col": "AA", "excel_header": "拨款卡号", "db_field": "payer_account_number", "required": False},
        {"excel_col": "AB", "excel_header": "拨款开户行", "db_field": "payer_bank", "required": False},
        {"excel_col": "AC", "excel_header": "拨款经办人", "db_field": "payer_handler", "required": False},
        {"excel_col": "AD", "excel_header": "拨款联系方式", "db_field": "payer_contact", "required": False},
        # 收款方
        {"excel_col": "AE", "excel_header": "收款账户名称", "db_field": "payee_account_name", "required": False},
        {"excel_col": "AF", "excel_header": "收款卡号", "db_field": "payee_account_number", "required": False},
        {"excel_col": "AG", "excel_header": "收款开户行", "db_field": "payee_bank", "required": False},
        {"excel_col": "AH", "excel_header": "收款经办人", "db_field": "payee_handler", "required": False},
        {"excel_col": "AI", "excel_header": "收款联系方式", "db_field": "payee_contact", "required": False},
    ],
    "rural_work": [
        # 基本信息
        {"excel_col": "A", "excel_header": "工作名称", "db_field": "name", "required": True},
        {"excel_col": "B", "excel_header": "工作类型", "db_field": "type", "required": False},
        {"excel_col": "C", "excel_header": "帮扶村", "db_field": "village_name", "required": False},
        {"excel_col": "D", "excel_header": "工作描述", "db_field": "description", "required": False},
        {"excel_col": "E", "excel_header": "工作目标", "db_field": "target", "required": False},
        # 组织管理
        {"excel_col": "F", "excel_header": "负责人", "db_field": "responsible_person", "required": False},
        {"excel_col": "G", "excel_header": "联系电话", "db_field": "contact_phone", "required": False},
        # 时间进度
        {"excel_col": "H", "excel_header": "开始日期", "db_field": "start_date", "required": False},
        {"excel_col": "I", "excel_header": "结束日期", "db_field": "end_date", "required": False},
        {"excel_col": "J", "excel_header": "当前进度(%)", "db_field": "progress", "required": False},
        {"excel_col": "K", "excel_header": "工作状态", "db_field": "status", "required": False},
    ],
}

# ---- Endpoints ----


@router.get("")
async def list_templates(
    type: Optional[str] = Query(default=None, description="模板类型: import/export"),
    module: Optional[str] = Query(default=None, description="关联模块"),
    is_active: Optional[bool] = Query(default=None),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取报表模板列表"""
    try:
        query = db.query(ReportTemplate)
        if type:
            query = query.filter(ReportTemplate.type == type)
        if module:
            query = query.filter(ReportTemplate.module == module)
        if is_active is not None:
            query = query.filter(ReportTemplate.is_active == is_active)
        templates = query.order_by(ReportTemplate.created_at.desc()).all()
        result = [_template_to_dict(t) for t in templates]
        return {"data": result}
    except Exception as e:
        logger.error("获取报表模板列表失败: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取模板列表失败: {str(e)}")


# 合法的模板类型和模块值
VALID_TEMPLATE_TYPES = {"import", "export"}
VALID_TEMPLATE_MODULES = {
    "village",
    "school",
    "fund",
    "project",
    "rural_work",
    "comprehensive",
}


@router.post("")
async def create_template(
    tpl_in: TemplateCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建报表模板"""
    # 校验 type 和 module 值
    if tpl_in.type not in VALID_TEMPLATE_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的模板类型: {tpl_in.type}，允许值: {', '.join(VALID_TEMPLATE_TYPES)}",
        )
    if tpl_in.module not in VALID_TEMPLATE_MODULES:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的关联模块: {tpl_in.module}，允许值: {', '.join(VALID_TEMPLATE_MODULES)}",
        )

    # 使用默认字段映射
    fields = tpl_in.fields
    if not fields and tpl_in.module in DEFAULT_FIELDS:
        fields = json.dumps(DEFAULT_FIELDS[tpl_in.module], ensure_ascii=False)

    try:
        template = ReportTemplate(
            name=tpl_in.name,
            type=tpl_in.type,
            module=tpl_in.module,
            fields=fields,
            format_config=tpl_in.format_config,
            description=tpl_in.description,
            created_by=current_user.id,
        )
        db.add(template)
        db.commit()
        db.refresh(template)
        return _template_to_dict(template)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error("创建报表模板失败: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建模板失败: {str(e)}")


@router.get("/{template_id}")
async def get_template(
    template_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取模板详情"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="模板不存在")
    return _template_to_dict(t)


@router.put("/{template_id}")
async def update_template(
    template_id: int,
    tpl_in: TemplateUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="模板不存在")
    try:
        for key, value in tpl_in.model_dump(exclude_unset=True).items():
            setattr(t, key, value)
        db.commit()
        db.refresh(t)
        return _template_to_dict(t)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error("更新报表模板失败 (id=%s): %s", template_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"更新模板失败: {str(e)}")


@router.delete("/{template_id}")
async def delete_template(
    template_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除模板"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="模板不存在")
    try:
        db.delete(t)
        db.commit()
        return {"code": 200, "data": {"id": template_id}, "detail": "模板已删除"}
    except Exception as e:
        db.rollback()
        logger.error("删除报表模板失败 (id=%s): %s", template_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除模板失败: {str(e)}")


@router.get("/{template_id}/download")
async def download_template(
    template_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载模板 Excel 文件（根据字段映射自动生成）"""
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="模板不存在")

    fields = json.loads(t.fields) if t.fields else []
    wb = Workbook()
    ws = wb.active
    ws.title = t.name[:31]  # Excel sheet name max 31 chars

    # 军绿色表头样式
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 标题行
    ws.merge_cells(f"A1:{chr(64 + max(len(fields), 1))}1")
    title_cell = ws.cell(row=1, column=1, value=t.name)
    title_cell.font = Font(bold=True, size=14, color="1B4332")
    title_cell.alignment = Alignment(horizontal="center")

    # 表头行
    for i, field in enumerate(fields, 1):
        cell = ws.cell(row=2, column=i, value=field.get("excel_header", ""))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[chr(64 + i)].width = 18
        # 必填字段标注
        if field.get("required"):
            note = ws.cell(row=3, column=i, value="(必填)")
            note.font = Font(color="FF0000", size=9)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    import urllib.parse

    filename = urllib.parse.quote(f"{t.name}.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ---- 辅助函数：解析 Excel ----


def _parse_template_excel(
    content: bytes, fields: List[Dict]
) -> Tuple[List[Dict[str, Any]], List[str], List[Dict[str, Any]]]:
    """解析模板 Excel 文件

    Args:
        content: Excel 文件字节内容
        fields: 字段映射配置

    Returns:
        (parsed_data, errors, rows) 三元组
        - parsed_data: 解析后的数据行列表
        - errors: 错误信息列表
        - rows: 原始数据行列表
    """
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # 跳过标题和表头

    parsed_data = []
    errors = []
    for row_idx, row in enumerate(rows, start=3):
        if all(v is None for v in row):
            continue  # 跳过空行
        record = {}
        row_errors = []
        for i, field in enumerate(fields):
            value = row[i] if i < len(row) else None
            db_field = field.get("db_field", "")
            if field.get("required") and (value is None or str(value).strip() == ""):
                row_errors.append(f"第{row_idx}行: {field.get('excel_header', '')}为必填项")
            record[db_field] = value
        if row_errors:
            errors.extend(row_errors)
        else:
            parsed_data.append(record)

    wb.close()
    return parsed_data, errors, rows


# ---- 帮扶村导入 ----


def _import_village_data(
    db: Session, parsed_data: List[Dict[str, Any]], user_id: Optional[int], mode: str
) -> Dict[str, Any]:
    """导入帮扶村数据

    Args:
        mode: "incremental"=跳过重复, "overwrite"=清空后全量导入
    """
    from app.models.supported_village import SupportedVillage

    created = 0
    skipped = 0
    deleted = 0
    errors = []

    # overwrite 模式：清空现有记录
    if mode == "overwrite":
        deleted = db.query(SupportedVillage).delete()
        db.commit()

    # 增量模式：仅查询本次导入涉及的村名（避免全表扫描）
    existing_names = set()
    if mode == "incremental":
        from sqlalchemy import func as sql_func
        imported_names = list({
            str(r.get("village_name", "")).lower()
            for r in parsed_data if r.get("village_name")
        })
        if imported_names:
            existing_names = {
                r[0] for r in
                db.query(sql_func.lower(SupportedVillage.village_name))
                .filter(sql_func.lower(SupportedVillage.village_name).in_(imported_names))
                .all() if r[0]
            }

    # 批量预加载 Village 名称映射（避免 N+1 逐行查询）
    from app.models.village import Village
    all_village_names = list({
        str(r.get("village_name", "")).strip()
        for r in parsed_data if r.get("village_name")
    })
    village_map = {}
    if all_village_names:
        villages = db.query(Village).filter(Village.name.in_(all_village_names)).all()
        village_map = {v.name: v for v in villages}

    for idx, data in enumerate(parsed_data):
        try:
            name = str(data.get("village_name", "")).strip()
            if not name:
                errors.append(f"第{idx + 3}行: 村名为空")
                continue

            if mode == "incremental" and name.lower() in existing_names:
                skipped += 1
                continue

            # 布尔字段转换
            def to_bool(val):
                if val is None:
                    return False
                s = str(val).strip()
                return s in ("是", "true", "True", "1", "yes", "Yes")

            def safe_str(val):
                if val is None:
                    return None
                s = str(val).strip()
                return s if s else None

            def safe_int(val, default=0):
                if val is None:
                    return default
                try:
                    return int(float(str(val)))
                except (ValueError, TypeError):
                    return default

            # O(1) 查找 village_name → village_id（已预加载）
            village_obj = village_map.get(name)

            record = SupportedVillage(
                village_name=name,
                province=safe_str(data.get("province")) or "贵州省",
                prefecture=safe_str(data.get("prefecture")) or "黔南布依族苗族自治州",
                county=safe_str(data.get("county")),
                region_scope=safe_str(data.get("region_scope")),
                is_three_regions=to_bool(data.get("is_three_regions")),
                is_key_county=to_bool(data.get("is_key_county")),
                support_unit=safe_str(data.get("support_unit")),
                department=safe_str(data.get("department")),
                longitude=safe_str(data.get("longitude")),
                latitude=safe_str(data.get("latitude")),
                total_households=safe_int(data.get("total_households")),
                registered_population=safe_int(data.get("registered_population")),
                created_by=user_id,
            )
            if village_obj:
                record.village_id = village_obj.id

            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"第{idx + 3}行: {str(e)}")

    db.commit()
    msg = (
        f"成功导入 {created} 条（全量覆盖，删除 {deleted} 条旧记录）"
        if mode == "overwrite"
        else f"成功导入 {created} 条，跳过 {skipped} 条重复记录"
    )

    # 自动记录工作日志
    try:
        from app.services.work_log_service import get_work_log_recorder

        recorder = get_work_log_recorder(db)
        recorder.record_batch_import("village", created, mode, user_id=user_id, username="系统")
    except Exception:
        logger.debug("记录工作日志失败")

    return {
        "success": True,
        "message": msg,
        "imported": created,
        "skipped": skipped,
        "deleted": deleted,
        "failed": len(errors),
        "errors": errors[:50],
    }


# ---- 帮扶学校导入 ----


def _import_school_data(
    db: Session, parsed_data: List[Dict[str, Any]], user_id: Optional[int], mode: str
) -> Dict[str, Any]:
    """导入帮扶学校数据"""
    from app.models.school import School, SchoolType, SupportStatus

    TYPE_MAP = {
        "小学": "primary",
        "初中": "middle",
        "中学": "middle",
        "高中": "high",
        "职业学校": "vocational",
        "其他": "other",
    }
    STATUS_MAP = {"帮扶中": "active", "未帮扶": "inactive", "已完成": "completed"}

    created = 0
    skipped = 0
    deleted = 0
    errors = []

    # overwrite 模式：清空现有记录
    if mode == "overwrite":
        deleted = db.query(School).delete()
        db.commit()

    existing_names = set()
    if mode == "incremental":
        from sqlalchemy import func as sql_func
        imported_names = list({
            str(r.get("name", "")).lower()
            for r in parsed_data if r.get("name")
        })
        if imported_names:
            existing_names = {
                r[0] for r in
                db.query(sql_func.lower(School.name))
                .filter(sql_func.lower(School.name).in_(imported_names))
                .all() if r[0]
            }

    for idx, data in enumerate(parsed_data):
        try:
            name = str(data.get("name", "")).strip()
            if not name:
                errors.append(f"第{idx + 3}行: 学校名称为空")
                continue

            if mode == "incremental" and name.lower() in existing_names:
                skipped += 1
                continue

            def safe_str(val):
                if val is None:
                    return None
                s = str(val).strip()
                return s if s else None

            def safe_int(val, default=0):
                if val is None:
                    return default
                try:
                    return int(float(str(val)))
                except (ValueError, TypeError):
                    return default

            school_type_str = TYPE_MAP.get(str(data.get("type", "")), data.get("type", "primary") or "primary")
            status_str = STATUS_MAP.get(str(data.get("support_status", "")), "inactive")

            record = School(
                name=name,
                code=safe_str(data.get("code")),
                type=SchoolType(school_type_str),
                province=safe_str(data.get("province")),
                city=safe_str(data.get("city")),
                district=safe_str(data.get("district")),
                address=safe_str(data.get("address")),
                student_count=safe_int(data.get("student_count")),
                teacher_count=safe_int(data.get("teacher_count")),
                support_status=SupportStatus(status_str),
                support_unit=safe_str(data.get("support_unit")),
                principal=safe_str(data.get("principal")),
                contact_phone=safe_str(data.get("contact_phone")),
            )
            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"第{idx + 3}行: {str(e)}")

    db.commit()
    msg = (
        f"成功导入 {created} 所学校（全量覆盖，删除 {deleted} 条旧记录）"
        if mode == "overwrite"
        else f"成功导入 {created} 所学校，跳过 {skipped} 条重复记录"
    )

    # 自动记录工作日志
    try:
        from app.services.work_log_service import get_work_log_recorder

        recorder = get_work_log_recorder(db)
        recorder.record_batch_import("school", created, mode, user_id=user_id, username="系统")
    except Exception:
        logger.debug("记录工作日志失败")

    return {
        "success": True,
        "message": msg,
        "imported": created,
        "skipped": skipped,
        "deleted": deleted,
        "failed": len(errors),
        "errors": errors[:50],
    }


# ---- 帮扶项目导入 ----


_PROJECT_TYPE_MAP = {
    "基础设施": "infrastructure",
    "教育帮扶": "education",
    "产业发展": "industry",
    "医疗卫生": "healthcare",
    "农业发展": "agriculture",
    "其他": "other",
}
_PROJECT_STATUS_MAP = {
    "草稿": "draft",
    "待审批": "pending",
    "已审批": "approved",
    "进行中": "in_progress",
    "已完成": "completed",
}
_PROJECT_URGENCY_MAP = {"紧急": "urgent", "重要": "important", "一般": "normal"}


def _import_project_data(
    db: Session, parsed_data: List[Dict[str, Any]], user_id: Optional[int], mode: str
) -> Dict[str, Any]:
    """导入帮扶项目数据"""
    from app.models.project import Project

    created = 0
    skipped = 0
    deleted = 0
    errors = []

    # overwrite 模式：清空现有记录
    if mode == "overwrite":
        deleted = db.query(Project).delete()
        db.commit()

    existing_names = set()
    if mode == "incremental":
        from sqlalchemy import func as sql_func
        imported_names = list({
            str(r.get("name", "")).lower()
            for r in parsed_data if r.get("name")
        })
        if imported_names:
            existing_names = {
                r[0] for r in
                db.query(sql_func.lower(Project.name))
                .filter(sql_func.lower(Project.name).in_(imported_names))
                .all() if r[0]
            }

    def safe_str(val):
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def safe_decimal(val, default="0"):
        if val is None:
            return Decimal(default)
        try:
            return Decimal(str(val))
        except Exception:
            return Decimal(default)

    def safe_int(val, default=0):
        if val is None:
            return default
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return default

    def parse_date(val):
        if val is None:
            return None
        s = str(val).strip()[:10]
        try:
            return date_type.fromisoformat(s)
        except ValueError:
            return None

    for idx, data in enumerate(parsed_data):
        try:
            name = safe_str(data.get("name"))
            if not name:
                errors.append(f"第{idx + 3}行: 项目名称为空")
                continue

            if mode == "incremental" and name.lower() in existing_names:
                skipped += 1
                continue

            # 类型/状态/紧急程度转换
            proj_type = _PROJECT_TYPE_MAP.get(str(data.get("type", "")), data.get("type") or "other")
            proj_status = _PROJECT_STATUS_MAP.get(str(data.get("status", "")), "draft")
            urgency = _PROJECT_URGENCY_MAP.get(str(data.get("urgency_level", "")), "normal")

            # 项目编号
            code = safe_str(data.get("code"))
            if not code:
                code = f"PRJ-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"

            # 查找帮扶村关联
            village_id = None
            if data.get("village_name"):
                from app.models.village import Village

                v = db.query(Village).filter(Village.name == data["village_name"]).first()
                if v:
                    village_id = v.id

            record = Project(
                name=name,
                code=code,
                type=proj_type,
                status=proj_status,
                description=safe_str(data.get("description")),
                objectives=safe_str(data.get("objectives")),
                budget=safe_decimal(data.get("budget")),
                invested_amount=safe_decimal(data.get("invested_amount")),
                progress=safe_int(data.get("progress")),
                start_date=parse_date(data.get("start_date")),
                end_date=parse_date(data.get("end_date")),
                responsible_unit=safe_str(data.get("responsible_unit")),
                responsible_person=safe_str(data.get("responsible_person")),
                contact_phone=safe_str(data.get("contact_phone")),
                contract_number=safe_str(data.get("contract_number")),
                fund_manager=safe_str(data.get("fund_manager")),
                fund_usage_plan=safe_str(data.get("fund_usage_plan")),
                fund_source=safe_str(data.get("fund_source")),
                urgency_level=urgency,
                is_delayed=str(data.get("is_delayed", "")).strip() in ("是", "true", "True"),
                delay_reason=safe_str(data.get("delay_reason")),
                expected_benefits=safe_str(data.get("expected_benefits")),
                achievements=safe_str(data.get("achievements")),
                remarks=safe_str(data.get("remarks")),
                payer_account_name=safe_str(data.get("payer_account_name")),
                payer_account_number=safe_str(data.get("payer_account_number")),
                payer_bank=safe_str(data.get("payer_bank")),
                payer_handler=safe_str(data.get("payer_handler")),
                payer_contact=safe_str(data.get("payer_contact")),
                payee_account_name=safe_str(data.get("payee_account_name")),
                payee_account_number=safe_str(data.get("payee_account_number")),
                payee_bank=safe_str(data.get("payee_bank")),
                payee_handler=safe_str(data.get("payee_handler")),
                payee_contact=safe_str(data.get("payee_contact")),
                village_id=village_id,
                created_by=user_id,
            )
            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"第{idx + 3}行: {str(e)}")

    db.commit()
    msg = (
        f"成功导入 {created} 个项目（全量覆盖，删除 {deleted} 条旧记录）"
        if mode == "overwrite"
        else f"成功导入 {created} 个项目，跳过 {skipped} 条重复记录"
    )

    # 自动记录工作日志
    try:
        from app.services.work_log_service import get_work_log_recorder

        recorder = get_work_log_recorder(db)
        recorder.record_batch_import("project", created, mode, user_id=user_id, username="系统")
    except Exception:
        logger.debug("记录工作日志失败")

    return {
        "success": True,
        "message": msg,
        "imported": created,
        "skipped": skipped,
        "deleted": deleted,
        "failed": len(errors),
        "errors": errors[:50],
    }


# ---- 乡村工作导入 ----


_RURAL_WORK_TYPE_MAP = {
    "基础设施": "infrastructure",
    "产业": "industry",
    "教育": "education",
    "医疗": "healthcare",
    "医疗卫生": "healthcare",
    "环境": "environment",
}
_RURAL_WORK_STATUS_MAP = {
    "待处理": "pending",
    "计划中": "planned",
    "进行中": "in_progress",
    "已完成": "completed",
    "已延期": "delayed",
}


def _import_rural_work_data(
    db: Session, parsed_data: List[Dict[str, Any]], user_id: Optional[int], mode: str
) -> Dict[str, Any]:
    """导入乡村工作数据"""
    from app.models.rural_work import RuralWork, WorkStatus, WorkType

    created = 0
    skipped = 0
    deleted = 0
    errors = []

    # overwrite 模式：清空现有记录
    if mode == "overwrite":
        deleted = db.query(RuralWork).delete()
        db.commit()

    existing_names = set()
    if mode == "incremental":
        from sqlalchemy import func as sql_func
        imported_names = list({
            str(r.get("name", "")).lower()
            for r in parsed_data if r.get("name")
        })
        if imported_names:
            existing_names = {
                r[0] for r in
                db.query(sql_func.lower(RuralWork.name))
                .filter(sql_func.lower(RuralWork.name).in_(imported_names))
                .all() if r[0]
            }

    def safe_str(val):
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def safe_int(val, default=0):
        if val is None:
            return default
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return default

    def parse_datetime(val):
        if val is None:
            return None
        s = str(val).strip()[:19]
        try:
            return datetime.fromisoformat(s.replace("/", "-"))
        except ValueError:
            try:
                return datetime.strptime(s, "%Y-%m-%d")
            except ValueError:
                return None

    for idx, data in enumerate(parsed_data):
        try:
            name = safe_str(data.get("name"))
            if not name:
                errors.append(f"第{idx + 3}行: 工作名称为空")
                continue

            if mode == "incremental" and name.lower() in existing_names:
                skipped += 1
                continue

            work_type_str = _RURAL_WORK_TYPE_MAP.get(str(data.get("type", "")), "infrastructure")
            status_str = _RURAL_WORK_STATUS_MAP.get(str(data.get("status", "")), "planned")

            # 查找帮扶村关联
            village_id = None
            if data.get("village_name"):
                from app.models.village import Village

                v = db.query(Village).filter(Village.name == data["village_name"]).first()
                if v:
                    village_id = v.id

            record = RuralWork(
                code=f"RW-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
                name=name,
                type=WorkType(work_type_str),
                status=WorkStatus(status_str),
                description=safe_str(data.get("description")),
                target=safe_str(data.get("target")),
                responsible_person=safe_str(data.get("responsible_person")),
                contact_phone=safe_str(data.get("contact_phone")),
                start_date=parse_datetime(data.get("start_date")),
                end_date=parse_datetime(data.get("end_date")),
                progress=safe_int(data.get("progress")),
                village_id=village_id,
                created_by=user_id,
                updated_by=user_id,
            )
            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"第{idx + 3}行: {str(e)}")

    db.commit()
    msg = (
        f"成功导入 {created} 条乡村工作（全量覆盖，删除 {deleted} 条旧记录）"
        if mode == "overwrite"
        else f"成功导入 {created} 条乡村工作记录，跳过 {skipped} 条重复记录"
    )

    # 自动记录工作日志
    try:
        from app.services.work_log_service import get_work_log_recorder

        recorder = get_work_log_recorder(db)
        recorder.record_batch_import("rural_work", created, mode, user_id=user_id, username="系统")
    except Exception:
        logger.debug("记录工作日志失败")

    return {
        "success": True,
        "message": msg,
        "imported": created,
        "skipped": skipped,
        "deleted": deleted,
        "failed": len(errors),
        "errors": errors[:50],
    }


# ---- 上传端点 ----


@router.post("/{template_id}/upload")
async def upload_filled_template(
    template_id: int,
    file: UploadFile = File(...),
    mode: str = Query("preview", description="preview=预览数据, confirm=确认导入数据库"),
    import_mode: str = Query("incremental", description="incremental=增量导入, overwrite=全量覆盖"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    上传已填写的模板 Excel 文件

    - mode=preview（默认）: 解析 Excel，返回预览数据（总行数、有效数据、错误列表）
    - mode=confirm: 解析 Excel，根据关联模块类型写入对应数据表
    - import_mode: incremental=增量导入（跳过重复）, overwrite=全量覆盖（先删除再导入）
    """
    t = db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="模板不存在")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    fields = json.loads(t.fields) if t.fields else []
    if not fields:
        raise HTTPException(status_code=400, detail="模板未配置字段映射")

    content = await file.read()

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        wb.close()
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件")

    # 解析 Excel
    parsed_data, errors, rows = _parse_template_excel(content, fields)

    # 预览模式
    if mode == "preview":
        return {
            "total_rows": len(rows),
            "success_count": len(parsed_data),
            "error_count": len(errors),
            "errors": errors[:50],
            "parsed_data": parsed_data[:10],
            "module": t.module,
        }

    # 确认导入模式
    user_id = getattr(current_user, "id", None)

    if t.module == "village":
        result = _import_village_data(db, parsed_data, user_id, import_mode)
    elif t.module == "school":
        result = _import_school_data(db, parsed_data, user_id, import_mode)
    elif t.module == "project":
        result = _import_project_data(db, parsed_data, user_id, import_mode)
    elif t.module == "rural_work":
        result = _import_rural_work_data(db, parsed_data, user_id, import_mode)
    elif t.module == "fund":
        raise HTTPException(
            status_code=400,
            detail="经费数据暂不支持Excel导入，请通过经费管理模块录入",
        )
    else:
        raise HTTPException(status_code=400, detail=f"模块「{t.module}」暂不支持导入")

    return {
        "total_rows": len(rows),
        "module": t.module,
        **result,
    }

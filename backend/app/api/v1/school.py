"""
学校管理API
支持完整CRUD、导入导出功能
"""

# NOTE: 数据权限过滤建议迁移到 app.core.data_scope_adapter.apply_scope_filter()
# 当前使用: OrgScopeFilter.filter_by_org_ids() (旧式 query + 组织树风格)

import io
import logging
import os
import tempfile
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.core.response import ok_list, success_response
from ...core.config import settings
from ...core.database import get_db
from ...core.security import get_current_user
from ...core.upload_security import validate_excel_upload
from ...models.school import (
    ProjectPhase,
    ScholarshipStatus,
    ScholarshipStudent,
    School,
    SchoolAttachment,
    SchoolProject,
    SchoolType,
    SupportStatus,
)
from ...core.data_permission import require_data_permission
from app.api.v1.deps import enforce_admin_include_deleted, build_viewable_because
from app.core.unified_data_scope import OrgScopeFilter, get_org_scope
from app.core.transaction import safe_commit
from ...services.work_log_service import write_work_log

logger = logging.getLogger(__name__)


def _fmt_row_error(row_idx: int, exc: Exception, hint: str, record_ident: str = "?") -> str:
    """统一格式化 Excel 导入行错误信息。"""
    return f"第{row_idx}行({record_ident}): {hint} — {exc}"


router = APIRouter(prefix="/schools", tags=["帮扶学校管理"])


# ==================== 权限检查依赖 ====================


def _get_school_and_check_permission(school_id: int, current_user, db: Session, action: str = "操作") -> School:
    """获取学校并检查数据权限（复用）"""
    school = db.query(School).filter(School.id == school_id, School.is_active == True).first()  # noqa: E712
    if not school:
        raise AppError.not_found("学校")
    require_data_permission(
        current_user,
        school.organization_id,
        school.created_by,
        db,
        error_message=f"无权在该学校{action}",
    )
    return school


def _validate_file_path(file_path: str) -> str:
    """验证并返回安全的文件路径"""
    if not os.path.isabs(file_path):
        file_path = os.path.abspath(file_path)
    allowed_base = os.path.abspath(settings.UPLOAD_DIR)
    _norm_file = os.path.normpath(file_path)
    _norm_base = os.path.normpath(allowed_base).rstrip(os.sep)
    if not (_norm_file == _norm_base or _norm_file.startswith(_norm_base + os.sep)):
        raise AppError.forbidden("非法文件路径")
    if not os.path.exists(file_path):
        raise AppError.not_found("文件")
    return file_path


# ==================== Pydantic模型 ====================


class SchoolBase(BaseModel):
    name: str
    code: Optional[str] = None
    type: Optional[str] = "primary"
    province: Optional[str] = None
    city: Optional[str] = None
    district: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    student_count: int = 0
    teacher_count: int = 0
    class_count: int = 0
    established_year: Optional[int] = None
    support_status: Optional[str] = "inactive"
    support_unit: Optional[str] = None
    support_start_date: Optional[datetime] = None
    support_end_date: Optional[datetime] = None
    principal: Optional[str] = None
    contact_phone: Optional[str] = None
    email: Optional[str] = None
    description: Optional[str] = None
    remarks: Optional[str] = None


class SchoolCreate(SchoolBase):
    pass


class SchoolUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    type: Optional[str] = None
    province: Optional[str] = None
    city: Optional[str] = None
    district: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    student_count: Optional[int] = None
    teacher_count: Optional[int] = None
    class_count: Optional[int] = None
    established_year: Optional[int] = None
    support_status: Optional[str] = None
    support_unit: Optional[str] = None
    support_start_date: Optional[datetime] = None
    support_end_date: Optional[datetime] = None
    principal: Optional[str] = None
    contact_phone: Optional[str] = None
    email: Optional[str] = None
    description: Optional[str] = None
    remarks: Optional[str] = None


class PaginatedSchoolResponse(BaseModel):
    items: List[dict]
    total: int
    page: int
    page_size: int


# ==================== 类型映射 ====================

TYPE_MAP = {
    "小学": "primary",
    "初中": "middle",
    "中学": "middle",
    "高中": "high",
    "职业学校": "vocational",
    "其他": "other",
    "primary": "primary",
    "middle": "middle",
    "high": "high",
    "vocational": "vocational",
    "other": "other",
}

TYPE_DISPLAY = {
    "primary": "小学",
    "middle": "初中",
    "high": "高中",
    "vocational": "职业学校",
    "other": "其他",
}


def _parse_scholarship_status(row):
    """安全解析奖学金状态。"""
    from app.models.scholarship import ScholarshipStatus
    status_map = {"pending": "pending", "approved": "approved", "rejected": "rejected"}
    if len(row) > 8 and row[8]:
        return ScholarshipStatus(status_map.get(str(row[8]), "pending"))
    return ScholarshipStatus.PENDING


# ==================== 导入导出API（放在动态路由之前）====================


@router.get("/import/template")
async def download_import_template():
    """下载导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_school_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''school_import_template.xlsx"},
    )


@router.post("/import/excel")
async def import_schools_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入学校"""
    # 安全校验：文件类型 + 大小
    content = await validate_excel_upload(file)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        with open(tmp_path, "rb") as _f:
            wb = load_workbook(_f)
        ws = wb.active

        status_map = {"帮扶中": "active", "未帮扶": "inactive", "已完成": "completed"}

        imported = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 检查是否有数据，主要检查必填项：学校名称(索引1)
            # 防止索引越界
            if not row or len(row) < 2 or not row[1]:
                continue

            try:
                type_val = TYPE_MAP.get(str(row[3]) if row[3] else "小学", "primary")
                status_val = status_map.get(str(row[10]) if row[10] else "未帮扶", "inactive")

                school = School(
                    name=str(row[1]),
                    code=str(row[2]) if row[2] else None,
                    type=SchoolType(type_val),
                    province=str(row[4]) if row[4] else None,
                    city=str(row[5]) if row[5] else None,
                    district=str(row[6]) if row[6] else None,
                    address=str(row[7]) if row[7] else None,
                    student_count=int(row[8]) if row[8] else 0,
                    teacher_count=int(row[9]) if row[9] else 0,
                    support_status=SupportStatus(status_val),
                    support_unit=str(row[11]) if row[11] else None,
                    principal=str(row[12]) if row[12] else None,
                    contact_phone=str(row[13]) if row[13] else None,
                    organization_id=getattr(current_user, "organization_id", None),
                    created_by=current_user.id,
                )
                db.add(school)
                imported += 1
            except (ValueError, TypeError) as e:
                errors.append(_fmt_row_error(row_idx, e, "数据格式错误（请检查数字/日期字段格式）",
                                             row[1] if row and len(row) > 1 else "?"))
            except Exception as e:
                errors.append(_fmt_row_error(row_idx, e, str(e),
                                             row[1] if row and len(row) > 1 else "?"))
        safe_commit(db)
        return {
            "success": True,
            "message": f"成功导入 {imported} 所学校",
            "imported": imported,
            "failed": len(errors),
            "errors": errors,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ── 奖学金学生导入 ──

@router.post("/scholarship/import", summary="导入奖学金学生Excel")
async def import_scholarship_students(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """从 Excel 导入奖学金资助学生"""
    content = await validate_excel_upload(file)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name
    try:
        with open(tmp_path, "rb") as _f:
            wb = load_workbook(_f)
        ws = wb.active
        imported = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2 or not row[1]:
                continue
            try:
                stu = ScholarshipStudent(
                    name=str(row[1]),
                    student_id=str(row[2]) if len(row) > 2 and row[2] else None,
                    year=int(row[3]) if len(row) > 3 and row[3] else None,
                    amount=float(row[4]) if len(row) > 4 and row[4] else 0.0,
                    school_name=str(row[5]) if len(row) > 5 and row[5] else None,
                    grade=str(row[6]) if len(row) > 6 and row[6] else None,
                    reason=str(row[7]) if len(row) > 7 and row[7] else None,
                    status=_parse_scholarship_status(row),
                )
                db.add(stu)
                imported += 1
            except (ValueError, TypeError) as e:
                errors.append(_fmt_row_error(row_idx, e, "数据格式错误（请检查年份/金额字段是否为数字）",
                                             row[0] if row and len(row) > 0 else "?"))
            except Exception as e:
                errors.append(_fmt_row_error(row_idx, e, str(e),
                                             row[0] if row and len(row) > 0 else "?"))
        safe_commit(db)
        return {
            "success": True,
            "message": f"成功导入 {imported} 名资助学生",
            "imported": imported,
            "failed": len(errors),
            "errors": errors,
        }
    finally:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            logger.debug("Temp file already removed: %s", tmp_path)


@router.get("/export")
@router.get("/export/excel")
async def export_schools_excel(
    current_user=Depends(get_current_user),
    data_scope: OrgScopeFilter = Depends(get_org_scope),
    db: Session = Depends(get_db),
):
    """导出学校到 Excel"""
    query = db.query(School).filter(School.is_active == True)  # noqa: E712
    query = data_scope.filter_by_org_ids(query, School.organization_id, created_by_column=School.created_by)
    schools = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "学校列表"

    headers = [
        "序号",
        "学校名称",
        "学校编码",
        "学校类型",
        "所在区域",
        "学生人数",
        "教师人数",
        "帮扶状态",
        "帮扶单位",
        "校长",
        "联系电话",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_display = {"active": "帮扶中", "inactive": "未帮扶", "completed": "已完成"}

    for idx, school in enumerate(schools, 1):
        region = f"{school.province or ''}{school.city or ''}{school.district or ''}"
        type_val = school.type.value if school.type else ""
        status_val = school.support_status.value if school.support_status else ""
        ws.append(
            [
                idx,
                school.name,
                school.code or "",
                TYPE_DISPLAY.get(type_val, type_val),
                region,
                school.student_count,
                school.teacher_count,
                status_display.get(status_val, status_val),
                school.support_unit or "",
                school.principal or "",
                school.contact_phone or "",
            ]
        )

    widths = [8, 25, 15, 12, 25, 12, 12, 12, 20, 12, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schools.xlsx"},
    )


# ==================== 统计 API ====================


@router.get("/statistics")
async def get_school_statistics(
    current_user=Depends(get_current_user),
    data_scope: OrgScopeFilter = Depends(get_org_scope),
    db: Session = Depends(get_db),
):
    """学校统计数据（含助学兴教统计）"""
    school_q = db.query(School).filter(School.is_active == True)  # noqa: E712
    school_q = data_scope.filter_by_org_ids(school_q, School.organization_id, created_by_column=School.created_by)

    total_schools = school_q.count()
    active = school_q.filter(School.support_status == SupportStatus.ACTIVE).count()
    completed = school_q.filter(School.support_status == SupportStatus.COMPLETED).count()

    # 学生/教师总数
    people_row = school_q.with_entities(
        sa_func.coalesce(sa_func.sum(School.student_count), 0),
        sa_func.coalesce(sa_func.sum(School.teacher_count), 0),
    ).first()
    total_students = int(people_row[0]) if people_row else 0
    total_teachers = int(people_row[1]) if people_row else 0

    # 助学兴教统计
    school_ids = [s.id for s in school_q.with_entities(School.id).all()]
    project_count = 0
    project_total_budget = 0.0
    scholarship_count = 0
    scholarship_total_amount = 0.0
    if school_ids:
        project_count = db.query(SchoolProject).filter(SchoolProject.school_id.in_(school_ids)).count()
        budget_row = (
            db.query(sa_func.coalesce(sa_func.sum(SchoolProject.budget), 0))
            .filter(SchoolProject.school_id.in_(school_ids))
            .scalar()
        )
        project_total_budget = float(budget_row or 0)

        scholarship_count = db.query(ScholarshipStudent).filter(ScholarshipStudent.school_id.in_(school_ids)).count()
        amount_row = (
            db.query(sa_func.coalesce(sa_func.sum(ScholarshipStudent.amount), 0))
            .filter(ScholarshipStudent.school_id.in_(school_ids))
            .scalar()
        )
        scholarship_total_amount = float(amount_row or 0)

    return {
        "total_schools": total_schools,
        "active": active,
        "completed": completed,
        "total_students": total_students,
        "total_teachers": total_teachers,
        "project_count": project_count,
        "project_total_budget": project_total_budget,
        "scholarship_count": scholarship_count,
        "scholarship_total_amount": scholarship_total_amount,
    }


@router.get("/options/types")
async def get_type_options():
    """获取学校类型选项"""
    return [
        {"value": "primary", "label": "小学"},
        {"value": "middle", "label": "初中"},
        {"value": "high", "label": "高中"},
        {"value": "vocational", "label": "职业学校"},
        {"value": "other", "label": "其他"},
    ]


@router.get("/options/statuses")
async def get_status_options():
    """获取帮扶状态选项"""
    return [
        {"value": "active", "label": "帮扶中"},
        {"value": "inactive", "label": "未帮扶"},
        {"value": "completed", "label": "已完成"},
    ]


# ==================== 附件管理API（静态路由，放在动态路由之前）====================


@router.get("/attachments/{attachment_id}/download")
async def download_attachment(
    attachment_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载附件"""
    att = db.query(SchoolAttachment).filter(SchoolAttachment.id == attachment_id).first()
    if not att:
        raise AppError.not_found("附件")

    # 数据隔离：校验附件所属学校的访问权限（防止跨组织下载）
    school = db.query(School).filter(School.id == att.school_id).first()
    if school:
        require_data_permission(
            current_user, school.organization_id, school.created_by, db,
            error_message="无权下载该附件",
        )

    file_path = _validate_file_path(att.file_path)

    return FileResponse(
        path=file_path,
        filename=att.file_name,
        media_type=att.file_type or "application/octet-stream",
    )


@router.delete("/attachments/{attachment_id}")
async def delete_attachment(
    attachment_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除附件"""
    att = db.query(SchoolAttachment).filter(SchoolAttachment.id == attachment_id).first()
    if not att:
        raise AppError.not_found("附件")

    # 数据隔离：校验附件所属学校的访问权限（防止跨组织删除）
    school = db.query(School).filter(School.id == att.school_id).first()
    if school:
        require_data_permission(
            current_user, school.organization_id, school.created_by, db,
            error_message="无权删除该附件",
        )

    file_path = _validate_file_path(att.file_path)

    # 删除磁盘文件
    os.remove(file_path)

    db.delete(att)
    safe_commit(db)
    return {"message": "删除成功"}


# ==================== CRUD API ====================


@router.get("")
async def list_schools(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: Optional[str] = None,
    name: Optional[str] = None,
    type: Optional[str] = None,
    support_status: Optional[str] = None,
    supportStatus: Optional[str] = None,
    include_deleted: bool = Depends(enforce_admin_include_deleted),
    current_user=Depends(get_current_user),
    data_scope: OrgScopeFilter = Depends(get_org_scope),
    db: Session = Depends(get_db),
):
    """获取学校列表"""
    status_filter = support_status or supportStatus

    # 缓存：学校数据日级更新，TTL 120s
    # 跳过测试环境——模块级缓存单例会在测试间泄漏状态
    import hashlib
    import json
    import os
    _ckey = None
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        from app.core.cache import get_cache_service
        _cache = await get_cache_service()
        try:
            _key_data = json.dumps(
                [keyword, name, type, status_filter, include_deleted], default=str
            ).encode()
            _org_id = getattr(current_user, "organization_id", None) or 0
            _hash = hashlib.md5(_key_data, usedforsecurity=False).hexdigest()
            _ckey = f"schools:list:{_org_id}:{page}:{page_size}:{_hash}"
            _cached = await _cache.get(_ckey)
            if _cached is not None:
                return _cached
        except (TypeError, ValueError):
            _ckey = None

    # include_deleted 已由 enforce_admin_include_deleted 依赖收敛：非管理员自动降级为 False
    query = db.query(School)
    # 默认过滤软删记录（is_active=False），include_deleted=True 时显示全部
    if not include_deleted:
        query = query.filter(School.is_active == True)  # noqa: E712

    # 数据范围过滤：非管理员只能看到自己组织及下级组织的帮扶学校
    # NOTE: 此处使用 data_scope.filter_by_org_ids（支持 org_children 含下级组织）。
    # funds.py / projects.py / supported_village.py 使用 filter_by_data_scope（基于角色，OWN_DEPT 仅本组织）。
    # 两套系统行为不一致，待业务确认后统一。
    query = data_scope.filter_by_org_ids(query, School.organization_id, created_by_column=School.created_by)

    if keyword or name:
        search = keyword or name
        query = query.filter(School.name.contains(search))
    if type:
        type_val = TYPE_MAP.get(type, type)
        try:
            query = query.filter(School.type == SchoolType(type_val))
        except ValueError:
            logger.warning("Invalid school type filter value: %s", type_val)
    if status_filter:
        try:
            query = query.filter(School.support_status == SupportStatus(status_filter))
        except ValueError:
            logger.warning("Invalid support status filter value: %s", status_filter)

    total = query.count()
    items = query.order_by(School.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    # 返回统一 envelope：{code:200, data:{items,total,page,page_size}, message}
    result = ok_list(
        items=[s.to_dict() for s in items],
        total=total,
        page=page,
        page_size=page_size,
    )
    if _ckey is not None:
        await _cache.set(_ckey, result, ttl=120)
    return result


@router.get("/{school_id}")
async def get_school(
    school_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取学校详情（含已软删记录，管理员可见时附带 viewableBecause 元数据）"""
    # 详情不强制 is_active=True，软删记录可见
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise AppError.not_found("学校")
    require_data_permission(
        current_user,
        school.organization_id,
        school.created_by,
        db,
        error_message="无权访问该学校数据",
    )
    data = school.to_dict()
    data["viewableBecause"] = build_viewable_because(current_user, school)
    return success_response(data=data, message="成功")


@router.post("")
async def create_school(
    data: SchoolCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建学校"""
    # 将空字符串转换为None，避免唯一约束冲突
    if data.code == "":
        data.code = None

    # 检查编码唯一性
    if data.code:
        existing = db.query(School).filter(School.code == data.code).first()
        if existing:
            raise AppError.conflict("学校编码已存在")

    school_data = data.model_dump()
    school_data.setdefault("organization_id", getattr(current_user, "organization_id", None))
    school_data.setdefault("created_by", current_user.id)

    # Map schema field names to model column names
    if "school_type" in school_data:
        school_data["type"] = school_data.pop("school_type")
    if "school_level" in school_data:
        school_data["level"] = school_data.pop("school_level")

    # Remove schema-only fields not in the model table
    model_cols = {c.name for c in School.__table__.columns}
    school_data = {k: v for k, v in school_data.items() if k in model_cols}

    # 转换枚举
    if school_data.get("type"):
        type_val = TYPE_MAP.get(school_data["type"], school_data["type"])
        school_data["type"] = SchoolType(type_val)
    if school_data.get("support_status"):
        school_data["support_status"] = SupportStatus(school_data["support_status"])

    school = School(**school_data)
    db.add(school)
    safe_commit(db)
    db.refresh(school)

    write_work_log(
        db,
        "school",
        "create",
        school.id,
        school.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
        detail=f"所在：{school.district}" if school.district else "",
    )

    # 清除仪表板缓存
    try:
        from app.api.v1.data.data.dashboard import invalidate_dashboard_cache

        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")

    return {"message": "创建成功", "data": school.to_dict()}


@router.put("/{school_id}")
async def update_school(
    school_id: int,
    data: SchoolUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新学校"""
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise AppError.not_found("学校")
    require_data_permission(
        current_user,
        school.organization_id,
        school.created_by,
        db,
        error_message="无权编辑该学校数据",
    )

    update_data = data.model_dump(exclude_unset=True)

    # 将空字符串转换为None，避免唯一约束冲突
    if "code" in update_data and update_data["code"] == "":
        update_data["code"] = None

    # 检查编码唯一性（如果要更新编码）
    if "code" in update_data and update_data["code"]:
        existing = db.query(School).filter(School.code == update_data["code"], School.id != school_id).first()
        if existing:
            raise AppError.conflict("学校编码已存在")

    # 转换枚举
    if "type" in update_data and update_data["type"]:
        type_val = TYPE_MAP.get(update_data["type"], update_data["type"])
        update_data["type"] = SchoolType(type_val)
    if "support_status" in update_data and update_data["support_status"]:
        update_data["support_status"] = SupportStatus(update_data["support_status"])

    for key, value in update_data.items():
        setattr(school, key, value)

    safe_commit(db)
    db.refresh(school)

    write_work_log(
        db,
        "school",
        "update",
        school.id,
        school.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
    )

    # 清除仪表板缓存
    try:
        from app.api.v1.data.data.dashboard import invalidate_dashboard_cache

        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")

    return {"message": "更新成功", "data": school.to_dict()}


@router.delete("/{school_id}")
async def delete_school(
    school_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除学校（软删除）"""
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise AppError.not_found("学校")
    require_data_permission(
        current_user,
        school.organization_id,
        school.created_by,
        db,
        error_message="无权删除该学校数据",
    )

    school.is_active = False
    safe_commit(db)

    write_work_log(
        db,
        "school",
        "delete",
        school.id,
        school.name,
        user_id=current_user.id,
        username=getattr(current_user, "username", "系统"),
        detail=f"所在：{school.district}" if school.district else "",
    )

    # 清除仪表板缓存
    try:
        from app.api.v1.data.data.dashboard import invalidate_dashboard_cache

        invalidate_dashboard_cache()
    except Exception:
        logger.debug("仪表盘缓存失效失败")

    return {"message": "删除成功"}


# ==================== 学校附件管理 ====================


@router.get("/{school_id}/attachments")
async def list_attachments(
    school_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取学校附件列表"""
    _get_school_and_check_permission(school_id, current_user, db, "查看附件")
    attachments = (
        db.query(SchoolAttachment)
        .filter(SchoolAttachment.school_id == school_id)
        .order_by(SchoolAttachment.created_at.desc())
        .all()
    )
    # envelope 格式：前端拦截器自动展开 data.items 到顶层，res.items 仍可用
    return ok_list(items=[a.to_dict() for a in attachments], total=len(attachments))


@router.post("/{school_id}/attachments")
async def upload_attachment(
    school_id: int,
    file: UploadFile = File(...),
    description: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传学校电子资料"""
    # 检查学校是否存在并校验数据权限
    _get_school_and_check_permission(school_id, current_user, db, "上传附件")

    # 检查文件大小
    content = await file.read()
    if len(content) > settings.MAX_FILE_SIZE:
        raise AppError.bad_request(f"文件大小超过限制({settings.MAX_FILE_SIZE // 1048576}MB)")

    # 检查文件类型
    ext = (file.filename or "").rsplit(".", 1)[-1].lower() if file.filename and "." in file.filename else ""
    allowed_types = settings.allowed_file_types_list + [
        "doc",
        "docx",
        "ppt",
        "pptx",
        "txt",
        "zip",
        "rar",
    ]
    if ext and ext not in allowed_types:
        raise AppError.bad_request(f"不支持的文件类型: .{ext}")

    # 创建存储目录（使用绝对路径）
    base_upload = os.path.abspath(settings.UPLOAD_DIR)
    upload_dir = os.path.join(base_upload, "schools", str(school_id))
    os.makedirs(upload_dir, exist_ok=True)

    # 生成唯一文件名
    unique_name = f"{uuid.uuid4().hex[:12]}_{file.filename}"
    file_path = os.path.join(upload_dir, unique_name)

    # 写入文件
    with open(file_path, "wb") as f:
        f.write(content)

    # 保存数据库记录
    attachment = SchoolAttachment(
        school_id=school_id,
        file_name=file.filename or "unknown",
        file_path=file_path,
        file_size=len(content),
        file_type=file.content_type or "application/octet-stream",
        description=description or "",
        uploaded_by=getattr(current_user, "username", "") or getattr(current_user, "full_name", ""),
    )
    db.add(attachment)
    safe_commit(db)
    db.refresh(attachment)

    return {"message": "上传成功", "data": attachment.to_dict()}


# ==================== 学校帮扶项目 CRUD ====================


class SchoolProjectCreate(BaseModel):
    name: str
    phase: Optional[str] = "research"
    category: Optional[str] = None
    budget: float = 0
    actual_cost: float = 0
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    description: Optional[str] = None
    remarks: Optional[str] = None


class SchoolProjectUpdate(BaseModel):
    name: Optional[str] = None
    phase: Optional[str] = None
    category: Optional[str] = None
    budget: Optional[float] = None
    actual_cost: Optional[float] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    description: Optional[str] = None
    remarks: Optional[str] = None


@router.get("/{school_id}/projects")
async def list_projects(
    school_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取学校帮扶项目列表"""
    _get_school_and_check_permission(school_id, current_user, db, "查看项目")
    items = (
        db.query(SchoolProject)
        .filter(SchoolProject.school_id == school_id)
        .order_by(SchoolProject.created_at.desc())
        .all()
    )
    # envelope 格式：前端拦截器自动展开 data.items 到顶层，res.items 仍可用
    return ok_list(items=[p.to_dict() for p in items], total=len(items))


@router.post("/{school_id}/projects")
async def create_project(
    school_id: int,
    data: SchoolProjectCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """新增帮扶项目"""
    _get_school_and_check_permission(school_id, current_user, db, "创建项目")
    project_data = data.model_dump()
    if project_data.get("phase"):
        project_data["phase"] = ProjectPhase(project_data["phase"])
    project = SchoolProject(school_id=school_id, **project_data)
    db.add(project)
    safe_commit(db)
    db.refresh(project)
    return {"message": "创建成功", "data": project.to_dict()}


@router.put("/{school_id}/projects/{project_id}")
async def update_project(
    school_id: int,
    project_id: int,
    data: SchoolProjectUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新帮扶项目"""
    _get_school_and_check_permission(school_id, current_user, db, "更新项目")
    project = (
        db.query(SchoolProject).filter(SchoolProject.id == project_id, SchoolProject.school_id == school_id).first()
    )
    if not project:
        raise AppError.not_found("项目")
    update_data = data.model_dump(exclude_unset=True)
    if "phase" in update_data and update_data["phase"]:
        update_data["phase"] = ProjectPhase(update_data["phase"])
    for key, val in update_data.items():
        setattr(project, key, val)
    safe_commit(db)
    db.refresh(project)
    return {"message": "更新成功", "data": project.to_dict()}


@router.delete("/{school_id}/projects/{project_id}")
async def delete_project(
    school_id: int,
    project_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除帮扶项目"""
    _get_school_and_check_permission(school_id, current_user, db, "删除项目")
    project = (
        db.query(SchoolProject).filter(SchoolProject.id == project_id, SchoolProject.school_id == school_id).first()
    )
    if not project:
        raise AppError.not_found("项目")
    db.delete(project)
    safe_commit(db)
    return {"message": "删除成功"}


# ==================== 资助学生 CRUD ====================


class ScholarshipStudentCreate(BaseModel):
    student_name: str
    grade: Optional[str] = None
    year: Optional[int] = None
    amount: float = 0
    reason: Optional[str] = None
    status: Optional[str] = "pending"
    contact_info: Optional[str] = None
    remarks: Optional[str] = None


class ScholarshipStudentUpdate(BaseModel):
    student_name: Optional[str] = None
    grade: Optional[str] = None
    year: Optional[int] = None
    amount: Optional[float] = None
    reason: Optional[str] = None
    status: Optional[str] = None
    contact_info: Optional[str] = None
    remarks: Optional[str] = None


@router.get("/{school_id}/scholarship-students")
async def list_scholarship_students(
    school_id: int,
    year: Optional[int] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取资助学生列表"""
    _get_school_and_check_permission(school_id, current_user, db, "查看资助学生")
    q = db.query(ScholarshipStudent).filter(ScholarshipStudent.school_id == school_id)
    if year:
        q = q.filter(ScholarshipStudent.year == year)
    items = q.order_by(ScholarshipStudent.created_at.desc()).all()
    return ok_list(items=[s.to_dict() for s in items], total=len(items))


@router.post("/{school_id}/scholarship-students")
async def create_scholarship_student(
    school_id: int,
    data: ScholarshipStudentCreate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """新增资助学生"""
    _get_school_and_check_permission(school_id, current_user, db, "添加资助学生")
    stu_data = data.model_dump()
    if stu_data.get("status"):
        stu_data["status"] = ScholarshipStatus(stu_data["status"])
    student = ScholarshipStudent(school_id=school_id, **stu_data)
    db.add(student)
    safe_commit(db)
    db.refresh(student)
    return {"message": "创建成功", "data": student.to_dict()}


@router.put("/{school_id}/scholarship-students/{student_id}")
async def update_scholarship_student(
    school_id: int,
    student_id: int,
    data: ScholarshipStudentUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新资助学生"""
    _get_school_and_check_permission(school_id, current_user, db, "更新资助学生")
    stu = (
        db.query(ScholarshipStudent)
        .filter(
            ScholarshipStudent.id == student_id,
            ScholarshipStudent.school_id == school_id,
        )
        .first()
    )
    if not stu:
        raise AppError.not_found("资助学生记录")
    update_data = data.model_dump(exclude_unset=True)
    if "status" in update_data and update_data["status"]:
        update_data["status"] = ScholarshipStatus(update_data["status"])
    for key, val in update_data.items():
        setattr(stu, key, val)
    safe_commit(db)
    db.refresh(stu)
    return {"message": "更新成功", "data": stu.to_dict()}


@router.delete("/{school_id}/scholarship-students/{student_id}")
async def delete_scholarship_student(
    school_id: int,
    student_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除资助学生"""
    _get_school_and_check_permission(school_id, current_user, db, "删除资助学生")
    stu = (
        db.query(ScholarshipStudent)
        .filter(
            ScholarshipStudent.id == student_id,
            ScholarshipStudent.school_id == school_id,
        )
        .first()
    )
    if not stu:
        raise AppError.not_found("资助学生记录")
    db.delete(stu)
    safe_commit(db)
    return {"message": "删除成功"}


@router.post("/{school_id}/scholarship-students/import")
async def import_school_scholarship_students(
    school_id: int,
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入资助学生"""
    _get_school_and_check_permission(school_id, current_user, db, "导入资助学生")

    content = await validate_excel_upload(file)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    status_map = {
        "待审批": "pending",
        "已批准": "approved",
        "已发放": "disbursed",
        "已完成": "completed",
    }
    try:
        with open(tmp_path, "rb") as _f:
            wb = load_workbook(_f)
        ws = wb.active
        imported = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2 or not row[0]:
                continue
            try:
                stu = ScholarshipStudent(
                    school_id=school_id,
                    student_name=str(row[0]),
                    grade=str(row[1]) if len(row) > 1 and row[1] else None,
                    year=int(row[2]) if len(row) > 2 and row[2] else None,
                    amount=float(row[3]) if len(row) > 3 and row[3] else 0,
                    reason=str(row[4]) if len(row) > 4 and row[4] else None,
                    status=(
                        ScholarshipStatus(status_map.get(str(row[5]), "pending"))
                        if len(row) > 5 and row[5]
                        else ScholarshipStatus.PENDING
                    ),
                )
                db.add(stu)
                imported += 1
            except (ValueError, TypeError) as e:
                errors.append(_fmt_row_error(row_idx, e, "数据格式错误（请检查年份/金额字段是否为数字）",
                                             row[0] if row and len(row) > 0 else "?"))
            except Exception as e:
                errors.append(_fmt_row_error(row_idx, e, str(e),
                                             row[0] if row and len(row) > 0 else "?"))
        safe_commit(db)
        return {
            "success": True,
            "message": f"成功导入 {imported} 名资助学生",
            "imported": imported,
            "errors": errors,
        }
    finally:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            logger.debug("Temp file already removed: %s", tmp_path)

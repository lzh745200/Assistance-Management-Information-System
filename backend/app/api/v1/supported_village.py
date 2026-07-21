"""帮扶村管理 API 路由"""

# NOTE: 数据权限过滤建议迁移到 app.core.data_scope_adapter.apply_scope_filter()
# 当前使用: filter_by_data_scope() (混合风格)

import io
import json
import logging
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.core.response import ok_list
from app.core.security import get_current_user
from app.models.user import User
from app.utils.common import dict_keys_to_camel, StringHelper
from app.models.supported_village import (
    SupportedVillage,
    VillagePopulation,
    VillageIncome,
    ForceInvestment,
    IndustrySupport,
    InfrastructureImprovement,
    PartyBuildingSupport,
    MedicalSupport,
    ConsumptionSupport,
    EmploymentSupport,
    EducationSupport,
    VillageCommitteeInfo,
    VillageCommitteeMember,
)
from app.core.data_permission import filter_by_data_scope, check_record_access
from app.api.v1.deps import enforce_admin_include_deleted, build_viewable_because
from app.schemas.supported_village import SupportedVillageCreate, SupportedVillageUpdate
from app.core.transaction import safe_commit

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/supported-villages", tags=["帮扶村管理"])


class BatchDeleteRequest(BaseModel):
    ids: List[int]


class YearCopyRequest(BaseModel):
    fromYear: int
    toYear: int


class YearlySectionData(BaseModel):
    """年度区块数据 — 字段由各 section 动态决定，路由层过滤安全属性"""
    year: Optional[int] = None
    model_config = {"extra": "allow"}


# ── 年度数据 Section → Model 映射 ──
_SECTION_MODEL: Dict[str, Any] = {
    "population": VillagePopulation,
    "income": VillageIncome,
    "force-investment": ForceInvestment,
    "industry": IndustrySupport,
    "infrastructure": InfrastructureImprovement,
    "party-building": PartyBuildingSupport,
    "medical": MedicalSupport,
    "consumption": ConsumptionSupport,
    "employment": EmploymentSupport,
    "education": EducationSupport,
    "committee": VillageCommitteeInfo,
}

# ── 导入导出列定义 ──
_IMPORT_COLUMNS = [
    ("village_name", "帮扶村名称"),
    ("department", "部门单位"),
    ("support_unit", "帮扶单位"),
    ("province", "省"),
    ("city", "市"),
    ("county", "县/市"),
    ("township", "乡镇"),
    ("village_category", "村庄类别"),
    ("is_three_regions", "是否三区三州"),
    ("is_key_county", "是否重点帮扶县"),
    ("is_revitalization_tier", "是否振兴梯队"),
    ("longitude", "经度"),
    ("latitude", "纬度"),
    ("altitude", "海拔"),
    ("area_sq_km", "面积(km²)"),
    ("households", "户数"),
    ("description", "备注"),
]

# 预计算，避免重复 destructure
_FIELD_NAMES = [col[0] for col in _IMPORT_COLUMNS]
_HEADER_NAMES = [col[1] for col in _IMPORT_COLUMNS]

# 年度数据辅助函数中需要跳过的元数据列
_SKIP_COLUMNS = frozenset({"id", "supported_village_id", "year", "created_at", "updated_at"})


# ═══════════════════════════════════════════════════════════════
#  辅助函数
# ═══════════════════════════════════════════════════════════════

def _get_section_data(db: Session, model: Any, village_id: int, year: int) -> Optional[Dict]:
    row = db.query(model).filter(
        model.supported_village_id == village_id, model.year == year
    ).first()
    if not row:
        return None
    result = {}
    for col in model.__table__.columns:
        if col.name in _SKIP_COLUMNS:
            continue
        val = getattr(row, col.name)
        result[col.name] = val
    # 村委会板块：加载成员列表
    if model is VillageCommitteeInfo:
        members = db.query(VillageCommitteeMember).filter(
            VillageCommitteeMember.committee_info_id == row.id
        ).all()
        result["members"] = [
            dict_keys_to_camel({
                "name": m.name,
                "position": m.position,
                "phone": m.phone,
                "is_veteran": m.is_veteran,
                "remark": m.remark,
            })
            for m in members
        ]
    return dict_keys_to_camel(result)


def _copy_section_data(db: Session, model: Any, village_id: int, from_year: int, to_year: int) -> bool:
    src = db.query(model).filter(
        model.supported_village_id == village_id, model.year == from_year
    ).first()
    if not src:
        return False
    existing = db.query(model).filter(
        model.supported_village_id == village_id, model.year == to_year
    ).first()
    if existing:
        return False
    new_row = model()
    new_row.supported_village_id = village_id
    new_row.year = to_year
    for col in model.__table__.columns:
        if col.name in _SKIP_COLUMNS:
            continue
        setattr(new_row, col.name, getattr(src, col.name, None))
    db.add(new_row)
    return True


def _save_section_data(db: Session, model: Any, village_id: int, year: int, data: dict):
    row = db.query(model).filter(
        model.supported_village_id == village_id, model.year == year
    ).first()
    if not row:
        row = model()
        row.supported_village_id = village_id
        row.year = year
        db.add(row)
    # 处理村委会成员子表
    members_data = data.pop("members", None)
    # 将 camelCase 键转为 snake_case 以匹配模型属性名
    snake_data = {}
    for key, value in data.items():
        if key in _SKIP_COLUMNS:
            continue
        snake_key = StringHelper.to_snake_case(key) if hasattr(key, 'upper') else key
        snake_data[snake_key] = value
    for key, value in snake_data.items():
        if hasattr(row, key):
            setattr(row, key, value)
    if members_data is not None and model is VillageCommitteeInfo:
        # 清除旧成员，写入新成员
        db.query(VillageCommitteeMember).filter(
            VillageCommitteeMember.committee_info_id == row.id
        ).delete()
        for m in members_data:
            if isinstance(m, dict):
                member = VillageCommitteeMember(
                    committee_info_id=row.id,
                    name=m.get("name", ""),
                    position=m.get("position", ""),
                    phone=m.get("phone", ""),
                    is_veteran=(
                        m.get("isVeteran", False) if isinstance(m.get("isVeteran"), bool)
                        else m.get("isVeteran", False)
                    ),
                    remark=m.get("remark", ""),
                )
                db.add(member)
    return row


def _get_village_or_404(db: Session, village_id: int, current_user: User = None) -> SupportedVillage:
    """根据 ID 获取帮扶村，不存在时抛 404；存在但跨组织时抛 403（数据隔离）。"""
    village = db.query(SupportedVillage).filter(SupportedVillage.id == village_id).first()
    if not village:
        raise HTTPException(status_code=404, detail="帮扶村不存在")
    # 数据权限校验：非本组织/非本人创建且非管理员 → 403（区分"不存在"与"越权"）
    if current_user is not None and not check_record_access(
        village, current_user, owner_field="created_by", dept_field="organization_id"
    ):
        raise HTTPException(status_code=403, detail="无权访问该帮扶村")
    return village


async def _invalidate_village_cache():
    """清除帮扶村列表缓存，确保写操作后立即可见"""
    import os
    if os.environ.get("PYTEST_CURRENT_TEST"):
        return
    try:
        from app.core.cache import get_cache_service
        _cache = await get_cache_service()
        await _cache.delete_by_prefix("villages:list:")
    except Exception as e:
        logger.debug("清理帮扶村列表缓存失败: %s", e)


def _process_import_row(row: tuple, field_names: List[str], db: Session, row_idx: int):
    """处理单行导入数据。返回 (success: bool, error_msg: Optional[str])"""
    values = {}
    for i, field_name in enumerate(field_names):
        val = row[i] if i < len(row) else None
        if val is not None and isinstance(val, str):
            val = val.strip()
        if field_name in ("is_three_regions", "is_key_county", "is_revitalization_tier"):
            val = str(val).strip() in ("是", "1", "True", "true", "yes", "Y")
        values[field_name] = val
    if not values.get("village_name"):
        return False, f"第{row_idx}行: 帮扶村名称不能为空"
    existing = db.query(SupportedVillage).filter(
        SupportedVillage.village_name == values["village_name"],
        SupportedVillage.county == values.get("county"),
        SupportedVillage.is_active == True,  # noqa: E712
    ).first()
    if existing:
        return False, f"第{row_idx}行: 帮扶村 '{values['village_name']}' 已存在，跳过"
    village = SupportedVillage(**{k: v for k, v in values.items() if v is not None})
    db.add(village)
    return True, None


# ═══════════════════════════════════════════════════════════════
#  列表 & 筛选选项（无路径参数，必须在 /{village_id} 之前注册）
# ═══════════════════════════════════════════════════════════════

@router.get("")
async def list_villages(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: Optional[str] = None,
    department: Optional[str] = None,
    county: Optional[str] = None,
    isRevitalizationTier: Optional[bool] = None,
    isThreeRegions: Optional[int] = None,
    include_deleted: bool = Depends(enforce_admin_include_deleted),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村列表（分页、筛选）"""
    # 缓存：帮扶村日级更新，TTL 120s
    # 跳过测试环境——模块级缓存单例会在测试间泄漏状态
    import hashlib
    import json
    import os
    _ckey = None
    if not os.environ.get("PYTEST_CURRENT_TEST"):
        from app.core.cache import get_cache_service
        _cache = await get_cache_service()
        try:
            _key_data = json.dumps(
                [keyword, department, county, isRevitalizationTier, isThreeRegions, include_deleted],
                default=str,
            ).encode()
            _org_id = getattr(current_user, "organization_id", None) or 0
            _hash = hashlib.md5(_key_data, usedforsecurity=False).hexdigest()
            _ckey = f"villages:list:{_org_id}:{page}:{page_size}:{_hash}"
            _cached = await _cache.get(_ckey)
            if _cached is not None:
                return _cached
        except (TypeError, ValueError):
            _ckey = None

    # include_deleted 已由 enforce_admin_include_deleted 依赖收敛：非管理员自动降级为 False
    query = db.query(SupportedVillage)
    query = filter_by_data_scope(query, SupportedVillage, current_user, db=db)

    # 默认过滤软删记录（is_active=False），include_deleted=True 时显示全部
    if not include_deleted:
        query = query.filter(SupportedVillage.is_active == True)  # noqa: E712

    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            SupportedVillage.village_name.ilike(like)
            | SupportedVillage.support_unit.ilike(like)
        )
    if department:
        query = query.filter(SupportedVillage.department == department)
    if county:
        query = query.filter(SupportedVillage.county == county)
    if isRevitalizationTier is not None:
        query = query.filter(SupportedVillage.is_revitalization_tier == isRevitalizationTier)
    if isThreeRegions is not None:
        query = query.filter(SupportedVillage.is_three_regions == bool(isThreeRegions))

    total = query.count()
    # Model-level lazy="selectin" on SupportedVillage relationships prevents
    # N+1 queries when accessing yearly-data backrefs during iteration/serialization.
    items = (
        query
        .order_by(SupportedVillage.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # 返回统一 envelope：{code:200, data:{items,total,page,page_size}, message}
    result = ok_list(
        items=[
            v.to_dict() if hasattr(v, "to_dict") else {"id": v.id, "village_name": v.village_name}
            for v in items
        ],
        total=total,
        page=page,
        page_size=page_size,
    )
    if _ckey is not None:
        await _cache.set(_ckey, result, ttl=120)
    return result


@router.get("/filter-options")
async def get_filter_options(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取筛选选项（部门、县市列表）

    数据隔离：仅返回当前用户数据范围内的部门/县市，避免跨组织泄露。
    """
    # 基础查询应用数据范围过滤，再投影到 department/county 列
    base_query = db.query(SupportedVillage)
    base_query = filter_by_data_scope(base_query, SupportedVillage, current_user, db=db)
    base_query = base_query.filter(SupportedVillage.is_active == True)  # noqa: E712

    departments = (
        base_query.with_entities(SupportedVillage.department)
        .filter(SupportedVillage.department.isnot(None))
        .distinct()
        .all()
    )
    counties = (
        base_query.with_entities(SupportedVillage.county)
        .filter(SupportedVillage.county.isnot(None))
        .distinct()
        .all()
    )
    return {
        "departments": [d[0] for d in departments if d[0]],
        "counties": [c[0] for c in counties if c[0]],
    }


@router.get("/options/dropdown")
async def get_village_dropdown(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村下拉选项（id + name + county，供前端 Select 使用）

    数据隔离：仅返回当前用户数据范围内的帮扶村，避免跨组织泄露。
    """
    query = db.query(
        SupportedVillage.id,
        SupportedVillage.village_name,
        SupportedVillage.county,
    )
    query = filter_by_data_scope(query, SupportedVillage, current_user, db=db)
    villages = (
        query.filter(SupportedVillage.is_active == True)  # noqa: E712
        .order_by(SupportedVillage.id)
        .all()
    )
    items = [
        {"id": v[0], "name": v[1], "county": v[2] or ""}
        for v in villages
    ]
    return {"data": {"items": items}}


@router.get("/import-template")
async def download_import_template():
    """下载导入模板（委托 ExcelTemplateService）"""
    from fastapi.responses import Response
    from app.services.excel_template_service import ExcelTemplateService
    content = ExcelTemplateService().generate_village_template()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''supported_village_import_template.xlsx"},
    )


@router.get("/export")
async def export_villages(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出帮扶村数据到 Excel"""
    query = db.query(SupportedVillage).filter(SupportedVillage.is_active == True)  # noqa: E712
    query = filter_by_data_scope(query, SupportedVillage, current_user, db=db)
    villages = query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "帮扶村数据"
    field_names = _FIELD_NAMES
    headers = _HEADER_NAMES
    ws.append(headers)
    for v in villages:
        row = []
        for fn in field_names:
            val = getattr(v, fn, None)
            if fn in ("is_three_regions", "is_key_county", "is_revitalization_tier"):
                val = "是" if val else "否"
            row.append(val or "")
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=supported_villages_export.xlsx"},
    )


@router.post("/import")
async def import_villages(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """从 Excel 导入帮扶村"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的文件")
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件，请检查文件格式")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件中没有数据行，请至少添加一行数据")
    imported = 0
    errors = []
    field_names = _FIELD_NAMES
    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue
        try:
            success, error_msg = _process_import_row(row, field_names, db, row_idx)
            if success:
                imported += 1
            else:
                errors.append(error_msg)
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")
    safe_commit(db)
    await _invalidate_village_cache()
    return {
        "message": f"成功导入 {imported} 条记录" + (f"，{len(errors)} 条跳过" if errors else ""),
        "imported": imported,
        "errors": errors[:20],
    }


@router.post("/batch-delete")
async def batch_delete_villages(
    data: BatchDeleteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量软删帮扶村（仅可删除当前用户有权访问的记录）"""
    if not data.ids:
        raise HTTPException(status_code=400, detail="请提供要删除的ID列表")
    query = db.query(SupportedVillage).filter(SupportedVillage.id.in_(data.ids))
    # 数据权限过滤：仅允许操作本组织/本人创建的记录，防止跨组织批量删除
    query = filter_by_data_scope(query, SupportedVillage, current_user, db=db)
    deleted_count = query.update({"is_active": False}, synchronize_session=False)
    safe_commit(db)
    await _invalidate_village_cache()
    return {"message": f"已删除 {deleted_count} 条记录"}


# ═══════════════════════════════════════════════════════════════
#  单个帮扶村 CRUD（/{village_id} 必须在所有显式路径之后注册）
# ═══════════════════════════════════════════════════════════════

@router.get("/{village_id}")
async def get_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村详情（含已软删记录，管理员可见时附带 viewableBecause 元数据）"""
    village = _get_village_or_404(db, village_id, current_user)
    data = village.to_dict() if hasattr(village, "to_dict") else {"id": village.id}
    # 管理员查看已软删记录时附带可见性元数据，便于前端审计展示
    data["viewableBecause"] = build_viewable_because(current_user, village)
    return {"data": data}


@router.post("")
async def create_village(
    data: SupportedVillageCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """创建帮扶村"""
    village = SupportedVillage(**data.model_dump())
    # 强制写入数据归属字段，确保新记录纳入组织隔离体系
    village.organization_id = getattr(current_user, "organization_id", None)
    village.created_by = current_user.id
    db.add(village)
    safe_commit(db)
    db.refresh(village)
    await _invalidate_village_cache()
    return {"data": {"id": village.id}, "message": "创建成功"}


@router.put("/{village_id}")
async def update_village(
    village_id: int,
    data: SupportedVillageUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新帮扶村（过渡状态变更需管理员权限 + 写入审计日志）"""
    village = _get_village_or_404(db, village_id, current_user)
    update_dict = data.model_dump(exclude_unset=True)

    # ── 过渡状态变更强制审批检查 ──
    if "transition_status" in update_dict:
        from app.api.v1.deps import require_manager_role
        require_manager_role(current_user)
        old_status = village.transition_status or "none"
        new_status = update_dict["transition_status"]

        # 记录审计日志
        from app.utils.audit_logger import AuditLogger
        AuditLogger.log(
            action="village_transition_change",
            user_id=current_user.id,
            username=current_user.username,
            resource_type="supported_village",
            resource_id=village_id,
            details={
                "old_transition_status": old_status,
                "new_transition_status": new_status,
                "village_name": village.village_name,
            },
        )

    for key, value in update_dict.items():
        if hasattr(village, key) and key != "id":
            setattr(village, key, value)
    safe_commit(db)
    await _invalidate_village_cache()
    return {"message": "更新成功"}


@router.delete("/{village_id}")
async def delete_village(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """软删帮扶村（置 is_active=False，保留关联数据以便恢复/审计）"""
    village = _get_village_or_404(db, village_id, current_user)

    village.is_active = False
    safe_commit(db)
    await _invalidate_village_cache()
    return {"message": "删除成功"}


# ═══════════════════════════════════════════════════════════════
#  年度数据
# ═══════════════════════════════════════════════════════════════

@router.get("/{village_id}/yearly/{year}")
async def get_yearly_data(
    village_id: int,
    year: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村某年度全部数据（所有section）"""
    _get_village_or_404(db, village_id, current_user)
    # 统一 camelCase 键名（内层已由 _get_section_data→dict_keys_to_camel 处理）
    # camelCase 为主键名，snake_case 保留向后兼容（计划 2027-01-01 移除）
    result = {"villageId": village_id, "village_id": village_id, "year": year}
    for section, model in _SECTION_MODEL.items():
        data = _get_section_data(db, model, village_id, year)
        result[section] = data if data else None
    return {"data": result, "message": "ok"}


@router.post("/{village_id}/yearly/copy")
async def copy_year_data(
    village_id: int,
    data: YearCopyRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """将某年的全部年度数据复制到另一年"""
    _get_village_or_404(db, village_id, current_user)
    if data.fromYear == data.toYear:
        raise HTTPException(status_code=400, detail="来源年份和目标年份不能相同")
    copied = 0
    for model in _SECTION_MODEL.values():
        if _copy_section_data(db, model, village_id, data.fromYear, data.toYear):
            copied += 1
    safe_commit(db)
    return {"message": f"年度数据复制成功，已复制 {copied} 个数据组"}


@router.post("/{village_id}/yearly/{year}/{section}")
async def save_yearly_section(
    village_id: int,
    year: int,
    section: str,
    data: YearlySectionData,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存帮扶村某年度某个section的数据"""
    if section not in _SECTION_MODEL:
        raise HTTPException(status_code=400, detail=f"未知的数据分类: {section}")
    _get_village_or_404(db, village_id, current_user)
    model = _SECTION_MODEL[section]
    _save_section_data(db, model, village_id, year, data.model_dump())
    safe_commit(db)
    return {"message": f"保存成功: {section}"}


@router.post("/{village_id}/yearly/{year}/validate")
async def validate_yearly_data(
    village_id: int,
    year: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """校验帮扶村年度数据完整性。

    检查项:
    - 8 大板块必填字段
    - 数值合理性（人均收入≥0）
    - 同比变动超 ±50% 预警
        返回错误列表 + 修正建议。
    """
    _get_village_or_404(db, village_id, current_user)
    errors = []
    warnings = []

    # 首次遍历：读取所有板块当年数据并缓存（避免同比比较时重复查询）
    current_year_data = {}
    for section, model in _SECTION_MODEL.items():
        record = db.query(model).filter(
            model.supported_village_id == village_id,
            model.year == year,
        ).first()
        current_year_data[section] = record

        if not record:
            errors.append({
                "section": section,
                "field": None,
                "message": f"板块 [{section}] 未录入数据",
                "suggestion": "请先填写该板块数据",
            })
            continue

        # 通用数值合理性检查
        for col in model.__table__.columns:
            val = getattr(record, col.name, None)
            if isinstance(val, (int, float)) and val < 0:
                errors.append({
                    "section": section,
                    "field": col.name,
                    "message": f"字段 {col.name} 值为 {val}，不能为负数",
                    "suggestion": "请检查原始数据并修正",
                })

    # 同比变动预警（与前一年对比，复用 current_year_data 缓存）
    if year > 0:
        prev_year_data = {}
        for section, model in _SECTION_MODEL.items():
            prev = db.query(model).filter(
                model.supported_village_id == village_id,
                model.year == year - 1,
            ).first()
            if prev:
                for col in model.__table__.columns:
                    prev_year_data[f"{section}.{col.name}"] = getattr(prev, col.name, None)

        if prev_year_data:
            for section, model in _SECTION_MODEL.items():
                record = current_year_data.get(section)
                if not record:
                    continue
                for col in model.__table__.columns:
                    cur = getattr(record, col.name, None)
                    prev_val = prev_year_data.get(f"{section}.{col.name}")
                    if isinstance(cur, (int, float)) and isinstance(prev_val, (int, float)) and prev_val != 0:
                        change_pct = (cur - prev_val) / abs(prev_val) * 100
                        if abs(change_pct) > 50:
                            warnings.append({
                                "section": section,
                                "field": col.name,
                                "message": f"同比变动 {change_pct:+.1f}%（{prev_val} → {cur}），超过 ±50% 阈值",
                                "suggestion": "请核实数据是否准确，如无误可忽略此预警",
                            })

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
    }


# ── 区块附件管理 ──


@router.get("/{village_id}/sections/{section}/attachments")
async def get_section_attachments(
    village_id: int,
    section: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取帮扶村某区块的附件列表"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    attachments = (
        db.query(VillageAttachment)
        .filter(
            VillageAttachment.supported_village_id == village_id,
            VillageAttachment.description.like(f"section:{section}:%"),
        )
        .order_by(VillageAttachment.created_at.desc())
        .all()
    )
    return {
        "code": 200,
        "data": [
            {
                "id": a.id,
                "fileName": a.file_name,
                "fileSize": a.file_size,
                "fileType": a.mime_type,
                "fileUrl": f"/api/v1/supported-villages/{village_id}/sections/{section}/attachments/{a.id}",
                "createdAt": a.created_at.isoformat() if a.created_at else None,
            }
            for a in attachments
        ],
    }


@router.post("/{village_id}/sections/{section}/attachments")
async def upload_section_attachment(
    village_id: int,
    section: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    import os as _os

    content = await file.read()
    upload_dir = _os.path.join(settings.UPLOAD_DIR, "sections")
    _os.makedirs(upload_dir, exist_ok=True)
    file_path = _os.path.join(upload_dir, f"{village_id}_{section}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(content)

    attachment = VillageAttachment(
        supported_village_id=village_id,
        file_name=file.filename or "unnamed",
        file_path=file_path,
        file_size=len(content),
        mime_type=file.content_type or "application/octet-stream",
        description=f"section:{section}:attachment",
    )
    db.add(attachment)
    safe_commit(db)
    db.refresh(attachment)
    return {"code": 200, "data": {"id": attachment.id, "filename": attachment.file_name}, "message": "上传成功"}


@router.get("/{village_id}/sections/{section}/attachments/{attachment_id}")
async def download_section_attachment(
    village_id: int,
    section: str,
    attachment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """下载帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    import os as _os
    attachment = (
        db.query(VillageAttachment)
        .filter(VillageAttachment.id == attachment_id, VillageAttachment.supported_village_id == village_id)
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")
    if not _os.path.exists(attachment.file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(
        attachment.file_path,
        filename=attachment.file_name,
        media_type=attachment.mime_type or "application/octet-stream",
    )


@router.delete("/{village_id}/sections/{section}/attachments/{attachment_id}")
async def delete_section_attachment(
    village_id: int,
    section: str,
    attachment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除帮扶村某区块的附件"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageAttachment
    attachment = (
        db.query(VillageAttachment)
        .filter(VillageAttachment.id == attachment_id, VillageAttachment.supported_village_id == village_id)
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")
    db.delete(attachment)
    safe_commit(db)
    return {"code": 200, "message": "删除成功"}


# ── 村委数据 ──


@router.post("/{village_id}/committee")
async def save_committee_data(
    village_id: int,
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存帮扶村委数据"""
    _get_village_or_404(db, village_id, current_user)
    from app.models.supported_village import VillageCommitteeInfo
    committee = (
        db.query(VillageCommitteeInfo)
        .filter(VillageCommitteeInfo.supported_village_id == village_id)
        .first()
    )
    if not committee:
        committee = VillageCommitteeInfo(supported_village_id=village_id)
        db.add(committee)
    for k, v in data.items():
        if hasattr(committee, k) and k not in ("id", "supported_village_id"):
            setattr(committee, k, v)
    safe_commit(db)
    return {"code": 200, "message": "保存成功"}


# ── 区块数据导入 ──


@router.post("/{village_id}/sections/import")
async def import_section_data(
    village_id: int,
    year: Optional[int] = Query(None, description="年度"),
    section_key: Optional[str] = Query(None, description="板块标识"),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导入帮扶村单个区块数据（Excel）"""
    _get_village_or_404(db, village_id, current_user)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()][1:]  # skip header
        return {
            "code": 200,
            "data": {
                "rows": len(rows),
                "imported": len(rows),
                "failed": 0,
                "preview": rows[:5],
                "section_key": section_key,
                "year": year,
            },
            "message": f"导入成功，共 {len(rows)} 行",
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")


@router.post("/{village_id}/sections/import-all")
async def import_all_sections_data(
    village_id: int,
    year: Optional[int] = Query(None, description="年度"),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导入帮扶村所有区块数据（Excel）"""
    _get_village_or_404(db, village_id, current_user)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
        sheets_imported = len(wb.sheetnames)
        total_rows = sum(max(0, ws.max_row - 1) for ws in [wb[s] for s in wb.sheetnames])
        return {
            "code": 200,
            "data": {
                "sheets": sheets_imported,
                "sections": [{"name": s, "rows": max(0, wb[s].max_row - 1)} for s in wb.sheetnames],
                "imported": total_rows,
                "failed": 0,
                "year": year,
            },
            "message": f"导入成功，共 {sheets_imported} 个工作表，{total_rows} 行数据",
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")


# ── 转移支付资金 ──


class TransitionFundingItem(BaseModel):
    year: int
    militaryInvestment: float = 0
    localInvestment: float = 0
    totalInvestment: float = 0
    remarks: Optional[str] = None


class TransitionFundingRequest(BaseModel):
    items: List[TransitionFundingItem]


@router.get("/{village_id}/transition-funding")
async def get_transition_funding(
    village_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """获取转移支付资金按年度明细"""
    village = _get_village_or_404(db, village_id, current_user)
    items = []
    if village.transition_fund_items:
        try:
            items = json.loads(village.transition_fund_items)
        except (json.JSONDecodeError, TypeError):
            items = []
    return {"data": items, "success": True}


@router.post("/{village_id}/transition-funding")
async def save_transition_funding(
    village_id: int,
    data: TransitionFundingRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """保存转移支付资金数据"""
    village = _get_village_or_404(db, village_id, current_user)
    # Sum up military and local totals from the submitted items
    military = sum(item.militaryInvestment for item in data.items)
    local = sum(item.localInvestment for item in data.items)
    village.transition_fund_military_total = military
    village.transition_fund_local_total = local
    # Store per-year breakdown as JSON for retrieval
    items_json = [{
        "year": item.year,
        "militaryInvestment": item.militaryInvestment,
        "localInvestment": item.localInvestment,
        "totalInvestment": item.totalInvestment,
    } for item in data.items]
    village.transition_fund_items = json.dumps(items_json, ensure_ascii=False)
    safe_commit(db)
    return {"success": True, "message": "转移支付资金已保存", "items": items_json}


# ── 模板下载 ──


@router.get("/templates/all")
async def download_all_templates(
    current_user: User = Depends(get_current_user),
):
    """下载所有区块模板（Excel 多工作表）"""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sections = ["income", "industry", "infrastructure", "education", "medical", "employment", "population"]
    for s in sections:
        ws = wb.create_sheet(title=s)
        ws.append(["年份", "项目", "数值", "单位", "备注"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=all_templates.xlsx"},
    )

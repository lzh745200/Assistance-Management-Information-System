import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, field_validator
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.response import ok_list
from app.core.security import get_current_user
from app.models.audit import AuditAction, AuditLevel, AuditStatus
from app.services.audit_service import AuditService, SecurityEventService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/audit", tags=["Audit Logs"])


# ─── 请求体 Schema ───────────────────────────────────────────────────────────


class BatchDeleteRequest(BaseModel):
    """批量删除审计日志请求体。

    支持三种模式（优先级从高到低）：
    1. ids: 指定要删除的日志 ID 列表（自动将字符串类型 coerce 为 int）
    2. actions: 按操作类型列表批量删除
    3. action: 按单个操作类型删除（向后兼容）
    可与 before_date 组合使用缩小范围。
    """

    ids: Optional[List[int]] = None
    actions: Optional[List[str]] = None  # 支持多个操作类型（前端 action_types → actions）
    action: Optional[str] = None  # 单操作类型（向后兼容）
    before_date: Optional[str] = None

    @field_validator("ids", mode="before")
    @classmethod
    def coerce_ids_to_int(cls, v):
        """将 ids 中的字符串元素自动转换为 int，非法值抛出 ValueError。"""
        if v is None:
            return v
        result: List[int] = []
        for item in v:
            try:
                result.append(int(item))
            except (TypeError, ValueError):
                logger.error("BatchDeleteRequest: ids 中包含非整数值 %r", item)
                raise ValueError(f"ids 中包含非整数值: {item!r}，log_id 必须为整数")
        return result

    @field_validator("actions", mode="before")
    @classmethod
    def normalize_actions(cls, v):
        """兼容前端旧字段名 action_types（保留向后兼容，由路由层适配）。"""
        if v is None:
            return v
        return [str(a).strip() for a in v if a is not None]


# ─── 辅助函数 ─────────────────────────────────────────────────────────────────


def get_audit_service(db: Session = Depends(get_db)) -> AuditService:
    return AuditService(db)


def get_security_service(db: Session = Depends(get_db)) -> SecurityEventService:
    return SecurityEventService(db)


# ─── 路由：静态路径必须在动态路径（{log_id}）之前注册 ──────────────────────────
# FastAPI 按注册顺序匹配路由，DELETE /logs/batch 若在 DELETE /logs/{log_id}
# 之后注册，"batch" 会被当作整数解析，触发 422 "log_id 必须为整数" 错误。


@router.delete("/logs/batch")
async def batch_delete_audit_logs(
    body: BatchDeleteRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """批量删除审计日志（仅管理员）。

    支持三种删除策略：
    - ids: 按 ID 列表删除（优先级最高）
    - actions / action: 按操作类型（列表或单个）删除
    - before_date: 按日期范围删除（可与 actions 组合）
    """
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import AuditLog as AuditLogModel

    logger.info(
        "批量删除审计日志: user=%s ids=%s actions=%s action=%s before_date=%s",
        getattr(current_user, "username", "?"),
        body.ids,
        body.actions,
        body.action,
        body.before_date,
    )

    query = db.query(AuditLogModel)

    if body.ids:
        # 模式1：按 ID 列表删除（ids 已由 Pydantic coerce 为 int）
        query = query.filter(AuditLogModel.id.in_(body.ids))
    else:
        # 模式2：按操作类型 + 可选日期范围删除
        # 兼容 actions（列表）和 action（单值）
        effective_actions: Optional[List[str]] = None
        if body.actions:
            effective_actions = body.actions
        elif body.action:
            effective_actions = [body.action]

        if effective_actions:
            query = query.filter(AuditLogModel.action.in_(effective_actions))
        elif not body.before_date:
            # 无任何过滤条件时不删除任何记录，防止误删全表
            logger.info("批量删除: 无有效过滤条件，跳过删除")
            return {"message": "已删除 0 条日志记录", "deleted_count": 0}

        if body.before_date:
            try:
                dt = datetime.fromisoformat(body.before_date)
                query = query.filter(AuditLogModel.created_at < dt)
            except ValueError:
                logger.warning(
                    "批量删除: before_date 格式无效 %r，已忽略日期过滤",
                    body.before_date,
                )

    deleted = query.delete(synchronize_session=False)
    safe_commit(db)
    logger.info("批量删除审计日志完成: 删除 %d 条", deleted)
    return {"message": f"已删除 {deleted} 条日志记录", "deleted_count": deleted}


# ─── 注意：含路径参数的路由必须在所有同前缀静态路由之后注册 ───────────────────


@router.delete("/logs/{log_id}")
async def delete_audit_log(
    log_id: int,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """删除单条审计日志"""
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import AuditLog as AuditLogModel

    logger.info("删除审计日志 id=%d, user=%s", log_id, getattr(current_user, "username", "?"))
    log = db.query(AuditLogModel).filter(AuditLogModel.id == log_id).first()
    if not log:
        logger.warning("删除审计日志失败: id=%d 不存在", log_id)
        raise HTTPException(status_code=404, detail="Audit log not found")

    db.delete(log)
    safe_commit(db)
    logger.info("删除审计日志成功 id=%d", log_id)
    return {"message": "删除成功"}


@router.patch("/logs/{log_id}/remark")
async def update_audit_log_remark(
    log_id: int,
    body: dict,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """更新审计日志备注"""
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import AuditLog as AuditLogModel

    log = db.query(AuditLogModel).filter(AuditLogModel.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")

    remark = body.get("remark", "")
    # Store remark in metadata JSON field
    import json

    meta = log.metadata_ or {}
    if isinstance(meta, str):
        try:
            meta = json.loads(meta)
        except (json.JSONDecodeError, TypeError) as e:
            logger.warning(f"解析审计日志元数据失败 (id={log_id}): {e}")
            meta = {}
    meta["remark"] = remark
    log.metadata_ = meta
    safe_commit(db)
    return {"message": "备注更新成功", "id": log_id, "remark": remark}


@router.get("/logs/export")
async def export_audit_logs(  # noqa: C901
    action: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    format: str = Query("json", pattern="^(json|excel|csv)$"),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """导出审计日志为 JSON / Excel / CSV 格式

    支持三种导出格式：
    - json: 返回 JSON 数据（默认）
    - excel: 返回 .xlsx 文件下载
    - csv: 返回 .csv 文件下载
    需要管理员权限。
    """
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import AuditLog as AuditLogModel

    query = db.query(AuditLogModel).order_by(AuditLogModel.created_at.desc())
    if action:
        query = query.filter(AuditLogModel.action == action)
    if start_date:
        query = query.filter(AuditLogModel.created_at >= start_date)
    if end_date:
        query = query.filter(AuditLogModel.created_at <= end_date)

    logs = query.limit(5000).all()
    items = []
    for log in logs:
        items.append(
            {
                "id": log.id,
                "time": str(log.created_at) if log.created_at else "",
                "user": log.username or "",
                "action": log.action or "",
                "resource_type": log.resource_type or "",
                "detail": ((log.metadata_ or {}).get("remark", "") if isinstance(log.metadata_, dict) else ""),
                "status": log.status or "success",
                "ip": log.user_ip or "",
            }
        )

    # JSON 格式直接返回数据
    if format == "json":
        return ok_list(items=items, total=len(items))

    # Excel / CSV 格式返回文件下载
    import io
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "审计日志"

            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
            headers = ["ID", "时间", "用户", "操作", "资源类型", "详情", "状态", "IP地址"]
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for item in items:
                ws.append([
                    item["id"],
                    item["time"],
                    item["user"],
                    item["action"],
                    item["resource_type"],
                    item["detail"],
                    item["status"],
                    item["ip"],
                ])

            # 自动列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        logger.debug("Excel 列宽计算跳过单元格: %s", e)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"audit_report_{timestamp}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
            )
        except ImportError:
            logger.warning("openpyxl 未安装，回退到 CSV 格式")

    # CSV 格式
    import csv
    output = io.StringIO()
    output.write('\ufeff')  # BOM for Excel UTF-8 compatibility
    field_names = ["id", "time", "user", "action",
                   "resource_type", "detail", "status", "ip"]
    writer = csv.DictWriter(output, fieldnames=field_names)
    writer.writeheader()
    writer.writerows(items)

    csv_bytes = output.getvalue().encode("utf-8")
    filename = f"audit_report_{timestamp}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/logs")
async def get_audit_logs(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    status: Optional[str] = None,
    level: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    audit = AuditService(db)

    try:
        action_enum = AuditAction(action) if action else None
        status_enum = AuditStatus(status) if status else None
        level_enum = AuditLevel(level) if level else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid parameter value")

    result = audit.query_audit_logs(
        user_id=user_id,
        action=action_enum,
        resource_type=resource_type,
        start_date=start_date,
        end_date=end_date,
        status=status_enum,
        level=level_enum,
        page=page,
        page_size=page_size,
    )

    return result


@router.get("/logs/{log_id}")
async def get_audit_log_detail(log_id: int, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import AuditLog as AuditLogModel

    log = db.query(AuditLogModel).filter(AuditLogModel.id == log_id).first()

    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")

    return log


@router.get("/stats")
async def get_audit_stats(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    audit = AuditService(db)
    return audit.get_audit_stats(start_date=start_date, end_date=end_date)


@router.get("/actions")
async def get_available_actions():
    return {"actions": [a.value for a in AuditAction]}


@router.get("/levels")
async def get_available_levels():
    return {"levels": [level.value for level in AuditLevel]}


@router.get("/security/events")
async def get_security_events(
    severity: Optional[str] = None,
    event_type: Optional[str] = None,
    resolved: Optional[bool] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    service = SecurityEventService(db)
    return service.get_events(
        severity=severity,
        event_type=event_type,
        resolved=resolved,
        start_date=start_date,
        end_date=end_date,
        page=page,
        page_size=page_size,
    )


@router.get("/security/stats")
async def get_security_stats(current_user=Depends(get_current_user), db: Session = Depends(get_db)):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    service = SecurityEventService(db)
    return service.get_security_stats()


@router.post("/security/events/{event_id}/resolve")
async def resolve_security_event(
    event_id: int,
    resolution_notes: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    service = SecurityEventService(db)
    event = service.resolve_event(
        event_id=event_id,
        resolved_by=current_user.id,
        resolution_notes=resolution_notes,
    )

    if not event:
        raise HTTPException(status_code=404, detail="Security event not found")

    return {"message": "Event resolved", "event": event}


@router.get("/login-attempts")
async def get_login_attempts(
    username: Optional[str] = None,
    ip_address: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import LoginAttempt

    query = db.query(LoginAttempt)

    if username:
        query = query.filter(LoginAttempt.username == username)
    if ip_address:
        query = query.filter(LoginAttempt.ip_address == ip_address)
    if start_date:
        query = query.filter(LoginAttempt.attempt_time >= start_date)
    if end_date:
        query = query.filter(LoginAttempt.attempt_time <= end_date)

    total = query.count()
    attempts = query.order_by(LoginAttempt.attempt_time.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return ok_list(items=[att.to_dict() for att in attempts], total=total, page=page, page_size=page_size)


@router.get("/api-access")
async def get_api_access_logs(
    user_id: Optional[int] = None,
    endpoint: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import APIAccessLog

    query = db.query(APIAccessLog)

    if user_id:
        query = query.filter(APIAccessLog.user_id == user_id)
    if endpoint:
        query = query.filter(APIAccessLog.endpoint.contains(endpoint))
    if start_date:
        query = query.filter(APIAccessLog.created_at >= start_date)
    if end_date:
        query = query.filter(APIAccessLog.created_at <= end_date)

    total = query.count()
    logs = query.order_by(APIAccessLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return ok_list(items=[log.to_dict() for log in logs], total=total, page=page, page_size=page_size)


@router.get("/exports")
async def get_export_logs(
    user_id: Optional[int] = None,
    export_type: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if getattr(current_user, "role", None) not in ("admin", "super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")

    from app.models.audit import DataExportLog

    query = db.query(DataExportLog)

    if user_id:
        query = query.filter(DataExportLog.user_id == user_id)
    if export_type:
        query = query.filter(DataExportLog.export_type == export_type)
    if start_date:
        query = query.filter(DataExportLog.created_at >= start_date)
    if end_date:
        query = query.filter(DataExportLog.created_at <= end_date)

    total = query.count()
    logs = query.order_by(DataExportLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "items": [log.to_dict() for log in logs],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/user-activity/{user_id}")
async def get_user_activity(
    user_id: int,
    days: int = Query(7, ge=1, le=90),
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if (
        getattr(current_user, "role", None) not in ("admin", "super_admin")
        and getattr(current_user, "id", None) != user_id
    ):
        raise HTTPException(status_code=403, detail="Access denied")

    from datetime import timedelta

    from app.models.audit import AuditLog
from app.core.transaction import safe_commit

    start_date = datetime.now() - timedelta(days=days)

    logs = (
        db.query(AuditLog)
        .filter(AuditLog.user_id == user_id, AuditLog.created_at >= start_date)
        .order_by(AuditLog.created_at.desc())
        .all()
    )

    action_counts = {}
    for log in logs:
        action = log.action
        action_counts[action] = action_counts.get(action, 0) + 1

    return {
        "user_id": user_id,
        "period_days": days,
        "total_actions": len(logs),
        "action_breakdown": action_counts,
        "recent_activity": [log.to_dict() for log in logs[:20]],
    }

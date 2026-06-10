"""
帮扶村模板API端点
提供Excel模板下载和模板列表查询功能

Requirements: 13.4 - WHEN a user clicks the template download button
THEN THE Template_Service SHALL generate and download the corresponding
Excel file
"""

from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from pydantic import BaseModel

from app.core.security import get_current_user
from app.services.excel_template_service import ExcelTemplateService

router = APIRouter(prefix="/templates", tags=["模板管理"])


# 模板服务实例
template_service = ExcelTemplateService()


# 可用模板模块定义
AVAILABLE_MODULES = {
    "village": {
        "name": "帮扶村基础数据",
        "description": "帮扶村基础信息导入模板，包含部门、帮扶单位、村庄名称等字段",
        "filename": "帮扶村数据导入模板",
    },
    "population": {
        "name": "人口数据",
        "description": "帮扶村人口数据导入模板，包含户数、人口、常住人口等字段",
        "filename": "人口数据导入模板",
    },
    "income": {
        "name": "收入数据",
        "description": "帮扶村收入数据导入模板，包含人均纯收入、村集体收入等字段",
        "filename": "收入数据导入模板",
    },
    "funding": {
        "name": "经费投入",
        "description": "帮扶经费投入数据导入模板",
        "filename": "经费投入导入模板",
    },
    "force": {
        "name": "力量投入",
        "description": "力量投入情况数据导入模板，包含领导干部到村人次等字段",
        "filename": "力量投入导入模板",
    },
    "industry": {
        "name": "产业帮扶",
        "description": "产业帮扶情况数据导入模板",
        "filename": "产业帮扶导入模板",
    },
    "infrastructure": {
        "name": "基础设施",
        "description": "基础设施改善情况数据导入模板",
        "filename": "基础设施导入模板",
    },
    "party": {
        "name": "党建帮扶",
        "description": "党建帮扶情况数据导入模板",
        "filename": "党建帮扶导入模板",
    },
    "medical": {
        "name": "医疗帮扶",
        "description": "医疗帮扶情况数据导入模板",
        "filename": "医疗帮扶导入模板",
    },
    "consumption": {
        "name": "消费帮扶",
        "description": "消费帮扶情况数据导入模板",
        "filename": "消费帮扶导入模板",
    },
    "employment": {
        "name": "就业帮扶",
        "description": "就业帮扶情况数据导入模板",
        "filename": "就业帮扶导入模板",
    },
    "education": {
        "name": "教育帮扶",
        "description": "教育帮扶情况数据导入模板",
        "filename": "教育帮扶导入模板",
    },
}


class TemplateInfo(BaseModel):
    """模板信息响应模型"""

    module: str
    name: str
    description: str
    filename: str


class TemplateListResponse(BaseModel):
    """模板列表响应模型"""

    total: int
    templates: List[TemplateInfo]


@router.get("", response_model=TemplateListResponse)
async def list_templates(current_user=Depends(get_current_user)):
    """
    获取可用模板列表

    Returns:
        模板列表，包含模块名称、描述和文件名
    """
    templates = [
        TemplateInfo(
            module=module_key,
            name=module_info["name"],
            description=module_info["description"],
            filename=module_info["filename"],
        )
        for module_key, module_info in AVAILABLE_MODULES.items()
    ]

    return TemplateListResponse(total=len(templates), templates=templates)


@router.get("/{module}")
async def download_template(
    module: str,
    include_example: bool = Query(True, description="是否包含示例数据"),
    year: Optional[int] = Query(None, description="数据年份（可选）"),
    current_user=Depends(get_current_user),
):
    """
    下载指定模块的Excel模板

    Args:
        module: 模块名称（village, population, income, funding,
                force, industry, infrastructure, party, medical,
                consumption, employment, education）
        include_example: 是否包含示例数据，默认为True
        year: 数据年份（可选）

    Returns:
        Excel模板文件

    Raises:
        HTTPException: 当模块不存在时返回404错误
    """
    if module not in AVAILABLE_MODULES:
        raise HTTPException(
            status_code=404,
            detail=f"模板模块 '{module}' 不存在，可用模块: {', '.join(AVAILABLE_MODULES.keys())}",
        )

    module_info = AVAILABLE_MODULES[module]

    # 根据模块类型生成对应的模板
    if module == "village":
        content = template_service.generate_village_template(include_example=include_example)
    else:
        # 对于其他模块，使用通用模板生成方法
        content = _generate_module_template(module, include_example, year)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    year_suffix = f"_{year}" if year else ""
    filename = f"{module_info['filename']}{year_suffix}_{timestamp}.xlsx"

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


def _generate_module_template(module: str, include_example: bool, year: Optional[int]) -> bytes:
    """
    生成指定模块的Excel模板

    Args:
        module: 模块名称
        include_example: 是否包含示例数据
        year: 数据年份

    Returns:
        Excel文件的字节内容
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # 模块字段定义
    MODULE_FIELDS = {
        "population": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "households",
                "label": "户数",
                "required": False,
                "type": "int",
                "example": "100",
                "description": "总户数",
            },
            {
                "name": "population",
                "label": "人数",
                "required": False,
                "type": "int",
                "example": "500",
                "description": "总人数",
            },
            {
                "name": "resident_population",
                "label": "常住人口",
                "required": False,
                "type": "int",
                "example": "400",
                "description": "常住人口数",
            },
            {
                "name": "unstable_poverty_households",
                "label": "脱贫不稳定户数",
                "required": False,
                "type": "int",
                "example": "5",
                "description": "脱贫不稳定户数",
            },
            {
                "name": "unstable_poverty_population",
                "label": "脱贫不稳定人数",
                "required": False,
                "type": "int",
                "example": "15",
                "description": "脱贫不稳定人数",
            },
            {
                "name": "marginal_poverty_households",
                "label": "边缘易致贫户数",
                "required": False,
                "type": "int",
                "example": "3",
                "description": "边缘易致贫户数",
            },
            {
                "name": "marginal_poverty_population",
                "label": "边缘易致贫人数",
                "required": False,
                "type": "int",
                "example": "10",
                "description": "边缘易致贫人数",
            },
            {
                "name": "sudden_difficulty_households",
                "label": "突发严重困难户数",
                "required": False,
                "type": "int",
                "example": "2",
                "description": "突发严重困难户数",
            },
            {
                "name": "sudden_difficulty_population",
                "label": "突发严重困难人数",
                "required": False,
                "type": "int",
                "example": "6",
                "description": "突发严重困难人数",
            },
            {
                "name": "village_secretary_veterans",
                "label": "村支书退役军人数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "村支书中退役军人数",
            },
            {
                "name": "village_committee_veterans",
                "label": "村委员退役军人数",
                "required": False,
                "type": "int",
                "example": "2",
                "description": "村委员中退役军人数",
            },
        ],
        "income": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "per_capita_income",
                "label": "村人均纯收入(万元)",
                "required": False,
                "type": "float",
                "example": "1.5",
                "description": "村人均纯收入，单位万元",
            },
            {
                "name": "county_per_capita_income",
                "label": "县区农村人均纯收入(万元)",
                "required": False,
                "type": "float",
                "example": "1.8",
                "description": "当年帮扶村所在县区农村群众人均纯收入，单位万元",
            },
            {
                "name": "collective_income",
                "label": "村集体收入(万元)",
                "required": False,
                "type": "float",
                "example": "50",
                "description": "村集体收入，单位万元",
            },
        ],
        "funding": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "military_investment",
                "label": "部队合计投入(万元)",
                "required": False,
                "type": "float",
                "example": "100",
                "description": "部队合计投入，单位万元",
            },
            {
                "name": "local_investment",
                "label": "协调地方投入(万元)",
                "required": False,
                "type": "float",
                "example": "50",
                "description": "协调地方投入，单位万元",
            },
        ],
        "force": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "senior_leader_visits",
                "label": "军以上领导干部到村人次",
                "required": False,
                "type": "int",
                "example": "5",
                "description": "军以上领导干部到村人次",
            },
            {
                "name": "unit_soldier_visits",
                "label": "帮扶单位官兵到村人次",
                "required": False,
                "type": "int",
                "example": "50",
                "description": "帮扶单位官兵到村人次",
            },
        ],
        "industry": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "industry_investment",
                "label": "产业帮扶投入(万元)",
                "required": False,
                "type": "float",
                "example": "30",
                "description": "产业帮扶投入，单位万元",
            },
            {
                "name": "planting_breeding_count",
                "label": "种植养殖项目数",
                "required": False,
                "type": "int",
                "example": "2",
                "description": "新增种植养殖项目数",
            },
            {
                "name": "workshop_count",
                "label": "帮扶车间数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "新增帮扶车间数",
            },
            {
                "name": "rural_tourism_count",
                "label": "乡村旅游项目数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "新增乡村旅游项目数",
            },
            {
                "name": "other_industry_count",
                "label": "其他产业项目数",
                "required": False,
                "type": "int",
                "example": "0",
                "description": "新增其他产业项目数",
            },
        ],
        "infrastructure": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "infrastructure_investment",
                "label": "基础设施投入(万元)",
                "required": False,
                "type": "float",
                "example": "50",
                "description": "基础设施投入，单位万元",
            },
            {
                "name": "road_km",
                "label": "乡村道路公里数",
                "required": False,
                "type": "float",
                "example": "5.5",
                "description": "新建乡村道路公里数",
            },
            {
                "name": "housing_renovation",
                "label": "住房改造户数",
                "required": False,
                "type": "int",
                "example": "10",
                "description": "住房改造户数",
            },
            {
                "name": "water_facilities",
                "label": "水利设施个数",
                "required": False,
                "type": "int",
                "example": "3",
                "description": "新建水利设施个数",
            },
            {
                "name": "cultural_plaza",
                "label": "文化广场个数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "新建文化广场个数",
            },
            {
                "name": "library_cafe",
                "label": "书屋网吧个数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "新建书屋网吧个数",
            },
        ],
        "party": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "party_building_investment",
                "label": "党建帮扶投入(万元)",
                "required": False,
                "type": "float",
                "example": "10",
                "description": "党建帮扶投入，单位万元",
            },
            {
                "name": "paired_branches",
                "label": "结对帮扶村党支部个数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "结对帮扶村党支部个数",
            },
            {
                "name": "party_instructors",
                "label": "部队兼职党建指导员人数",
                "required": False,
                "type": "int",
                "example": "2",
                "description": "部队兼职党建指导员人数",
            },
            {
                "name": "joint_activities",
                "label": "支部联建共促活动次数",
                "required": False,
                "type": "int",
                "example": "4",
                "description": "支部联建共促活动次数",
            },
            {
                "name": "civilization_activities",
                "label": "乡风文明建设活动次数",
                "required": False,
                "type": "int",
                "example": "6",
                "description": "乡风文明建设活动次数",
            },
        ],
        "medical": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "medical_investment",
                "label": "医疗帮扶投入(万元)",
                "required": False,
                "type": "float",
                "example": "20",
                "description": "医疗帮扶投入，单位万元",
            },
            {
                "name": "clinics_built",
                "label": "帮建乡村卫生院室个数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "帮建乡村卫生院室个数",
            },
            {
                "name": "patients_served",
                "label": "巡诊群众人次",
                "required": False,
                "type": "int",
                "example": "200",
                "description": "巡诊群众人次",
            },
        ],
        "consumption": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "village_products_purchase",
                "label": "采购带销帮扶村产品(万元)",
                "required": False,
                "type": "float",
                "example": "15",
                "description": "采购带销帮扶村产品，单位万元",
            },
            {
                "name": "other_products_purchase",
                "label": "采购带销帮扶村以外产品(万元)",
                "required": False,
                "type": "float",
                "example": "5",
                "description": "采购带销帮扶村以外产品，单位万元",
            },
            {
                "name": "sales_counters",
                "label": "营区销售专柜个数",
                "required": False,
                "type": "int",
                "example": "2",
                "description": "营区销售专柜个数",
            },
            {
                "name": "benefited_population",
                "label": "惠及脱贫地区群众人数",
                "required": False,
                "type": "int",
                "example": "100",
                "description": "惠及脱贫地区群众人数",
            },
        ],
        "employment": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "hired_population",
                "label": "聘用脱贫地区群众人数",
                "required": False,
                "type": "int",
                "example": "10",
                "description": "聘用脱贫地区群众人数",
            },
            {
                "name": "trained_population",
                "label": "实用技能培训人数人次",
                "required": False,
                "type": "int",
                "example": "50",
                "description": "实用技能培训人数人次",
            },
            {
                "name": "recommended_employment",
                "label": "推荐就业人数人次",
                "required": False,
                "type": "int",
                "example": "20",
                "description": "推荐就业人数人次",
            },
        ],
        "education": [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
            {
                "name": "education_investment",
                "label": "教育帮扶总投入(万元)",
                "required": False,
                "type": "float",
                "example": "25",
                "description": "教育帮扶总投入，单位万元",
            },
            {
                "name": "donated_schools",
                "label": "捐赠帮扶村学校所数",
                "required": False,
                "type": "int",
                "example": "1",
                "description": "捐赠帮扶村学校所数",
            },
            {
                "name": "aided_external_schools",
                "label": "援建帮扶村以外学校所数",
                "required": False,
                "type": "int",
                "example": "0",
                "description": "援建帮扶村以外学校所数",
            },
            {
                "name": "education_activities",
                "label": "开展助学兴教活动次数",
                "required": False,
                "type": "int",
                "example": "4",
                "description": "开展助学兴教活动次数",
            },
            {
                "name": "aided_students",
                "label": "资助困难学生人数",
                "required": False,
                "type": "int",
                "example": "15",
                "description": "资助困难学生人数",
            },
            {
                "name": "volunteer_counselors",
                "label": "官兵兼职校外辅导员人数",
                "required": False,
                "type": "int",
                "example": "3",
                "description": "官兵兼职校外辅导员人数",
            },
        ],
    }

    fields = MODULE_FIELDS.get(module, [])
    if not fields:
        # 如果没有定义字段，返回基础模板
        fields = [
            {
                "name": "village_name",
                "label": "帮扶村名称",
                "required": True,
                "type": "str",
                "example": "某某村",
                "description": "帮扶村名称（必填）",
            },
            {
                "name": "year",
                "label": "年份",
                "required": True,
                "type": "int",
                "example": str(year or 2024),
                "description": "数据年份（必填）",
            },
        ]

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    REQUIRED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = Workbook()

    # 创建数据表
    ws_data = wb.active
    ws_data.title = AVAILABLE_MODULES[module]["name"]

    # 写入表头
    for col_idx, field in enumerate(fields, 1):
        cell = ws_data.cell(row=1, column=col_idx)
        label = field["label"]
        if field["required"]:
            label = f"*{label}"
        cell.value = label
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

        # 设置列宽
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = max(len(label) * 2, 15)

    # 写入示例数据
    if include_example:
        for col_idx, field in enumerate(fields, 1):
            cell = ws_data.cell(row=2, column=col_idx)
            cell.value = field["example"]
            cell.fill = EXAMPLE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # 冻结首行
    ws_data.freeze_panes = "A2"

    # 创建说明表
    ws_help = wb.create_sheet("填写说明")

    # 标题
    ws_help.merge_cells("A1:D1")
    title_cell = ws_help["A1"]
    title_cell.value = f"{AVAILABLE_MODULES[module]['name']}导入模板填写说明"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 表头
    headers = ["字段名称", "是否必填", "数据类型", "说明"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_help.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 字段说明
    type_map = {"str": "文本", "int": "整数", "float": "数字", "bool": "是 / 否"}
    for row_idx, field in enumerate(fields, 4):
        ws_help.cell(row=row_idx, column=1).value = field["label"]
        ws_help.cell(row=row_idx, column=2).value = "是" if field["required"] else "否"
        ws_help.cell(row=row_idx, column=3).value = type_map.get(field["type"], "文本")
        ws_help.cell(row=row_idx, column=4).value = field["description"]

        for col_idx in range(1, 5):
            cell = ws_help.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if field["required"] and col_idx == 2:
                cell.fill = REQUIRED_FILL

    # 设置列宽
    ws_help.column_dimensions["A"].width = 35
    ws_help.column_dimensions["B"].width = 12
    ws_help.column_dimensions["C"].width = 12
    ws_help.column_dimensions["D"].width = 50

    # 添加注意事项
    note_row = len(fields) + 6
    ws_help.merge_cells(f"A{note_row}:D{note_row}")
    note_cell = ws_help.cell(row=note_row, column=1)
    note_cell.value = "注意事项："
    note_cell.font = Font(bold=True, size=12)

    notes = [
        "1. 带*号的字段为必填字段，不能为空",
        "2. 单次导入最多支持1000条记录",
        "3. 请勿修改表头名称和顺序",
        "4. 示例数据行（绿色背景）在导入时会被忽略",
    ]

    for idx, note in enumerate(notes, 1):
        cell = ws_help.cell(row=note_row + idx, column=1)
        cell.value = note
        ws_help.merge_cells(f"A{note_row + idx}:D{note_row + idx}")

    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

"""
批量导入优化器 — pandas 加速 Excel 导入

替换 openpyxl 逐行处理为 pandas 向量化操作，
配合 SQLAlchemy bulk_insert_mappings 大幅提升导入性能。

方案 #9 — 数据导入批量化
"""

import logging
from typing import Any, Callable, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# 单次批量插入的批次大小
BATCH_SIZE = 500

# pandas 类型映射（dtype=str 可避免科学记数法等转换问题）
PANDAS_READ_KWARGS = {
    "engine": "openpyxl",
    "dtype": str,
    "keep_default_na": False,
}


def read_excel_fast(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """使用 pandas 快速读取 Excel，比 openpyxl 逐行读取快 3-5 倍。

    Args:
        file_content: Excel 文件字节内容

    Returns:
        (列名列表, 行数据字典列表)
    """
    import pandas as pd
    from io import BytesIO

    try:
        df = pd.read_excel(BytesIO(file_content), **PANDAS_READ_KWARGS)
    except ImportError:
        logger.warning("pandas 不可用，回退到 openpyxl 逐行模式")
        return _read_excel_fallback(file_content)

    # 清理列名
    df.columns = [str(col).strip() for col in df.columns]
    headers = list(df.columns)

    # 向量化 NaN 替换
    df = df.fillna("")

    # 转为字典列表（列名已在 df.columns 中清理过，无需再次 strip）
    rows = df.to_dict(orient="records")

    return rows, headers


def _read_excel_fallback(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """openpyxl 回退读取（pandas 不可用时）"""
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 第一行是表头
    try:
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
    except StopIteration:
        wb.close()
        return [], []

    rows = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)})

    wb.close()
    return rows, headers


def validate_rows(
    rows: List[Dict[str, Any]],
    required_fields: Optional[List[str]] = None,
    max_lengths: Optional[Dict[str, int]] = None,
) -> List[Dict[str, Any]]:
    """逐行数据验证（必填字段 + 最大长度检查）。

    Args:
        rows: 数据行列表
        required_fields: 必填字段列表
        max_lengths: 字段名→最大长度映射

    Returns:
        验证错误列表（空列表 = 全部通过）
    """
    errors = []
    for i, row in enumerate(rows, start=2):  # Excel 行号从 2 开始（1 是表头）
        if required_fields:
            for field in required_fields:
                val = row.get(field, "")
                if val is None or str(val).strip() == "":
                    errors.append({
                        "row_number": i,
                        "field_name": field,
                        "error_code": "REQUIRED",
                        "message": f"必填字段「{field}」为空",
                    })
        if max_lengths:
            for field, max_len in max_lengths.items():
                val = str(row.get(field, ""))
                if len(val) > max_len:
                    errors.append({
                        "row_number": i,
                        "field_name": field,
                        "error_code": "MAX_LENGTH",
                        "message": f"字段「{field}」超出最大长度 {max_len}（当前 {len(val)}）",
                    })
    return errors


def batch_insert_optimized(
    db: Session,
    model_class: Any,
    rows: List[Dict[str, Any]],
    batch_size: int = BATCH_SIZE,
    preprocessor: Optional[Callable[[Dict[str, Any]], Dict[str, Any]]] = None,
) -> int:
    """使用 bulk_insert_mappings 批量插入。

    Args:
        db: 数据库会话
        model_class: SQLAlchemy 模型类
        rows: 行数据列表
        batch_size: 每批次大小
        preprocessor: 可选的预处理函数（如字段映射、类型转换）

    Returns:
        成功插入的行数
    """
    total = 0

    for i in range(0, len(rows), batch_size):
        batch = rows[i: i + batch_size]
        if preprocessor:
            batch = [preprocessor(row) for row in batch]

        db.bulk_insert_mappings(model_class, batch)
        db.flush()
        total += len(batch)
        logger.debug(f"  批量插入: {total}/{len(rows)} 条")

    return total

"""
Excel导入服务
提供帮扶村数据的Excel批量导入功能

Requirements: 1.4, 1.6, 1.7, 1.10 - 批量创建帮扶村记录，支持事务回滚，增量 / 全量模式，1000条限制
"""

import logging
from dataclasses import dataclass, field
from datetime import timezone, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models.fund import Fund
from app.models.import_history import ImportHistory, ImportMode, ImportStatus
from app.models.organization import Organization
from app.models.project import Project
from app.models.school import School
from app.models.supported_village import SupportedVillage
from app.services.data_validator_service import (
    DataValidatorService,
    ValidationErrorCode,
)
from app.services.entity_import_validator import EntityImportValidator

logger = logging.getLogger(__name__)

# 尝试导入 pandas 加速模块（可选依赖）
try:
    from app.services.batch_import_optimizer import read_excel_fast as _pandas_read
    _HAS_PANDAS_FAST_READ = True
except ImportError:
    _HAS_PANDAS_FAST_READ = False

# 注意: ImportMode 枚举已在 app.models.import_history 中定义，此处统一使用该定义，避免重复


@dataclass
class ImportError:
    """导入错误"""

    row_number: int
    field_name: str
    error_code: str
    message: str
    value: Optional[Any] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_number": self.row_number,
            "field_name": self.field_name,
            "error_code": self.error_code,
            "message": self.message,
            "value": str(self.value) if self.value is not None else None,
        }


@dataclass
class ImportResult:
    """导入结果"""

    success: bool
    total_rows: int = 0
    success_rows: int = 0
    failed_rows: int = 0
    skipped_rows: int = 0  # 增量模式下跳过的重复记录
    errors: List[ImportError] = field(default_factory=list)
    created_ids: List[int] = field(default_factory=list)
    import_history_id: Optional[int] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "total_rows": self.total_rows,
            "success_rows": self.success_rows,
            "failed_rows": self.failed_rows,
            "skipped_rows": self.skipped_rows,
            "error_count": len(self.errors),
            "errors": [e.to_dict() for e in self.errors],
            "created_ids": self.created_ids,
            "import_history_id": self.import_history_id,
        }


class ExcelImporterService:
    """
    Excel导入服务

    提供帮扶村数据的Excel批量导入功能，支持：
    - Excel文件解析
    - 增量导入模式（跳过已存在记录）
    - 全量覆盖模式（删除现有记录后导入）
    - 事务回滚机制
    - 1000条记录限制
    """

    # 最大导入行数
    MAX_ROWS = 1000

    # 示例数据行的标识（绿色背景行）
    EXAMPLE_ROW_INDEX = 2  # Excel中的第2行是示例数据

    def __init__(self, db: Session):
        self.db = db
        self.validator = DataValidatorService()

    def parse_excel(
        self, file_content: bytes, entity_type: str = "supported_village"
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        解析Excel文件内容

        优先使用 pandas 加速读取（比 openpyxl 快 3-5 倍），
        pandas 不可用时自动回退到 openpyxl 逐行模式。

        Args:
            file_content: Excel文件的字节内容
            entity_type: 实体类型

        Returns:
            Tuple[List[Dict], List[str]]: (解析后的数据行列表, 表头列表)
        """
        # 快速路径：pandas 向量化读取
        if _HAS_PANDAS_FAST_READ:
            try:
                return _pandas_read(file_content)
            except Exception as e:
                logger.warning("pandas 快速读取失败，回退到 openpyxl: %s", e)

        # 回退路径：openpyxl 逐行读取
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active

        rows = []
        headers = []
        header_mapping = {}

        # 根据实体类型选择验证器
        if entity_type == "supported_village":
            header_parser = self.validator.parse_excel_headers
            example_markers = ["某某部门", "示例部门"]
        else:
            entity_validator = EntityImportValidator(entity_type)
            header_parser = entity_validator.parse_excel_headers
            example_markers = ["某某村", "示例名称", "某某希望小学", "某某帮扶单位"]

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # 解析表头
                headers = [str(cell).strip() if cell else "" for cell in row]
                header_mapping = header_parser(headers)
                continue

            # 跳过示例数据行（第2行）
            if row_idx == self.EXAMPLE_ROW_INDEX:
                first_value = row[0] if len(row) > 0 else None
                if first_value and str(first_value).strip() in example_markers:
                    continue

            # 检查是否为空行
            if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                continue

            # 将行数据转换为字典
            row_data = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx in header_mapping:
                    field_name = header_mapping[col_idx]
                    row_data[field_name] = cell_value

            if row_data:
                rows.append(row_data)

        wb.close()
        return rows, headers

    def import_data(
        self,
        file_content: bytes,
        file_name: str,
        file_size: int,
        user_id: int,
        mode: ImportMode = ImportMode.INCREMENTAL,
        entity_type: str = "supported_village",
    ) -> ImportResult:
        """
        导入Excel数据

        Args:
            file_content: Excel文件的字节内容
            file_name: 文件名
            file_size: 文件大小
            user_id: 操作用户ID
            mode: 导入模式（增量 / 全量）
            entity_type: 实体类型

        Returns:
            ImportResult: 导入结果
        """
        # 创建导入历史记录
        import_history = ImportHistory(
            user_id=user_id,
            file_name=file_name,
            file_size=file_size,
            import_mode=mode.value,
            entity_type=entity_type,
            status=ImportStatus.PROCESSING.value,
            started_at=datetime.now(timezone.utc),
        )
        self.db.add(import_history)
        self.db.flush()  # 获取ID但不提交

        result = ImportResult(success=False, import_history_id=import_history.id)

        try:
            # 解析Excel
            rows, headers = self.parse_excel(file_content, entity_type=entity_type)
            result.total_rows = len(rows)

            # 检查行数限制
            if len(rows) > self.MAX_ROWS:
                result.errors.append(
                    ImportError(
                        row_number=0,
                        field_name="",
                        error_code=ValidationErrorCode.ROW_LIMIT_EXCEEDED.value,
                        message=f"数据行数超过限制，单次最多导入 {self.MAX_ROWS} 条记录，当前 {len(rows)} 条",
                    )
                )
                self._update_import_history(import_history, result, ImportStatus.FAILED)
                self.db.commit()
                return result

            # 根据实体类型选择验证器和导入逻辑
            if entity_type == "supported_village":
                validation_result = self.validator.validate_batch(rows)
                duplicate_errors = self.validator.check_duplicates(rows)
                self.validator.convert_row_types
            else:
                entity_validator = EntityImportValidator(entity_type)
                validation_result = entity_validator.validate_batch(rows)
                duplicate_errors = entity_validator.check_duplicates(rows)
                entity_validator.convert_row_types

            # 验证数据
            if not validation_result.is_valid:
                for error in validation_result.errors:
                    result.errors.append(
                        ImportError(
                            row_number=error.row_number,
                            field_name=error.field_name,
                            error_code=error.error_code.value,
                            message=error.message,
                            value=error.value,
                        )
                    )
                result.failed_rows = len(rows)
                self._update_import_history(import_history, result, ImportStatus.FAILED)
                self.db.commit()
                return result

            # 检查重复数据
            if duplicate_errors:
                for error in duplicate_errors:
                    result.errors.append(
                        ImportError(
                            row_number=error.row_number,
                            field_name=error.field_name,
                            error_code=error.error_code.value,
                            message=error.message,
                            value=error.value,
                        )
                    )
                result.failed_rows = len(rows)
                self._update_import_history(import_history, result, ImportStatus.FAILED)
                self.db.commit()
                return result

            # 执行导入
            if entity_type == "supported_village":
                if mode == ImportMode.FULL:
                    result = self._import_full_mode(rows, result, import_history)
                else:
                    result = self._import_incremental_mode(rows, result, import_history)
            elif entity_type == "project":
                result = self._import_projects(rows, result, import_history, mode)
            elif entity_type == "fund":
                result = self._import_funds(rows, result, import_history, mode)
            elif entity_type == "school":
                result = self._import_schools(rows, result, import_history, mode)

            # 更新导入历史
            status = ImportStatus.COMPLETED if result.success else ImportStatus.FAILED
            self._update_import_history(import_history, result, status)

            self.db.commit()
            return result

        except SQLAlchemyError as e:
            # 数据库错误，回滚事务
            self.db.rollback()
            result.errors.append(
                ImportError(
                    row_number=0,
                    field_name="",
                    error_code=ValidationErrorCode.DATABASE_ERROR.value,
                    message=f"数据库错误: {str(e)}",
                )
            )
            result.failed_rows = result.total_rows

            # 重新创建导入历史记录（因为回滚了）
            import_history = ImportHistory(
                user_id=user_id,
                file_name=file_name,
                file_size=file_size,
                import_mode=mode.value,
                entity_type=entity_type,
                status=ImportStatus.FAILED.value,
                started_at=datetime.now(timezone.utc),
                completed_at=datetime.now(timezone.utc),
                total_rows=result.total_rows,
                success_rows=0,
                failed_rows=result.total_rows,
                error_details=[e.to_dict() for e in result.errors],
            )
            self.db.add(import_history)
            self.db.commit()
            result.import_history_id = import_history.id

            return result

        except Exception as e:
            # 其他错误
            self.db.rollback()
            result.errors.append(
                ImportError(
                    row_number=0,
                    field_name="",
                    error_code="IMPORT_999",
                    message=f"导入失败: {str(e)}",
                )
            )
            result.failed_rows = result.total_rows
            return result

    def _import_incremental_mode(
        self,
        rows: List[Dict[str, Any]],
        result: ImportResult,
        import_history: ImportHistory,
    ) -> ImportResult:
        """
        增量导入模式
        跳过已存在的记录（根据village_name判断）
        """
        # 批量查询：仅检查本次导入涉及的名称，避免全表扫描
        from sqlalchemy import func as sql_func
        imported_names = list({
            row.get("village_name", "").lower()
            for row in rows
            if row.get("village_name", "")
        })
        existing_names = set()
        if imported_names:
            existing = (
                self.db.query(sql_func.lower(SupportedVillage.village_name))
                .filter(sql_func.lower(SupportedVillage.village_name).in_(imported_names))
                .all()
            )
            existing_names = {r[0] for r in existing if r[0]}

        for row_idx, row in enumerate(rows, 1):
            # 转换数据类型
            converted_row = self.validator.convert_row_types(row)

            village_name = converted_row.get("village_name", "")
            if village_name and village_name.lower() in existing_names:
                # 跳过已存在的记录
                result.skipped_rows += 1
                continue

            # 创建新记录
            try:
                village = self._create_village(converted_row)
                result.created_ids.append(village.id)
                result.success_rows += 1

                # 添加到已存在集合，避免同批次重复
                if village_name:
                    existing_names.add(village_name.lower())

            except Exception as e:
                result.errors.append(
                    ImportError(
                        row_number=row_idx,
                        field_name="",
                        error_code=ValidationErrorCode.DATABASE_ERROR.value,
                        message=f"创建记录失败: {str(e)}",
                    )
                )
                result.failed_rows += 1

        result.success = result.failed_rows == 0
        return result

    def _import_full_mode(
        self,
        rows: List[Dict[str, Any]],
        result: ImportResult,
        import_history: ImportHistory,
    ) -> ImportResult:
        """
        全量覆盖模式
        删除现有所有记录后导入新数据
        """
        # 删除现有所有记录
        self.db.query(SupportedVillage).delete()

        for row_idx, row in enumerate(rows, 1):
            # 转换数据类型
            converted_row = self.validator.convert_row_types(row)

            try:
                village = self._create_village(converted_row)
                result.created_ids.append(village.id)
                result.success_rows += 1

            except Exception as e:
                result.errors.append(
                    ImportError(
                        row_number=row_idx,
                        field_name="",
                        error_code=ValidationErrorCode.DATABASE_ERROR.value,
                        message=f"创建记录失败: {str(e)}",
                    )
                )
                result.failed_rows += 1

        result.success = result.failed_rows == 0
        return result

    def _create_village(self, data: Dict[str, Any]) -> SupportedVillage:
        """
        创建帮扶村记录

        Args:
            data: 转换后的数据字典

        Returns:
            SupportedVillage: 创建的记录
        """
        # 过滤掉None值和非模型字段
        village_fields = {
            "sequence_no",
            "department",
            "support_unit",
            "village_name",
            "region_scope",
            "is_three_regions",
            "is_border_area",
            "is_ethnic_area",
            "is_revolutionary_area",
            "is_key_county",
            "revitalization_tier",
            "is_provincial_demo",
            "is_hundred_village_demo",
            "is_tiered_development",
            "is_cross_province",
            "is_cross_city",
            "is_cross_unit_cooperation",
            "is_in_overall_plan",
            "honors",
        }

        filtered_data = {k: v for k, v in data.items() if k in village_fields and v is not None}

        # 设置布尔字段的默认值
        bool_fields = [
            "is_three_regions",
            "is_border_area",
            "is_ethnic_area",
            "is_revolutionary_area",
            "is_key_county",
            "is_provincial_demo",
            "is_hundred_village_demo",
            "is_tiered_development",
            "is_cross_province",
            "is_cross_city",
            "is_cross_unit_cooperation",
            "is_in_overall_plan",
        ]
        for bool_field in bool_fields:
            if bool_field not in filtered_data:
                filtered_data[bool_field] = False

        village = SupportedVillage(**filtered_data)
        self.db.add(village)
        self.db.flush()  # 获取ID
        return village

    def _update_import_history(self, import_history: ImportHistory, result: ImportResult, status: ImportStatus):
        """更新导入历史记录"""
        import_history.status = status.value
        import_history.total_rows = result.total_rows
        import_history.success_rows = result.success_rows
        import_history.failed_rows = result.failed_rows
        import_history.error_details = [e.to_dict() for e in result.errors] if result.errors else None
        import_history.completed_at = datetime.now(timezone.utc)

    # ============== 项目导入 ==============

    def _import_projects(
        self,
        rows: List[Dict[str, Any]],
        result: ImportResult,
        import_history: ImportHistory,
        mode: ImportMode,
    ) -> ImportResult:
        validator = EntityImportValidator("project")
        existing_names = set()
        if mode == ImportMode.INCREMENTAL:
            existing_names = set(name[0].strip().lower() for name in self.db.query(Project.name).all() if name[0])
        else:
            self.db.query(Project).delete()

        # 预加载组织编码映射
        org_map = {o.code: o.id for o in self.db.query(Organization).all() if o.code}

        for row_idx, row in enumerate(rows, 1):
            converted = validator.convert_row_types(row)
            name = converted.get("name", "")
            if mode == ImportMode.INCREMENTAL and name and name.lower() in existing_names:
                result.skipped_rows += 1
                continue

            try:
                project_data = {
                    "name": converted.get("name"),
                    "code": converted.get("code"),
                    "type": converted.get("type") or "other",
                    "status": converted.get("status") or "draft",
                    "description": converted.get("description"),
                    "budget": converted.get("budget") or 0,
                    "start_date": converted.get("start_date"),
                    "end_date": converted.get("end_date"),
                    "leader": converted.get("leader"),
                    "contact_phone": converted.get("contact_phone"),
                    "responsible_unit": converted.get("responsible_unit"),
                }
                org_code = converted.get("organization_code")
                if org_code and org_code in org_map:
                    project_data["organization_id"] = org_map[org_code]

                project = Project(**project_data)
                self.db.add(project)
                self.db.flush()
                result.created_ids.append(project.id)
                result.success_rows += 1
                if name:
                    existing_names.add(name.lower())
            except Exception as e:
                result.errors.append(
                    ImportError(
                        row_number=row_idx,
                        field_name="",
                        error_code=ValidationErrorCode.DATABASE_ERROR.value,
                        message=f"创建项目失败: {str(e)}",
                    )
                )
                result.failed_rows += 1

        result.success = result.failed_rows == 0
        return result

    # ============== 资金导入 ==============

    def _import_funds(
        self,
        rows: List[Dict[str, Any]],
        result: ImportResult,
        import_history: ImportHistory,
        mode: ImportMode,
    ) -> ImportResult:
        validator = EntityImportValidator("fund")
        existing_names = set()
        if mode == ImportMode.INCREMENTAL:
            existing_names = set(name[0].strip().lower() for name in self.db.query(Fund.name).all() if name[0])
        else:
            self.db.query(Fund).delete()

        # 预加载项目编号映射
        project_map = {p.code: p.id for p in self.db.query(Project).all() if p.code}

        for row_idx, row in enumerate(rows, 1):
            converted = validator.convert_row_types(row)
            name = converted.get("name", "")
            if mode == ImportMode.INCREMENTAL and name and name.lower() in existing_names:
                result.skipped_rows += 1
                continue

            try:
                fund_data = {
                    "name": converted.get("name"),
                    "code": converted.get("code"),
                    "amount": converted.get("amount") or 0,
                    "fund_type": converted.get("fund_type") or "other",
                    "fund_source": converted.get("fund_source") or "other",
                    "purpose": converted.get("purpose"),
                    "status": converted.get("status") or "pending",
                    "operator": converted.get("operator"),
                    "receiver": converted.get("receiver"),
                    "date": converted.get("date"),
                    "remarks": converted.get("remarks"),
                }
                project_code = converted.get("project_code")
                if project_code and project_code in project_map:
                    fund_data["project_id"] = project_map[project_code]

                fund = Fund(**fund_data)
                self.db.add(fund)
                self.db.flush()
                result.created_ids.append(fund.id)
                result.success_rows += 1
                if name:
                    existing_names.add(name.lower())
            except Exception as e:
                result.errors.append(
                    ImportError(
                        row_number=row_idx,
                        field_name="",
                        error_code=ValidationErrorCode.DATABASE_ERROR.value,
                        message=f"创建资金记录失败: {str(e)}",
                    )
                )
                result.failed_rows += 1

        result.success = result.failed_rows == 0
        return result

    # ============== 学校导入 ==============

    def _import_schools(
        self,
        rows: List[Dict[str, Any]],
        result: ImportResult,
        import_history: ImportHistory,
        mode: ImportMode,
    ) -> ImportResult:
        validator = EntityImportValidator("school")
        existing_names = set()
        if mode == ImportMode.INCREMENTAL:
            existing_names = set(name[0].strip().lower() for name in self.db.query(School.name).all() if name[0])
        else:
            self.db.query(School).delete()

        for row_idx, row in enumerate(rows, 1):
            converted = validator.convert_row_types(row)
            name = converted.get("name", "")
            if mode == ImportMode.INCREMENTAL and name and name.lower() in existing_names:
                result.skipped_rows += 1
                continue

            try:
                school_data = {
                    "name": converted.get("name"),
                    "code": converted.get("code"),
                    "type": converted.get("type") or "other",
                    "province": converted.get("province"),
                    "city": converted.get("city"),
                    "district": converted.get("district"),
                    "address": converted.get("address"),
                    "principal": converted.get("principal"),
                    "contact_phone": converted.get("contact_phone"),
                    "student_count": converted.get("student_count") or 0,
                    "teacher_count": converted.get("teacher_count") or 0,
                    "support_unit": converted.get("support_unit"),
                    "support_status": converted.get("support_status") or "inactive",
                    "description": converted.get("description"),
                    "is_active": True,
                }
                school = School(**school_data)
                self.db.add(school)
                self.db.flush()
                result.created_ids.append(school.id)
                result.success_rows += 1
                if name:
                    existing_names.add(name.lower())
            except Exception as e:
                result.errors.append(
                    ImportError(
                        row_number=row_idx,
                        field_name="",
                        error_code=ValidationErrorCode.DATABASE_ERROR.value,
                        message=f"创建学校记录失败: {str(e)}",
                    )
                )
                result.failed_rows += 1

        result.success = result.failed_rows == 0
        return result

    async def import_data_async(
        self,
        file_bytes: bytes,
        filename: str,
        content_type: str,
        user_id: int,
        mode: ImportMode = ImportMode.INCREMENTAL,
        entity_type: str = "supported_village",
    ) -> ImportResult:
        """
        异步导入Excel数据（可用于API端点、CLI或后台任务）

        Args:
            file_bytes: 文件内容字节
            filename: 原始文件名
            content_type: MIME 类型
            user_id: 操作用户ID
            mode: 导入模式
            entity_type: 实体类型

        Returns:
            ImportResult: 导入结果
        """
        # 验证文件格式
        is_valid, error_msg = self.validator.validate_file_format(filename or "")
        if not is_valid:
            return ImportResult(
                success=False,
                errors=[
                    ImportError(
                        row_number=0,
                        field_name="file",
                        error_code=ValidationErrorCode.INVALID_FILE_FORMAT.value,
                        message=error_msg or "文件格式错误",
                    )
                ],
            )

        # 文件内容已在调用方读取
        file_size = len(file_bytes)

        # 验证文件大小
        is_valid, error_msg = self.validator.validate_file_size(file_size)
        if not is_valid:
            return ImportResult(
                success=False,
                errors=[
                    ImportError(
                        row_number=0,
                        field_name="file",
                        error_code=ValidationErrorCode.FILE_TOO_LARGE.value,
                        message=error_msg or "文件过大",
                    )
                ],
            )

        # 执行导入
        return self.import_data(
            file_content=file_bytes,
            file_name=filename or "unknown.xlsx",
            file_size=file_size,
            user_id=user_id,
            mode=mode,
            entity_type=entity_type,
        )

    def get_import_history(
        self, user_id: Optional[int] = None, page: int = 1, page_size: int = 20
    ) -> Tuple[List[ImportHistory], int]:
        """
        获取导入历史记录

        Args:
            user_id: 用户ID（可选，不传则获取所有）
            page: 页码
            page_size: 每页数量

        Returns:
            Tuple[List[ImportHistory], int]: (历史记录列表, 总数)
        """
        query = self.db.query(ImportHistory)

        if user_id:
            query = query.filter(ImportHistory.user_id == user_id)

        total = query.count()

        histories = (
            query.order_by(ImportHistory.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
        )

        return histories, total

    def get_import_history_by_id(self, history_id: int) -> Optional[ImportHistory]:
        """
        根据ID获取导入历史记录

        Args:
            history_id: 历史记录ID

        Returns:
            ImportHistory: 历史记录或None
        """
        return self.db.query(ImportHistory).filter(ImportHistory.id == history_id).first()

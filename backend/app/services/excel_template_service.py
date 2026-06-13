"""
Excel模板生成服务 — 军事正式风格 A4 打印版
提供帮扶村/项目/资金/学校数据导入模板的生成和下载功能

Requirements: 1.1, 1.9 - 标准 A4 打印模板，军事主题
"""

from io import BytesIO
from typing import Any, Dict, List
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

from app.services.entity_import_validator import EntityImportValidator

# ═══════════════════════════════════════════════════════════════
# 军事正式风格颜色常量
# ═══════════════════════════════════════════════════════════════
MILITARY_DARK_GREEN = "1b4332"
MILITARY_GOLD = "d4af37"
MILITARY_LIGHT_BG = "f0f4f0"
MILITARY_WHITE = "FFFFFF"
MILITARY_DARK_TEXT = "1e293b"
MILITARY_GRAY_TEXT = "64748b"
MILITARY_RED = "e63946"
MILITARY_ZEBRA_ODD = "FFFFFF"
MILITARY_ZEBRA_EVEN = "e8f0eb"
MILITARY_EXAMPLE_BG = "e2efda"
MILITARY_REQUIRED_BG = "fde8e8"

# ═══════════════════════════════════════════════════════════════
# 样式对象（复用避免重复创建）
# ═══════════════════════════════════════════════════════════════
_title_fill = PatternFill(start_color=MILITARY_DARK_GREEN, end_color=MILITARY_DARK_GREEN, fill_type="solid")
_title_font = Font(name="SimHei", bold=True, color=MILITARY_GOLD, size=22)
_subtitle_font = Font(name="SimHei", bold=True, color=MILITARY_WHITE, size=14)
_info_font = Font(name="SimSun", color=MILITARY_WHITE, size=10)
_header_fill = PatternFill(start_color=MILITARY_DARK_GREEN, end_color=MILITARY_DARK_GREEN, fill_type="solid")
_header_font = Font(name="SimHei", bold=True, color=MILITARY_WHITE, size=11)
_data_font = Font(name="SimSun", color=MILITARY_DARK_TEXT, size=10)
_example_fill = PatternFill(start_color=MILITARY_EXAMPLE_BG, end_color=MILITARY_EXAMPLE_BG, fill_type="solid")
_example_font = Font(name="SimSun", color=MILITARY_GRAY_TEXT, size=10, italic=True)
_required_fill = PatternFill(start_color=MILITARY_REQUIRED_BG, end_color=MILITARY_REQUIRED_BG, fill_type="solid")
_zebra_fills = [
    PatternFill(start_color=MILITARY_ZEBRA_ODD, end_color=MILITARY_ZEBRA_ODD, fill_type="solid"),
    PatternFill(start_color=MILITARY_ZEBRA_EVEN, end_color=MILITARY_ZEBRA_EVEN, fill_type="solid"),
]
_decor_line_fill = PatternFill(start_color=MILITARY_GOLD, end_color=MILITARY_GOLD, fill_type="solid")
_footer_font = Font(name="SimSun", color=MILITARY_GRAY_TEXT, size=9)
_section_font = Font(name="SimHei", bold=True, color=MILITARY_DARK_GREEN, size=12)
_note_font = Font(name="SimSun", color=MILITARY_GRAY_TEXT, size=10)

_thin_side = Side(style="thin", color="cbd5e1")
_thin_border = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
_bottom_gold = Border(
    left=_thin_side,
    right=_thin_side,
    top=_thin_side,
    bottom=Side(style="medium", color=MILITARY_GOLD),
)
_header_border = Border(
    left=_thin_side,
    right=_thin_side,
    top=Side(style="medium", color=MILITARY_GOLD),
    bottom=Side(style="medium", color=MILITARY_GOLD),
)
_center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

A4_PAGE_SETUP = {
    "paperSize": 9,  # A4
    "orientation": "landscape",
    "fitToWidth": 1,
    "fitToHeight": 0,
}

# 各模板的中文名称映射
TEMPLATE_NAMES = {
    "village": "帮扶村数据导入模板",
    "project": "帮扶项目数据导入模板",
    "fund": "经费台账数据导入模板",
    "school": "学校信息数据导入模板",
    "policy": "政策法规数据导入模板",
}


class ExcelTemplateService:
    """Excel模板生成服务 — 军事正式风格"""

    VILLAGE_FIELDS: List[Dict[str, Any]] = [
        {
            "name": "sequence_no",
            "label": "序号",
            "required": False,
            "type": "int",
            "example": "1",
            "description": "自动生成的序号",
        },
        {
            "name": "department",
            "label": "各部门各单位",
            "required": True,
            "type": "str",
            "example": "某某部门",
            "description": "所属部门名称（必填）",
        },
        {
            "name": "support_unit",
            "label": "具体帮扶单位",
            "required": True,
            "type": "str",
            "example": "某某帮扶单位",
            "description": "具体帮扶单位名称（必填）",
        },
        {
            "name": "village_name",
            "label": "定点帮扶村",
            "required": True,
            "type": "str",
            "example": "某某村",
            "description": "帮扶村名称（必填）",
        },
        {
            "name": "region_scope",
            "label": "地区范围",
            "required": False,
            "type": "str",
            "example": "西部地区",
            "description": "所属地区范围",
        },
        {
            "name": "is_three_regions",
            "label": "是否属于三区三州",
            "required": False,
            "type": "bool",
            "example": "是",
            "description": "是 / 否",
        },
        {
            "name": "is_border_area",
            "label": "是否属于边疆地区",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_ethnic_area",
            "label": "是否属于民族地区",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_revolutionary_area",
            "label": "是否属于革命地区",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_key_county",
            "label": "是否属于160个国家乡村振兴重点帮扶县",
            "required": False,
            "type": "bool",
            "example": "是",
            "description": "是 / 否",
        },
        {
            "name": "is_revitalization_tier",
            "label": "是否振兴梯队",
            "required": False,
            "type": "bool",
            "example": "是",
            "description": "是/否",
        },
        {
            "name": "is_provincial_demo",
            "label": "省级乡村振兴示范创建对象",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_hundred_village_demo",
            "label": "百村示范创建对象",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_tiered_development",
            "label": "梯次振兴发展对象",
            "required": False,
            "type": "bool",
            "example": "是",
            "description": "是 / 否",
        },
        {
            "name": "is_cross_province",
            "label": "是否跨省",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_cross_city",
            "label": "是否跨市",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_cross_unit_cooperation",
            "label": "是否跨大单位协作帮扶",
            "required": False,
            "type": "bool",
            "example": "否",
            "description": "是 / 否",
        },
        {
            "name": "is_in_overall_plan",
            "label": "是否纳入总盘子",
            "required": False,
            "type": "bool",
            "example": "是",
            "description": "是 / 否",
        },
        {
            "name": "honors",
            "label": "2021年以来获得的国家或省级表彰",
            "required": False,
            "type": "str",
            "example": "全国先进帮扶村",
            "description": "获得的表彰荣誉",
        },
    ]

    def __init__(self):
        pass

    # ══════════════════════════════════════════════════════════
    # 公共方法
    # ══════════════════════════════════════════════════════════

    def generate_village_template(self, include_example: bool = True) -> bytes:
        return self._build_template("village", include_example)

    def generate_project_template(self, include_example: bool = True) -> bytes:
        return self._build_template("project", include_example)

    def generate_fund_template(self, include_example: bool = True) -> bytes:
        return self._build_template("fund", include_example)

    def generate_school_template(self, include_example: bool = True) -> bytes:
        return self._build_template("school", include_example)

    def generate_policy_template(self, include_example: bool = True) -> bytes:
        return self._build_template("policy", include_example)

    # 政策字段定义（EntityImportValidator 不支持 policy，直接硬编码）
    POLICY_FIELDS: List[Dict[str, Any]] = [
        {"name": "sequence_no", "label": "序号", "required": False, "type": "int",
         "example": "1", "description": "自动生成的序号"},
        {"name": "title", "label": "政策标题", "required": True, "type": "str",
         "example": "关于进一步加强乡村振兴帮扶工作的指导意见", "description": "政策法规的完整标题（必填）"},
        {"name": "doc_number", "label": "政策文号", "required": False, "type": "str",
         "example": "国办发〔2024〕15号", "description": "政策发布文号，如无则留空"},
        {"name": "level", "label": "政策级别", "required": True, "type": "str",
         "example": "国家级", "description": "国家级/省级/市级/县级"},
        {"name": "issuing_authority", "label": "发布机关", "required": False, "type": "str",
         "example": "国务院办公厅", "description": "发布政策的具体机关名称"},
        {"name": "publish_date", "label": "发布日期", "required": False, "type": "date",
         "example": "2024-03-15", "description": "格式 YYYY-MM-DD"},
        {"name": "effective_date", "label": "生效日期", "required": False, "type": "date",
         "example": "2024-04-01", "description": "格式 YYYY-MM-DD"},
        {"name": "status", "label": "状态", "required": True, "type": "str",
         "example": "现行有效", "description": "现行有效/已修订/已废止/即将实施"},
        {"name": "keywords", "label": "关键词", "required": False, "type": "str",
         "example": "乡村振兴,帮扶工作,指导意见", "description": "多个关键词用逗号分隔"},
        {"name": "content", "label": "政策内容摘要", "required": False, "type": "str",
         "example": "为深入贯彻落实党中央关于乡村振兴战略部署...", "description": "政策正文摘要或全文"},
    ]

    def _build_template(self, entity_type: str, include_example: bool) -> bytes:
        """统一构建模板：标题页 + 数据表 + 填写说明"""
        title = TEMPLATE_NAMES.get(entity_type, f"{entity_type}数据导入模板")
        if entity_type == "village":
            fields = self.VILLAGE_FIELDS
            validator = None
        elif entity_type == "policy":
            fields = self.POLICY_FIELDS
            validator = None
        else:
            validator = EntityImportValidator(entity_type)
            fields = validator.get_field_definitions()

        wb = Workbook()
        ws = wb.active
        ws.title = "数据导入"
        self._setup_page(ws)
        data_start_row = self._write_title_header(ws, title, len(fields))
        self._write_data_table(ws, fields, data_start_row, include_example, validator)

        # 填写说明表
        ws_help = wb.create_sheet("填写说明")
        self._write_help_sheet(ws_help, title, fields)

        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        return output.getvalue()

    # ══════════════════════════════════════════════════════════
    # A4 页面设置
    # ══════════════════════════════════════════════════════════

    @staticmethod
    def _setup_page(ws):
        """设置 A4 打印参数"""
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(
            left=0.6,
            right=0.6,
            top=0.7,
            bottom=0.7,
            header=0.3,
            footer=0.3,
        )
        ws.sheet_properties.pageSetUpPr = None  # reset
        ws.print_title_rows = None

    # ══════════════════════════════════════════════════════════
    # 标题区
    # ══════════════════════════════════════════════════════════

    def _write_title_header(self, ws, title: str, col_count: int) -> int:
        """写入军事风格标题区，返回数据起始行号"""
        max_col = get_column_letter(max(col_count, 6))

        # ── 第1行：主标题（深绿底色 + 金色大字）──
        ws.merge_cells(f"A1:{max_col}1")
        c = ws["A1"]
        c.value = "★ 军队乡村振兴管理系统 ★"
        c.font = _title_font
        c.fill = _title_fill
        c.alignment = _center_align
        ws.row_dimensions[1].height = 42

        # ── 第2行：模板名称 ──
        ws.merge_cells(f"A2:{max_col}2")
        c = ws["A2"]
        c.value = title
        c.font = _subtitle_font
        c.fill = _title_fill
        c.alignment = _center_align
        ws.row_dimensions[2].height = 30

        # ── 第3行：金色装饰线 ──
        for col in range(1, col_count + 1):
            ws.cell(row=3, column=col).fill = _decor_line_fill
        ws.row_dimensions[3].height = 3

        # ── 第4行：信息栏 ──
        ws.merge_cells(f"A4:{max_col}4")
        today = date.today().strftime("%Y年%m月%d日")
        c = ws["A4"]
        c.value = f"填报单位：____________________    填报人：__________    填报日期：{today}    年度：__________"
        c.font = Font(name="SimSun", size=10, color=MILITARY_DARK_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[4].height = 24

        # ── 第5行：分隔空白 ──
        ws.row_dimensions[5].height = 6

        return 6  # 数据表从第6行开始

    # ══════════════════════════════════════════════════════════
    # 数据表
    # ══════════════════════════════════════════════════════════

    def _write_data_table(self, ws, fields, start_row: int, include_example: bool, validator=None):
        """写入数据表头 + 示例行 + 下拉验证"""
        col_count = len(fields)

        # ── 表头行 ──
        for col_idx, field in enumerate(fields, 1):
            label = field["label"]
            if field.get("required"):
                label = f"*{label}"
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = label
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = _center_align
            if field.get("required"):
                cell.border = _header_border
            else:
                cell.border = _thin_border
            # 列宽
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(label) * 2.2, 14)
        ws.row_dimensions[start_row].height = 30

        # ── 示例数据行 ──
        if include_example:
            ex_row = start_row + 1
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=ex_row, column=col_idx)
                cell.value = field.get("example", "")
                cell.font = _example_font
                cell.fill = _example_fill
                cell.alignment = _center_align
                cell.border = _thin_border
            ws.row_dimensions[ex_row].height = 22

            # 标注
            ex_label_col = get_column_letter(col_count + 1) if col_count < 26 else "Z"
            ws.merge_cells(f"{ex_label_col}{ex_row}:{get_column_letter(col_count + 2)}{ex_row}")
            note_cell = ws.cell(row=ex_row, column=col_count + 1)
            note_cell.value = "← 示例行（导入时自动跳过）"
            note_cell.font = Font(name="SimSun", color=MILITARY_GRAY_TEXT, size=9, italic=True)

        # ── 空数据行（预格式化3行供填写）──
        data_start = start_row + 2 if include_example else start_row + 1
        for row_idx in range(data_start, data_start + 3):
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = _thin_border
                cell.alignment = _center_align
                cell.font = _data_font
                # 斑马纹
                zebra_idx = (row_idx - data_start) % 2
                cell.fill = _zebra_fills[zebra_idx]
            ws.row_dimensions[row_idx].height = 22

        # ── 下拉验证 ──
        last_data_row = data_start + 999
        self._add_validation(ws, fields, data_start, last_data_row, validator)

        # ── 冻结表头 ──
        ws.freeze_panes = f"A{data_start}"

        # ── 底部填写说明 ──
        note_start = data_start + 4
        ws.merge_cells(f"A{note_start}:{get_column_letter(col_count)}{note_start}")
        ws.cell(row=note_start, column=1).value = (
            "【填写说明】1. 带 * 为必填字段  2. 日期格式：YYYY-MM-DD  3. 请勿修改表头  4. 单次最多导入1000条"
        )
        ws.cell(row=note_start, column=1).font = Font(name="SimSun", color=MILITARY_GRAY_TEXT, size=8)
        ws.row_dimensions[note_start].height = 18

        # ── 页脚 ──
        footer_row = note_start + 2
        ws.merge_cells(f"A{footer_row}:{get_column_letter(col_count)}{footer_row}")
        ws.cell(row=footer_row, column=1).value = (
            f"— 军队乡村振兴管理系统 v1.3.0 — {date.today().strftime('%Y-%m-%d')} —"
        )
        ws.cell(row=footer_row, column=1).font = _footer_font
        ws.cell(row=footer_row, column=1).alignment = _center_align

    # ══════════════════════════════════════════════════════════
    # 下拉验证
    # ══════════════════════════════════════════════════════════

    @staticmethod
    def _add_validation(ws, fields, start_row: int, end_row: int, validator=None):
        """为枚举字段添加下拉列表"""
        for col_idx, field in enumerate(fields, 1):
            ftype = field.get("type", "")
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

            if ftype == "bool":
                formula = '"是,否"'
            elif validator and ftype in getattr(validator, "ENUM_VALUES", {}):
                formula = f'"{",".join(validator.ENUM_VALUES[ftype])}"'
            elif ftype == "project_type":
                formula = '"基础设施,产业发展,公共服务,其他"'
            elif ftype == "project_status":
                formula = '"draft,planning,in_progress,completed,cancelled"'
            elif ftype == "fund_type":
                formula = '"project,operation,education,infrastructure,emergency,other"'
            elif ftype == "fund_source":
                formula = '"military,government,donation,enterprise,other"'
            elif ftype == "fund_status":
                formula = '"pending,planned,approved,allocated,in_use,completed,audited"'
            elif ftype == "school_type":
                formula = '"primary,middle,high,vocational,other"'
            elif ftype == "support_status":
                formula = '"active,inactive,completed"'
            else:
                continue

            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = "请从下拉列表中选择"
            dv.errorTitle = "输入错误"
            dv.add(cell_range)
            ws.add_data_validation(dv)

    # ══════════════════════════════════════════════════════════
    # 填写说明表
    # ══════════════════════════════════════════════════════════

    def _write_help_sheet(self, ws, title: str, fields):
        """写入填写说明"""
        col_count = 5
        max_col = get_column_letter(col_count)

        # 标题
        ws.merge_cells(f"A1:{max_col}1")
        c = ws["A1"]
        c.value = f"★ {title} — 填写说明 ★"
        c.font = Font(name="SimHei", bold=True, color=MILITARY_GOLD, size=16)
        c.fill = _title_fill
        c.alignment = _center_align
        ws.row_dimensions[1].height = 36

        # 金色装饰线
        for col in range(1, col_count + 1):
            ws.cell(row=2, column=col).fill = _decor_line_fill
        ws.row_dimensions[2].height = 3

        # 表头
        headers = ["序号", "字段名称", "是否必填", "数据类型", "填写说明"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = _center_align
            cell.border = _thin_border
        ws.row_dimensions[4].height = 26

        # 字段列表
        for row_idx, field in enumerate(fields):
            r = row_idx + 5
            ws.cell(row=r, column=1, value=row_idx + 1).font = _data_font
            ws.cell(row=r, column=2, value=field["label"]).font = _data_font
            req_text = "★ 必填" if field.get("required") else "可选"
            req_cell = ws.cell(row=r, column=3, value=req_text)
            req_cell.font = Font(
                name="SimSun", bold=True, color=MILITARY_RED if field.get("required") else MILITARY_GRAY_TEXT, size=10
            )
            ws.cell(row=r, column=4, value=self._get_type_label(field.get("type", "str"))).font = _data_font
            ws.cell(row=r, column=5, value=field.get("description", "")).font = _data_font
            for c_idx in range(1, col_count + 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = _thin_border
                cell.alignment = _left_align
            ws.row_dimensions[r].height = 22

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 48

        # 注意事项
        note_start = len(fields) + 7
        ws.merge_cells(f"A{note_start}:{max_col}{note_start}")
        ws.cell(row=note_start, column=1, value="▎注意事项").font = _section_font

        notes = [
            "1. 带 ★ 必填 标记的字段必须填写，否则导入失败",
            "2. 日期字段请使用 YYYY-MM-DD 格式（如 2024-06-15）",
            "3. 是 / 否字段请从下拉列表中选择",
            "4. 单次导入最多支持 1000 条记录",
            "5. 请勿修改表头名称和列顺序",
            "6. 示例数据行（浅绿底色）在导入时自动跳过",
            "7. 填写完成后可直接打印（已设定为 A4 纸张大小）",
        ]
        for idx, note in enumerate(notes, 1):
            row = note_start + idx
            ws.merge_cells(f"A{row}:{max_col}{row}")
            ws.cell(row=row, column=1, value=note).font = _note_font

        # 页脚
        footer_row = note_start + len(notes) + 2
        ws.merge_cells(f"A{footer_row}:{max_col}{footer_row}")
        ws.cell(row=footer_row, column=1).value = "— 军队乡村振兴管理系统 v1.3.0 —"
        ws.cell(row=footer_row, column=1).font = _footer_font
        ws.cell(row=footer_row, column=1).alignment = _center_align

    # ══════════════════════════════════════════════════════════
    # 工具方法
    # ══════════════════════════════════════════════════════════

    @staticmethod
    def _get_type_label(type_str: str) -> str:
        return {
            "str": "文本",
            "int": "整数",
            "float": "数字",
            "bool": "是 / 否",
            "date": "日期",
            "phone": "手机号",
            "county": "区县",
            "project_type": "项目类型",
            "project_status": "项目状态",
            "fund_type": "资金类型",
            "fund_source": "资金来源",
            "fund_status": "资金状态",
            "school_type": "学校类型",
            "support_status": "帮扶状态",
        }.get(type_str, "文本")

    def get_field_mapping(self) -> Dict[str, str]:
        return {f["label"]: f["name"] for f in self.VILLAGE_FIELDS}

    def get_required_fields(self) -> List[str]:
        return [f["name"] for f in self.VILLAGE_FIELDS if f["required"]]

    def get_field_types(self) -> Dict[str, str]:
        return {f["name"]: f["type"] for f in self.VILLAGE_FIELDS}

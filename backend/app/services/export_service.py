"""Excel 导出服务 - 实现真实的数据导出功能"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelExportService:
    """Excel 导出服务"""

    _HEADER_FONT = Font(bold=True, size=11)
    _HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    _HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
    _THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    _CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    def _create_workbook(self, sheet_name: str, headers: list, rows: list[dict]) -> Workbook:
        """创建通用 Excel 工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self._HEADER_FONT_WHITE
            cell.fill = self._HEADER_FILL
            cell.alignment = self._CENTER_ALIGN
            cell.border = self._THIN_BORDER

        # 写数据行
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
                cell.border = self._THIN_BORDER
                cell.alignment = self._CENTER_ALIGN

        # 自适应列宽
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(rows) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

        return wb

    @staticmethod
    def _to_bytes(wb: Workbook) -> bytes:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_user_list(self, data: list[dict]) -> bytes:
        headers = ["ID", "用户名", "邮箱", "姓名", "角色", "状态", "最后登录"]
        wb = self._create_workbook("用户列表", headers, data)
        return self._to_bytes(wb)

    def export_village_list(self, data: list[dict]) -> bytes:
        headers = ["ID", "名称", "编码", "省份", "城市", "区县", "人口", "状态", "创建时间"]
        wb = self._create_workbook("村庄列表", headers, data)
        return self._to_bytes(wb)

    def export_school_list(self, data: list[dict]) -> bytes:
        headers = ["ID", "名称", "编码", "类型", "城市", "学生数", "教师数", "状态"]
        wb = self._create_workbook("学校列表", headers, data)
        return self._to_bytes(wb)

    def export_project_list(self, data: list[dict]) -> bytes:
        headers = ["ID", "名称", "编码", "类型", "状态", "预算", "进度", "开始日期", "结束日期"]
        wb = self._create_workbook("项目列表", headers, data)
        return self._to_bytes(wb)

    def export_fund_list(self, data: list[dict]) -> bytes:
        headers = ["ID", "名称", "类型", "金额", "来源", "用途", "状态", "经办人", "使用日期"]
        wb = self._create_workbook("经费列表", headers, data)
        return self._to_bytes(wb)

    def export_comprehensive_report(
        self,
        summary: dict,
        village_data: list[dict],
        project_data: list[dict],
        fund_data: list[dict],
    ) -> bytes:
        """导出综合报表（多 sheet）"""
        wb = Workbook()

        # 汇总 sheet
        ws_summary = wb.active
        ws_summary.title = "汇总"
        for row_idx, (key, value) in enumerate(summary.items(), 1):
            cell_a = ws_summary.cell(row=row_idx, column=1, value=key)
            cell_a.font = Font(bold=True)
            cell_a.border = self._THIN_BORDER
            cell_b = ws_summary.cell(row=row_idx, column=2, value=value)
            cell_b.border = self._THIN_BORDER
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 30

        # 村庄 sheet
        if village_data:
            ws_v = wb.create_sheet("村庄")
            v_headers = ["ID", "名称", "人口", "项目数", "产业数"]
            for col_idx, h in enumerate(v_headers, 1):
                ws_v.cell(row=1, column=col_idx, value=h).font = self._HEADER_FONT
            for row_idx, row in enumerate(village_data, 2):
                for col_idx, h in enumerate(v_headers, 1):
                    ws_v.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

        # 项目 sheet
        if project_data:
            ws_p = wb.create_sheet("项目")
            p_headers = ["ID", "名称", "状态", "预算", "进度"]
            for col_idx, h in enumerate(p_headers, 1):
                ws_p.cell(row=1, column=col_idx, value=h).font = self._HEADER_FONT
            for row_idx, row in enumerate(project_data, 2):
                for col_idx, h in enumerate(p_headers, 1):
                    ws_p.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

        # 经费 sheet
        if fund_data:
            ws_f = wb.create_sheet("经费")
            f_headers = ["ID", "名称", "金额", "状态", "使用日期"]
            for col_idx, h in enumerate(f_headers, 1):
                ws_f.cell(row=1, column=col_idx, value=h).font = self._HEADER_FONT
            for row_idx, row in enumerate(fund_data, 2):
                for col_idx, h in enumerate(f_headers, 1):
                    ws_f.cell(row=row_idx, column=col_idx, value=row.get(h, ""))

        return self._to_bytes(wb)


export_service = ExcelExportService()

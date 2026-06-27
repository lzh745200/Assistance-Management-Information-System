"""Policy import from Excel — extracted from api/v1/policy.py to reduce C901 complexity."""

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.policy import Policy
from app.core.upload_security import validate_excel_upload
from app.core.logging import logger


async def import_policies_from_excel(
    file: UploadFile,
    db: Session,
    current_user=None,
) -> Dict[str, Any]:
    """从 Excel 导入政策

    返回格式: { imported: int, errors: list, errorRows: list, total: int }
    """
    try:
        content = await validate_excel_upload(file)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件校验失败: {str(e)}")

    try:
        wb = load_workbook(BytesIO(content))
        ws = wb.active

        level_map = {
            "国家级": "national",
            "省级": "provincial",
            "市级": "municipal",
            "县级": "county",
            "军队": "military",
            "中央军委": "central_military",
            "战区": "theater",
            "军": "army",
            "师": "division",
        }
        status_map = {
            "草稿": "draft",
            "有效": "active",
            "失效": "invalid",
            "已发布": "active",
            "已归档": "invalid",
            "已过期": "invalid",
        }

        imported = 0
        errors: List[Dict[str, Any]] = []
        error_rows: List[int] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            if not row[1]:
                errors.append({"row": row_idx, "title": "", "error": "政策标题不能为空"})
                error_rows.append(row_idx)
                continue

            try:
                result = _process_policy_row(row, row_idx, level_map, status_map, db)
                if result.get("error"):
                    errors.append(result)
                    error_rows.append(row_idx)
                elif result.get("duplicate"):
                    errors.append(result)
                    error_rows.append(row_idx)
                else:
                    imported += 1
            except (ValueError, TypeError, KeyError) as e:
                errors.append(_make_row_error(row, row_idx, str(e)))
                error_rows.append(row_idx)

        db.commit()

        return {
            "imported": imported,
            "errors": errors,
            "errorRows": error_rows,
            "total": imported + len(errors),
        }
    except Exception as e:
        db.rollback()
        logger.error(f"导入政策失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


def _process_policy_row(row, row_idx, level_map, status_map, db) -> Dict[str, Any]:
    """处理单行 Excel 数据，返回结果字典。"""
    issue_date = _parse_date_cell(row, 5, row_idx, "发布日期")
    effective_date = _parse_date_cell(row, 6, row_idx, "生效日期")

    level_str = _safe_str(row, 3)
    level_val = level_map.get(level_str, "national") if level_str else "national"

    status_str = _safe_str(row, 7)
    status_val = status_map.get(status_str, "active") if status_str else "active"

    category_val = "military" if level_val == "military" else "local"
    content_val = _safe_str(row, 9, default="")
    title_val = _safe_str(row, 1, default=f"未命名政策{row_idx}")
    code_val = _safe_str(row, 2)
    authority_val = _safe_str(row, 4)
    keywords_val = _safe_str(row, 8)

    if code_val:
        existing = db.query(Policy).filter(Policy.code == code_val).first()
        if existing:
            return {"row": row_idx, "title": title_val, "error": f"文号\u201c{code_val}\u201d已存在，已跳过"}

    policy = Policy(
        title=title_val,
        code=code_val,
        level=level_val,
        issuing_authority=authority_val,
        issue_date=issue_date,
        effective_date=effective_date,
        status=status_val,
        keywords=keywords_val,
        category=category_val,
        content=content_val,
    )
    nested = db.begin_nested()
    try:
        db.add(policy)
        db.flush()
        nested.commit()
    except Exception as row_err:
        nested.rollback()
        logger.warning(f"保存政策失败 (行{row_idx}): {row_err}")
        raise row_err
    return {"ok": True}


def _parse_date_cell(row, index: int, row_idx: int, label: str) -> Optional[datetime]:
    """安全解析 Excel 单元格中的日期值。"""
    if len(row) <= index:
        return None
    try:
        val = row[index]
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            try:
                return datetime.strptime(val.strip(), "%Y-%m-%d")
            except ValueError:
                return None
    except (TypeError, AttributeError) as e:
        logger.debug(f"解析{label}失败 (行{row_idx}): {e}")
    return None


def _safe_str(row, index: int, default: Optional[str] = None) -> Optional[str]:
    """安全提取字符串字段，索引越界或类型错误时返回默认值。"""
    try:
        val = row[index] if len(row) > index and row[index] is not None else None
        return str(val).strip() if val is not None else default
    except (TypeError, ValueError, IndexError) as e:
        logger.debug(f"提取字段失败 (索引{index}): {e}")
        return default


def _make_row_error(row, row_idx: int, msg: str) -> Dict[str, Any]:
    """构造行错误字典。"""
    title = _safe_str(row, 1, default=f"未命名政策{row_idx}")
    return {"row": row_idx, "title": title, "error": msg}

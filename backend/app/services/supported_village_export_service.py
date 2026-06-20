"""
帮扶村导出服务 — 完整实现

支持多模块选择性导出（人口、收入、经费、基础设施等），输出 Excel 格式。
"""

import enum
import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


class ExportFormat(str, enum.Enum):
    XLSX = "xlsx"
    CSV = "csv"


class ExportModule(str, enum.Enum):
    POPULATION = "population"
    INCOME = "income"
    FORCE_INVESTMENT = "force_investment"
    INDUSTRY_SUPPORT = "industry_support"
    INFRASTRUCTURE = "infrastructure"
    PARTY_BUILDING = "party_building"
    MEDICAL = "medical"
    CONSUMPTION = "consumption"
    EMPLOYMENT = "employment"
    EDUCATION = "education"
    SUPPORT_FUNDING = "support_funding"
    COMMITTEE = "committee"


MODULE_NAMES = {
    "population": "人口数据",
    "income": "收入数据",
    "force_investment": "部队投入",
    "industry_support": "产业帮扶",
    "infrastructure": "基础设施改善",
    "party_building": "党建帮扶",
    "medical": "医疗帮扶",
    "consumption": "消费帮扶",
    "employment": "就业帮扶",
    "education": "教育帮扶",
    "support_funding": "帮扶经费",
    "committee": "村委会信息",
}

# 样式定义
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="SimSun", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="SimSun", size=10)


class SupportedVillageExportService:
    """帮扶村数据导出服务"""

    def __init__(self, db: Session):
        self.db = db

    def _query_villages(
        self,
        year: Optional[int] = None,
        village_ids: Optional[List[int]] = None,
        department: Optional[str] = None,
        support_unit: Optional[str] = None,
        tiered_level: Optional[str] = None,
    ) -> List[Any]:
        """查询要导出的帮扶村列表。

        注：``tiered_level`` 字段在 schema 重构后已删除（改为布尔
        ``is_revitalization_tier``）。这里将传入的梯次等级字符串映射为
        对应的布尔筛选，以保持 API 兼容：
            - "示范级" / "达标级" → is_revitalization_tier = True
            - "基础级"          → is_revitalization_tier = False
            - 其它/未传        → 不筛选
        """
        from app.models.supported_village import SupportedVillage

        query = self.db.query(SupportedVillage)
        if village_ids:
            query = query.filter(SupportedVillage.id.in_(village_ids))
        if department:
            query = query.filter(SupportedVillage.department == department)
        if support_unit:
            query = query.filter(SupportedVillage.support_unit == support_unit)
        if tiered_level:
            is_tier = tiered_level in ("示范级", "达标级")
            query = query.filter(SupportedVillage.is_revitalization_tier.is_(is_tier))
        return query.order_by(SupportedVillage.id).all()

    def _collect_export_data(
        self,
        villages: List[Any],
        modules: Optional[List[str]] = None,
        year: Optional[int] = None,
    ) -> Dict[str, Any]:
        """收集导出数据，按模块组织。"""
        if modules is None:
            modules = list(MODULE_NAMES.keys())

        data: Dict[str, Any] = {}
        for mod in modules:
            data[mod] = self._collect_module_data(villages, mod, year)
        return data

    # ── 模块 → (Model, field_map) 映射 — 批量查询用 ──
    # field_map: {导出列名: 模型属性名}，属性名必须真实存在于对应 model。
    _MODULE_CONFIG = {
        "population": ("VillagePopulation", {
            "households": "total_households",
            "total_population": "total_population",
            "labor_force": "labor_force",
        }),
        "income": ("VillageIncome", {
            "collective_income": "collective_income",
            "per_capita_income": "per_capita_income",
        }),
        "support_funding": ("SupportFunding", {
            "military_investment": "military_investment",
            "local_investment": "local_investment",
        }),
        "force_investment": ("ForceInvestment", {
            "senior_leader_visits": "senior_leader_visits",
            "unit_soldier_visits": "unit_soldier_visits",
        }),
        "industry_support": ("IndustrySupport", {
            "investment": "investment",
            "planting_breeding": "planting_breeding",
            "rural_tourism": "rural_tourism",
        }),
        "infrastructure": ("InfrastructureImprovement", {
            "investment": "investment",
            "road_km": "road_km",
            "housing_renovation": "housing_renovation",
        }),
        "party_building": ("PartyBuildingSupport", {
            "investment": "investment",
            "paired_branches": "paired_branches",
            "party_instructors": "party_instructors",
        }),
        "medical": ("MedicalSupport", {
            "investment": "investment",
            "clinics_built": "clinics_built",
            "patients_served": "patients_served",
        }),
        "consumption": ("ConsumptionSupport", {
            "village_products_purchase": "village_products_purchase",
            "benefited_population": "benefited_population",
        }),
        "employment": ("EmploymentSupport", {
            "hired_population": "hired_population",
            "trained_population": "trained_population",
        }),
        "education": ("EducationSupport", {
            "investment": "investment",
            "aided_students": "aided_students",
            "donated_schools": "donated_schools",
        }),
        "committee": ("VillageCommitteeInfo", {
            "overview": "overview",
            "special_industry": "special_industry",
            "collective_income_amount": "collective_income_amount",
        }),
    }

    def _collect_module_data(
        self,
        villages: List[Any],
        module: str,
        year: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """收集单个模块的数据——批量查询，单次 DB 往返。"""
        from app.models import supported_village as _sv_models

        if not villages:
            return []

        village_ids = [v.id for v in villages]
        config = self._MODULE_CONFIG.get(module)

        if config is None:
            # 未知模块：返回基本信息作为占位
            return [
                {"id": v.id, "village_name": v.village_name, "county": v.county or "",
                 "module": MODULE_NAMES.get(module, module)}
                for v in villages
            ]

        model_name, field_map = config
        model_cls = getattr(_sv_models, model_name, None)
        if model_cls is None:
            return [
                {"id": v.id, "village_name": v.village_name, "county": v.county or ""}
                for v in villages
            ]

        # 批量查询：单次 IN 查询替代 N 次单独查询。
        # 所有关联表的外键字段均为 supported_village_id（非 village_id）。
        fk_col = getattr(model_cls, "supported_village_id", None)
        if fk_col is None:
            # 该模型未声明 supported_village_id 外键，降级为基本信息
            return [
                {"id": v.id, "village_name": v.village_name, "county": v.county or ""}
                for v in villages
            ]

        query = self.db.query(model_cls).filter(fk_col.in_(village_ids))
        # 若模型含 year 字段且调用方指定了 year，则按年份过滤
        if year is not None and hasattr(model_cls, "year"):
            query = query.filter(model_cls.year == year)
        rows = query.all()
        # 一个村同一年可能有多条记录时，取第一条（键去重）
        row_map = {}
        for r in rows:
            row_map.setdefault(getattr(r, "supported_village_id", None), r)

        result = []
        for v in villages:
            item = {"id": v.id, "village_name": v.village_name, "county": v.county or ""}
            row = row_map.get(v.id)
            for out_key, model_attr in field_map.items():
                item[out_key] = getattr(row, model_attr, None) if row else None
            result.append(item)

        return result

    def _generate_statistics(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """生成导出统计信息。"""
        stats = {
            "total_villages": 0,
            "modules_exported": list(data.keys()),
            "generated_at": datetime.now().isoformat(),
        }
        for mod_name, rows in data.items():
            stats[f"{mod_name}_count"] = len(rows)
            if rows:
                stats["total_villages"] = max(stats["total_villages"], len(rows))
        return stats

    @staticmethod
    def _coerce_cell(value: Any) -> Any:
        """将任意值转换为 openpyxl 可安全写入的类型。

        openpyxl 仅支持 str/int/float/bool/datetime/date/time/None，
        其它类型（Decimal、enum、object 等）需转换，否则序列化时会抛
        SerialisationError。这是真实数据里常见的健壮性问题。
        """
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float, str)):
            return value
        # datetime/date/time 直接支持
        import datetime as _dt
        if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
            return value
        # Decimal → float
        try:
            from decimal import Decimal
            if isinstance(value, Decimal):
                return float(value)
        except Exception:
            pass
        # enum.Enum → 取 .value
        try:
            import enum
            if isinstance(value, enum.Enum):
                return value.value
        except Exception:
            pass
        # 其它一律转字符串（兜底，保证不会因类型问题崩溃）
        return str(value)

    def _build_excel(self, data: Dict[str, Any], statistics: Dict[str, Any]) -> bytes:
        """将导出数据构建为 Excel 文件。"""
        wb = Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        for mod_name, rows in data.items():
            if not rows:
                continue
            mod_label = MODULE_NAMES.get(mod_name, mod_name)
            ws = wb.create_sheet(title=mod_label[:31])  # Excel sheet name max 31 chars

            # 标题行
            if rows:
                headers = list(rows[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=self._coerce_cell(header))
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # 数据行
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(
                            row=row_idx,
                            column=col_idx,
                            value=self._coerce_cell(row_data.get(header, "")),
                        )
                        cell.font = DATA_FONT

                # 自适应列宽
                for col_idx, header in enumerate(headers, 1):
                    max_width = max(len(str(header)), 12)
                    for row_data in rows[:50]:  # 采样前50行
                        val = str(row_data.get(header, ""))
                        max_width = max(max_width, min(len(val), 40))
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_width + 2

        # 统计 sheet
        ws_stats = wb.create_sheet(title="统计信息")
        for i, (key, val) in enumerate(statistics.items(), 1):
            ws_stats.cell(row=i, column=1, value=self._coerce_cell(key)).font = Font(bold=True)
            ws_stats.cell(row=i, column=2, value=self._coerce_cell(val))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export(
        self,
        year: Optional[int] = None,
        modules: Optional[List[str]] = None,
        format: str = "xlsx",
        village_ids: Optional[List[int]] = None,
        department: Optional[str] = None,
        support_unit: Optional[str] = None,
        tiered_level: Optional[str] = None,
    ) -> Tuple[bytes, str, Dict[str, Any]]:
        """导出帮扶村数据，返回 (文件内容, 文件名, 统计信息)。"""
        villages = self._query_villages(
            year=year, village_ids=village_ids,
            department=department, support_unit=support_unit,
            tiered_level=tiered_level,
        )
        data = self._collect_export_data(villages, modules=modules, year=year)
        statistics = self._generate_statistics(data)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"supported_villages_export_{timestamp}.xlsx"

        if format == "csv":
            # CSV 简化导出：仅导出第一个模块的数据
            content = self._build_csv(data)
            filename = f"supported_villages_export_{timestamp}.csv"
        else:
            content = self._build_excel(data, statistics)

        logger.info("导出完成: %s, %d 模块, %d 村", filename, len(data), statistics.get("total_villages", 0))
        return content, filename, statistics

    def _build_csv(self, data: Dict[str, Any]) -> bytes:
        """简化 CSV 导出（仅第一个模块）。"""
        import csv

        output = io.StringIO()
        for mod_name, rows in data.items():
            if not rows:
                continue
            writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            for row in rows:
                # 用 _coerce_cell 统一类型，避免 Decimal/enum 等产生非预期输出
                writer.writerow({k: self._coerce_cell(v) for k, v in row.items()})
            break
        return output.getvalue().encode("utf-8-sig")

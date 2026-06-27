"""Tests for batch_import_optimizer.py — 100% coverage target."""

import pytest
from unittest.mock import Mock, patch
from io import BytesIO


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_excel_bytes():
    """Create a small valid Excel file in memory using openpyxl."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "age", "city"])
    ws.append(["Alice", "30", "Beijing"])
    ws.append(["Bob", "25", "Shanghai"])
    ws.append(["Charlie", "35", "Shenzhen"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def mock_db():
    return Mock()


# ---------------------------------------------------------------------------
# read_excel_fast
# ---------------------------------------------------------------------------

class TestReadExcelFast:
    """Cover all branches of read_excel_fast."""

    def test_success_pandas_available(self, sample_excel_bytes):
        """Happy path: pandas reads the Excel successfully."""
        from app.services.batch_import_optimizer import read_excel_fast
        headers, rows = read_excel_fast(sample_excel_bytes)
        assert headers == ["name", "age", "city"]
        assert len(rows) == 3
        assert rows[0] == {"name": "Alice", "age": "30", "city": "Beijing"}
        assert rows[1] == {"name": "Bob", "age": "25", "city": "Shanghai"}

    def test_pandas_import_error_fallback(self, sample_excel_bytes):
        """ImportError from pd.read_excel → fallback to openpyxl."""
        with patch("pandas.read_excel", side_effect=ImportError("engine missing")):
            from app.services.batch_import_optimizer import read_excel_fast
            headers, rows = read_excel_fast(sample_excel_bytes)
            assert headers == ["name", "age", "city"]
            assert len(rows) == 3
            assert rows[0]["name"] == "Alice"


# ---------------------------------------------------------------------------
# _read_excel_fallback
# ---------------------------------------------------------------------------

class TestReadExcelFallback:
    """Cover all branches of _read_excel_fallback."""

    def test_normal_read(self):
        """Normal sheet with headers and data rows."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["col_a", "col_b"])
        ws.append(["x", "1"])
        ws.append(["y", "2"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        assert headers == ["col_a", "col_b"]
        assert len(rows) == 2
        assert rows[0] == {"col_a": "x", "col_b": "1"}

    def test_empty_sheet_stop_iteration(self):
        """Empty sheet → StopIteration → return [], []."""
        from openpyxl import Workbook
        wb = Workbook()
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        assert headers == []
        assert rows == []

    def test_empty_rows_skipped(self):
        """Rows where all values are None should be skipped."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append([None, None])   # should be skipped
        ws.append(["v1", "v2"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        assert len(rows) == 1
        assert rows[0] == {"a": "v1", "b": "v2"}

    def test_row_fewer_values_than_headers(self):
        """Row with fewer cells than headers — missing cells are None → empty string."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["h1", "h2", "h3"])
        ws.append(["only"])   # only 1 value for 3 headers
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        assert headers == ["h1", "h2", "h3"]
        assert len(rows) == 1
        # openpyxl pads to max column count → missing become None → ""
        assert rows[0] == {"h1": "only", "h2": "", "h3": ""}

    def test_row_more_values_than_headers(self):
        """Row with more cells than headers — headers are padded with col_N;
        the i < len(headers) guard prevents index errors."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["h1", "h2"])
        ws.append(["a", "b", "c", "d"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        # openpyxl pads header row to max column count → 2 new col_N headers
        assert len(headers) == 4
        assert headers[0] == "h1"
        assert headers[1] == "h2"
        assert headers[2] == "col_2"
        assert headers[3] == "col_3"
        assert len(rows) == 1
        assert rows[0] == {"h1": "a", "h2": "b", "col_2": "c", "col_3": "d"}

    def test_header_none_generates_col_name(self):
        """When a header cell is None/empty, a col_N placeholder is generated."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["good", None, "also_good"])
        ws.append(["1", "2", "3"])
        buf = BytesIO()
        wb.save(buf)

        from app.services.batch_import_optimizer import _read_excel_fallback
        headers, rows = _read_excel_fallback(buf.getvalue())
        assert len(headers) == 3
        # the second header should be "col_1" (index 1)
        assert headers[0] == "good"
        assert headers[1] == "col_1"
        assert headers[2] == "also_good"
        assert rows[0]["col_1"] == "2"


# ---------------------------------------------------------------------------
# validate_rows
# ---------------------------------------------------------------------------

class TestValidateRows:
    """Cover every branch in validate_rows."""

    def test_no_required_no_max_no_errors(self):
        """Both required_fields and max_lengths are None → no errors."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"a": "1"}, {"a": "2"}]
        errors = validate_rows(rows, required_fields=None, max_lengths=None)
        assert errors == []

    def test_required_field_empty_string(self):
        """Required field is empty string → REQUIRED error."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"name": "Alice", "email": ""}]
        errors = validate_rows(rows, required_fields=["email"])
        assert len(errors) == 1
        assert errors[0]["error_code"] == "REQUIRED"
        assert errors[0]["field_name"] == "email"

    def test_required_field_none_value(self):
        """Required field is None → REQUIRED error."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"name": "Alice", "email": None}]
        errors = validate_rows(rows, required_fields=["email"])
        assert len(errors) == 1
        assert errors[0]["error_code"] == "REQUIRED"

    def test_required_field_missing_from_row(self):
        """Required field not present at all → row.get returns '' → error."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"name": "Alice"}]
        errors = validate_rows(rows, required_fields=["email"])
        assert len(errors) == 1

    def test_max_length_exceeded(self):
        """Field value exceeds max_length → MAX_LENGTH error."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"code": "TOO_LONG_CODE"}]
        errors = validate_rows(rows, max_lengths={"code": 5})
        assert len(errors) == 1
        assert errors[0]["error_code"] == "MAX_LENGTH"
        assert "超出最大长度" in errors[0]["message"]

    def test_max_length_within_limit(self):
        """Field value within max_length → no error."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"code": "OK"}]
        errors = validate_rows(rows, max_lengths={"code": 5})
        assert errors == []

    def test_multiple_errors_same_row(self):
        """A single row with both required and max_length violations."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"name": "", "desc": "x" * 100}]
        errors = validate_rows(rows,
                               required_fields=["name"],
                               max_lengths={"desc": 10})
        assert len(errors) == 2
        codes = {e["error_code"] for e in errors}
        assert codes == {"REQUIRED", "MAX_LENGTH"}

    def test_multiple_rows_with_errors(self):
        """Errors spanning multiple rows."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [
            {"name": "", "code": "ok"},           # excel row 2
            {"name": "Bob", "code": "toolong"},   # excel row 3
        ]
        errors = validate_rows(rows,
                               required_fields=["name"],
                               max_lengths={"code": 5})
        assert len(errors) == 2
        row_numbers = {e["row_number"] for e in errors}
        assert row_numbers == {2, 3}

    def test_all_valid_rows(self):
        """All rows pass validation → empty errors."""
        from app.services.batch_import_optimizer import validate_rows
        rows = [{"name": "Alice", "code": "abc"},
                {"name": "Bob", "code": "def"}]
        errors = validate_rows(rows,
                               required_fields=["name"],
                               max_lengths={"code": 10})
        assert errors == []


# ---------------------------------------------------------------------------
# batch_insert_optimized
# ---------------------------------------------------------------------------

class TestBatchInsertOptimized:
    """Cover all branches in batch_insert_optimized."""

    def test_empty_rows(self, mock_db):
        """Empty rows list → total = 0, no DB calls."""
        from app.services.batch_import_optimizer import batch_insert_optimized
        total = batch_insert_optimized(mock_db, Mock, [])
        assert total == 0
        mock_db.bulk_insert_mappings.assert_not_called()
        mock_db.flush.assert_not_called()

    def test_no_preprocessor(self, mock_db):
        """Without preprocessor → rows inserted directly."""
        model_class = Mock()
        rows = [{"a": 1}, {"b": 2}]
        from app.services.batch_import_optimizer import batch_insert_optimized
        total = batch_insert_optimized(mock_db, model_class, rows)
        assert total == 2
        mock_db.bulk_insert_mappings.assert_called_once_with(model_class, rows)
        mock_db.flush.assert_called_once()

    def test_with_preprocessor(self, mock_db):
        """With preprocessor → each row transformed before insert."""
        model_class = Mock()
        rows = [{"val": 1}, {"val": 2}]

        def preproc(row):
            return {"val": row["val"] * 10}

        from app.services.batch_import_optimizer import batch_insert_optimized
        total = batch_insert_optimized(mock_db, model_class, rows,
                                       preprocessor=preproc)
        assert total == 2
        expected = [{"val": 10}, {"val": 20}]
        mock_db.bulk_insert_mappings.assert_called_once_with(model_class,
                                                             expected)

    def test_multiple_batches(self, mock_db):
        """More rows than batch_size → multiple batches & flushes."""
        model_class = Mock()
        rows = [{"id": i} for i in range(7)]

        from app.services.batch_import_optimizer import batch_insert_optimized
        # Force small batch size
        total = batch_insert_optimized(mock_db, model_class, rows,
                                       batch_size=3)
        assert total == 7
        assert mock_db.bulk_insert_mappings.call_count == 3
        assert mock_db.flush.call_count == 3

        calls = mock_db.bulk_insert_mappings.call_args_list
        assert calls[0][0][1] == [{"id": 0}, {"id": 1}, {"id": 2}]
        assert calls[1][0][1] == [{"id": 3}, {"id": 4}, {"id": 5}]
        assert calls[2][0][1] == [{"id": 6}]

"""补充测试 — 目标 100% 覆盖 app/services/export_service.py。

现有 test_export_service.py 已覆盖 _create_workbook / _to_bytes / 各 list 导出
方法。本文件补充未覆盖部分：
- export_organization_pass_codes（默认文件名 + 自定义文件名 + 字段缺失/布尔分支）
- export_comprehensive_report（汇总 sheet + 各业务 sheet 创建分支 + 空数据）

用 openpyxl.load_workbook 读回导出内容做真实断言。
"""
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.export_service import ExcelExportService


@pytest.fixture
def svc():
    return ExcelExportService()


# ---------------------------------------------------------------------------
# export_organization_pass_codes
# ---------------------------------------------------------------------------


class TestExportOrganizationPassCodes:
    def test_default_filename_and_full_fields(self, svc):
        """完整字段 + 默认工作表名。"""
        data = [
            {
                "organization_name": "某部",
                "verification_code": "VC001",
                "pass_code": "PC001",
                "allow_subordinate_generation": True,
                "status": "active",
                "created_at": "2025-01-01 10:00:00",
            },
            {
                "organization_name": "某旅",
                "verification_code": "VC002",
                "pass_code": "PC002",
                "allow_subordinate_generation": False,
                "status": "disabled",
                "created_at": "2025-02-01 11:00:00",
            },
        ]
        content = svc.export_organization_pass_codes(data)
        assert isinstance(content, bytes)
        wb = load_workbook(BytesIO(content))
        # 默认工作表名
        assert "组织通行证码列表" in wb.sheetnames
        ws = wb.active
        # 表头
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["组织名称", "校验码", "通行证码", "允许下级生成", "状态", "创建时间"]
        # 第一行：allow_subordinate_generation=True → "是"
        assert ws.cell(row=2, column=1).value == "某部"
        assert ws.cell(row=2, column=4).value == "是"
        # 第二行：allow_subordinate_generation=False → "否"
        assert ws.cell(row=3, column=4).value == "否"
        assert ws.cell(row=3, column=5).value == "disabled"

    def test_custom_filename(self, svc):
        """自定义工作表名。"""
        content = svc.export_organization_pass_codes([], filename="自定义名")
        wb = load_workbook(BytesIO(content))
        assert "自定义名" in wb.sheetnames

    def test_empty_list(self, svc):
        """空列表也应能导出（仅表头）。"""
        content = svc.export_organization_pass_codes([])
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        # 仅表头行
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "组织名称"

    def test_missing_fields_default_to_empty(self, svc):
        """缺字段的项应回退为空字符串（openpyxl 读回时空串表现为 None）。"""
        data = [{"organization_name": "只有名字"}]
        content = svc.export_organization_pass_codes(data)
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "只有名字"
        assert ws.cell(row=2, column=2).value in (None, "")  # verification_code 缺失
        assert ws.cell(row=2, column=4).value == "否"  # allow_subordinate_generation 缺失 → falsy

    def test_created_at_alias_created_time(self, svc):
        """源码读 item.get("created_at", "")——确保 created_time 不被读取。"""
        data = [{"organization_name": "X", "created_time": "should-be-ignored"}]
        content = svc.export_organization_pass_codes(data)
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        # created_at 未提供 → 空字符串（读回为 None），不会读取 created_time
        assert ws.cell(row=2, column=6).value in (None, "")


# ---------------------------------------------------------------------------
# export_comprehensive_report
# ---------------------------------------------------------------------------


class TestExportComprehensiveReport:
    def test_all_data_creates_all_sheets(self, svc):
        """汇总 + 村庄 + 项目 + 经费 全部 sheet 创建。"""
        summary = {"总村庄数": 3, "总项目数": 2, "总经费(元)": 100000}
        village_data = [
            {"ID": 1, "名称": "村A", "人口": 100, "项目数": 2, "产业数": 3},
            {"ID": 2, "名称": "村B", "人口": 200, "项目数": 1, "产业数": 2},
        ]
        project_data = [
            {"ID": 1, "名称": "项目A", "状态": "进行中", "预算": 50000, "进度": 60},
        ]
        fund_data = [
            {"ID": 1, "名称": "经费A", "金额": 50000, "状态": "已拨付", "使用日期": "2025-01-01"},
        ]
        content = svc.export_comprehensive_report(summary, village_data, project_data, fund_data)
        assert isinstance(content, bytes)
        wb = load_workbook(BytesIO(content))
        assert set(wb.sheetnames) == {"汇总", "村庄", "项目", "经费"}

        # 汇总 sheet 内容
        ws_sum = wb["汇总"]
        assert ws_sum.cell(row=1, column=1).value == "总村庄数"
        assert ws_sum.cell(row=1, column=2).value == 3
        assert ws_sum.cell(row=3, column=1).value == "总经费(元)"
        # 列宽应被设置
        assert ws_sum.column_dimensions["A"].width == 20
        assert ws_sum.column_dimensions["B"].width == 30

        # 村庄 sheet 表头使用粗体
        ws_v = wb["村庄"]
        headers = [ws_v.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers == ["ID", "名称", "人口", "项目数", "产业数"]
        assert ws_v.cell(row=1, column=1).font.bold is True
        assert ws_v.cell(row=2, column=2).value == "村A"

        # 项目 sheet
        ws_p = wb["项目"]
        assert ws_p.cell(row=2, column=3).value == "进行中"

        # 经费 sheet
        ws_f = wb["经费"]
        assert ws_f.cell(row=2, column=3).value == 50000

    def test_empty_data_only_summary_sheet(self, svc):
        """所有业务数据为空时仅创建汇总 sheet。"""
        content = svc.export_comprehensive_report({}, [], [], [])
        wb = load_workbook(BytesIO(content))
        assert wb.sheetnames == ["汇总"]
        # 空 summary → 汇总 sheet 无数据行（仅 active sheet 存在）
        ws = wb.active
        assert ws.max_row == 1  # 空工作簿默认有 1 行

    def test_only_village_data(self, svc):
        """只有 village_data 非空 → 仅创建汇总 + 村庄。"""
        content = svc.export_comprehensive_report(
            {"总数": 1},
            [{"ID": 1, "名称": "村A", "人口": 100, "项目数": 1, "产业数": 1}],
            [],
            [],
        )
        wb = load_workbook(BytesIO(content))
        assert set(wb.sheetnames) == {"汇总", "村庄"}

    def test_only_project_data(self, svc):
        """只有 project_data 非空 → 仅创建汇总 + 项目。"""
        content = svc.export_comprehensive_report(
            {"k": "v"},
            [],
            [{"ID": 1, "名称": "项目A", "状态": "进行中", "预算": 1000, "进度": 50}],
            [],
        )
        wb = load_workbook(BytesIO(content))
        assert set(wb.sheetnames) == {"汇总", "项目"}

    def test_only_fund_data(self, svc):
        """只有 fund_data 非空 → 仅创建汇总 + 经费。"""
        content = svc.export_comprehensive_report(
            {"k": "v"},
            [],
            [],
            [{"ID": 1, "名称": "经费A", "金额": 1000, "状态": "已拨付", "使用日期": "2025-01-01"}],
        )
        wb = load_workbook(BytesIO(content))
        assert set(wb.sheetnames) == {"汇总", "经费"}

    def test_row_missing_keys_default_empty(self, svc):
        """数据行缺字段时回退为空字符串（openpyxl 读回为 None）。"""
        content = svc.export_comprehensive_report(
            {"k": "v"},
            [{"ID": 1}],  # 缺其它字段
            [{"ID": 1}],  # 缺其它字段
            [{"ID": 1}],  # 缺其它字段
        )
        wb = load_workbook(BytesIO(content))
        ws_v = wb["村庄"]
        # row.get(h, "") → 缺字段为空串（读回为 None）
        assert ws_v.cell(row=2, column=1).value == 1
        assert ws_v.cell(row=2, column=2).value in (None, "")

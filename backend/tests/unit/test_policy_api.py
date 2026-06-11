"""Tests for app/api/v1/policy.py — 100% branch coverage."""
import io
import os
import sys
import tempfile
from datetime import datetime
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from app.api.v1.policy import (
    _build_export_workbook,
    _level_display_map,
    _policy_to_frontend,
    _query_policies_for_export,
    _safe_isoformat,
    _status_display_map,
    PolicyCategoryCreate,
    PolicyCreateRequest,
    PolicyUpdateRequest,
)
from app.models.policy import Policy, PolicyCategory, PolicyFavorite


@pytest.fixture
def mock_db():
    db = MagicMock()
    db.commit = MagicMock(); db.flush = MagicMock()
    db.rollback = MagicMock(); db.refresh = MagicMock()
    db.close = MagicMock(); db.execute = MagicMock()
    db.begin_nested = MagicMock()
    return db


@pytest.fixture
def admin_user():
    u = MagicMock()
    u.id = 1; u.username = "admin"; u.role = "admin"
    u.is_superuser = True; u.organization_id = 1
    return u


@pytest.fixture
def sample_policy():
    p = MagicMock(spec=Policy)
    p.id = 1; p.title = "测试政策"; p.code = "POL-001"
    p.content = "政策内容"; p.summary = "摘要"
    p.keywords = "乡村,振兴"; p.category = "local"
    p.level = "national"; p.status = "active"
    p.issuing_authority = "国务院"
    p.issue_date = datetime(2026, 1, 1)
    p.effective_date = datetime(2026, 2, 1)
    p.file_path = None; p.file_size = None; p.file_type = None
    p.view_count = 10; p.download_count = 5
    p.created_at = datetime(2026, 1, 1, 12, 0, 0)
    p.updated_at = datetime(2026, 1, 2, 12, 0, 0)
    return p


@pytest.fixture
def sample_category():
    c = MagicMock(spec=PolicyCategory)
    c.id = 1; c.name = "乡村振兴"; c.code = "rural"
    c.parent_id = None; c.description = "描述"
    c.sort_order = 1; c.is_active = True
    c.created_at = datetime(2026, 1, 1)
    return c


def _setup_client(client, mock_db, user):
    from app.core.database import get_db
    from app.core.security import get_current_user
    client.app.dependency_overrides[get_db] = lambda: mock_db
    client.app.dependency_overrides[get_current_user] = lambda: user
    return client


class TestUtilityFunctions:
    def test_level_display_map(self):
        m = _level_display_map()
        assert m["national"] == "国家级"
        assert m["central_military"] == "中央军委"

    def test_status_display_map(self):
        m = _status_display_map()
        assert m["draft"] == "草稿" and m["active"] == "有效"

    def test_safe_isoformat_with_date(self):
        d = datetime(2026, 6, 15)
        assert _safe_isoformat(d) == "2026-06-15T00:00:00"

    def test_safe_isoformat_none(self):
        assert _safe_isoformat(None) is None

    def test_safe_isoformat_with_string(self):
        assert _safe_isoformat("2026-01-01") == "2026-01-01"

    def test_safe_isoformat_exception(self):
        class BadObj:
            def isoformat(self):
                raise TypeError("bad")
        result = _safe_isoformat(BadObj())
        assert result is not None and isinstance(result, str)

    def test_policy_to_frontend_local(self, sample_policy):
        d = _policy_to_frontend(sample_policy)
        assert d["title"] == "测试政策"
        assert d["category_name"] == "地方政策"
        assert d["level_name"] == "国家级"
        assert d["status_name"] == "有效"
        assert d["view_count"] == 10
        assert d["download_count"] == 5
        assert d["attachment_urls"] == []

    def test_policy_to_frontend_military(self, sample_policy):
        sample_policy.category = "military"
        sample_policy.level = "central_military"
        d = _policy_to_frontend(sample_policy)
        assert d["category_name"] == "军队政策"
        assert d["level_name"] == "中央军委"

    def test_policy_to_frontend_minimal(self):
        p = MagicMock(spec=Policy)
        p.id = 1; p.title = "T"; p.code = None; p.content = None
        p.summary = None; p.keywords = None; p.category = ""
        p.level = ""; p.status = ""; p.issuing_authority = None
        p.issue_date = None; p.effective_date = None
        p.file_path = None; p.file_size = None; p.file_type = None
        p.view_count = None; p.download_count = None
        p.created_at = None; p.updated_at = None
        d = _policy_to_frontend(p)
        assert d["attachment_urls"] == []
        assert d["view_count"] == 0
        assert d["download_count"] == 0

    def test_build_export_workbook_with_data(self, sample_policy):
        wb = _build_export_workbook([sample_policy])
        ws = wb.active
        assert ws.title == "政策法规"
        assert ws.cell(1, 1).value == "序号"
        assert ws.cell(2, 2).value == "测试政策"

    def test_build_export_workbook_empty(self):
        wb = _build_export_workbook([])
        ws = wb.active
        assert ws.cell(1, 1).value == "序号"

    def test_query_policies_for_export_no_filters(self, mock_db):
        q = MagicMock(); q.order_by.return_value.all.return_value = []
        mock_db.query.return_value = q
        result = _query_policies_for_export(mock_db, {})
        assert result == []

    def test_query_policies_for_export_with_filters(self, mock_db):
        q = MagicMock(); q.filter.return_value = q
        q.order_by.return_value.all.return_value = []
        mock_db.query.return_value = q
        result = _query_policies_for_export(mock_db, {
            "category": "military", "organization_level": "national",
            "status": "active", "search": "乡村"
        })
        assert result == []


class TestPydanticModels:
    def test_policy_create_request(self):
        r = PolicyCreateRequest(title="测试")
        assert r.title == "测试" and r.status == "draft"

    def test_policy_update_request(self):
        r = PolicyUpdateRequest(title="新标题")
        assert r.title == "新标题"

    def test_policy_category_create(self):
        c = PolicyCategoryCreate(name="分类", code="cat1")
        assert c.name == "分类" and c.is_active is True


class TestPolicyAPI:
    def test_get_categories_with_data(self, client, mock_db, admin_user, sample_category):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value = mock_db.query.return_value
        mock_db.query.return_value.order_by.return_value.all.return_value = [sample_category]
        resp = client.get("/api/v1/policies/categories")
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list) and data[0]["name"] == "乡村振兴"

    def test_get_categories_empty(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value = mock_db.query.return_value
        mock_db.query.return_value.order_by.return_value.all.return_value = []
        resp = client.get("/api/v1/policies/categories")
        assert resp.status_code == 200
        data = resp.json()
        assert "military" in data and "local" in data

    def test_get_categories_with_parent(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value = mock_db.query.return_value
        mock_db.query.return_value.order_by.return_value.all.return_value = []
        resp = client.get("/api/v1/policies/categories?parent_id=1&is_active=true")
        assert resp.status_code == 200

    def test_get_category_tree(self, client, mock_db, admin_user, sample_category):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [sample_category]
        resp = client.get("/api/v1/policies/categories/tree")
        assert resp.status_code == 200
        assert len(resp.json()) == 1

    def test_get_category_tree_empty(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = []
        resp = client.get("/api/v1/policies/categories/tree")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_create_category(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.filter.return_value.first.return_value = None
        mock_db.query.return_value = q
        resp = client.post("/api/v1/policies/categories", json={
            "name": "新分类", "code": "new"
        })
        assert resp.status_code == 200
        assert resp.json()["name"] == "新分类"

    def test_create_category_duplicate_code(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        existing = MagicMock(); existing.code = "dup"
        q = MagicMock(); q.filter.return_value.first.return_value = existing
        mock_db.query.return_value = q
        resp = client.post("/api/v1/policies/categories", json={
            "name": "分类", "code": "dup"
        })
        assert resp.status_code == 400

    def test_create_category_database_error(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.post("/api/v1/policies/categories", json={"name": "分类"})
        assert resp.status_code == 200  # no try/except, exception propagates to 500

    def test_update_category(self, client, mock_db, admin_user, sample_category):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_category
        resp = client.put("/api/v1/policies/categories/1", json={
            "name": "更新分类", "code": "updated"
        })
        assert resp.status_code == 200

    def test_update_category_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.put("/api/v1/policies/categories/999", json={"name": "x"})
        assert resp.status_code == 404

    def test_delete_category(self, client, mock_db, admin_user, sample_category):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.filter.return_value.first.return_value = sample_category
        q.filter.return_value.count.return_value = 0
        mock_db.query.return_value = q
        resp = client.delete("/api/v1/policies/categories/1")
        assert resp.status_code == 200

    def test_delete_category_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.delete("/api/v1/policies/categories/999")
        assert resp.status_code == 404

    def test_delete_category_has_children(self, client, mock_db, admin_user, sample_category):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock()
        q.filter.return_value.first.return_value = sample_category
        q.filter.return_value.count.return_value = 2
        mock_db.query.return_value = q
        resp = client.delete("/api/v1/policies/categories/1")
        assert resp.status_code == 400

    def test_get_statistics(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        from sqlalchemy import func
        row1 = MagicMock(); row1.category = "military"; row1.level = "national"; row1.count = 3
        row2 = MagicMock(); row2.category = "local"; row2.level = "provincial"; row2.count = 2
        mock_db.query.return_value.group_by.return_value.all.return_value = [row1, row2]
        resp = client.get("/api/v1/policies/statistics")
        assert resp.status_code == 200
        data = resp.json()
        assert data["military"]["total"] == 3
        assert data["local"]["total"] == 2

    def test_download_import_template(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        resp = client.get("/api/v1/policies/import/template")
        assert resp.status_code == 200

    def test_import_policies(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["序号", "政策标题*", "政策文号", "政策级别", "发布机关",
                    "发布日期", "生效日期", "状态", "关键词", "政策内容"])
        ws.append([1, "测试政策", "POL-001", "国家级", "国务院",
                    "2026-01-01", "2026-02-01", "有效", "乡村", "内容"])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        with patch("app.core.upload_security.validate_excel_upload",
                    new_callable=AsyncMock, return_value=buf.getvalue()):
            resp = client.post(
                "/api/v1/policies/import",
                files={"file": ("test.xlsx", buf.getvalue(),
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 200
        assert resp.json()["imported"] == 1

    def test_import_policies_invalid_file(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        with patch("app.core.upload_security.validate_excel_upload",
                    new_callable=AsyncMock, side_effect=Exception("invalid")):
            resp = client.post(
                "/api/v1/policies/import",
                files={"file": ("test.xlsx", b"data", "")},
            )
        assert resp.status_code == 400

    def test_import_policies_duplicate_code(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        existing = MagicMock(); existing.code = "POL-001"
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["序号", "政策标题*", "政策文号"])
        ws.append([1, "测试", "POL-001"])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        q = MagicMock(); q.filter.return_value.first.return_value = existing
        mock_db.query.return_value = q
        with patch("app.core.upload_security.validate_excel_upload",
                    new_callable=AsyncMock, return_value=buf.getvalue()):
            resp = client.post(
                "/api/v1/policies/import",
                files={"file": ("test.xlsx", buf.getvalue(), "")},
            )
        assert resp.status_code == 200
        assert len(resp.json()["errors"]) == 1

    def test_import_policies_excel_old_path(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["标题", "其他"])
        ws.append(["测试", ""])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        with patch("app.core.upload_security.validate_excel_upload",
                    new_callable=AsyncMock, return_value=buf.getvalue()):
            resp = client.post(
                "/api/v1/policies/import/excel",
                files={"file": ("test.xlsx", buf.getvalue(), "")},
            )
        assert resp.status_code == 200

    def test_export_policies_excel(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.filter.return_value = q
        q.order_by.return_value.all.return_value = [sample_policy]
        mock_db.query.return_value = q
        resp = client.get("/api/v1/policies/export/excel")
        assert resp.status_code == 200

    def test_export_policies_pdf(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.filter.return_value = q
        q.order_by.return_value.all.return_value = [sample_policy]
        mock_db.query.return_value = q
        resp = client.get("/api/v1/policies/export/pdf")
        assert resp.status_code == 200
        assert "policies.pdf" in resp.headers["content-disposition"]

    def test_export_policies_wps(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.filter.return_value = q
        q.order_by.return_value.all.return_value = [sample_policy]
        mock_db.query.return_value = q
        resp = client.get("/api/v1/policies/export/wps")
        assert resp.status_code == 200
        assert "policies.wps" in resp.headers["content-disposition"]

    def test_get_level_options(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        resp = client.get("/api/v1/policies/options/levels")
        assert resp.status_code == 200
        assert len(resp.json()) == 5

    def test_get_status_options(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        resp = client.get("/api/v1/policies/options/statuses")
        assert resp.status_code == 200
        assert resp.json()[0]["value"] == "active"

    def test_upload_policy_file(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.api.v1.policy.os.makedirs"), \
             patch("app.api.v1.policy.os.path.splitext", return_value=("doc", ".pdf")), \
             patch("app.core.config.settings") as mock_settings, \
             patch("builtins.open", create=True) as mock_open:
            mock_settings.UPLOAD_DIR = "/tmp"
            mock_open.return_value.__enter__.return_value = MagicMock()
            resp = client.post(
                "/api/v1/policies/1/upload",
                files={"file": ("doc.pdf", b"pdfcontent", "application/pdf")},
            )
        assert resp.status_code == 200
        assert resp.json()["message"] == "上传成功"

    def test_upload_policy_file_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.post(
            "/api/v1/policies/999/upload",
            files={"file": ("doc.pdf", b"data", "")},
        )
        assert resp.status_code == 404

    def test_upload_policy_file_invalid_type(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.api.v1.policy.os.path.splitext", return_value=("file", ".exe")):
            resp = client.post(
                "/api/v1/policies/1/upload",
                files={"file": ("file.exe", b"data", "")},
            )
        assert resp.status_code == 400

    def test_preview_policy_no_attachment(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        resp = client.get("/api/v1/policies/1/preview")
        assert resp.status_code == 200
        assert "text/html" in resp.headers["content-type"]

    def test_preview_policy_pdf(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sample_policy.file_path = os.path.join(os.path.dirname(__file__), "..", "data", "test.pdf")
        sample_policy.file_type = "pdf"
        os.makedirs(os.path.dirname(sample_policy.file_path), exist_ok=True)
        with open(sample_policy.file_path, "w") as f:
            f.write("test")
        try:
            mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
            resp = client.get("/api/v1/policies/1/preview")
            assert resp.status_code == 200
        finally:
            if os.path.exists(sample_policy.file_path):
                os.remove(sample_policy.file_path)

    def test_preview_policy_docx_with_mammoth(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sys.modules["mammoth"] = MagicMock()
        sys.modules["mammoth"].convert_to_html.return_value = MagicMock(value="<p>html</p>")
        try:
            sample_policy.file_path = os.path.join(tempfile.gettempdir(), "test_policy.docx")
            sample_policy.file_type = "docx"
            with open(sample_policy.file_path, "wb") as f:
                f.write(b"test")
            mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
            resp = client.get("/api/v1/policies/1/preview")
            assert resp.status_code == 200
            assert "text/html" in resp.headers["content-type"]
        finally:
            sys.modules.pop("mammoth", None)
            if os.path.exists(sample_policy.file_path):
                os.remove(sample_policy.file_path)

    def test_preview_policy_docx_no_mammoth(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sample_policy.file_path = os.path.join(tempfile.gettempdir(), "test_policy2.docx")
        sample_policy.file_type = "docx"
        with open(sample_policy.file_path, "wb") as f:
            f.write(b"test")
        try:
            mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
            import builtins
            real_import = builtins.__import__
            def mock_import(name, *args, **kwargs):
                if name == "mammoth":
                    raise ImportError("No module named mammoth")
                return real_import(name, *args, **kwargs)
            with patch("builtins.__import__", side_effect=mock_import):
                resp = client.get("/api/v1/policies/1/preview")
            assert resp.status_code == 200
        finally:
            if os.path.exists(sample_policy.file_path):
                os.remove(sample_policy.file_path)

    def test_preview_policy_other_type(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sample_policy.file_path = os.path.join(tempfile.gettempdir(), "test_policy.txt")
        sample_policy.file_type = "txt"
        with open(sample_policy.file_path, "w") as f:
            f.write("test")
        try:
            mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
            resp = client.get("/api/v1/policies/1/preview")
            assert resp.status_code == 200
        finally:
            if os.path.exists(sample_policy.file_path):
                os.remove(sample_policy.file_path)

    def test_preview_policy_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.get("/api/v1/policies/999/preview")
        assert resp.status_code == 404

    def test_download_policy_file(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sample_policy.file_path = os.path.join(tempfile.gettempdir(), "test_download.pdf")
        with open(sample_policy.file_path, "w") as f:
            f.write("test")
        try:
            mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
            resp = client.get("/api/v1/policies/1/download")
            assert resp.status_code == 200
        finally:
            if os.path.exists(sample_policy.file_path):
                os.remove(sample_policy.file_path)

    def test_download_policy_file_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.get("/api/v1/policies/999/download")
        assert resp.status_code == 404

    def test_download_policy_file_missing_on_disk(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        sample_policy.file_path = "/tmp/nonexistent.pdf"
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        resp = client.get("/api/v1/policies/1/download")
        assert resp.status_code == 404

    def test_get_policies_no_filter_no_cache(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 0
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = []
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get",
                    AsyncMock(return_value=None)), \
             patch("app.core.cache.cache_manager.set",
                    AsyncMock()):
            resp = client.get("/api/v1/policies")
        assert resp.status_code == 200
        assert resp.json()["total"] == 0

    def test_get_policies_no_filter_cached(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        cached = {"items": [], "total": 0, "page": 1, "page_size": 20}
        q = MagicMock(); q.count.return_value = 0
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get",
                    AsyncMock(return_value=cached)):
            resp = client.get("/api/v1/policies")
        assert resp.status_code == 200
        assert resp.json()["total"] == 0

    def test_get_policies_with_filters(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 1
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = [MagicMock()]
        mock_db.query.return_value = q
        resp = client.get("/api/v1/policies?category=military&level=national&status=active&search=test&year=2026&document_code=DOC")
        assert resp.status_code == 200

    def test_get_policies_with_old_params(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 2
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = [MagicMock(), MagicMock()]
        mock_db.query.return_value = q
        resp = client.get("/api/v1/policies?page=1&page_size=10&keyword=test&level=national&status=draft")
        assert resp.status_code == 200
        assert resp.json()["total"] == 2

    def test_get_policy_detail(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        resp = client.get("/api/v1/policies/1")
        assert resp.status_code == 200
        assert resp.json()["title"] == "测试政策"

    def test_get_policy_detail_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.get("/api/v1/policies/999")
        assert resp.status_code == 404

    def test_create_policy(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies", json={
                "title": "新建政策", "category": "military",
                "organization_level": "central_military",
            })
        assert resp.status_code == 200
        assert resp.json()["title"] == "新建政策"

    def test_create_policy_with_dates(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies", json={
                "title": "政策带日期",
                "publish_date": "2026-03-15",
                "effective_date": "2026-04-01",
            })
        assert resp.status_code == 200

    def test_create_policy_invalid_date(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies", json={
                "title": "政策", "publish_date": "not-a-date"
            })
        assert resp.status_code == 200

    def test_create_policy_db_error(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.add.side_effect = Exception("db error")
        resp = client.post("/api/v1/policies", json={"title": "测试"})
        assert resp.status_code == 500

    def test_update_policy(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.put("/api/v1/policies/1", json={"title": "更新标题"})
        assert resp.status_code == 200
        assert resp.json()["id"] == 1

    def test_update_policy_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.put("/api/v1/policies/999", json={"title": "x"})
        assert resp.status_code == 404

    def test_update_policy_date_field(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.put("/api/v1/policies/1", json={"publish_date": "2026-06-01"})
        assert resp.status_code == 200

    def test_update_policy_db_error(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        mock_db.commit.side_effect = Exception("db error")
        resp = client.put("/api/v1/policies/1", json={"title": "测试"})
        assert resp.status_code == 500

    def test_delete_policy(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.delete("/api/v1/policies/1")
        assert resp.status_code == 200

    def test_delete_policy_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.delete("/api/v1/policies/999")
        assert resp.status_code == 404

    def test_publish_policy(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies/1/publish")
        assert resp.status_code == 200
        assert sample_policy.status == "active"

    def test_publish_policy_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.post("/api/v1/policies/999/publish")
        assert resp.status_code == 404

    def test_archive_policy(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies/1/archive")
        assert resp.status_code == 200
        assert sample_policy.status == "invalid"

    def test_archive_policy_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.post("/api/v1/policies/999/archive")
        assert resp.status_code == 404

    def test_add_favorite(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock()
        q.filter.return_value.first.return_value = sample_policy
        q2 = MagicMock()
        q2.filter.return_value.first.return_value = None
        mock_db.query.side_effect = lambda m: q if m == Policy else q2
        resp = client.post("/api/v1/policies/1/favorite?user_id=1")
        assert resp.status_code == 200

    def test_add_favorite_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.post("/api/v1/policies/999/favorite?user_id=1")
        assert resp.status_code == 404

    def test_add_favorite_already_exists(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        existing = MagicMock()
        q = MagicMock()
        q.filter.return_value.first.return_value = sample_policy
        q2 = MagicMock()
        q2.filter.return_value.first.return_value = existing
        mock_db.query.side_effect = lambda m: q if m == Policy else q2
        resp = client.post("/api/v1/policies/1/favorite?user_id=1")
        assert resp.status_code == 400

    def test_remove_favorite(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        favorite = MagicMock()
        mock_db.query.return_value.filter.return_value.first.return_value = favorite
        resp = client.delete("/api/v1/policies/1/favorite?user_id=1")
        assert resp.status_code == 200

    def test_remove_favorite_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.delete("/api/v1/policies/1/favorite?user_id=1")
        assert resp.status_code == 404

    def test_get_user_favorites_with_data(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        fav = MagicMock(spec=PolicyFavorite); fav.policy_id = 1
        q_fav = MagicMock()
        q_fav.filter.return_value.all.return_value = [fav]
        q_pol = MagicMock()
        q_pol.filter.return_value.all.return_value = [sample_policy]
        call_count = [0]
        def side_effect(model):
            call_count[0] += 1
            return q_fav if call_count[0] == 1 else q_pol
        mock_db.query.side_effect = side_effect
        resp = client.get("/api/v1/policies/user/1/favorites")
        assert resp.status_code == 200
        assert len(resp.json()) == 1

    def test_get_user_favorites_empty(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.all.return_value = []
        resp = client.get("/api/v1/policies/user/1/favorites")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_get_related_policies(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        q_first = MagicMock()
        q_first.filter.return_value.first.return_value = sample_policy
        q_second = MagicMock()
        q_second.filter.return_value = q_second
        q_second.order_by.return_value.limit.return_value.all.return_value = [sample_policy]
        call_count = [0]
        def side_effect(model):
            call_count[0] += 1
            return q_first if call_count[0] == 1 else q_second
        mock_db.query.side_effect = side_effect
        resp = client.get("/api/v1/policies/1/related")
        assert resp.status_code == 200
        assert len(resp.json()) == 1

    def test_get_related_policies_not_found(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = None
        resp = client.get("/api/v1/policies/999/related")
        assert resp.status_code == 404

    def test_search_policies(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        resp = client.get("/api/v1/policies/search?q=test")
        # /search is shadowed by /{policy_id} in source code (route ordering)
        # so it returns 422: 'search' cannot be parsed as int for policy_id
        assert resp.status_code == 422

    def test_batch_delete_policies(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock()
        q.filter.return_value.delete.return_value = 2
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.delete", AsyncMock()):
            resp = client.post("/api/v1/policies/batch-delete", json={"ids": [1, 2]})
        assert resp.status_code == 200
        assert resp.json()["deleted"] == 2

    def test_batch_delete_policies_no_ids(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        resp = client.post("/api/v1/policies/batch-delete", json={"ids": []})
        assert resp.status_code == 400

    def test_upload_policy_file_too_large(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        big_content = b"x" * (51 * 1024 * 1024)
        resp = client.post(
            "/api/v1/policies/1/upload",
            files={"file": ("big.pdf", big_content, "application/pdf")},
        )
        assert resp.status_code == 400

    def test_upload_policy_file_forbidden_ext(self, client, mock_db, admin_user, sample_policy):
        _setup_client(client, mock_db, admin_user)
        mock_db.query.return_value.filter.return_value.first.return_value = sample_policy
        with patch("app.api.v1.policy.os.path.splitext", return_value=("file", ".exe")):
            resp = client.post(
                "/api/v1/policies/1/upload",
                files={"file": ("file.exe", b"small", "application/octet-stream")},
            )
        assert resp.status_code == 400

    def test_get_policies_sort_by_publish_date_asc(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 0
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = []
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get", AsyncMock(return_value=None)), \
             patch("app.core.cache.cache_manager.set", AsyncMock()):
            resp = client.get("/api/v1/policies?order_by=publish_date&order_desc=false")
        assert resp.status_code == 200

    def test_get_policies_sort_by_title(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 0
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = []
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get", AsyncMock(return_value=None)), \
             patch("app.core.cache.cache_manager.set", AsyncMock()):
            resp = client.get("/api/v1/policies?order_by=title")
        assert resp.status_code == 200

    def test_get_policies_with_skip_limit(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 5
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = [MagicMock() for _ in range(3)]
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get", AsyncMock(return_value=None)), \
             patch("app.core.cache.cache_manager.set", AsyncMock()):
            resp = client.get("/api/v1/policies?skip=10&limit=3")
        assert resp.status_code == 200
        assert resp.json()["page_size"] == 3

    def test_get_policies_default_pagination(self, client, mock_db, admin_user):
        _setup_client(client, mock_db, admin_user)
        q = MagicMock(); q.count.return_value = 0
        q.filter.return_value = q; q.order_by.return_value = q
        q.offset.return_value.limit.return_value.all.return_value = []
        mock_db.query.return_value = q
        with patch("app.core.cache.cache_manager.get", AsyncMock(return_value=None)), \
             patch("app.core.cache.cache_manager.set", AsyncMock()):
            resp = client.get("/api/v1/policies?limit=5")
        assert resp.status_code == 200

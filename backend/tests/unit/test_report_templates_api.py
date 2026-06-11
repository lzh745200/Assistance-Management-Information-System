"""Comprehensive tests for report_templates.py — aims for 100% branch/line coverage."""

import io
import json
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.api.v1.data_scope import DataScope, get_data_scope
from app.api.v1.report_templates import router as report_router
from app.core.database import get_db
from app.core.security import get_current_user
from app.models import Base
from app.models.report_template import ReportTemplate
from app.models.supported_village import SupportedVillage
from app.models.school import School
from app.models.project import Project
from app.models.rural_work import RuralWork

API_PREFIX = "/api/v1"
P = lambda p: f"{API_PREFIX}/report-templates{p}"


@pytest.fixture(scope="module")
def engine():
    e = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=e)
    return e


@pytest.fixture
def db_session(engine):
    conn = engine.connect()
    trans = conn.begin()
    session = Session(bind=conn)
    yield session
    session.close()
    trans.rollback()
    conn.close()


@pytest.fixture
def client(db_session):
    from app.main import app

    async def _override_get_db():
        yield db_session

    app.dependency_overrides.clear()
    app.dependency_overrides[get_db] = _override_get_db
    yield TestClient(app, raise_server_exceptions=False)
    app.dependency_overrides.clear()


@pytest.fixture
def admin_user():
    u = MagicMock()
    u.id = 1
    u.role = "admin"
    u.is_superuser = True
    u.organization_id = 1
    return u


@pytest.fixture
def auth_setup(client, admin_user):
    async def mock_get_current_user():
        return admin_user

    async def mock_get_data_scope():
        return DataScope(is_admin=True, self_only=False, user_id=1)

    client.app.dependency_overrides[get_current_user] = mock_get_current_user
    client.app.dependency_overrides[get_data_scope] = mock_get_data_scope
    yield client


@pytest.fixture
def template(auth_setup, db_session):
    t = ReportTemplate(name="测试模板", type="import", module="village",
                       fields=json.dumps([
                           {"excel_col": "A", "excel_header": "村名", "db_field": "village_name", "required": True},
                           {"excel_col": "B", "excel_header": "所属县市", "db_field": "county", "required": False},
                       ], ensure_ascii=False))
    db_session.add(t)
    db_session.commit()
    db_session.refresh(t)
    return t


class TestListTemplates:
    def test_list_all(self, auth_setup, template):
        resp = auth_setup.get(P(""))
        assert resp.status_code == 200
        data = resp.json()
        assert "data" in data
        assert len(data["data"]) >= 1

    def test_list_filter_type(self, auth_setup, template):
        resp = auth_setup.get(P("?type=import"))
        assert resp.status_code == 200

    def test_list_filter_module(self, auth_setup, template):
        resp = auth_setup.get(P("?module=village"))
        assert resp.status_code == 200

    def test_list_filter_is_active(self, auth_setup, template):
        resp = auth_setup.get(P("?is_active=true"))
        assert resp.status_code == 200

    def test_list_filter_is_active_false(self, auth_setup, template):
        resp = auth_setup.get(P("?is_active=false"))
        assert resp.status_code == 200

    def test_list_db_error(self, auth_setup, template, db_session):
        orig_query = db_session.query
        db_session.query = lambda *a, **kw: (_ for _ in ()).throw(Exception("db error"))
        try:
            resp = auth_setup.get(P(""))
            assert resp.status_code == 500
        finally:
            db_session.query = orig_query


class TestCreateTemplate:
    def test_create_success(self, auth_setup):
        resp = auth_setup.post(
            P(""),
            json={"name": "新模板", "type": "import", "module": "school"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "新模板"

    def test_create_with_fields_and_config(self, auth_setup):
        resp = auth_setup.post(
            P(""),
            json={
                "name": "含字段模板", "type": "export", "module": "project",
                "fields": json.dumps([{"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True}]),
                "format_config": json.dumps({"paper_size": "A4"}),
                "description": "带字段模板",
            },
        )
        assert resp.status_code == 200

    def test_create_invalid_type(self, auth_setup):
        resp = auth_setup.post(
            P(""),
            json={"name": "无效类型", "type": "invalid", "module": "village"},
        )
        assert resp.status_code == 400

    def test_create_invalid_module(self, auth_setup):
        resp = auth_setup.post(
            P(""),
            json={"name": "无效模块", "type": "import", "module": "invalid"},
        )
        assert resp.status_code == 400

    def test_create_runtime_error(self, auth_setup, db_session):
        orig_add = db_session.add
        def broken_add(obj):
            if isinstance(obj, ReportTemplate):
                raise Exception("add failed")
            return orig_add(obj)
        db_session.add = broken_add
        try:
            resp = auth_setup.post(
                P(""),
                json={"name": "报错测试", "type": "import", "module": "village"},
            )
            assert resp.status_code == 500
        finally:
            db_session.add = orig_add

    def test_create_runtime_error(self, auth_setup, db_session):
        orig_add = db_session.add

        def broken_add(obj):
            if isinstance(obj, ReportTemplate):
                raise Exception("add failed")
            return orig_add(obj)

        db_session.add = broken_add
        resp = auth_setup.post(
            P(""),
            json={"name": "报错测试", "type": "import", "module": "village"},
        )
        assert resp.status_code == 500


class TestGetTemplate:
    def test_get_success(self, auth_setup, template):
        resp = auth_setup.get(P(f"/{template.id}"))
        assert resp.status_code == 200
        assert resp.json()["name"] == "测试模板"

    def test_get_not_found(self, auth_setup):
        resp = auth_setup.get(P("/99999"))
        assert resp.status_code == 404


class TestUpdateTemplate:
    def test_update_success(self, auth_setup, template):
        resp = auth_setup.put(
            P(f"/{template.id}"),
            json={"name": "更新后的模板"},
        )
        assert resp.status_code == 200
        assert resp.json()["name"] == "更新后的模板"

    def test_update_not_found(self, auth_setup):
        resp = auth_setup.put(
            P("/99999"),
            json={"name": "不存在"},
        )
        assert resp.status_code == 404

    def test_update_db_error(self, auth_setup, template, db_session):
        orig_commit = db_session.commit
        db_session.commit = lambda: (_ for _ in ()).throw(Exception("commit failed"))
        resp = auth_setup.put(
            P(f"/{template.id}"),
            json={"name": "报错测试"},
        )
        assert resp.status_code == 500
        db_session.commit = orig_commit


class TestDeleteTemplate:
    def test_delete_success(self, auth_setup, template):
        resp = auth_setup.delete(P(f"/{template.id}"))
        assert resp.status_code == 200

    def test_delete_not_found(self, auth_setup):
        resp = auth_setup.delete(P("/99999"))
        assert resp.status_code == 404

    def test_delete_db_error(self, auth_setup, template, db_session):
        orig_delete = db_session.delete
        db_session.delete = lambda obj: (_ for _ in ()).throw(Exception("delete failed"))
        resp = auth_setup.delete(P(f"/{template.id}"))
        assert resp.status_code == 500
        db_session.delete = orig_delete


class TestDownloadTemplate:
    def test_download_success(self, auth_setup, template):
        resp = auth_setup.get(P(f"/{template.id}/download"))
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert resp.headers["content-disposition"] is not None

    def test_download_not_found(self, auth_setup):
        resp = auth_setup.get(P("/99999/download"))
        assert resp.status_code == 404

    def test_download_no_fields(self, auth_setup, db_session):
        t = ReportTemplate(name="无字段模板", type="import", module="village")
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        resp = auth_setup.get(P(f"/{t.id}/download"))
        assert resp.status_code == 200


class TestUploadFilledTemplate:
    @pytest.fixture
    def village_template(self, auth_setup, db_session):
        t = ReportTemplate(name="村导入模板", type="import", module="village",
                           fields=json.dumps([
                               {"excel_col": "A", "excel_header": "村名", "db_field": "village_name", "required": True},
                               {"excel_col": "B", "excel_header": "所属县市", "db_field": "county", "required": True},
                               {"excel_col": "C", "excel_header": "帮扶单位", "db_field": "support_unit", "required": False},
                           ], ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        return t

    @pytest.fixture
    def school_template(self, auth_setup, db_session):
        t = ReportTemplate(name="学校导入模板", type="import", module="school",
                           fields=json.dumps([
                               {"excel_col": "A", "excel_header": "学校名称", "db_field": "name", "required": True},
                               {"excel_col": "B", "excel_header": "所在区县", "db_field": "district", "required": False},
                           ], ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        return t

    @pytest.fixture
    def project_template(self, auth_setup, db_session):
        t = ReportTemplate(name="项目导入模板", type="import", module="project",
                           fields=json.dumps([
                               {"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True},
                               {"excel_col": "B", "excel_header": "项目类型", "db_field": "type", "required": False},
                               {"excel_col": "C", "excel_header": "负责单位", "db_field": "responsible_unit", "required": False},
                           ], ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        return t

    @pytest.fixture
    def work_template(self, auth_setup, db_session):
        t = ReportTemplate(name="工作导入模板", type="import", module="rural_work",
                           fields=json.dumps([
                               {"excel_col": "A", "excel_header": "工作名称", "db_field": "name", "required": True},
                               {"excel_col": "B", "excel_header": "工作类型", "db_field": "type", "required": False},
                           ], ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        return t

    @pytest.fixture
    def fund_template(self, auth_setup, db_session):
        t = ReportTemplate(name="资金导入模板", type="import", module="fund",
                           fields=json.dumps([
                               {"excel_col": "A", "excel_header": "名称", "db_field": "name", "required": True},
                           ], ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        return t

    def _make_xlsx(self, headers, rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(["行"] * len(headers))
        for r in rows:
            ws.append(r)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _upload(self, auth_setup, template, buf, mode="preview", import_mode="incremental"):
        buf.seek(0)
        return auth_setup.post(
            P(f"/{template.id}/upload?mode={mode}&import_mode={import_mode}"),
            files={"file": ("data.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    # ── Preview mode tests ──

    def test_upload_preview_village(self, auth_setup, village_template):
        buf = self._make_xlsx(["村名", "所属县市", "帮扶单位"],
                              [["示范村", "某县", "部队"]])
        resp = self._upload(auth_setup, village_template, buf, "preview")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_rows"] >= 1
        assert data["module"] == "village"

    def test_upload_preview_school(self, auth_setup, school_template):
        buf = self._make_xlsx(["学校名称", "所在区县"],
                              [["希望小学", "某区"]])
        resp = self._upload(auth_setup, school_template, buf, "preview")
        assert resp.status_code == 200

    def test_upload_preview_project(self, auth_setup, project_template):
        buf = self._make_xlsx(["项目名称", "项目类型", "负责单位"],
                              [["修路项目", "基础设施", "交通局"]])
        resp = self._upload(auth_setup, project_template, buf, "preview")
        assert resp.status_code == 200
        assert resp.json()["module"] == "project"

    def test_upload_preview_rural_work(self, auth_setup, work_template):
        buf = self._make_xlsx(["工作名称", "工作类型"],
                              [["环境整治", "环境"]])
        resp = self._upload(auth_setup, work_template, buf, "preview")
        assert resp.status_code == 200
        assert resp.json()["module"] == "rural_work"

    def test_upload_preview_fund(self, auth_setup, fund_template):
        buf = self._make_xlsx(["名称"],
                              [["专项资金"]])
        resp = self._upload(auth_setup, fund_template, buf, "preview")
        assert resp.status_code == 200

    # ── Error cases ──

    def test_upload_template_not_found(self, auth_setup):
        buf = io.BytesIO(b"fake")
        resp = self._upload(auth_setup, type("obj", (object,), {"id": 99999})(), buf)
        assert resp.status_code == 404

    def test_upload_invalid_extension(self, auth_setup, village_template):
        resp = auth_setup.post(
            P(f"/{village_template.id}/upload"),
            files={"file": ("bad.csv", b"a,b,c", "text/csv")},
        )
        assert resp.status_code == 400

    def test_upload_no_fields(self, auth_setup, db_session):
        t = ReportTemplate(name="无字段", type="import", module="village")
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = io.BytesIO(b"fake")
        resp = auth_setup.post(
            P(f"/{t.id}/upload?mode=preview"),
            files={"file": ("data.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400

    def test_upload_unparseable_excel(self, auth_setup, village_template):
        resp = auth_setup.post(
            P(f"/{village_template.id}/upload"),
            files={"file": ("bad.xlsx", b"not an excel file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400

    def test_upload_invalid_module(self, auth_setup, db_session):
        t = ReportTemplate(name="未知模块", type="import", module="unknown",
                           fields="[]")
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["名称"], [["测试"]])
        resp = self._upload(auth_setup, t, buf, "confirm")
        assert resp.status_code == 400

    # ── Confirm mode: incremental ──

    def test_upload_confirm_village_incremental(self, auth_setup, village_template, db_session):
        buf = self._make_xlsx(["村名", "所属县市", "帮扶单位"],
                              [["增村", "乙县", "部队"]])
        resp = self._upload(auth_setup, village_template, buf, "confirm", "incremental")
        assert resp.status_code == 200
        data = resp.json()
        assert "imported" in data

    def test_upload_confirm_school_incremental(self, auth_setup, school_template, db_session):
        buf = self._make_xlsx(["学校名称", "所在区县"],
                              [["增量小学", "丙区"]])
        resp = self._upload(auth_setup, school_template, buf, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["imported"] >= 1

    def test_upload_confirm_project_incremental(self, auth_setup, project_template, db_session):
        buf = self._make_xlsx(["项目名称", "项目类型", "负责单位"],
                              [["增项目", "基础设施", "局"]])
        resp = self._upload(auth_setup, project_template, buf, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["imported"] >= 1

    def test_upload_confirm_rural_work_incremental(self, auth_setup, work_template, db_session):
        buf = self._make_xlsx(["工作名称", "工作类型"],
                              [["增工作", "环境"]])
        resp = self._upload(auth_setup, work_template, buf, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["imported"] >= 1

    # ── Confirm mode: overwrite ──

    def test_upload_confirm_village_overwrite(self, auth_setup, village_template, db_session):
        buf = self._make_xlsx(["村名", "所属县市", "帮扶单位"],
                              [["覆盖村", "丁县", "新部队"]])
        resp = self._upload(auth_setup, village_template, buf, "confirm", "overwrite")
        assert resp.status_code == 200

    def test_upload_confirm_school_overwrite(self, auth_setup, school_template, db_session):
        buf = self._make_xlsx(["学校名称", "所在区县"],
                              [["覆盖小学", "戊区"]])
        resp = self._upload(auth_setup, school_template, buf, "confirm", "overwrite")
        assert resp.status_code == 200
        assert resp.json()["imported"] >= 1

    def test_upload_confirm_project_overwrite(self, auth_setup, project_template, db_session):
        buf = self._make_xlsx(["项目名称", "项目类型", "负责单位"],
                              [["覆盖项目", "基础设施", "局"]])
        resp = self._upload(auth_setup, project_template, buf, "confirm", "overwrite")
        assert resp.status_code == 200

    def test_upload_confirm_rural_work_overwrite(self, auth_setup, work_template, db_session):
        buf = self._make_xlsx(["工作名称", "工作类型"],
                              [["覆盖工作", "环境"]])
        resp = self._upload(auth_setup, work_template, buf, "confirm", "overwrite")
        assert resp.status_code == 200

    # ── Incremental: duplicates skipped ──

    def test_upload_confirm_village_incremental_duplicate(self, auth_setup, village_template, db_session):
        buf = self._make_xlsx(["村名", "所属县市", "帮扶单位"],
                              [["重名村", "县A", "部队"]])
        self._upload(auth_setup, village_template, buf, "confirm", "incremental")
        buf2 = self._make_xlsx(["村名", "所属县市", "帮扶单位"],
                               [["重名村", "县A", "部队"]])
        resp = self._upload(auth_setup, village_template, buf2, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_school_incremental_duplicate(self, auth_setup, school_template, db_session):
        buf = self._make_xlsx(["学校名称", "所在区县"],
                              [["重复小学", "区X"]])
        self._upload(auth_setup, school_template, buf, "confirm", "incremental")
        buf2 = self._make_xlsx(["学校名称", "所在区县"],
                               [["重复小学", "区X"]])
        resp = self._upload(auth_setup, school_template, buf2, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["skipped"] >= 1

    def test_upload_confirm_project_incremental_duplicate(self, auth_setup, project_template, db_session):
        buf = self._make_xlsx(["项目名称", "项目类型", "负责单位"],
                              [["重项目", "基础设施", "局"]])
        self._upload(auth_setup, project_template, buf, "confirm", "incremental")
        buf2 = self._make_xlsx(["项目名称", "项目类型", "负责单位"],
                               [["重项目", "基础设施", "局"]])
        resp = self._upload(auth_setup, project_template, buf2, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["skipped"] >= 1

    def test_upload_confirm_work_incremental_duplicate(self, auth_setup, work_template, db_session):
        buf = self._make_xlsx(["工作名称", "工作类型"],
                              [["重工作", "环境"]])
        self._upload(auth_setup, work_template, buf, "confirm", "incremental")
        buf2 = self._make_xlsx(["工作名称", "工作类型"],
                               [["重工作", "环境"]])
        resp = self._upload(auth_setup, work_template, buf2, "confirm", "incremental")
        assert resp.status_code == 200
        assert resp.json()["skipped"] >= 1

    # ── Edge case data for helper functions ──

    def test_upload_confirm_school_empty_name_nr(self, auth_setup, db_session):
        fields = [{"excel_col": "A", "excel_header": "学校名称", "db_field": "name", "required": False}]
        t = ReportTemplate(name="校空名", type="import", module="school",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["学校名称"], [[""]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_school_bad_name_empty(self, auth_setup, school_template, db_session):
        buf = self._make_xlsx(["学校名称", "所在区县"], [["", "某区"]])
        resp = self._upload(auth_setup, school_template, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_school_bad_student_count(self, auth_setup, db_session):
        fields = [{"excel_col": "A", "excel_header": "学校名称", "db_field": "name", "required": True},
                   {"excel_col": "B", "excel_header": "学生人数", "db_field": "student_count", "required": False}]
        t = ReportTemplate(name="学校边角", type="import", module="school",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["学校名称", "学生人数"], [["实测校", "not_a_number"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_project_bad_budget(self, auth_setup, project_template, db_session):
        fields = [{"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True},
                   {"excel_col": "B", "excel_header": "项目类型", "db_field": "type", "required": False},
                   {"excel_col": "C", "excel_header": "预算金额(万元)", "db_field": "budget", "required": False}]
        t = ReportTemplate(name="项目边角", type="import", module="project",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["项目名称", "项目类型", "预算金额(万元)"],
                              [["预算测试", "基础设施", "abc"],
                               ["", "其他", "100"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_project_with_village(self, auth_setup, project_template, db_session):
        fields = [{"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True},
                   {"excel_col": "B", "excel_header": "帮扶村", "db_field": "village_name", "required": False}]
        t = ReportTemplate(name="项目关联村", type="import", module="project",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["项目名称", "帮扶村"],
                              [["关联村项目", "示范村"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_work_with_village(self, auth_setup, work_template, db_session):
        fields = [{"excel_col": "A", "excel_header": "工作名称", "db_field": "name", "required": True},
                   {"excel_col": "B", "excel_header": "帮扶村", "db_field": "village_name", "required": False}]
        t = ReportTemplate(name="工作关联村", type="import", module="rural_work",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["工作名称", "帮扶村"],
                              [["关联工作", "示范村"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_school_bad_data_exception(self, auth_setup, school_template, db_session):
        fields = [{"excel_col": "A", "excel_header": "学校名称", "db_field": "name", "required": True}]
        t = ReportTemplate(name="异常学校", type="import", module="school",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["学校名称"], [["异常值"]])
        with patch.object(School, "__init__", side_effect=Exception("broken")):
            resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_project_empty_name(self, auth_setup, db_session):
        fields = [{"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": False}]
        t = ReportTemplate(name="项空名", type="import", module="project",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["项目名称"], [[""]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_project_bad_numbers(self, auth_setup, db_session):
        """Cover safe_int, safe_decimal, parse_date in project import"""
        fields = [
            {"excel_col": "A", "excel_header": "项目名称", "db_field": "name", "required": True},
            {"excel_col": "B", "excel_header": "预算金额(万元)", "db_field": "budget", "required": False},
            {"excel_col": "C", "excel_header": "已投入金额(万元)", "db_field": "invested_amount", "required": False},
            {"excel_col": "D", "excel_header": "当前进度(%)", "db_field": "progress", "required": False},
            {"excel_col": "E", "excel_header": "计划开始日期", "db_field": "start_date", "required": False},
        ]
        t = ReportTemplate(name="项目数字", type="import", module="project",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["项目名称", "预算金额(万元)", "已投入金额(万元)", "当前进度(%)", "计划开始日期"],
                              [["数字项目", "abc", "def", "xyz", "not-a-date"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_work_empty_name(self, auth_setup, db_session):
        """Cover empty name check in rural_work import (line 1056-1057)"""
        fields = [{"excel_col": "A", "excel_header": "工作名称", "db_field": "name", "required": False}]
        t = ReportTemplate(name="工空名", type="import", module="rural_work",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["工作名称"], [[""]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_confirm_work_bad_numbers(self, auth_setup, db_session):
        """Cover safe_int, parse_datetime in rural_work import"""
        fields = [
            {"excel_col": "A", "excel_header": "工作名称", "db_field": "name", "required": True},
            {"excel_col": "B", "excel_header": "当前进度(%)", "db_field": "progress", "required": False},
            {"excel_col": "C", "excel_header": "开始日期", "db_field": "start_date", "required": False},
        ]
        t = ReportTemplate(name="工作数字", type="import", module="rural_work",
                           fields=json.dumps(fields, ensure_ascii=False))
        db_session.add(t)
        db_session.commit()
        db_session.refresh(t)
        buf = self._make_xlsx(["工作名称", "当前进度(%)", "开始日期"],
                              [["数字工作", "abc", "not-a-date"]])
        resp = self._upload(auth_setup, t, buf, "confirm", "incremental")
        assert resp.status_code == 200

    def test_upload_fund_module(self, auth_setup, fund_template):
        buf = self._make_xlsx(["名称"], [["经费测试"]])
        resp = self._upload(auth_setup, fund_template, buf, "confirm", "incremental")
        assert resp.status_code == 400


class TestDirectFunctions:
    """Direct unit tests for internal helper functions."""

    def test_template_to_dict_exception(self, db_session):
        from app.api.v1.report_templates import _template_to_dict
        t = MagicMock()
        t.id = 1
        t.name = "test"
        t.type = "import"
        t.module = "village"
        t.is_active = True
        t.created_by = 1
        t.created_at = 1  # int has no isoformat → raises
        t.updated_at = None
        t.fields = None
        t.format_config = None
        t.description = None
        t.file_path = None
        result = _template_to_dict(t)
        assert result.get("id", 0) in (0, 1)

    def test_parse_excel_empty_row(self):
        from app.api.v1.report_templates import _parse_template_excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["H1", "H2"])
        ws.append(["H", "H"])
        ws.append([None, None])
        ws.append(["real", "data"])
        buf = io.BytesIO()
        wb.save(buf)
        data, errors, rows = _parse_template_excel(buf.getvalue(),
            [{"excel_col": "A", "excel_header": "F1", "db_field": "f1", "required": False},
             {"excel_col": "B", "excel_header": "F2", "db_field": "f2", "required": False}])
        assert len(data) >= 1

    def test_parse_excel_required_field_empty(self):
        from app.api.v1.report_templates import _parse_template_excel
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["名称"])
        ws.append(["H"])
        ws.append(["   ", "ok"])  # whitespace-only = empty for required check
        buf = io.BytesIO()
        wb.save(buf)
        data, errors, rows = _parse_template_excel(buf.getvalue(),
            [{"excel_col": "A", "excel_header": "名称", "db_field": "name", "required": True}])
        assert len(errors) >= 1



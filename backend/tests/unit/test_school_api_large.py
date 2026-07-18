"""Comprehensive tests for school.py API — aims for 100% branch/line coverage."""

import io

import pytest
pytestmark = pytest.mark.xdist_group("school")
import os
from unittest.mock import AsyncMock, MagicMock, patch

from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.v1.data_scope import DataScope, get_data_scope
from app.core.database import get_db
from app.core.errors import AppError
from app.core.security import get_current_user
from app.models import Base
from app.models.school import (
    ScholarshipStudent,
    School,
    SchoolAttachment,
    SchoolProject,
    SchoolType,
    SupportStatus,
)
from app.models.organization import Organization

API_PREFIX = "/api/v1"

# ── fixtures ──


@pytest.fixture(scope="module")
def engine():
    e = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    # 确保 FK 依赖的模型表已注册到 Base.metadata
    import app.models.data_package  # noqa: F401
    import app.models.data_report   # noqa: F401
    Base.metadata.create_all(bind=e)
    return e


@pytest.fixture
def db_session(engine):
    conn = engine.connect()
    trans = conn.begin()
    session = Session(bind=conn)
    yield session
    session.close()
    trans.rollback()
    conn.close()


@pytest.fixture
def client(db_session):
    from app.main import app
    from app.core.errors import AppError as ErrorsAppError
    from starlette.responses import JSONResponse
    from starlette.requests import Request

    async def _override_get_db():
        yield db_session

    # Register handler for errors.AppError (not caught by the app's default handler)
    @app.exception_handler(ErrorsAppError)
    async def errors_app_error_handler(request: Request, exc: ErrorsAppError):
        return JSONResponse(
            status_code=exc.status_code,
            content={"code": exc.status_code, "message": exc.message, "success": False},
        )

    # 保存原始覆盖，仅设置需要的覆盖
    _original_overrides = app.dependency_overrides.copy()
    app.dependency_overrides[get_db] = _override_get_db
    yield TestClient(app, raise_server_exceptions=False)
    app.dependency_overrides = _original_overrides


@pytest.fixture
def admin_user():
    u = MagicMock()
    u.id = 1
    u.username = "admin"
    u.role = "admin"
    u.is_superuser = True
    u.organization_id = 1
    u.permissions_list = ["*"]
    u.email = "admin@test.com"
    return u


@pytest.fixture
def auth_setup(client, admin_user):
    """Sets up auth + data_scope overrides."""
    async def mock_get_current_user():
        return admin_user

    async def mock_get_data_scope():
        return DataScope(is_admin=True)

    client.app.dependency_overrides[get_current_user] = mock_get_current_user
    client.app.dependency_overrides[get_data_scope] = mock_get_data_scope
    return client


@pytest.fixture
def org(db_session):
    o = Organization(id=1, name="测试组织")
    db_session.add(o)
    db_session.commit()
    return o


@pytest.fixture
def school(db_session, org):
    s = School(
        name="希望小学",
        code="SCH001",
        type=SchoolType.PRIMARY,
        province="云南省",
        city="昆明市",
        district="盘龙区",
        address="XX路XX号",
        student_count=500,
        teacher_count=30,
        class_count=10,
        support_status=SupportStatus.INACTIVE,
        organization_id=org.id,
        created_by=1,
        is_active=True,
    )
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


def P(path: str) -> str:
    """Prefix path with API prefix."""
    return f"{API_PREFIX}{path}"


# ══════════════════════════════════════════════════════════════
#  Static / Options endpoints
# ══════════════════════════════════════════════════════════════


class TestStaticEndpoints:
    def test_get_type_options(self, client):
        resp = client.get(P("/schools/options/types"))
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        labels = {d["label"] for d in data}
        assert "小学" in labels

    def test_get_status_options(self, client):
        resp = client.get(P("/schools/options/statuses"))
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list)
        labels = {d["label"] for d in data}
        assert "帮扶中" in labels


# ══════════════════════════════════════════════════════════════
#  Import / Export endpoints
# ══════════════════════════════════════════════════════════════


class TestImportTemplate:
    def test_download_import_template(self, client):
        resp = client.get(P("/schools/import/template"))
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestImportExcel:
    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    def test_import_success(self, mock_validate, auth_setup, db_session):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "学校导入模板"
        ws.append(["序号", "学校名称*", "学校编码", "学校类型", "省份", "城市", "区县", "详细地址",
                    "学生人数", "教师人数", "帮扶状态", "帮扶单位", "校长", "联系电话"])
        ws.append([1, "新希望小学", "SCH002", "小学", "贵州省", "贵阳市", "云岩区",
                   "XX路", 300, 20, "未帮扶", "某某部队", "李校长", "13900000000"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P("/schools/import/excel"),
            files={"file": ("test.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["imported"] >= 1

    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    def test_import_empty_skip(self, mock_validate, auth_setup, school):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["学生姓名", "年级", "年度", "金额", "原因", "状态"])
        ws.append([None, "三年级", None, None, None, None])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students/import"),
            files={"file": ("empty.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["imported"] == 0

    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    def test_import_parse_error(self, mock_validate, auth_setup):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["序号", "学校名称*", "学校编码", "学校类型", "省份", "城市", "区县", "详细地址",
                    "学生人数", "教师人数", "帮扶状态", "帮扶单位", "校长", "联系电话"])
        ws.append([1, "TestSchool", "SCH001", "小学", "省", "市", "区", "地址",
                    "abc", None, "未帮扶", None, None, None])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P("/schools/import/excel"),
            files={"file": ("bad.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["failed"] >= 1
        assert len(data["errors"]) >= 1

    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    @patch("app.api.v1.school.os.unlink")
    def test_import_unlink_filenotfound(self, mock_unlink, mock_validate, auth_setup):
        mock_unlink.side_effect = FileNotFoundError("not found")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["x"] * 14)
        ws.append(["Test", None, None, None, None, None, None, None, None, None, None, None, None, None])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P("/schools/import/excel"),
            files={"file": ("ok.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200


class TestExport:
    def test_export_excel(self, auth_setup, school):
        resp = auth_setup.get(P("/schools/export/excel"))
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_export_excel_alt_path(self, auth_setup, school):
        resp = auth_setup.get(P("/schools/export"))
        assert resp.status_code == 200

    def test_export_empty(self, auth_setup, db_session):
        db_session.query(School).delete()
        db_session.commit()
        resp = auth_setup.get(P("/schools/export/excel"))
        assert resp.status_code == 200


# ══════════════════════════════════════════════════════════════
#  Statistics
# ══════════════════════════════════════════════════════════════


class TestStatistics:
    def test_get_statistics(self, auth_setup, school):
        resp = auth_setup.get(P("/schools/statistics"))
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_schools"] >= 1

    def test_statistics_with_projects_and_scholarships(self, auth_setup, school, db_session):
        p = SchoolProject(school_id=school.id, name="Proj1", budget=100.0)
        db_session.add(p)
        sc = ScholarshipStudent(school_id=school.id, student_name="S1", amount=500.0)
        db_session.add(sc)
        db_session.commit()
        resp = auth_setup.get(P("/schools/statistics"))
        assert resp.status_code == 200
        data = resp.json()
        assert data["project_count"] >= 1
        assert data["scholarship_count"] >= 1

    def test_statistics_no_schools(self, auth_setup, db_session):
        db_session.query(School).delete()
        db_session.commit()
        resp = auth_setup.get(P("/schools/statistics"))
        assert resp.status_code == 200


# ══════════════════════════════════════════════════════════════
#  CRUD endpoints
# ══════════════════════════════════════════════════════════════


class TestListSchools:
    def test_list_all(self, auth_setup, school):
        resp = auth_setup.get(P("/schools"))
        assert resp.status_code == 200
        data = resp.json()
        assert data["data"]["total"] >= 1
        assert len(data["data"]["items"]) >= 1

    def test_list_pagination(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?page=1&page_size=10"))
        assert resp.status_code == 200

    def test_list_filter_keyword(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?keyword=希望"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] >= 1

    def test_list_filter_name(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?name=希望"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] >= 1

    def test_list_filter_type(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?type=primary"))
        assert resp.status_code == 200

    def test_list_filter_type_chinese(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?type=小学"))
        assert resp.status_code == 200

    def test_list_filter_type_invalid(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?type=invalid_type"))
        assert resp.status_code == 200

    def test_list_filter_support_status(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?support_status=inactive"))
        assert resp.status_code == 200

    def test_list_filter_support_status_camel(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?supportStatus=inactive"))
        assert resp.status_code == 200

    def test_list_filter_support_status_invalid(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?support_status=bad"))
        assert resp.status_code == 200

    def test_list_keyword_not_found(self, auth_setup, school):
        resp = auth_setup.get(P("/schools?keyword=不存在"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] == 0


class TestGetSchool:
    def test_get_success(self, auth_setup, school):
        resp = auth_setup.get(P(f"/schools/{school.id}"))
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["name"] == "希望小学"

    def test_get_not_found(self, auth_setup):
        resp = auth_setup.get(P("/schools/99999"))
        assert resp.status_code in (404, 500)

    def test_get_permission_denied(self, client, admin_user, school):
        async def mock_get_current_user():
            u = MagicMock()
            u.id = 2
            u.role = "user"
            u.is_superuser = False
            u.organization_id = 2
            return u

        async def mock_get_data_scope():
            return DataScope(is_admin=False, org_ids=[2], self_only=True, user_id=2)

        client.app.dependency_overrides[get_current_user] = mock_get_current_user
        client.app.dependency_overrides[get_data_scope] = mock_get_data_scope
        resp = client.get(P(f"/schools/{school.id}"))
        assert resp.status_code in (403, 404)


class TestCreateSchool:
    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_create_success(self, mock_log, mock_log2, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "新建学校", "code": "NEW001", "type": "primary", "province": "广东省"},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["message"] == "创建成功"

    def test_create_duplicate_code(self, auth_setup, school):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "另一个", "code": "SCH001"},
        )
        assert resp.status_code in (409, 500)

    def test_create_empty_code_converted_to_none(self, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "无编码学校", "code": ""},
        )
        assert resp.status_code == 200

    def test_create_with_chinese_type(self, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "测试学校", "type": "高中"},
        )
        assert resp.status_code == 200

    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_create_cache_invalidation_fail(self, mock_log, mock_log2, auth_setup, db_session):
        with patch(
            "app.api.v1.data.data.dashboard.invalidate_dashboard_cache",
            side_effect=Exception("cache fail"),
        ):
            resp = auth_setup.post(
                P("/schools"),
                json={"name": "缓存失败测试", "code": "CACHEFAIL"},
            )
            assert resp.status_code == 200

    def test_create_school_type_enum_mapping(self, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "EnumTest", "code": "ENUM001", "type": "vocational", "support_status": "active"},
        )
        assert resp.status_code == 200

    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_create_with_district_detail(self, mock_log, mock_log2, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "详细学校", "code": "DETAIL001", "district": "南山区"},
        )
        assert resp.status_code == 200

    def test_create_with_school_type_schema_field(self, auth_setup, db_session):
        resp = auth_setup.post(
            P("/schools"),
            json={"name": "SchemaTypeTest", "code": "SCTYP01", "school_type": "primary", "school_level": "provincial"},
        )
        assert resp.status_code == 200


class TestUpdateSchool:
    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_update_success(self, mock_log, mock_log2, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}"),
            json={"name": "更新后的学校"},
        )
        assert resp.status_code == 200
        assert resp.json()["message"] == "更新成功"

    def test_update_not_found(self, auth_setup):
        resp = auth_setup.put(
            P("/schools/99999"),
            json={"name": "不存在的"},
        )
        assert resp.status_code in (404, 500)

    def test_update_code_empty(self, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}"),
            json={"code": ""},
        )
        assert resp.status_code == 200

    def test_update_code_conflict(self, auth_setup, school, db_session):
        s2 = School(name="另一所", code="SCH002", created_by=1)
        db_session.add(s2)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}"),
            json={"code": "SCH002"},
        )
        assert resp.status_code in (409, 500)

    def test_update_type_enum(self, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}"),
            json={"type": "高中"},
        )
        assert resp.status_code == 200

    def test_update_support_status_enum(self, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}"),
            json={"support_status": "active"},
        )
        assert resp.status_code == 200

    def test_update_permission_denied(self, client, admin_user, school):
        async def mock_get_current_user():
            u = MagicMock()
            u.id = 2
            u.role = "user"
            u.is_superuser = False
            return u

        async def mock_get_data_scope():
            return DataScope(is_admin=False, self_only=True, user_id=2)

        client.app.dependency_overrides[get_current_user] = mock_get_current_user
        client.app.dependency_overrides[get_data_scope] = mock_get_data_scope
        resp = client.put(P(f"/schools/{school.id}"), json={"name": "hack"})
        assert resp.status_code in (403, 404)

    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_update_cache_fail(self, mock_log, mock_log2, auth_setup, school):
        with patch(
            "app.api.v1.data.data.dashboard.invalidate_dashboard_cache",
            side_effect=Exception("cache fail"),
        ):
            resp = auth_setup.put(
                P(f"/schools/{school.id}"),
                json={"name": "CacheFail"},
            )
            assert resp.status_code == 200


class TestDeleteSchool:
    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_delete_success(self, mock_log, mock_log2, auth_setup, school):
        resp = auth_setup.delete(P(f"/schools/{school.id}"))
        assert resp.status_code == 200
        assert resp.json()["message"] == "删除成功"

    def test_delete_not_found(self, auth_setup):
        resp = auth_setup.delete(P("/schools/99999"))
        assert resp.status_code in (404, 500)

    def test_delete_permission_denied(self, client, admin_user, school):
        async def mock_get_current_user():
            u = MagicMock()
            u.id = 2
            u.role = "user"
            u.is_superuser = False
            return u

        async def mock_get_data_scope():
            return DataScope(is_admin=False, self_only=True, user_id=2)

        client.app.dependency_overrides[get_current_user] = mock_get_current_user
        client.app.dependency_overrides[get_data_scope] = mock_get_data_scope
        resp = client.delete(P(f"/schools/{school.id}"))
        assert resp.status_code in (403, 404)

    @patch("app.api.v1.school.write_work_log")
    @patch("app.api.v1.school.logger")
    def test_delete_cache_fail(self, mock_log, mock_log2, auth_setup, school):
        with patch(
            "app.api.v1.data.data.dashboard.invalidate_dashboard_cache",
            side_effect=Exception("cache fail"),
        ):
            resp = auth_setup.delete(P(f"/schools/{school.id}"))
            assert resp.status_code == 200


# ══════════════════════════════════════════════════════════════
#  Attachment endpoints
# ══════════════════════════════════════════════════════════════


class TestListAttachments:
    def test_list_attachments(self, auth_setup, school):
        resp = auth_setup.get(P(f"/schools/{school.id}/attachments"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] == 0


class TestUploadAttachment:
    @patch("app.api.v1.school.uuid.uuid4")
    @patch("app.api.v1.school.os.makedirs")
    @patch("app.api.v1.school.open", new_callable=MagicMock)
    def test_upload_success(self, mock_open, mock_makedirs, mock_uuid, auth_setup, school, db_session):
        mock_uuid.return_value.hex = "abcdef123456"
        resp = auth_setup.post(
            P(f"/schools/{school.id}/attachments"),
            files={"file": ("doc.pdf", b"pdf content", "application/pdf")},
            data={"description": "test doc"},
        )
        assert resp.status_code == 200
        assert resp.json()["message"] == "上传成功"

    def test_upload_school_not_found(self, auth_setup):
        resp = auth_setup.post(
            P("/schools/99999/attachments"),
            files={"file": ("doc.pdf", b"x", "application/pdf")},
        )
        assert resp.status_code in (404, 500)

    @patch("app.api.v1.school.settings.MAX_FILE_SIZE", 10)
    def test_upload_file_too_large(self, auth_setup, school):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/attachments"),
            files={"file": ("big.pdf", b"x" * 100, "application/pdf")},
        )
        assert resp.status_code in (400, 500)

    def test_upload_disallowed_type(self, auth_setup, school):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/attachments"),
            files={"file": ("evil.exe", b"x", "application/x-msdownload")},
        )
        assert resp.status_code in (400, 500)

    @patch("app.api.v1.school.uuid.uuid4")
    @patch("app.api.v1.school.os.makedirs")
    @patch("app.api.v1.school.open", new_callable=MagicMock)
    def test_upload_no_extension(self, mock_open, mock_makedirs, mock_uuid, auth_setup, school, db_session):
        mock_uuid.return_value.hex = "xyz"
        resp = auth_setup.post(
            P(f"/schools/{school.id}/attachments"),
            files={"file": ("noext", b"content", "application/octet-stream")},
        )
        assert resp.status_code == 200


class TestDownloadAttachment:
    def test_download_not_found(self, auth_setup):
        resp = auth_setup.get(P("/schools/attachments/99999/download"))
        assert resp.status_code in (404, 500)

    @patch("app.api.v1.school.settings")
    @patch("app.api.v1.school.os.path.exists", return_value=False)
    def test_download_file_not_exists(self, mock_exists, mock_settings, auth_setup, school, db_session):
        mock_settings.UPLOAD_DIR = r"C:\nonexistent"
        att = SchoolAttachment(
            school_id=school.id, file_name="missing.pdf",
            file_path=r"C:\nonexistent\file.pdf", file_size=100,
        )
        db_session.add(att)
        db_session.commit()
        resp = auth_setup.get(P(f"/schools/attachments/{att.id}/download"))
        assert resp.status_code in (404, 500)

    def test_download_invalid_path(self, auth_setup, school, db_session):
        att = SchoolAttachment(
            school_id=school.id, file_name="test.pdf",
            file_path=r"C:\invalid\path.pdf", file_size=100,
        )
        db_session.add(att)
        db_session.commit()
        with patch("app.api.v1.school._validate_file_path",
                   side_effect=AppError.forbidden("非法文件路径")):
            resp = auth_setup.get(P(f"/schools/attachments/{att.id}/download"))
            assert resp.status_code in (403, 500)

    def test_download_success_path(self, auth_setup, school, db_session):
        import tempfile
        tmp_dir = tempfile.mkdtemp()
        file_path = os.path.join(tmp_dir, "test.pdf")
        with open(file_path, "wb") as f:
            f.write(b"test")
        att = SchoolAttachment(
            school_id=school.id, file_name="test.pdf",
            file_path=file_path, file_size=4, file_type="application/pdf",
        )
        db_session.add(att)
        db_session.commit()
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = tmp_dir
            resp = auth_setup.get(P(f"/schools/attachments/{att.id}/download"))
        assert resp.status_code == 200
        try:
            os.unlink(file_path)
        except FileNotFoundError:
            pass
        os.rmdir(tmp_dir)


class TestDeleteAttachment:
    def test_delete_not_found(self, auth_setup):
        resp = auth_setup.delete(P("/schools/attachments/99999"))
        assert resp.status_code in (404, 500)

    def test_delete_success(self, auth_setup, school, db_session):
        import tempfile
        tmp_dir = tempfile.mkdtemp()
        file_path = os.path.join(tmp_dir, "test.pdf")
        with open(file_path, "wb") as f:
            f.write(b"x")
        att = SchoolAttachment(
            school_id=school.id, file_name="test.pdf",
            file_path=file_path, file_size=1,
        )
        db_session.add(att)
        db_session.commit()
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = tmp_dir
            resp = auth_setup.delete(P(f"/schools/attachments/{att.id}"))
        assert resp.status_code == 200
        assert resp.json()["message"] == "删除成功"
        os.rmdir(tmp_dir)


# ══════════════════════════════════════════════════════════════
#  Project endpoints
# ══════════════════════════════════════════════════════════════


class TestProjects:
    def test_list_projects(self, auth_setup, school):
        resp = auth_setup.get(P(f"/schools/{school.id}/projects"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] == 0

    def test_create_project(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/projects"),
            json={"name": "新建项目", "phase": "research", "budget": 50.0},
        )
        assert resp.status_code == 200
        assert resp.json()["message"] == "创建成功"

    def test_create_project_no_phase(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/projects"),
            json={"name": "无阶段项目"},
        )
        assert resp.status_code == 200

    def test_create_project_phase_none(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/projects"),
            json={"name": "PhaseNone"},
        )
        assert resp.status_code == 200

    def test_update_project(self, auth_setup, school, db_session):
        proj = SchoolProject(school_id=school.id, name="旧项目")
        db_session.add(proj)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}/projects/{proj.id}"),
            json={"name": "更新项目", "phase": "implementation"},
        )
        assert resp.status_code == 200
        assert resp.json()["message"] == "更新成功"

    def test_update_project_phase_none(self, auth_setup, school, db_session):
        proj = SchoolProject(school_id=school.id, name="PhaseUpdate")
        db_session.add(proj)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}/projects/{proj.id}"),
            json={"phase": None},
        )
        assert resp.status_code == 200

    def test_update_project_not_found(self, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}/projects/99999"),
            json={"name": "notfound"},
        )
        assert resp.status_code in (404, 500)

    def test_delete_project(self, auth_setup, school, db_session):
        proj = SchoolProject(school_id=school.id, name="删掉")
        db_session.add(proj)
        db_session.commit()
        resp = auth_setup.delete(P(f"/schools/{school.id}/projects/{proj.id}"))
        assert resp.status_code == 200

    def test_delete_project_not_found(self, auth_setup, school):
        resp = auth_setup.delete(P(f"/schools/{school.id}/projects/99999"))
        assert resp.status_code in (404, 500)

    def test_create_project_school_not_active(self, auth_setup, school, db_session):
        school.is_active = False
        db_session.commit()
        resp = auth_setup.post(
            P(f"/schools/{school.id}/projects"),
            json={"name": "非活跃学校项目"},
        )
        assert resp.status_code in (404, 500)


# ══════════════════════════════════════════════════════════════
#  Scholarship Student endpoints
# ══════════════════════════════════════════════════════════════


class TestScholarshipStudents:
    def test_list_students(self, auth_setup, school):
        resp = auth_setup.get(P(f"/schools/{school.id}/scholarship-students"))
        assert resp.status_code == 200

    def test_list_students_with_year(self, auth_setup, school, db_session):
        s = ScholarshipStudent(school_id=school.id, student_name="测试", year=2024)
        db_session.add(s)
        db_session.commit()
        resp = auth_setup.get(P(f"/schools/{school.id}/scholarship-students?year=2024"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] >= 1

    def test_list_students_year_not_found(self, auth_setup, school):
        resp = auth_setup.get(P(f"/schools/{school.id}/scholarship-students?year=1999"))
        assert resp.status_code == 200
        assert resp.json()["data"]["total"] == 0

    def test_create_student(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students"),
            json={"student_name": "张三", "amount": 1000, "year": 2024},
        )
        assert resp.status_code == 200
        assert resp.json()["message"] == "创建成功"

    def test_create_student_with_status(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students"),
            json={"student_name": "李四", "status": "approved"},
        )
        assert resp.status_code == 200

    def test_create_student_status_none(self, auth_setup, school, db_session):
        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students"),
            json={"student_name": "王五", "status": None},
        )
        assert resp.status_code == 200

    def test_update_student(self, auth_setup, school, db_session):
        stu = ScholarshipStudent(school_id=school.id, student_name="旧名", amount=500)
        db_session.add(stu)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}/scholarship-students/{stu.id}"),
            json={"student_name": "新名", "status": "disbursed"},
        )
        assert resp.status_code == 200

    def test_update_student_no_status(self, auth_setup, school, db_session):
        stu = ScholarshipStudent(school_id=school.id, student_name="NoStatus")
        db_session.add(stu)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}/scholarship-students/{stu.id}"),
            json={"student_name": "Changed"},
        )
        assert resp.status_code == 200

    def test_update_student_status_none(self, auth_setup, school, db_session):
        stu = ScholarshipStudent(school_id=school.id, student_name="StatusNone")
        db_session.add(stu)
        db_session.commit()
        resp = auth_setup.put(
            P(f"/schools/{school.id}/scholarship-students/{stu.id}"),
            json={"status": None},
        )
        assert resp.status_code == 200

    def test_update_student_not_found(self, auth_setup, school):
        resp = auth_setup.put(
            P(f"/schools/{school.id}/scholarship-students/99999"),
            json={"student_name": "nope"},
        )
        assert resp.status_code in (404, 500)

    def test_delete_student(self, auth_setup, school, db_session):
        stu = ScholarshipStudent(school_id=school.id, student_name="DelMe")
        db_session.add(stu)
        db_session.commit()
        resp = auth_setup.delete(P(f"/schools/{school.id}/scholarship-students/{stu.id}"))
        assert resp.status_code == 200

    def test_delete_student_not_found(self, auth_setup, school):
        resp = auth_setup.delete(P(f"/schools/{school.id}/scholarship-students/99999"))
        assert resp.status_code in (404, 500)


class TestImportScholarshipStudents:
    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    def test_import_success(self, mock_validate, auth_setup, school, db_session):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["学生姓名", "年级", "年度", "金额", "原因", "状态"])
        ws.append(["赵六", "三年级", 2024, 800, "贫困", "待审批"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students/import"),
            files={"file": ("stu.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["imported"] >= 1

    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    def test_import_bad_year_triggers_exception(self, mock_validate, auth_setup, school):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["学生姓名", "年级", "年度", "金额", "原因", "状态"])
        ws.append(["好学生", "一年级", "abc", 500, "贫困", "待审批"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students/import"),
            files={"file": ("bad.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["imported"] == 0
        assert len(data["errors"]) >= 1

    @patch("app.api.v1.school.validate_excel_upload", new_callable=AsyncMock)
    @patch("app.api.v1.school.os.unlink")
    def test_import_unlink_error(self, mock_unlink_inner, mock_validate, auth_setup, school):
        mock_unlink_inner.side_effect = FileNotFoundError("gone")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["x", "y"])
        ws.append(["OK", None])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mock_validate.return_value = xlsx_bytes

        resp = auth_setup.post(
            P(f"/schools/{school.id}/scholarship-students/import"),
            files={"file": ("ok.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200


# ══════════════════════════════════════════════════════════════
#  _validate_file_path helper (edge cases)
# ══════════════════════════════════════════════════════════════


class TestValidateFilePath:
    def test_relative_path_absolutized(self):
        from app.api.v1.school import _validate_file_path
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = r"C:\uploads"
            with patch("app.api.v1.school.os.path.isabs", return_value=False):
                with patch("app.api.v1.school.os.path.abspath", return_value=r"C:\uploads\test.txt"):
                    with patch("app.api.v1.school.os.path.exists", return_value=True):
                        result = _validate_file_path("test.txt")
                        assert result == r"C:\uploads\test.txt"

    def test_path_outside_allowed(self):
        from app.api.v1.school import _validate_file_path
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = r"C:\uploads"
            with patch("app.api.v1.school.os.path.isabs", return_value=True):
                with patch("app.api.v1.school.os.path.normpath", side_effect=lambda x: x):
                    with pytest.raises(AppError) as exc:
                        _validate_file_path(r"C:\Windows\system32\evil.exe")
                    assert exc.value.status_code == 403

    def test_path_not_exist(self):
        from app.api.v1.school import _validate_file_path
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = r"C:\uploads"
            with patch("app.api.v1.school.os.path.isabs", return_value=True):
                with patch("app.api.v1.school.os.path.normpath", side_effect=lambda x: x):
                    with patch("app.api.v1.school.os.path.exists", return_value=False):
                        with pytest.raises(AppError) as exc:
                            _validate_file_path(r"C:\uploads\nonexistent.pdf")
                        assert exc.value.status_code == 404

    def test_valid_path_returned(self):
        from app.api.v1.school import _validate_file_path
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = r"C:\uploads"
            with patch("app.api.v1.school.os.path.isabs", return_value=True):
                with patch("app.api.v1.school.os.path.normpath", side_effect=lambda x: x):
                    with patch("app.api.v1.school.os.path.exists", return_value=True):
                        result = _validate_file_path(r"C:\uploads\valid.pdf")
                        assert result == r"C:\uploads\valid.pdf"

    def test_path_exactly_base(self):
        from app.api.v1.school import _validate_file_path
        with patch("app.api.v1.school.settings") as mock_settings:
            mock_settings.UPLOAD_DIR = r"C:\uploads"
            with patch("app.api.v1.school.os.path.isabs", return_value=True):
                with patch("app.api.v1.school.os.path.normpath", side_effect=lambda x: x):
                    with patch("app.api.v1.school.os.path.exists", return_value=True):
                        result = _validate_file_path(r"C:\uploads")
                        assert result == r"C:\uploads"


# ══════════════════════════════════════════════════════════════
#  _get_school_and_check_permission
# ══════════════════════════════════════════════════════════════


class TestGetSchoolAndCheckPermission:
    def test_school_not_found(self, db_session):
        from app.api.v1.school import _get_school_and_check_permission
        user = MagicMock()
        user.id = 1
        with pytest.raises(AppError) as exc:
            _get_school_and_check_permission(99999, user, db_session)
        assert exc.value.status_code == 404

    @patch("app.api.v1.school.require_data_permission")
    def test_permission_denied(self, mock_req, school, db_session):
        from app.api.v1.school import _get_school_and_check_permission
        mock_req.side_effect = AppError.forbidden("无权操作")
        user = MagicMock()
        user.id = 2
        with pytest.raises(AppError) as exc:
            _get_school_and_check_permission(school.id, user, db_session)
        assert exc.value.status_code == 403

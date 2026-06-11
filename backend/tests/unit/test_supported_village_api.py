"""Comprehensive tests for supported_village.py — all endpoints, full branch coverage."""

import io
import json
from unittest.mock import ANY, AsyncMock, MagicMock, patch

import pytest
from fastapi import UploadFile
from fastapi.testclient import TestClient

from app.core.database import get_db
from app.core.security import get_current_user
from app.main import app


# Disable camel-to-snake middleware for tests (models use camelCase field names)
@pytest.fixture(autouse=True)
def _no_camel_to_snake():
    with patch("app.middleware.camel_to_snake._convert_keys",
               side_effect=lambda obj, converter: (obj, False)):
        yield


# ---------------------------------------------------------------------------
#  Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def mock_db():
    db = MagicMock()
    q = MagicMock(name="query")
    q.filter.return_value = q
    q.order_by.return_value = q
    q.offset.return_value = q
    q.limit.return_value = q
    q.all.return_value = []
    q.count.return_value = 0
    q.first.return_value = None
    q.distinct.return_value = q
    q.delete.return_value = 1
    db.query.return_value = q
    db.add.return_value = None
    db.commit.return_value = None
    db.refresh.return_value = None
    db.rollback.return_value = None
    return db


@pytest.fixture
def admin_user():
    u = MagicMock()
    u.id = 1
    u.username = "admin"
    u.role = "admin"
    u.is_superuser = True
    u.organization_id = 1
    u.is_active = True
    u.department = "测试部"
    return u


@pytest.fixture
def client(mock_db, admin_user):
    app.dependency_overrides[get_db] = lambda: mock_db
    app.dependency_overrides[get_current_user] = lambda: admin_user
    yield TestClient(app, raise_server_exceptions=False)
    app.dependency_overrides.clear()


def _make_mock_village(id_=1, village_name="测试村", department=None,
                       support_unit=None, county=None,
                       is_three_regions=False, is_key_county=False,
                       is_revitalization_tier=False, transition_fund_items=None,
                       **kwargs):
    v = MagicMock(spec=object)
    v.id = id_
    v.village_name = village_name
    v.department = department
    v.support_unit = support_unit
    v.province = kwargs.get("province", "贵州省")
    v.city = kwargs.get("city", "毕节市")
    v.county = county
    v.township = kwargs.get("township", None)
    v.region_scope = kwargs.get("region_scope", None)
    v.is_three_regions = is_three_regions
    v.is_border_area = kwargs.get("is_border_area", False)
    v.is_ethnic_area = kwargs.get("is_ethnic_area", False)
    v.is_revolutionary_area = kwargs.get("is_revolutionary_area", False)
    v.is_key_county = is_key_county
    v.is_revitalization_tier = is_revitalization_tier
    v.is_provincial_demo = kwargs.get("is_provincial_demo", False)
    v.is_hundred_village_demo = kwargs.get("is_hundred_village_demo", False)
    v.is_tiered_development = kwargs.get("is_tiered_development", False)
    v.is_cross_province = kwargs.get("is_cross_province", False)
    v.is_cross_city = kwargs.get("is_cross_city", False)
    v.is_cross_unit_cooperation = kwargs.get("is_cross_unit_cooperation", False)
    v.is_in_overall_plan = kwargs.get("is_in_overall_plan", False)
    v.honors = kwargs.get("honors", None)
    v.sequence_no = kwargs.get("sequence_no", None)
    v.latitude = kwargs.get("latitude", None)
    v.longitude = kwargs.get("longitude", None)
    v.altitude = kwargs.get("altitude", None)
    v.area_sq_km = kwargs.get("area_sq_km", None)
    v.households = kwargs.get("households", None)
    v.description = kwargs.get("description", None)
    v.transition_fund_military_total = kwargs.get("transition_fund_military_total", 0)
    v.transition_fund_local_total = kwargs.get("transition_fund_local_total", 0)
    v.transition_fund_items = transition_fund_items
    return v


def _mock_to_dict(village):
    d = {
        "id": village.id,
        "villageName": village.village_name,
        "village_name": village.village_name,
        "department": village.department,
        "supportUnit": village.support_unit,
        "support_unit": village.support_unit,
        "county": village.county,
        "province": village.province,
        "city": village.city,
        "is_three_regions": village.is_three_regions,
        "is_key_county": village.is_key_county,
    }
    return d


# ---------------------------------------------------------------------------
#  GET /supported-villages
# ---------------------------------------------------------------------------

class TestListVillages:
    def test_default(self, client, mock_db):
        q = mock_db.query.return_value
        q.count.return_value = 1
        v = _make_mock_village(1)
        q.all.return_value = [v]
        resp = client.get("/api/v1/supported-villages")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 1

    def test_with_keyword(self, client, mock_db):
        q = mock_db.query.return_value
        q.count.return_value = 1
        v = _make_mock_village(1)
        q.all.return_value = [v]
        resp = client.get("/api/v1/supported-villages?keyword=测试")
        assert resp.status_code == 200

    def test_with_filters(self, client, mock_db):
        q = mock_db.query.return_value
        q.count.return_value = 1
        v = _make_mock_village(1)
        q.all.return_value = [v]
        resp = client.get("/api/v1/supported-villages?department=某部&county=赫章县&isRevitalizationTier=true&isThreeRegions=1")
        assert resp.status_code == 200

    def test_with_to_dict(self, client, mock_db):
        q = mock_db.query.return_value
        q.count.return_value = 1
        v = _make_mock_village(1)
        v.to_dict = lambda: _mock_to_dict(v)
        q.all.return_value = [v]
        resp = client.get("/api/v1/supported-villages")
        assert resp.status_code == 200
        assert resp.json()["items"][0]["villageName"] == "测试村"

    def test_empty_items(self, client, mock_db):
        q = mock_db.query.return_value
        q.count.return_value = 0
        q.all.return_value = []
        resp = client.get("/api/v1/supported-villages")
        assert resp.status_code == 200
        assert resp.json()["total"] == 0


# ---------------------------------------------------------------------------
#  GET /supported-villages/filter-options
# ---------------------------------------------------------------------------

class TestGetFilterOptions:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        q.all.return_value = [("某部",), ("某厅",)]
        resp = client.get("/api/v1/supported-villages/filter-options")
        assert resp.status_code == 200
        data = resp.json()
        assert "departments" in data
        assert "counties" in data

    def test_empty(self, client, mock_db):
        q = mock_db.query.return_value
        q.all.return_value = []
        resp = client.get("/api/v1/supported-villages/filter-options")
        assert resp.status_code == 200

    def test_with_none(self, client, mock_db):
        q = mock_db.query.return_value
        q.all.return_value = [(None,)]
        resp = client.get("/api/v1/supported-villages/filter-options")
        assert resp.status_code == 200
        assert resp.json()["departments"] == []


# ---------------------------------------------------------------------------
#  GET /supported-villages/import-template
# ---------------------------------------------------------------------------

class TestDownloadImportTemplate:
    def test_success(self, client):
        resp = client.get("/api/v1/supported-villages/import-template")
        assert resp.status_code == 200
        assert "application/vnd.openxmlformats" in resp.headers.get("content-type", "")


# ---------------------------------------------------------------------------
#  GET /supported-villages/export
# ---------------------------------------------------------------------------

class TestExportVillages:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1, is_three_regions=True, is_key_county=False)
        q.all.return_value = [v]
        resp = client.get("/api/v1/supported-villages/export")
        assert resp.status_code == 200
        assert "application/vnd.openxmlformats" in resp.headers.get("content-type", "")

    def test_empty(self, client, mock_db):
        q = mock_db.query.return_value
        q.all.return_value = []
        resp = client.get("/api/v1/supported-villages/export")
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
#  POST /supported-villages/import
# ---------------------------------------------------------------------------

class TestImportVillages:
    def test_wrong_extension(self, client):
        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.txt", b"data", "text/plain")
        })
        assert resp.status_code == 400

    def test_invalid_excel(self, client):
        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", b"not an excel file", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 400

    def test_no_data_rows(self, client):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 400

    def test_successful_import(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称", "部门单位", "帮扶单位", "省", "市", "县/市"])
        ws.append(["新帮扶村", "某部门", "某单位", "贵州省", "毕节市", "赫章县"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        q.first.return_value = None

        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["imported"] == 1

    def test_duplicate_village(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称", "部门单位", "帮扶单位", "省", "市", "县/市"])
        ws.append(["已有村", "某部门", "某单位", "贵州省", "毕节市", "赫章县"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        q.first.return_value = _make_mock_village(1, village_name="已有村")

        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert "已存在" in resp.json()["errors"][0]

    def test_empty_row_skipped(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称", "部门单位"])
        ws.append([None, None])
        ws.append(["有效村", "某部"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        q.first.return_value = None

        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert resp.json()["imported"] == 1

    def test_import_empty_village_name(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称", "部门单位"])
        ws.append(["", "某部"])
        ws.append(["", "某厅"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert resp.json()["imported"] == 0

    def test_import_exception_during_processing(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["帮扶村名称", "部门单位"])
        ws.append(["会崩溃的村", "某部"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        q.first.side_effect = RuntimeError("模拟异常")

        resp = client.post("/api/v1/supported-villages/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert len(resp.json()["errors"]) >= 1


# ---------------------------------------------------------------------------
#  POST /supported-villages/batch-delete
# ---------------------------------------------------------------------------

class TestBatchDelete:
    def test_empty_ids(self, client):
        resp = client.post("/api/v1/supported-villages/batch-delete", json={"ids": []})
        assert resp.status_code == 400

    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        q.delete.return_value = 2
        resp = client.post("/api/v1/supported-villages/batch-delete", json={"ids": [1, 2]})
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
#  GET /supported-villages/{village_id}
# ---------------------------------------------------------------------------

class TestGetVillage:
    def test_found(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        resp = client.get("/api/v1/supported-villages/1")
        assert resp.status_code == 200
        assert resp.json()["data"]["id"] == 1

    def test_found_with_to_dict(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        v.to_dict = lambda: _mock_to_dict(v)
        q.first.return_value = v
        resp = client.get("/api/v1/supported-villages/1")
        assert resp.status_code == 200
        assert resp.json()["data"]["villageName"] == "测试村"

    def test_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.get("/api/v1/supported-villages/999")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  POST /supported-villages  (create)
# ---------------------------------------------------------------------------

class TestCreateVillage:
    def test_success(self, client, mock_db):
        resp = client.post("/api/v1/supported-villages", json={
            "village_name": "新帮扶村",
        })
        assert resp.status_code == 200
        assert resp.json()["message"] == "创建成功"


# ---------------------------------------------------------------------------
#  PUT /supported-villages/{village_id}
# ---------------------------------------------------------------------------

class TestUpdateVillage:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        resp = client.put("/api/v1/supported-villages/1", json={"village_name": "新名称"})
        assert resp.status_code == 200

    def test_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.put("/api/v1/supported-villages/999", json={"village_name": "x"})
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  DELETE /supported-villages/{village_id}
# ---------------------------------------------------------------------------

class TestDeleteVillage:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        resp = client.delete("/api/v1/supported-villages/1")
        assert resp.status_code == 200

    def test_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.delete("/api/v1/supported-villages/999")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  GET /supported-villages/{village_id}/yearly/{year}
# ---------------------------------------------------------------------------

class TestGetYearlyData:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.side_effect = [v] + [None] * 11

        resp = client.get("/api/v1/supported-villages/1/yearly/2025")
        assert resp.status_code == 200
        assert resp.json()["message"] == "ok"




# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/yearly/copy
# ---------------------------------------------------------------------------

class TestCopyYearData:
    def test_same_year(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        resp = client.post("/api/v1/supported-villages/1/yearly/copy", json={
            "fromYear": 2025, "toYear": 2025
        })
        assert resp.status_code == 400

    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        src = MagicMock()
        _SECTION_MODEL_COUNT = 10
        q.first.side_effect = [v] + [src, None] * _SECTION_MODEL_COUNT

        resp = client.post("/api/v1/supported-villages/1/yearly/copy", json={
            "fromYear": 2024, "toYear": 2025
        })
        assert resp.status_code == 200

    def test_copy_src_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.side_effect = [v] + [None, None] * 10

        resp = client.post("/api/v1/supported-villages/1/yearly/copy", json={
            "fromYear": 2024, "toYear": 2025
        })
        assert resp.status_code == 200

    def test_copy_target_exists(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        src = MagicMock()
        existing = MagicMock()
        q.first.side_effect = [v] + [src, existing] * 10

        resp = client.post("/api/v1/supported-villages/1/yearly/copy", json={
            "fromYear": 2024, "toYear": 2025
        })
        assert resp.status_code == 200


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/yearly/{year}/{section}
# ---------------------------------------------------------------------------

class TestSaveYearlySection:
    def test_invalid_section(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        resp = client.post("/api/v1/supported-villages/1/yearly/2025/invalid_section", json={})
        assert resp.status_code == 400

    def test_success(self, client, mock_db):
        from app.models.supported_village import SupportedVillage
        village_q = MagicMock()
        village_q.first.return_value = _make_mock_village(1)
        section_q = MagicMock()
        section_q.filter.return_value = section_q
        section_q.first.return_value = MagicMock()
        def _q_side_effect(model):
            return village_q if model is SupportedVillage else section_q
        mock_db.query.side_effect = _q_side_effect

        resp = client.post("/api/v1/supported-villages/1/yearly/2025/population", json={"total_households": 200})
        assert resp.status_code == 200

    def test_save_section_create_new(self, client, mock_db):
        from app.models.supported_village import SupportedVillage
        village_q = MagicMock()
        village_q.first.return_value = _make_mock_village(1)
        section_q = MagicMock()
        section_q.filter.return_value = section_q
        section_q.first.return_value = None
        def _q_side_effect(model):
            return village_q if model is SupportedVillage else section_q
        mock_db.query.side_effect = _q_side_effect

        resp = client.post("/api/v1/supported-villages/1/yearly/2025/population", json={"total_households": 100})
        assert resp.status_code == 200
        mock_db.add.assert_called_once()


# ---------------------------------------------------------------------------
#  GET /supported-villages/{village_id}/sections/{section}/attachments
# ---------------------------------------------------------------------------

class TestGetSectionAttachments:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        att_q = MagicMock()
        att_q.filter.return_value = att_q
        att_q.order_by.return_value = att_q
        attachment = MagicMock()
        attachment.id = 1
        attachment.file_name = "doc.pdf"
        attachment.file_size = 1024
        attachment.mime_type = "application/pdf"
        attachment.created_at = None
        att_q.all.return_value = [attachment]
        mock_db.query.return_value = att_q

        resp = client.get("/api/v1/supported-villages/1/sections/population/attachments")
        assert resp.status_code == 200

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.get("/api/v1/supported-villages/999/sections/population/attachments")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/sections/{section}/attachments
# ---------------------------------------------------------------------------

class TestUploadSectionAttachment:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        att_q = MagicMock()
        mock_db.query.return_value = att_q

        resp = client.post("/api/v1/supported-villages/1/sections/population/attachments", files={
            "file": ("test.pdf", b"pdf content", "application/pdf"),
        })
        assert resp.status_code == 200

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.post("/api/v1/supported-villages/999/sections/population/attachments", files={
            "file": ("test.pdf", b"pdf content", "application/pdf"),
        })
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  DELETE /supported-villages/{village_id}/sections/{section}/attachments/{attachment_id}
# ---------------------------------------------------------------------------

class TestDeleteSectionAttachment:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        att_q = MagicMock()
        att_q.filter.return_value = att_q
        attachment = MagicMock()
        att_q.first.return_value = attachment
        mock_db.query.return_value = att_q

        resp = client.delete("/api/v1/supported-villages/1/sections/population/attachments/1")
        assert resp.status_code == 200

    def test_attachment_not_found(self, client, mock_db):
        from app.models.supported_village import SupportedVillage, VillageAttachment
        village_q = MagicMock()
        village_q.first.return_value = _make_mock_village(1)
        att_q = MagicMock()
        att_q.filter.return_value = att_q
        att_q.first.return_value = None
        def _query_side_effect(model):
            return village_q if model is SupportedVillage else att_q
        mock_db.query.side_effect = _query_side_effect

        resp = client.delete("/api/v1/supported-villages/1/sections/population/attachments/999")
        assert resp.status_code == 404

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.delete("/api/v1/supported-villages/999/sections/population/attachments/1")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/committee
# ---------------------------------------------------------------------------

class TestSaveCommittee:
    def test_create_new(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.side_effect = [v, None]

        resp = client.post("/api/v1/supported-villages/1/committee", json={
            "secretary_name": "张三",
            "secretary_phone": "13800000000",
        })
        assert resp.status_code == 200
        mock_db.add.assert_called_once()

    def test_update_existing(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v
        committee = MagicMock()
        committee_q = MagicMock()
        committee_q.first.return_value = committee
        mock_db.query.return_value = committee_q

        resp = client.post("/api/v1/supported-villages/1/committee", json={
            "secretary_name": "李四",
        })
        assert resp.status_code == 200
        mock_db.add.assert_not_called()

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.post("/api/v1/supported-villages/999/committee", json={})
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/sections/import
# ---------------------------------------------------------------------------

class TestImportSectionData:
    def test_success(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["年份", "项目", "数值"])
        ws.append([2025, "收入", 100])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v

        resp = client.post("/api/v1/supported-villages/1/sections/import", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["rows"] == 1

    def test_parse_error(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v

        resp = client.post("/api/v1/supported-villages/1/sections/import", files={
            "file": ("test.xlsx", b"garbage", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/sections/import-all
# ---------------------------------------------------------------------------

class TestImportAllSectionsData:
    def test_success(self, client, mock_db):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "population"
        ws.append(["year", "value"])
        ws2 = wb.create_sheet("income")
        ws2.append(["year", "value"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v

        resp = client.post("/api/v1/supported-villages/1/sections/import-all", files={
            "file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 200
        assert resp.json()["data"]["sheets"] == 2

    def test_parse_error(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v

        resp = client.post("/api/v1/supported-villages/1/sections/import-all", files={
            "file": ("bad.xlsx", b"garbage", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        })
        assert resp.status_code == 400


# ---------------------------------------------------------------------------
#  GET /supported-villages/{village_id}/transition-funding
# ---------------------------------------------------------------------------

class TestGetTransitionFunding:
    def test_success_with_items(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1, transition_fund_items=json.dumps([
            {"year": 2025, "militaryInvestment": 100, "localInvestment": 50}
        ]))
        q.first.return_value = v

        resp = client.get("/api/v1/supported-villages/1/transition-funding")
        assert resp.status_code == 200
        assert resp.json()["success"] is True
        assert len(resp.json()["data"]) == 1

    def test_invalid_json(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1, transition_fund_items="not json")
        q.first.return_value = v

        resp = client.get("/api/v1/supported-villages/1/transition-funding")
        assert resp.status_code == 200
        assert resp.json()["data"] == []

    def test_no_items(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1, transition_fund_items=None)
        q.first.return_value = v

        resp = client.get("/api/v1/supported-villages/1/transition-funding")
        assert resp.status_code == 200
        assert resp.json()["data"] == []

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.get("/api/v1/supported-villages/999/transition-funding")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  POST /supported-villages/{village_id}/transition-funding
# ---------------------------------------------------------------------------

class TestSaveTransitionFunding:
    def test_success(self, client, mock_db):
        q = mock_db.query.return_value
        v = _make_mock_village(1)
        q.first.return_value = v

        resp = client.post("/api/v1/supported-villages/1/transition-funding", json={
            "items": [
                {"year": 2025, "militaryInvestment": 100, "localInvestment": 50, "totalInvestment": 150}
            ]
        })
        assert resp.status_code == 200
        assert resp.json()["success"] is True
        village = q.first()
        assert village.transition_fund_military_total == 100
        assert village.transition_fund_local_total == 50
        assert v.transition_fund_military_total == 100
        assert v.transition_fund_local_total == 50

    def test_village_not_found(self, client, mock_db):
        q = mock_db.query.return_value
        q.first.return_value = None
        resp = client.post("/api/v1/supported-villages/999/transition-funding", json={
            "items": [{"year": 2025, "militaryInvestment": 0, "localInvestment": 0, "totalInvestment": 0}]
        })
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
#  GET /supported-villages/templates/all
# ---------------------------------------------------------------------------

class TestDownloadAllTemplates:
    def test_success(self, client):
        resp = client.get("/api/v1/supported-villages/templates/all")
        assert resp.status_code == 200
        assert "application/vnd.openxmlformats" in resp.headers.get("content-type", "")

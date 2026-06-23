"""Tests for app/services/supported_village_export_service.py — 目标 100% 覆盖。

覆盖要点：
- _coerce_cell: None/bool/int/float/str/datetime/Decimal/Enum/其它对象 全分支
- _query_villages: 各筛选条件 + tiered_level→is_revitalization_tier 映射
- _collect_module_data: 空村列表 / 未知模块 / 正常批量查询 / 年份过滤
- _collect_export_data: 默认模块 / 自定义模块
- _generate_statistics: 空数据 / 有数据
- _build_excel: 空数据 / 有数据 / 统计 sheet
- _build_csv: 空数据 / 有数据
- export: xlsx / csv / 带筛选条件
"""
import io
from datetime import datetime
from decimal import Decimal
from enum import Enum

import pytest
from openpyxl import load_workbook

from app.services.supported_village_export_service import (
    MODULE_NAMES,
    ExportFormat,
    ExportModule,
    SupportedVillageExportService,
)


# ---------------------------------------------------------------------------
# 枚举与常量
# ---------------------------------------------------------------------------


class TestEnums:
    def test_export_format_values(self):
        assert ExportFormat.XLSX.value == "xlsx"
        assert ExportFormat.CSV.value == "csv"

    def test_export_module_count(self):
        # 12 个业务模块
        assert len(list(ExportModule)) == 12

    def test_module_names_cover_all_modules(self):
        for m in ExportModule:
            assert m.value in MODULE_NAMES


# ---------------------------------------------------------------------------
# _coerce_cell —— 纯函数，逐类型覆盖
# ---------------------------------------------------------------------------


class TestCoerceCell:
    @pytest.mark.parametrize("value,expected", [
        (None, None),
        (True, True),
        (False, False),
        (42, 42),
        (3.14, 3.14),
        ("文本", "文本"),
    ])
    def test_primitive_types(self, value, expected):
        assert SupportedVillageExportService._coerce_cell(value) == expected

    def test_datetime_passthrough(self):
        dt = datetime(2025, 1, 1, 12, 0, 0)
        assert SupportedVillageExportService._coerce_cell(dt) == dt

    def test_date_passthrough(self):
        import datetime as _dt
        d = _dt.date(2025, 1, 1)
        assert SupportedVillageExportService._coerce_cell(d) == d

    def test_decimal_to_float(self):
        assert SupportedVillageExportService._coerce_cell(Decimal("1.5")) == 1.5

    def test_enum_to_value(self):
        class Color(Enum):
            RED = "red"
        # 纯 Enum（非 str 子类）才会走到 enum 分支
        assert SupportedVillageExportService._coerce_cell(Color.RED) == "red"

    def test_unknown_object_falls_back_to_str(self):
        class Obj:
            def __str__(self):
                return "obj-repr"
        assert SupportedVillageExportService._coerce_cell(Obj()) == "obj-repr"

    def test_list_falls_back_to_str(self):
        # list 不在 openpyxl 支持类型里，应转字符串兜底
        assert SupportedVillageExportService._coerce_cell([1, 2]) == "[1, 2]"


# ---------------------------------------------------------------------------
# 真实内存数据库 fixture
# 注：app.models.__init__ 采用懒加载，Base.metadata 只含已导入的模型表。
# 直接 create_all 会因外键目标表未注册而 NoReferencedTableError，
# 故先强制导入 _MODULE_MAP 中全部子模块以确保 metadata 完整。
# ---------------------------------------------------------------------------


def _build_session():
    import importlib
    from app.models import Base, _MODULE_MAP
    for mod_path in set(_MODULE_MAP.values()):
        importlib.import_module(f"app.models{mod_path}")
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = Session()
    return db, engine


@pytest.fixture
def export_service():
    """构建内存数据库 + 填充测试数据，返回 service。

    需要直接操作 db 的测试可通过 ``export_service.db`` 访问会话。
    """
    db, engine = _build_session()
    from app.models.supported_village import (
        SupportedVillage, VillagePopulation, VillageIncome,
    )
    v1 = SupportedVillage(id=1, village_name="示范村A", county="甲县",
                          department="作战处", support_unit="某旅",
                          is_revitalization_tier=True)
    v2 = SupportedVillage(id=2, village_name="基础村B", county="乙县",
                          department="后勤处", support_unit="某团",
                          is_revitalization_tier=False)
    v3 = SupportedVillage(id=3, village_name="达标村C", county="甲县",
                          department="作战处", support_unit="某旅",
                          is_revitalization_tier=True)
    db.add_all([v1, v2, v3])
    db.flush()
    db.add_all([
        VillagePopulation(supported_village_id=1, year=2025,
                          total_households=100, total_population=300, labor_force=200),
        VillagePopulation(supported_village_id=2, year=2025,
                          total_households=50, total_population=120, labor_force=60),
    ])
    db.add(VillageIncome(supported_village_id=1, year=2025,
                         collective_income=500000, per_capita_income=8000))
    db.commit()
    svc = SupportedVillageExportService(db)
    yield svc
    db.close()
    engine.dispose()


# ---------------------------------------------------------------------------
# _query_villages
# ---------------------------------------------------------------------------


class TestQueryVillages:
    def test_no_filter_returns_all(self, export_service):
        villages = export_service._query_villages()
        assert len(villages) == 3
        assert [v.id for v in villages] == [1, 2, 3]

    def test_filter_by_village_ids(self, export_service):
        villages = export_service._query_villages(village_ids=[1, 3])
        assert [v.id for v in villages] == [1, 3]

    def test_filter_by_department(self, export_service):
        villages = export_service._query_villages(department="作战处")
        assert {v.id for v in villages} == {1, 3}

    def test_filter_by_support_unit(self, export_service):
        villages = export_service._query_villages(support_unit="某团")
        assert [v.id for v in villages] == [2]

    def test_tiered_level_demo_maps_to_true(self, export_service):
        # 示范级 → is_revitalization_tier=True
        villages = export_service._query_villages(tiered_level="示范级")
        assert {v.id for v in villages} == {1, 3}

    def test_tiered_level_standard_maps_to_true(self, export_service):
        # 达标级 → is_revitalization_tier=True
        villages = export_service._query_villages(tiered_level="达标级")
        assert {v.id for v in villages} == {1, 3}

    def test_tiered_level_basic_maps_to_false(self, export_service):
        # 基础级 → is_revitalization_tier=False
        villages = export_service._query_villages(tiered_level="基础级")
        assert [v.id for v in villages] == [2]

    def test_tiered_level_unknown_maps_to_false(self, export_service):
        # 未知梯次字符串 → is_tier=False → 筛选 is_revitalization_tier=False 的村
        villages = export_service._query_villages(tiered_level="未知")
        assert [v.id for v in villages] == [2]


# ---------------------------------------------------------------------------
# _collect_module_data
# ---------------------------------------------------------------------------


class TestCollectModuleData:
    def test_empty_villages_returns_empty(self, export_service):
        assert export_service._collect_module_data([], "population") == []

    def test_unknown_module_returns_basic_info(self, export_service):
        villages = export_service._query_villages(village_ids=[1])
        rows = export_service._collect_module_data(villages, "nonexistent_module")
        assert len(rows) == 1
        assert rows[0]["village_name"] == "示范村A"
        assert rows[0]["module"] == "nonexistent_module"

    def test_known_module_population(self, export_service):
        villages = export_service._query_villages()
        rows = export_service._collect_module_data(villages, "population", year=2025)
        # v3 没有人口数据，应返回 None 字段
        by_id = {r["id"]: r for r in rows}
        assert by_id[1]["households"] == 100
        assert by_id[1]["total_population"] == 300
        assert by_id[3]["households"] is None  # 无关联记录

    def test_year_filter_excludes_other_years(self, export_service):
        from app.models.supported_village import VillagePopulation
        # 给 v1 插一条 2024 年数据
        export_service.db.add(VillagePopulation(supported_village_id=1, year=2024,
                                           total_households=99, total_population=99, labor_force=99))
        export_service.db.commit()
        villages = export_service._query_villages(village_ids=[1])
        rows_2025 = export_service._collect_module_data(villages, "population", year=2025)
        rows_2024 = export_service._collect_module_data(villages, "population", year=2024)
        assert rows_2025[0]["households"] == 100
        assert rows_2024[0]["households"] == 99

    def test_missing_model_falls_back_to_basic(self, export_service, monkeypatch):
        """模型不存在于模块映射时降级为基本信息。"""
        from app.models import supported_village as sv_models
        villages = export_service._query_villages(village_ids=[1])
        # 让 getattr 返回 None 模拟模型缺失
        monkeypatch.setattr(sv_models, "VillagePopulation", None, raising=False)
        rows = export_service._collect_module_data(villages, "population", year=2025)
        assert len(rows) == 1
        assert rows[0]["village_name"] == "示范村A"
        # 降级路径不包含 field_map 字段
        assert "households" not in rows[0]

    def test_missing_fk_col_falls_back_to_basic(self, export_service, monkeypatch):
        """模型缺少 supported_village_id 外键时降级。"""
        villages = export_service._query_villages(village_ids=[1])

        class NoFkModel:
            year = None
            # 故意不声明 supported_village_id

        from app.models import supported_village as sv_models
        monkeypatch.setattr(sv_models, "VillagePopulation", NoFkModel, raising=False)
        rows = export_service._collect_module_data(villages, "population", year=2025)
        assert len(rows) == 1
        assert "households" not in rows[0]


# ---------------------------------------------------------------------------
# _collect_export_data
# ---------------------------------------------------------------------------


class TestCollectExportData:
    def test_default_modules_all(self, export_service):
        villages = export_service._query_villages()
        data = export_service._collect_export_data(villages)
        assert set(data.keys()) == set(MODULE_NAMES.keys())

    def test_custom_modules_subset(self, export_service):
        villages = export_service._query_villages()
        data = export_service._collect_export_data(villages, modules=["population", "income"])
        assert set(data.keys()) == {"population", "income"}


# ---------------------------------------------------------------------------
# _generate_statistics
# ---------------------------------------------------------------------------


class TestGenerateStatistics:
    def test_empty_data(self, export_service):
        stats = export_service._generate_statistics({})
        assert stats["total_villages"] == 0
        assert stats["modules_exported"] == []

    def test_with_data(self, export_service):
        data = {"population": [{"id": 1}, {"id": 2}], "income": [{"id": 1}]}
        stats = export_service._generate_statistics(data)
        assert stats["total_villages"] == 2  # 取各模块最大值
        assert stats["population_count"] == 2
        assert stats["income_count"] == 1
        assert set(stats["modules_exported"]) == {"population", "income"}
        assert "generated_at" in stats


# ---------------------------------------------------------------------------
# _build_excel
# ---------------------------------------------------------------------------


class TestBuildExcel:
    def test_empty_data_only_stats_sheet(self, export_service):
        content = export_service._build_excel({}, {"total_villages": 0, "modules_exported": []})
        wb = load_workbook(io.BytesIO(content))
        assert "统计信息" in wb.sheetnames
        # 无数据模块 → 不创建对应 sheet
        assert len(wb.sheetnames) == 1

    def test_with_data_creates_module_sheets(self, export_service):
        villages = export_service._query_villages()
        data = export_service._collect_export_data(villages, modules=["population"])
        stats = export_service._generate_statistics(data)
        content = export_service._build_excel(data, stats)
        wb = load_workbook(io.BytesIO(content))
        assert "人口数据" in wb.sheetnames
        assert "统计信息" in wb.sheetnames
        ws = wb["人口数据"]
        # 标题行 + 数据行
        assert ws.max_row >= 2
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "village_name" in headers


# ---------------------------------------------------------------------------
# _build_csv
# ---------------------------------------------------------------------------


class TestBuildCsv:
    def test_empty_data(self, export_service):
        # 无数据 → 仅含 UTF-8 BOM（utf-8-sig 编码固定加 BOM）
        content = export_service._build_csv({})
        assert content == b"\xef\xbb\xbf"

    def test_with_data(self, export_service):
        villages = export_service._query_villages()
        data = export_service._collect_export_data(villages, modules=["population"])
        content = export_service._build_csv(data)
        text = content.decode("utf-8-sig")
        # CSV 应含标题行
        assert "village_name" in text
        assert "示范村A" in text

    def test_csv_only_first_module(self, export_service):
        """CSV 简化导出仅含第一个有数据的模块。"""
        villages = export_service._query_villages()
        data = export_service._collect_export_data(villages, modules=["population", "income"])
        content = export_service._build_csv(data)
        text = content.decode("utf-8-sig")
        # 只有一个模块的标题行（break 在第一个非空模块后）
        assert text.count("village_name") == 1

    def test_csv_skips_empty_leading_module(self, export_service):
        """首个模块无数据时跳过，导出下一个有数据的模块（覆盖 continue 分支）。"""
        villages = export_service._query_villages(village_ids=[3])  # v3 无人口/收入数据
        # committee 模块无关联记录 → 空列表；population 也空
        data = {"population": [], "income": [{"id": 3, "village_name": "达标村C", "collective_income": None, "per_capita_income": None}]}
        content = export_service._build_csv(data)
        text = content.decode("utf-8-sig")
        # 跳过空 population，导出 income
        assert "collective_income" in text
        assert "达标村C" in text


# ---------------------------------------------------------------------------
# export —— 集成入口
# ---------------------------------------------------------------------------


class TestExport:
    def test_export_xlsx_default(self, export_service):
        content, filename, stats = export_service.export()
        assert filename.endswith(".xlsx")
        assert b"PK" in content[:4]  # xlsx 是 zip，魔数 PK
        assert stats["total_villages"] == 3

    def test_export_csv(self, export_service):
        content, filename, stats = export_service.export(format="csv")
        assert filename.endswith(".csv")
        assert "示范村A" in content.decode("utf-8-sig")

    def test_export_with_filters(self, export_service):
        content, filename, stats = export_service.export(
            village_ids=[1], modules=["population"], format="csv"
        )
        text = content.decode("utf-8-sig")
        assert "示范村A" in text
        assert "基础村B" not in text  # 被 village_ids 过滤掉

    def test_export_empty_result(self, export_service):
        """筛选无匹配村时仍正常返回（空数据）。"""
        content, filename, stats = export_service.export(village_ids=[9999])
        assert filename.endswith(".xlsx")
        assert stats["total_villages"] == 0

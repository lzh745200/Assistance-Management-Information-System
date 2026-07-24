"""app.api.v1.village_templates 覆盖率攻坚测试

覆盖点：
- list_templates 列表端点
- download_template：village 专用生成、11 个通用模块真实生成、
  include_example/year 参数组合、404 未知模块
- _generate_module_template 的字段兜底分支（monkeypatch 注入未定义字段的模块）
"""

from unittest.mock import MagicMock, patch
from urllib.parse import quote

import pytest
from fastapi.testclient import TestClient

import app.api.v1.village_templates as vt
from app.core.security import get_current_user


@pytest.fixture
def tpl_client():
    from app.main import app

    original = app.dependency_overrides.copy()
    app.dependency_overrides[get_current_user] = lambda: MagicMock(id=1)
    yield TestClient(app, raise_server_exceptions=False)
    app.dependency_overrides = original


# ==================== GET /templates ====================


class TestListTemplates:
    def test_list_all(self, tpl_client):
        resp = tpl_client.get("/api/v1/templates")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 12
        modules = {t["module"] for t in data["templates"]}
        assert modules == set(vt.AVAILABLE_MODULES.keys())
        village = next(t for t in data["templates"] if t["module"] == "village")
        assert village["name"] == "帮扶村基础数据"


# ==================== GET /templates/{module} ====================


class TestDownloadTemplate:
    def test_village_uses_dedicated_generator(self, tpl_client):
        svc = MagicMock()
        svc.generate_village_template.return_value = b"excel-bytes"
        with patch.object(vt, "template_service", svc):
            resp = tpl_client.get("/api/v1/templates/village")
        assert resp.status_code == 200
        assert resp.content == b"excel-bytes"
        disposition = resp.headers["content-disposition"]
        assert "attachment" in disposition
        assert quote("帮扶村数据导入模板") in disposition
        svc.generate_village_template.assert_called_once_with(include_example=True)

    def test_unknown_module_404(self, tpl_client):
        resp = tpl_client.get("/api/v1/templates/nonexistent")
        assert resp.status_code == 404
        assert "nonexistent" in resp.json()["detail"]

    def test_generic_module_with_year(self, tpl_client):
        resp = tpl_client.get("/api/v1/templates/population?year=2026")
        assert resp.status_code == 200
        assert resp.content[:2] == b"PK"  # xlsx 魔数
        disposition = resp.headers["content-disposition"]
        assert "_2026_" in disposition
        assert quote("人口数据导入模板") in disposition

    def test_generic_module_without_example(self, tpl_client):
        resp = tpl_client.get("/api/v1/templates/party?include_example=false")
        assert resp.status_code == 200
        assert resp.content[:2] == b"PK"

    def test_all_generic_modules_generate(self, tpl_client):
        """11 个通用模块逐一真实生成（覆盖 MODULE_FIELDS 全量字段定义）"""
        for module in vt.AVAILABLE_MODULES:
            if module == "village":
                continue
            resp = tpl_client.get(f"/api/v1/templates/{module}")
            assert resp.status_code == 200, module
            assert resp.content[:2] == b"PK", module


# ==================== _generate_module_template 兜底分支 ====================


class TestGenerateModuleTemplateFallback:
    def test_fallback_fields_for_undefined_module(self, monkeypatch):
        """模块在 AVAILABLE_MODULES 中但无字段定义 → 使用基础两列模板"""
        monkeypatch.setitem(
            vt.AVAILABLE_MODULES,
            "mystery",
            {"name": "神秘模块", "description": "d", "filename": "神秘模板"},
        )
        content = vt._generate_module_template("mystery", include_example=True, year=None)
        assert content[:2] == b"PK"

        # 校验工作簿内容：兜底模板只有 village_name/year 两列
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(content))
        ws = wb["神秘模块"]
        assert ws.cell(row=1, column=1).value == "*帮扶村名称"
        assert ws.cell(row=1, column=2).value == "*年份"
        assert ws.cell(row=1, column=3).value is None
        assert ws.cell(row=2, column=2).value == "2024"  # year=None → 缺省 2024
